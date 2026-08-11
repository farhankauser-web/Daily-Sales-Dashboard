"""
apps/dashboard/views_marketing.py — Marketing Optimizer (read-only).

Two tabs — Search Terms (per campaign) and Budget Pacing — plus an Excel export.
Nothing here writes to Amazon; every "action" is a recommendation for a human.
"""
import re
from collections import defaultdict
from datetime import date, timedelta

from django.db.models import Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.contrib.auth.decorators import login_required

from apps.core.decorators import permission_required

from .views import (_allowed_marketplaces, _resolve_campaign_period,
                    _CAMPAIGN_PERIODS)

# Presets offered in the window dropdown (plus a custom From/To calendar).
_PERIOD_PRESETS = [('7d', 'Last 7 days'), ('14d', 'Last 14 days'),
                   ('30d', 'Last 30 days'), ('60d', 'Last 60 days'),
                   ('65d', 'Last 65 days'), ('90d', 'Last 90 days'),
                   ('mtd', 'Month to date')]
_MAX_DAYS = 120

# Heuristic thresholds for the hourly-spend "capped day" estimate.
_CAP_LAST_HOUR, _CAP_MIN_HOURS, _CAP_MIN_SPEND = 20, 3, 2.0

# Search-term tag → suggested action (first match wins; mirrors the UI).
_ST_ACTION = [
    ('high_spend_no_sales', 'Add negative keyword'),
    ('losing_money',        'Lower bid or negate'),
    ('high_ctr_low_cvr',    'Fix listing / price'),
    ('scaling_opportunity', 'Scale bid up'),
    ('high_profit',         'Protect & scale'),
]


def _st_action_label(tags) -> str:
    for t, label in _ST_ACTION:
        if t in (tags or []):
            return label
    return 'Monitor'


def _pacing_action(capped_days, cap_rate, acos):
    if capped_days == 0:
        return {'level': 'ok', 'label': 'Within budget',
                'note': 'No capped days in this window.'}
    profitable = acos is not None and acos < 0.35
    high_acos  = acos is not None and acos >= 0.50
    if cap_rate >= 0.30 and profitable:
        return {'level': 'raise', 'label': 'Raise daily budget',
                'note': 'Caps out often while profitable — you are losing sales in '
                        'the blocked hours.'}
    if cap_rate >= 0.30 and high_acos:
        return {'level': 'fix', 'label': 'Fix targeting first',
                'note': 'Caps out often but ACOS is high — more budget would buy more '
                        'waste. Tighten targeting/negatives before adding budget.'}
    return {'level': 'watch', 'label': 'Review pacing',
            'note': 'Capped on some days — raise budget only if those hours convert.'}


def _window(request):
    """Resolve (start, end, label) from ?start&end (custom calendar) or ?period."""
    def _d(x):
        try:
            return date.fromisoformat(x)
        except (TypeError, ValueError):
            return None
    ps, pe = _d(request.GET.get('start')), _d(request.GET.get('end'))
    if ps and pe and ps <= pe:
        if (pe - ps).days > _MAX_DAYS:
            ps = pe - timedelta(days=_MAX_DAYS)
        return ps, pe, f'{ps.isoformat()} → {pe.isoformat()}'
    period = (request.GET.get('period') or '30d').lower()
    today = date.today()
    m = re.fullmatch(r'(\d+)d', period)
    if m:
        n = max(1, min(int(m.group(1)), _MAX_DAYS))
        end = today - timedelta(days=1)
        return end - timedelta(days=n - 1), end, f'Last {n} days'
    if period in _CAMPAIGN_PERIODS and period != 'today':
        s, e, _ = _resolve_campaign_period(period, today)
        return s, e, _CAMPAIGN_PERIODS[period][0]
    end = today - timedelta(days=1)
    return end - timedelta(days=29), end, 'Last 30 days'


# ── Shared data builders (used by JSON endpoints AND the Excel export) ────────
def _search_term_rows(marketplace, start, end, limit):
    from .models import AdsSearchTermDailySnapshot, CampaignProfitDaily, Campaign
    from .views import _tag_search_term

    cm_map = {}
    for r in (CampaignProfitDaily.objects
              .filter(marketplace=marketplace, date__gte=start, date__lte=end)
              .values('campaign_id', 'date', 'ad_revenue', 'contribution_margin')):
        rev = float(r['ad_revenue'] or 0)
        if rev > 0:
            cm_map[(r['campaign_id'], r['date'])] = float(r['contribution_margin'] or 0) / rev
    id2name = {str(cid): nm for cid, nm in
               Campaign.objects.filter(marketplace=marketplace)
               .values_list('campaign_id', 'campaign_name')}

    raw = (AdsSearchTermDailySnapshot.objects
           .filter(marketplace=marketplace, date__gte=start, date__lte=end)
           .values('campaign_id', 'date', 'search_term', 'search_term_hash',
                   'spend', 'sales_7d', 'orders_7d', 'clicks', 'impressions')
           .order_by('-spend')[:max(limit * 8, 6000)])
    agg = defaultdict(lambda: {'term': '', 'cid': '', 'spend': 0.0, 'sales': 0.0,
                               'orders': 0, 'clicks': 0, 'impr': 0, 'profit': 0.0})
    for r in raw:
        cid = str(r['campaign_id'])
        b = agg[(r['search_term_hash'], cid)]
        b['term'] = r['search_term']; b['cid'] = cid
        s = float(r['spend'] or 0); v = float(r['sales_7d'] or 0)
        b['spend'] += s; b['sales'] += v
        b['orders'] += int(r['orders_7d'] or 0)
        b['clicks'] += int(r['clicks'] or 0)
        b['impr'] += int(r['impressions'] or 0)
        b['profit'] += v * cm_map.get((cid, r['date']), 0.0) - s

    rows = []
    for b in agg.values():
        spend, sales = b['spend'], b['sales']
        ctr = (b['clicks'] / b['impr']) if b['impr'] else None
        cvr = (b['orders'] / b['clicks']) if b['clicks'] else None
        acos = (spend / sales) if sales else None
        roas = (sales / spend) if spend else None
        tags = _tag_search_term(spend=spend, sales=sales, orders=b['orders'],
                                clicks=b['clicks'], impr=b['impr'],
                                est_profit=b['profit'], acos=acos)
        rows.append({
            'search_term': b['term'],
            'campaign': id2name.get(b['cid']) or f'#{b["cid"]}',
            'campaign_id': b['cid'],
            'spend': round(spend, 2), 'sales': round(sales, 2),
            'orders': b['orders'], 'clicks': b['clicks'], 'impressions': b['impr'],
            'ctr':  round(ctr * 100, 2) if ctr is not None else None,
            'cvr':  round(cvr * 100, 2) if cvr is not None else None,
            'acos': round(acos * 100, 1) if acos is not None else None,
            'roas': round(roas, 2) if roas is not None else None,
            'est_profit': round(b['profit'], 2), 'tags': tags,
        })
    rows.sort(key=lambda r: r['spend'], reverse=True)
    return rows[:limit]


def _budget_rows(marketplace, start, end):
    """Returns (rows, method, note). Prefers exact budget-usage; else estimates."""
    from .models import (PPCCampaignHourlySnapshot, Campaign, CampaignProfitDaily,
                         CampaignBudgetUsageDaily)
    id2name = {str(cid): nm for cid, nm in
               Campaign.objects.filter(marketplace=marketplace)
               .values_list('campaign_id', 'campaign_name')}
    rev = {}
    for r in (CampaignProfitDaily.objects
              .filter(marketplace=marketplace, date__gte=start, date__lte=end)
              .values('campaign_id').annotate(sp=Sum('spend'), rv=Sum('ad_revenue'))):
        rev[str(r['campaign_id'])] = (float(r['sp'] or 0), float(r['rv'] or 0))

    def _mk(cid, active, capped, spend, detail):
        sp, rv = rev.get(cid, (spend, 0.0))
        acos = (sp / rv) if rv > 0 else None
        cap_rate = capped / active if active else 0
        return {'campaign_id': cid, 'campaign': id2name.get(cid) or f'#{cid}',
                'active_days': active, 'capped_days': capped,
                'cap_rate': round(cap_rate * 100, 1),
                'spend': round(sp if rv else spend, 2),
                'acos': round(acos * 100, 1) if acos is not None else None,
                'detail': detail, 'action': _pacing_action(capped, cap_rate, acos)}

    rows = []
    bu = CampaignBudgetUsageDaily.objects.filter(
        marketplace=marketplace, date__gte=start, date__lte=end)
    if bu.exists():
        method = 'exact'
        agg = defaultdict(lambda: {'active': 0, 'capped': 0, 'usum': 0.0})
        for r in bu.values('campaign_id', 'usage_pct'):
            cid = str(r['campaign_id']); u = float(r['usage_pct'] or 0)
            a = agg[cid]; a['active'] += 1; a['usum'] += u
            if u >= CampaignBudgetUsageDaily.OUT_OF_BUDGET_PCT:
                a['capped'] += 1
        for cid, a in agg.items():
            avg = a['usum'] / a['active'] if a['active'] else 0
            rows.append(_mk(cid, a['active'], a['capped'], 0.0, f'avg peak {avg:.0f}% of budget'))
        note = ('Exact — from Amazon\'s budget-usage stream (peak % of daily budget '
                'consumed; a day that reached 100% ran out of budget).')
    else:
        method = 'estimated'
        hrows = (PPCCampaignHourlySnapshot.objects
                 .filter(marketplace=marketplace, campaign_type='sp',
                         date__gte=start, date__lte=end)
                 .values('campaign_id', 'date', 'hour').annotate(sp=Sum('spend')))
        day = defaultdict(dict)
        for r in hrows:
            day[(r['campaign_id'], r['date'])][r['hour']] = float(r['sp'] or 0)
        camp = defaultdict(lambda: {'active': 0, 'capped': 0, 'spend': 0.0, 'last': []})
        for (cid, _d), hours in day.items():
            dsp = sum(hours.values()); c = camp[str(cid)]; c['spend'] += dsp
            if dsp < _CAP_MIN_SPEND:
                continue
            active = [h for h, v in hours.items() if v > 0]
            c['active'] += 1; last = max(active) if active else 0; c['last'].append(last)
            if len(active) >= _CAP_MIN_HOURS and last <= _CAP_LAST_HOUR:
                c['capped'] += 1
        for cid, c in camp.items():
            if c['active'] == 0:
                continue
            avg_last = round(sum(c['last']) / len(c['last'])) if c['last'] else None
            rows.append(_mk(cid, c['active'], c['capped'], c['spend'],
                            f'stops ~{avg_last}:00' if avg_last is not None else ''))
        note = ('Estimated from the hourly SP spend curve — a campaign that spends '
                'through the day then flatlines to $0 for the last hours has almost '
                'certainly hit its daily budget. Connect Amazon\'s budget-usage '
                'stream (ingest_ams_s3) for exact figures.')
    rows.sort(key=lambda r: (r['capped_days'], r['cap_rate'], r['spend']), reverse=True)
    return rows, method, note


# ── Filter application (shared by the Excel export) ───────────────────────────
def _num_ok(val, op, thr):
    try:
        t = float(thr)
    except (TypeError, ValueError):
        return True
    if val is None:
        return False
    return {'gt': val > t, 'lt': val < t, 'gte': val >= t,
            'lte': val <= t, 'eq': val == t}.get(op, True)


def _parse_conds(raw):
    out = []
    for part in (raw or '').split(';'):
        bits = part.split(':', 2)
        if len(bits) == 3 and bits[0].strip():
            out.append((bits[0].strip(), bits[1].strip(), bits[2].strip()))
    return out


# ── Views ─────────────────────────────────────────────────────────────────────
@login_required
@permission_required('can_view_dashboard')
def marketing_optimizer(request):
    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        marketplace = _allowed_marketplaces(request.user)[0]
    return render(request, 'dashboard/marketing_optimizer.html', {
        'marketplace':          marketplace,
        'allowed_marketplaces': _allowed_marketplaces(request.user),
        'period_options':       [{'id': i, 'label': l} for i, l in _PERIOD_PRESETS],
        'today':                date.today(),
    })


@login_required
@permission_required('can_view_dashboard')
def api_mkt_search_terms(request):
    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        return JsonResponse({'error': 'forbidden'}, status=403)
    start, end, label = _window(request)
    try:
        limit = max(1, min(int(request.GET.get('limit') or 1200), 3000))
    except ValueError:
        limit = 1200
    rows = _search_term_rows(marketplace, start, end, limit)
    return JsonResponse({
        'marketplace': marketplace,
        'period': {'label': label, 'start': start.isoformat(), 'end': end.isoformat()},
        'kpi': {'rows': len(rows), 'terms': len({r['search_term'] for r in rows}),
                'spend': round(sum(r['spend'] for r in rows), 2),
                'sales': round(sum(r['sales'] for r in rows), 2),
                'orders': sum(r['orders'] for r in rows),
                'est_profit': round(sum(r['est_profit'] for r in rows), 2)},
        'campaigns': sorted({r['campaign'] for r in rows}),
        'rows': rows,
    })


@login_required
@permission_required('can_view_dashboard')
def api_budget_pacing(request):
    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        return JsonResponse({'error': 'forbidden'}, status=403)
    start, end, label = _window(request)
    rows, method, note = _budget_rows(marketplace, start, end)
    capped = [r for r in rows if r['capped_days'] > 0]
    return JsonResponse({
        'marketplace': marketplace,
        'period': {'label': label, 'start': start.isoformat(), 'end': end.isoformat()},
        'method': method,
        'kpi': {'campaigns': len(rows), 'campaigns_capped': len(capped),
                'capped_days': sum(r['capped_days'] for r in rows),
                'spend_capped': round(sum(r['spend'] for r in capped), 2)},
        'campaigns': sorted({r['campaign'] for r in rows}),
        'rows': rows, 'estimate_note': note,
    })


@login_required
@permission_required('can_view_dashboard')
def mkt_export(request):
    """Excel (.xlsx) export of the current tab, honouring the active filters."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        return JsonResponse({'error': 'forbidden'}, status=403)
    start, end, label = _window(request)
    tab = 'budget' if request.GET.get('tab') == 'budget' else 'terms'
    q = (request.GET.get('q') or '').strip().lower()
    campaign = request.GET.get('campaign') or 'all'
    action = request.GET.get('action') or 'all'
    conds = _parse_conds(request.GET.get('cond'))

    wb = Workbook(); ws = wb.active

    def conds_pass(row, action_of):
        for metric, op, val in conds:
            if metric == 'action':
                if action_of and action_of(row) != val:
                    return False
            elif not _num_ok(row.get(metric), op, val):
                return False
        return True

    if tab == 'budget':
        ws.title = 'Budget pacing'
        rows, method, _ = _budget_rows(marketplace, start, end)
        headers = ['Campaign', 'Capped days', 'Active days', 'Cap rate %', 'Spend',
                   'ACOS %', 'Detail', 'Recommendation', 'Why']
        ws.append(headers)
        for r in rows:
            if q and q not in r['campaign'].lower():
                continue
            if campaign != 'all' and r['campaign'] != campaign:
                continue
            if not conds_pass(r, None):
                continue
            ws.append([r['campaign'], r['capped_days'], r['active_days'], r['cap_rate'],
                       r['spend'], r['acos'], r['detail'], r['action']['label'],
                       r['action']['note']])
    else:
        ws.title = 'Search terms'
        rows = _search_term_rows(marketplace, start, end, 3000)
        action_of = lambda r: _st_action_label(r['tags'])
        headers = ['Search term', 'Campaign', 'Spend', 'Sales', 'Orders', 'Clicks',
                   'Impressions', 'ACOS %', 'CVR %', 'CTR %', 'ROAS', 'Est profit',
                   'Suggested action']
        ws.append(headers)
        for r in rows:
            if q and q not in r['search_term'].lower() and q not in r['campaign'].lower():
                continue
            if campaign != 'all' and r['campaign'] != campaign:
                continue
            if action != 'all' and action_of(r) != action:
                continue
            if not conds_pass(r, action_of):
                continue
            ws.append([r['search_term'], r['campaign'], r['spend'], r['sales'],
                       r['orders'], r['clicks'], r['impressions'], r['acos'],
                       r['cvr'], r['ctr'], r['roas'], r['est_profit'], action_of(r)])

    fill = PatternFill('solid', fgColor='232F3E')
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF'); c.fill = fill
    ws.freeze_panes = 'A2'

    resp = HttpResponse(content_type=(
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))
    fn = f'marketing-{tab}-{marketplace}-{start.isoformat()}_{end.isoformat()}.xlsx'
    resp['Content-Disposition'] = f'attachment; filename="{fn}"'
    wb.save(resp)
    return resp
