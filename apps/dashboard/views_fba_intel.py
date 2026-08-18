"""FBA Fee Intelligence — page shell + JSON. All computation lives in
apps.dashboard.fba_intel; this layer only handles auth, params and shaping."""
from datetime import date

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render

from apps.core.decorators import permission_required

from . import fba_intel
from .views import _allowed_marketplaces

_PERIODS = [7, 14, 30, 60, 90]


def _params(request):
    mp = request.GET.get('mp', 'usa')
    try:
        days = int(request.GET.get('days') or 30)
    except ValueError:
        days = 30
    days = days if days in _PERIODS else max(1, min(days, 365))
    anchor = None
    raw = (request.GET.get('anchor') or '').strip()
    if raw:
        try:
            anchor = date.fromisoformat(raw)
        except ValueError:
            anchor = None
    return (mp, days, anchor,
            (request.GET.get('category') or '').strip(),
            (request.GET.get('q') or '').strip())


@login_required
@permission_required('can_view_dashboard')
def fba_intelligence(request):
    mp = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(mp):
        mp = _allowed_marketplaces(request.user)[0]
    return render(request, 'dashboard/fba_intelligence.html', {
        'marketplace': mp,
        'allowed_marketplaces': _allowed_marketplaces(request.user),
        'period_options': _PERIODS,
    })


@login_required
@permission_required('can_view_dashboard')
def api_fba_intelligence(request):
    mp, days, anchor, category, q = _params(request)
    if mp and not request.user.can_access_marketplace(mp):
        return JsonResponse({'error': 'forbidden'}, status=403)
    return JsonResponse(fba_intel.compute(marketplace=mp, days=days,
                                          anchor=anchor, category=category,
                                          search=q))


@login_required
@permission_required('can_view_dashboard')
def fba_intelligence_export(request):
    """Export the SKU impact table exactly as filtered."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    mp, days, anchor, category, q = _params(request)
    if mp and not request.user.can_access_marketplace(mp):
        return JsonResponse({'error': 'forbidden'}, status=403)
    d = fba_intel.compute(marketplace=mp, days=days, anchor=anchor,
                          category=category, search=q)

    _SIG = {'fee_up_cover_down': 'Fee up while cover fell',
            'fee_down_cover_up': 'Fee down while cover rose',
            'no_clear_relationship': 'No clear relationship',
            'insufficient_data': 'Insufficient data'}
    wb = Workbook()
    ws = wb.active
    ws.title = f'FBA fee {days}d'
    ws.append(['SKU', 'ASIN', 'Product', 'Category',
               'Current fee/unit', 'Previous fee/unit', 'Delta/unit', 'Delta %',
               '7d drift', '14d drift', '30d drift',
               'Billed units', 'Units sold',
               'Incremental cost', 'Savings', 'Net impact',
               'Current inventory', 'Days cover', 'Inventory signal'])
    for r in d.get('rows', []):
        ws.append([r['sku'], r['asin'], r['title'], r['category'],
                   r['current_fee'], r['previous_fee'], r['fee_delta'],
                   r['fee_delta_pct'], r['drift_7d'], r['drift_14d'],
                   r['drift_30d'], r['billed_units'], r['units_sold'],
                   r['incremental_cost'], r['savings'], r['net_impact'],
                   r['current_inventory'], r['current_days_cover'],
                   _SIG.get(r['inventory_signal'], r['inventory_signal'])])
    fill = PatternFill('solid', fgColor='232F3E')
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = fill
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:S{max(ws.max_row, 1)}'
    for col, w in zip('ABCDEFGHIJKLMNOPQRS',
                      [22, 14, 40, 18, 15, 16, 12, 10, 10, 10, 10,
                       12, 11, 16, 12, 13, 16, 11, 26]):
        ws.column_dimensions[col].width = w
    note = wb.create_sheet('Methodology')
    note.append(['Methodology']); note.append([d.get('methodology', '')])
    note.append([]); note.append(['Scope']); note.append([d.get('scope_note', '')])
    note.column_dimensions['A'].width = 120

    resp = HttpResponse(content_type=(
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))
    resp['Content-Disposition'] = (
        f'attachment; filename="fba-fee-intelligence-{mp}-{days}d-'
        f'{date.today().isoformat()}.xlsx"')
    wb.save(resp)
    return resp
