"""
apps/dashboard/views.py — All dashboard views
"""
import csv
import io
import json
import logging
from datetime import date, timedelta
from decimal import Decimal
import re
from datetime import datetime
from zoneinfo import ZoneInfo

from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.http import StreamingHttpResponse, JsonResponse, HttpResponse
from django.conf import settings
from django.db import models
from django.db.models import Sum, Avg
from django.utils import timezone

from apps.core.decorators import permission_required
from apps.users.models import AuditLog
from apps.amazon_api.models import AmazonAPIConfig, AnthropicConfig
from .models import Product, COGSEntry, MonthlyTarget, DailyMetric, ProductTypePackMonthlyTarget, FBAFeeRate
from .forms import COGSBulkUploadForm, COGSEntryForm, MonthlyTargetForm, ProductForm, FBARateBulkUploadForm

logger = logging.getLogger(__name__)


def _allowed_marketplaces(user) -> list:
    """Superusers and users with empty allowed_marketplaces list see all markets."""
    if user.is_superuser or not user.allowed_marketplaces:
        return list(settings.AMAZON_MARKETPLACES.keys())
    return user.allowed_marketplaces


# ── AMZN.GR.* SKU variant → parent-Product resolver ─────────────────────────
# Amazon auto-creates SKU variants like
#     AMZN.GR.WSH-CLT-WHT-12-5FPIYFKB7IF1JA-LN
# for Vine review units, Launchpad listings, and gift-promotion variants.
# The parent SKU (`WSH-CLT-WHT-12`) is in our catalog with the right brand/
# title, but the variant has no Product row of its own → it would otherwise
# land in the "(unbranded)" bucket on the Brand P&L page.
#
# Implementation: strip the AMZN.GR. prefix, then try progressively shorter
# dash-prefix candidates against the Product table until one matches. The
# cache key is (mp, sku) so repeated lookups within a request are free.
_AMZN_PARENT_CACHE: dict = {}


def _resolve_amzn_parent_product(marketplace: str, sku: str):
    """
    Return the parent Product row for an AMZN.GR.* variant SKU, or None.
    For plain (non-variant) SKUs also does a direct Product lookup so callers
    can use this as a single source of truth for "which Product describes this
    SKU".
    """
    key = (marketplace, sku)
    if key in _AMZN_PARENT_CACHE:
        return _AMZN_PARENT_CACHE[key]
    from .models import Product as _P
    p = _P.objects.filter(marketplace=marketplace, sku=sku).only(
        'sku', 'brand', 'title', 'asin').first()
    if p:
        _AMZN_PARENT_CACHE[key] = p
        return p
    if sku.startswith('AMZN.GR.'):
        body  = sku[len('AMZN.GR.'):]
        parts = body.split('-')
        # Try 2, 3, then 4 trailing segments stripped — Amazon's hash format
        # isn't perfectly uniform (sometimes the hash itself contains a dash).
        for n in (2, 3, 4):
            if len(parts) <= n:
                break
            candidate = '-'.join(parts[:-n])
            p = _P.objects.filter(marketplace=marketplace, sku=candidate).only(
                'sku', 'brand', 'title', 'asin').first()
            if p:
                _AMZN_PARENT_CACHE[key] = p
                return p
    _AMZN_PARENT_CACHE[key] = None
    return None


def _build_initial_today(marketplace: str) -> dict:
    """
    Build a server-side snapshot of today's cumulative metrics + last-hour
    delta + same-hour-yesterday comparison, computed from HourlyMetricSnapshot.

    Returns dict ready for {{ initial_today.* }} placeholders in the template.
    If no hourly data exists yet, returns minimal stub so the page still renders.
    """
    from datetime import datetime as _dt, timedelta as _td
    from zoneinfo import ZoneInfo as _ZI
    from django.db.models import Sum as _Sum
    from .models import HourlyMetricSnapshot

    tz_name = settings.AMAZON_MARKETPLACES.get(marketplace, {}).get('timezone', settings.TIME_ZONE)
    now_local = _dt.now(tz=_ZI(tz_name))
    today_local = now_local.date()

    today_rows = list(
        HourlyMetricSnapshot.objects
        .filter(marketplace=marketplace, date=today_local)
        .order_by('hour')
        .values('hour', 'revenue', 'units', 'orders',
                'cgs', 'amazon_fee', 'fba_fee',
                'gross_margin', 'contribution_margin', 'ppc_spend',
                'synced_at')
    )

    if not today_rows:
        return {
            'has_data':       False,
            'marketplace':    marketplace,
            'tz_label':       now_local.strftime('%Z'),
            'today_local':    today_local.isoformat(),
        }

    # Cumulative totals across every hour we have for today so far
    totals = {
        'revenue':   0.0, 'units': 0, 'orders': 0,
        'cgs':       0.0, 'amz_fee': 0.0, 'fba_fee': 0.0,
        'gm':        0.0, 'cm': 0.0, 'ppc': 0.0,
    }
    last_synced_at = None
    hourly_revenue_series = [0.0] * 24
    hourly_units_series   = [0]   * 24

    for r in today_rows:
        totals['revenue'] += float(r['revenue'])
        totals['units']   += int(r['units'])
        totals['orders']  += int(r['orders'])
        totals['cgs']     += float(r['cgs'])
        totals['amz_fee'] += float(r['amazon_fee'])
        totals['fba_fee'] += float(r['fba_fee'])
        totals['gm']      += float(r['gross_margin'])
        totals['cm']      += float(r['contribution_margin'])
        totals['ppc']     += float(r['ppc_spend'])
        h = r['hour']
        if 0 <= h < 24:
            hourly_revenue_series[h] = float(r['revenue'])
            hourly_units_series[h]   = int(r['units'])
        if (last_synced_at is None) or (r['synced_at'] > last_synced_at):
            last_synced_at = r['synced_at']

    # Last hour = the row with the highest hour number
    last_hour_row = today_rows[-1]
    last_hour     = int(last_hour_row['hour'])
    last_hour_rev = float(last_hour_row['revenue'])
    last_hour_units = int(last_hour_row['units'])

    # Same-hour-yesterday for the comparison badge
    yesterday_local = today_local - _td(days=1)
    yest_same_hour = (
        HourlyMetricSnapshot.objects
        .filter(marketplace=marketplace, date=yesterday_local, hour=last_hour)
        .values('revenue', 'units')
        .first()
    )
    yest_rev = float(yest_same_hour['revenue']) if yest_same_hour else 0.0
    if yest_rev > 0:
        vs_yest_pct = round((last_hour_rev - yest_rev) / yest_rev * 100, 1)
    else:
        vs_yest_pct = None  # no comparable data

    rev = totals['revenue']
    # Ratios are measured against revenue ex-VAT, matching the API the page
    # refreshes from — otherwise the server-rendered tiles flash one number and
    # the first fetch replaces it with another.
    from .sync import net_factor as _net_factor
    rev_net = rev * _net_factor(marketplace)
    gm_pct = round((totals['gm'] / rev_net * 100), 2) if rev_net else 0.0
    cm_pct = round((totals['cm'] / rev_net * 100), 2) if rev_net else 0.0
    tacos  = round((totals['ppc'] / rev_net * 100), 2) if rev_net else 0.0
    arpu   = round(rev / totals['units'], 2) if totals['units'] else 0.0

    # As-of timestamp in marketplace TZ
    if last_synced_at:
        as_of = last_synced_at.astimezone(_ZI(tz_name))
        as_of_str = as_of.strftime('%-I:%M %p %Z')
    else:
        as_of_str = '—'

    return {
        'has_data':         True,
        'marketplace':      marketplace,
        'tz_label':         now_local.strftime('%Z'),
        'today_local':      today_local.isoformat(),
        # Cumulative KPIs
        'revenue':          round(rev, 2),
        'units':            totals['units'],
        'orders':           totals['orders'],
        'cgs':              round(totals['cgs'], 2),
        'amz_fee':          round(totals['amz_fee'], 2),
        'fba_fee':          round(totals['fba_fee'], 2),
        'gross_margin':     round(totals['gm'], 2),
        'gm_pct':           gm_pct,
        'contribution_margin': round(totals['cm'], 2),
        'cm_pct':           cm_pct,
        'ppc_spend':        round(totals['ppc'], 2),
        'tacos':            tacos,
        'arpu':             arpu,
        # Last-hour delta + comparison
        'last_hour':        last_hour,
        'last_hour_revenue':  round(last_hour_rev, 2),
        'last_hour_units':    last_hour_units,
        'vs_yesterday_pct':   vs_yest_pct,         # None when no comparison
        'yesterday_same_hour_rev': round(yest_rev, 2),
        # Intra-day chart series (24 ints — index = hour)
        'hourly_revenue':   hourly_revenue_series,
        'hourly_units':     hourly_units_series,
        # "As of"
        'as_of':            as_of_str,
        'as_of_iso':        last_synced_at.isoformat() if last_synced_at else None,
    }


@login_required
@permission_required('can_view_dashboard')
def index(request):
    import json as _json
    configs = AmazonAPIConfig.objects.filter(is_active=True).values(
        'marketplace', 'label', 'last_test_status', 'last_tested_at'
    )
    allowed = _allowed_marketplaces(request.user)
    default_mp = (allowed[0] if allowed else 'usa')

    initial_today = _build_initial_today(default_mp)

    vat_rates = {mp: v.get('vat', 0)
                 for mp, v in settings.AMAZON_MARKETPLACES.items()}
    ctx = {
        'configs': {c['marketplace']: c for c in configs},
        'vat_rates_json': _json.dumps(vat_rates),
        'allowed_marketplaces': allowed,
        'show_financials': request.user.has_perm_flag('can_view_financials'),
        'show_ppc':        request.user.has_perm_flag('can_view_ppc'),
        'show_inventory':  request.user.has_perm_flag('can_view_inventory'),
        'can_ai_summary':  request.user.has_perm_flag('can_generate_ai_summary'),
        'today':           date.today(),
        'initial_today':   initial_today,
        # Pre-serialised so the template can drop directly into <script>.
        'initial_today_json': _json.dumps(initial_today, default=str),
    }
    return render(request, 'dashboard/index.html', ctx)


@login_required
@permission_required('can_view_historical')
def historical(request):
    from .sync import sync_window, apply_ppc_from_snapshots, days_missing_ppc

    marketplace   = request.GET.get('mp', 'usa')
    period        = request.GET.get('period', '30d')
    backfill_days = request.GET.get('backfill_days')

    if not request.user.can_access_marketplace(marketplace):
        marketplace = _allowed_marketplaces(request.user)[0]

    today     = date.today()
    yesterday = today - timedelta(days=1)   # historical view ends at yesterday
    end       = yesterday                   # never show today's incomplete row

    days_map = {
        '7d':  7,
        '30d': 30,
        '90d': 90,
        'ytd': (end - end.replace(month=1, day=1)).days + 1,
    }
    days  = days_map.get(period, 30)
    start = end - timedelta(days=days - 1) if period != 'ytd' else end.replace(month=1, day=1)

    cfg = AmazonAPIConfig.objects.filter(marketplace=marketplace, is_active=True).first()
    has_config = bool(cfg and cfg.has_sp_api_credentials())

    sync_status = None

    # UI-triggered backfill. Synchronous, capped at 90s. If Amazon's report
    # isn't ready by then, the in-flight reportId persists in memory — clicking
    # the button again will re-poll the SAME report and download as soon as it's
    # built.
    if backfill_days and has_config:
        try:
            n = max(1, min(int(backfill_days), 90))
            bf_end   = end
            bf_start = bf_end - timedelta(days=n - 1)

            # Step 1: sync order data (revenue / units / COGS)
            res = sync_window(marketplace, bf_start, bf_end, max_wait_seconds=90)

            # Step 2: apply PPC from any snapshots already in DB (instant)
            ppc_updated = apply_ppc_from_snapshots(marketplace, bf_start, bf_end)

            # Step 3: if some days still have no PPC, launch backfill_ppc in the
            # background so Amazon's campaign reports are fetched asynchronously.
            missing = days_missing_ppc(marketplace, bf_start, bf_end)
            ppc_bg_launched = False
            if missing:
                try:
                    import subprocess, sys, os
                    subprocess.Popen(
                        [
                            sys.executable,
                            'manage.py', 'backfill_ppc',
                            '--marketplace', marketplace,
                            '--start', str(bf_start),
                            '--end',   str(bf_end),
                        ],
                        cwd=settings.BASE_DIR,
                        env=os.environ.copy(),
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL,
                    )
                    ppc_bg_launched = True
                except Exception as bg_err:
                    logger.warning('Failed to launch background backfill_ppc: %s', bg_err)

            sync_status = (
                f"Backfill {bf_start:%Y-%m-%d} → {bf_end:%Y-%m-%d}  "
                f"status={res['status']}  rows={res['rows']}  "
                f"days_written={res.get('days_written', 0)}  "
                f"ppc_days_from_cache={ppc_updated}"
                + (f"  days_with_orders={res['days_with_orders']}" if 'days_with_orders' in res else '')
            )
            if ppc_bg_launched:
                sync_status += (
                    f"  · PPC backfill started in background for {len(missing)} day(s) "
                    f"— refresh this page in 1-3 minutes to see PPC data."
                )
            if res['status'] not in ('OK', 'CACHED', 'FRESH'):
                sync_status += (
                    f"  · Amazon hasn't finished building the orders report yet "
                    f"(reportId={res.get('report_id')}). Click the backfill button again in "
                    f"1-2 minutes — the same reportId will be polled and downloaded when ready."
                )
        except Exception as e:
            logger.error('UI backfill failed: %s', e, exc_info=True)
            sync_status = f'backfill error: {e}'

    # Auto-sync yesterday if it has no row yet (fills the gap at start of each new day).
    # Also apply PPC from existing snapshots so yesterday's PPC shows immediately
    # even before the 6am cron runs.
    if has_config and not backfill_days:
        yest_exists = DailyMetric.objects.filter(
            marketplace=marketplace, date=yesterday
        ).exclude(revenue=0).exists()
        if not yest_exists:
            try:
                sync_window(marketplace, yesterday, yesterday, max_wait_seconds=30)
            except Exception:
                pass  # silent — user can trigger manual backfill if needed
        # Always apply PPC snapshots for yesterday (handles midnight→6am gap)
        try:
            apply_ppc_from_snapshots(marketplace, yesterday, yesterday)
        except Exception:
            pass

    # Read from DailyMetric — ends at yesterday (today is always incomplete)
    import calendar
    rows = list(DailyMetric.objects.filter(
        marketplace=marketplace, date__gte=start, date__lte=end,
    ).order_by('date'))
    by_date = {r.date: r for r in rows}

    # Pre-load monthly targets for any month in the range so we can compute
    # the per-day TAR GM/Day = monthly_revenue_target / days_in_month
    # Source priority:
    #   1. MonthlyTarget.revenue_target  (marketplace-level, if user set it)
    #   2. SUM(ProductTypePackMonthlyTarget.revenue_target)  (per-product targets)
    month_starts = set()
    _m = start.replace(day=1)
    while _m <= today:
        month_starts.add(_m)
        _m = (_m + timedelta(days=32)).replace(day=1)

    # Layer 1 — marketplace-level targets
    monthly_target_revenue = {}     # date(month) → revenue target (float)
    for t in MonthlyTarget.objects.filter(marketplace=marketplace, month__in=month_starts):
        if t.revenue_target:
            monthly_target_revenue[t.month] = float(t.revenue_target)

    # Layer 2 — fall back to summing per-product targets for any month not in layer 1
    missing_months = [m for m in month_starts if m not in monthly_target_revenue]
    if missing_months:
        per_product_sums = (
            ProductTypePackMonthlyTarget.objects
            .filter(marketplace=marketplace, month__in=missing_months)
            .values('month')
            .annotate(total=Sum('revenue_target'))
        )
        for row in per_product_sums:
            if row['total']:
                monthly_target_revenue[row['month']] = float(row['total'])

    chart_dates, chart_rev, chart_units, chart_ppc, chart_tacos, chart_gm_pct = [], [], [], [], [], []
    table_rows = []
    tot_rev = tot_rev_net = tot_units = tot_orders = 0
    tot_cgs = tot_amz = tot_fba = tot_cm = tot_gm = tot_ppc = tot_tar = 0.0

    cursor = start
    while cursor <= end:
        m = by_date.get(cursor)
        rev    = float(m.revenue)             if m else 0.0
        rev_net = float(m.revenue_net or m.revenue) if m else 0.0
        units  = int(m.units)                 if m else 0
        orders = int(m.orders)                if m else 0
        cgs    = float(m.cgs or 0)            if m else 0.0
        amz    = float(m.amazon_fee or 0)     if m else 0.0
        fba    = float(m.fba_fee or 0)        if m else 0.0
        cm_amt = float(m.contribution_margin) if m else 0.0
        cm_pct = float(m.cm_pct) * 100        if m else 0.0
        ppc    = float(m.ppc_spend)           if m else 0.0  # 0 until Ads API
        # GM = CM − PPC (compute on-the-fly so it's always correct)
        gm_amt = cm_amt - ppc
        gm_pct = (gm_amt / rev * 100) if rev else 0.0
        tacos  = float(m.tacos) * 100         if m else 0.0

        # CPA: spend per order (0 when PPC not yet connected)
        cpa = (ppc / orders) if orders else 0.0

        # TAR GM/Day = monthly revenue target ÷ days_in_month
        # (uses MonthlyTarget if set, else summed ProductTypePackMonthlyTarget)
        target_rev = monthly_target_revenue.get(cursor.replace(day=1), 0.0)
        if target_rev:
            days_in_mo = calendar.monthrange(cursor.year, cursor.month)[1]
            tar_day    = target_rev / days_in_mo
        else:
            tar_day = 0.0
        gm_minus_tar = gm_amt - tar_day if tar_day else 0.0

        chart_dates.append(cursor.isoformat())
        chart_rev.append(round(rev, 2))
        chart_units.append(units)
        chart_ppc.append(round(ppc, 2))
        chart_tacos.append(round(tacos, 2))
        chart_gm_pct.append(round(gm_pct, 2))

        table_rows.append({
            'date':         cursor,
            'revenue':      rev,
            'revenue_net':  rev_net,
            'units':        units,
            'orders':       orders,
            'cgs':          cgs,
            'amazon_fee':   amz,
            'fba_fee':      fba,
            'cm':           cm_amt,
            'cm_pct':       cm_pct,
            'ppc_spend':    ppc,
            'gm':           gm_amt,
            'gm_pct':       gm_pct,
            'cpa':          cpa,
            'tacos':        tacos,
            'tar_gm_day':   tar_day,
            'gm_minus_tar': gm_minus_tar,
        })

        tot_rev    += rev
        tot_rev_net += rev_net
        tot_units  += units
        tot_orders += orders
        tot_cgs    += cgs
        tot_amz    += amz
        tot_fba    += fba
        tot_cm     += cm_amt
        tot_gm     += gm_amt
        tot_ppc    += ppc
        tot_tar    += tar_day
        cursor     += timedelta(days=1)

    totals = {
        'total_revenue': round(tot_rev, 2),
        'total_revenue_net': round(tot_rev_net, 2),
        'total_units':   int(tot_units),
        'total_orders':  int(tot_orders),
        'total_cgs':     round(tot_cgs, 2),
        'total_amz':     round(tot_amz, 2),
        'total_fba':     round(tot_fba, 2),
        'total_cm':      round(tot_cm, 2),
        'total_gm':      round(tot_gm, 2),
        'total_ppc':     round(tot_ppc, 2),
        'total_tar':     round(tot_tar, 2),
        'avg_cm_pct':    round((tot_cm / tot_rev * 100) if tot_rev else 0, 2),
        'avg_gm_pct':    round((tot_gm / tot_rev * 100) if tot_rev else 0, 2),
        'avg_tacos':     round((tot_ppc / tot_rev * 100) if tot_rev else 0, 2),
        'avg_acos':      0,
        'cpa':           round((tot_ppc / tot_orders) if tot_orders else 0, 2),
        'gm_minus_tar':  round(tot_gm - tot_tar, 2) if tot_tar else 0,
    }
    has_data = bool(rows)
    chart_data = json.dumps({
        'dates': chart_dates, 'revenue': chart_rev, 'units': chart_units,
        'ppc':   chart_ppc,   'tacos':   chart_tacos, 'gm_pct': chart_gm_pct,
    })

    last_sync = (
        DailyMetric.objects
        .filter(marketplace=marketplace, date=yesterday)
        .values_list('synced_at', flat=True)
        .first()
    )

    target = MonthlyTarget.objects.filter(
        marketplace=marketplace, month=end.replace(day=1)
    ).first()

    ctx = {
        'metrics':     table_rows,
        'totals':      totals,
        'chart_data':  chart_data,
        'has_data':    has_data,
        'has_config':  has_config,
        'sync_status': sync_status,
        'last_sync_yest': last_sync,
        'marketplace': marketplace,
        'vat_rate':    settings.AMAZON_MARKETPLACES.get(marketplace, {}).get('vat', 0),
        'period':      period,
        'start':       start,
        'end':         end,        # yesterday
        'today':       today,
        'target':      target,
        'allowed_marketplaces': _allowed_marketplaces(request.user),
        'show_financials': request.user.has_perm_flag('can_view_financials'),
        'show_ppc':        request.user.has_perm_flag('can_view_ppc'),
    }
    return render(request, 'dashboard/historical.html', ctx)


# ── AJAX: Product-line cumulative analysis ────────────────────────────────────
@login_required
def product_line_analysis(request):
    """
    Returns per-product-group P&L for a historical date range.
    Uses:
      • FlatFileAllOrdersReport  → revenue / qty / cogs / fees per SKU
      • PPCProductSnapshot (DB)  → PPC spend per ASIN
      • ProductTypePackMonthlyTarget → TAR GM/Day
    """
    from collections import defaultdict
    from apps.amazon_api.services import SPAPIClient, AdsAPIClient
    from apps.amazon_api.models import AmazonAPIConfig
    from apps.dashboard.models import (
        Product, COGSEntry, PPCProductSnapshot, ProductTypePackMonthlyTarget
    )
    import calendar as _cal
    import re

    marketplace = request.GET.get('mp', 'usa')
    start_str   = request.GET.get('start')
    end_str     = request.GET.get('end')

    if not request.user.can_access_marketplace(marketplace):
        return JsonResponse({'error': 'Access denied'}, status=403)

    try:
        s_d = date.fromisoformat(start_str)
        e_d = date.fromisoformat(end_str)
    except Exception:
        yesterday = date.today() - timedelta(days=1)
        s_d = yesterday - timedelta(days=29)
        e_d = yesterday

    # Never let end bleed into today
    yesterday = date.today() - timedelta(days=1)
    e_d = min(e_d, yesterday)

    cfg = AmazonAPIConfig.objects.filter(marketplace=marketplace, is_active=True).first()
    if not cfg or not cfg.has_sp_api_credentials():
        return JsonResponse({'error': 'No SP-API config'}, status=400)

    period_days   = (e_d - s_d).days + 1
    days_in_month = _cal.monthrange(e_d.year, e_d.month)[1]

    # ── 1. Fetch order report ────────────────────────────────────────────────
    client = SPAPIClient(cfg)
    tz_name = settings.AMAZON_MARKETPLACES.get(marketplace, {}).get('timezone', settings.TIME_ZONE)
    local_zone = ZoneInfo(tz_name)
    report_result = client.fetch_orders_report_sync(
        'custom', start_date=str(s_d), end_date=str(e_d), max_wait_seconds=90,
    )
    rows = report_result.get('rows') or []

    # ── 2. Build catalog lookups ─────────────────────────────────────────────
    prods_by_sku  = {}
    prods_by_asin = {}
    for p in Product.objects.filter(marketplace=marketplace):
        if p.sku:  prods_by_sku[p.sku.upper()]  = p
        if p.asin: prods_by_asin[p.asin.upper()] = p

    month_start = e_d.replace(day=1)
    cogs_by_sku  = {}
    cogs_by_asin = {}
    for c in COGSEntry.objects.filter(
        product__marketplace=marketplace, month=month_start,
    ).select_related('product'):
        if c.product.sku:  cogs_by_sku[c.product.sku.upper()]  = c
        if c.product.asin: cogs_by_asin[c.product.asin.upper()] = c

    # ── 3. Aggregate order rows per SKU ─────────────────────────────────────
    agg = {}
    for row in rows:
        if (row.get('order-status') or '').strip().lower() == 'cancelled': continue
        if (row.get('item-status')  or '').strip().lower() == 'cancelled': continue
        ch = (row.get('sales-channel') or '').strip().lower()
        if ch and ch != 'amazon.com': continue

        sku  = (row.get('sku')  or '').strip()
        asin = (row.get('asin') or '').strip()
        key  = sku or asin
        if not key: continue
        try:   qty   = int(float(row.get('quantity') or 0))
        except: qty  = 0
        try:   price = float(row.get('item-price') or 0)
        except: price = 0.0
        try:   promo = float(row.get('item-promotion-discount') or 0)
        except: promo = 0.0
        rev = max(0.0, price - promo)

        a = agg.setdefault(key, {'qty': 0, 'revenue': 0.0, 'asin': asin, 'sku': sku,
                                  'title': (row.get('product-name') or '')})
        a['qty']     += qty
        a['revenue'] += rev

    # ── 4. Build per-group breakdown ─────────────────────────────────────────
    def split_title(title):
        parts = [p.strip() for p in (title or '').split(' - ') if p.strip()]
        pt   = parts[0] if parts else 'Other'
        pack = parts[1] if len(parts) > 1 else '—'
        var  = parts[2] if len(parts) > 2 else ''
        return pt, pack, var

    # VAT-exclusive basis for every margin/percentage below. 1.0 for the USA.
    from .sync import _vat_rate
    vat_rate   = _vat_rate(marketplace)
    net_factor = 1.0 / (1.0 + vat_rate)

    grouped = {}
    for key, m in agg.items():
        sku  = m['sku']
        asin = m['asin']
        qty  = m['qty']
        rev  = m['revenue']
        product = prods_by_sku.get(sku.upper())  or prods_by_asin.get(asin.upper())
        cogs    = cogs_by_sku.get(sku.upper())   or cogs_by_asin.get(asin.upper())

        if rev == 0 and product:
            rev = float(product.sale_price or product.list_price or 0) * qty

        if cogs:
            cgs_u = (float(cogs.unit_cost or 0) + float(cogs.duties_cost or 0)
                     + float(cogs.prep_cost or 0) + float(cogs.other_cost or 0))
            fba_u = float(cogs.shipping_cost or 0)
        else:
            cgs_u = fba_u = 0.0

        total_cgs = cgs_u * qty
        fulfill   = fba_u * qty
        amz_fee   = rev * 0.15
        # Margin is measured on revenue EXCLUDING VAT. In the UK/AE/SA the
        # item price Amazon reports is VAT-inclusive, and that VAT is never
        # ours — it is collected on behalf of the tax authority. Dividing a
        # margin by a VAT-inclusive top line understates every percentage, and
        # by a different amount per region (UK 20%, SA 15%, AE 5%), so the
        # regions were not comparable either. sync.py already does this for
        # the daily aggregates; this path had not caught up.
        # USA has no VAT, so net_factor is 1.0 and nothing changes there.
        rev_net   = rev * net_factor
        cm        = rev_net - total_cgs - amz_fee - fulfill

        if product and product.title:
            pt, pack, var = split_title(product.title)
        else:
            pt, pack, var = (m.get('title') or '')[:30] or sku, '—', ''

        gk = (pt, pack)
        if gk not in grouped:
            grouped[gk] = {
                'pt': pt, 'pack': pack,
                'qty': 0, 'revenue': 0.0, 'revenue_net': 0.0, 'cgs': 0.0,
                'amz_fee': 0.0, 'fulfill': 0.0, 'cm': 0.0,
                'ppc': 0.0, '_sku_set': set(), '_asin_set': set(),
            }
        g = grouped[gk]
        g['qty']     += qty
        g['revenue'] += rev
        g['revenue_net'] += rev_net
        g['cgs']     += total_cgs
        g['amz_fee'] += amz_fee
        g['fulfill'] += fulfill
        g['cm']      += cm
        if sku:  g['_sku_set'].add(sku.upper())
        if asin: g['_asin_set'].add(asin.upper())

    # ── 5. Merge PPC from DB snapshots ───────────────────────────────────────
    # PPCProductSnapshot is SP-only (~10% of real spend, product-attributed).
    # Scale it to the SP campaign total — NOT the SP+SB+SD total, which would
    # inflate every SKU's per-product spend by the SB/SD portion.
    from apps.dashboard.models import PPCCampaignSnapshot
    asin_ppc = defaultdict(float)
    sku_ppc  = defaultdict(float)
    for snap in PPCProductSnapshot.objects.filter(
        marketplace=marketplace, date__gte=s_d, date__lte=e_d, campaign_type='sp',
    ).values('asin', 'sku', 'spend'):
        if snap['asin']: asin_ppc[snap['asin'].upper()] += float(snap['spend'] or 0)
        if snap['sku']:  sku_ppc[snap['sku'].upper()]   += float(snap['spend'] or 0)

    # SP-only campaign total — the correct denominator for product scaling
    sp_camp_total_ppc = sum(
        float(v or 0) for v in
        PPCCampaignSnapshot.objects.filter(
            marketplace=marketplace, date__gte=s_d, date__lte=e_d, campaign_type='sp',
        ).values_list('spend', flat=True)
    )
    prod_total_ppc = sum(asin_ppc.values()) or 0
    ppc_scale = (sp_camp_total_ppc / prod_total_ppc) if prod_total_ppc and sp_camp_total_ppc > prod_total_ppc else 1.0
    if ppc_scale > 1.0:
        asin_ppc = defaultdict(float, {k: v * ppc_scale for k, v in asin_ppc.items()})
        sku_ppc  = defaultdict(float, {k: v * ppc_scale for k, v in sku_ppc.items()})

    # ── Allocate SB/SD spend at product-group level ───────────────────────────
    # Was a hard-coded 10-entry copy of the canonical prefix map, which had
    # drifted: it was missing 19 prefixes (every UK/UAE/KSA/DE one, plus 12KTH,
    # 3KTH, 4DT), so SB/SD spend on those campaigns was silently dropped from
    # this page. Now reads the single CampaignPrefixMap config. The grouping
    # and every calculation below are unchanged — only the lookup source moved.
    from .prefix_map import get_prefix_map
    _CAMP_PREFIX_GROUP = get_prefix_map()
    sb_sd_rows = (
        PPCCampaignSnapshot.objects
        .filter(marketplace=marketplace, date__gte=s_d, date__lte=e_d,
                campaign_type__in=['sb', 'sd'])
        .values('campaign_name', 'campaign_type')
        .annotate(spend=Sum('spend'))
    )
    # Group SB/SD spend by (product_group, ad_type)
    _sb_sd_by_group = {}
    for _r in sb_sd_rows:
        _pfx   = (_r['campaign_name'] or '').split('-')[0].strip().upper()
        _group = _CAMP_PREFIX_GROUP.get(_pfx)
        if _group:
            if _group not in _sb_sd_by_group:
                _sb_sd_by_group[_group] = {'sb': 0.0, 'sd': 0.0}
            _ad_type = _r['campaign_type']
            _sb_sd_by_group[_group][_ad_type] = (
                _sb_sd_by_group[_group].get(_ad_type, 0.0) + float(_r['spend'] or 0)
            )

    for (pt, pack), g in grouped.items():
        sp = 0.0
        for sku_ in g['_sku_set']:
            sp += sku_ppc.get(sku_, 0)
        if not sp:
            for asin_ in g['_asin_set']:
                sp += asin_ppc.get(asin_, 0)
        _grp_sb_sd = _sb_sd_by_group.get((pt, pack), {})
        g['ppc']    = sp + _grp_sb_sd.get('sb', 0) + _grp_sb_sd.get('sd', 0)
        g['ppc_sp'] = sp
        g['ppc_sb'] = _grp_sb_sd.get('sb', 0)
        g['ppc_sd'] = _grp_sb_sd.get('sd', 0)

    # ── 6. Load targets ───────────────────────────────────────────────────────
    def _norm_key(pt, pack):
        digits = re.search(r'\d+', str(pack or ''))
        pn = digits.group(0) if digits else str(pack or '').strip().lower()
        return f"{(pt or '').strip().lower()}::{pn}"

    tar_rows = ProductTypePackMonthlyTarget.objects.filter(
        marketplace=marketplace, month=month_start,
    )
    tar_by_key = {}
    for t in tar_rows:
        monthly = float(t.revenue_target or 0)
        day_tar = (monthly / days_in_month) * period_days if days_in_month else 0
        k = _norm_key(t.product_type, t.pack_size)
        tar_by_key[k] = tar_by_key.get(k, 0) + day_tar

    # ── 7. Finalise rows ──────────────────────────────────────────────────────
    out = []
    for (pt, pack), g in sorted(grouped.items(), key=lambda x: -x[1]['revenue']):
        rev    = g['revenue']
        # Every ratio below divides by revenue EX-VAT, matching the margin
        # numerators. Falls back to gross if a group somehow has no net.
        rev_net = g.get('revenue_net') or rev
        cm     = g['cm']
        ppc    = g['ppc']       # SP + SB + SD total
        ppc_sp = g.get('ppc_sp', ppc)
        ppc_sb = g.get('ppc_sb', 0)
        ppc_sd = g.get('ppc_sd', 0)
        qty  = g['qty']
        gm   = cm - ppc
        tar  = tar_by_key.get(_norm_key(pt, pack), 0)
        out.append({
            'group':     (pt[:6].upper().replace(' ', '') + pack[:4].upper().replace(' ', ''))[:8],
            'groupName': f'{pt} · {pack}' if pack != '—' else pt,
            'qty':        qty,
            'revenue':    round(rev, 2),
            'revenueNet': round(rev_net, 2),
            'vatRate':    round(vat_rate * 100, 2),
            'cgs':        round(g['cgs'], 2),
            'amzFee':     round(g['amz_fee'], 2),
            'fulfill':    round(g['fulfill'], 2),
            'cm':         round(cm, 2),
            'cmPct':      round((cm / rev_net * 100) if rev_net else 0, 2),
            'arpu':       round((rev / qty) if qty else 0, 2),
            'ppcSpend':   round(ppc, 2),
            'spSpend':    round(ppc_sp, 2),
            'sbSpend':    round(ppc_sb, 2),
            'sdSpend':    round(ppc_sd, 2),
            'grossMargin': round(gm, 2),
            'gmPerUnit':  round((gm / qty) if qty else 0, 2),
            'gmPct':      round((gm / rev_net * 100) if rev_net else 0, 2),
            'cpa':        round((ppc / qty) if qty else 0, 2),
            'tacos':      round((ppc / rev_net * 100) if rev_net else 0, 2),
            'tarGmDay':   round(tar, 2),
            'gmMinusTar': round(gm - tar, 2),
        })

    return JsonResponse({
        'groups':      out,
        'period':      f'{s_d} → {e_d}',
        'period_days': period_days,
        'report_used': report_result.get('status'),
    })


@login_required
@permission_required('can_manage_cogs')
def cogs(request):
    upload_form    = COGSBulkUploadForm()
    manual_form    = COGSEntryForm()
    fba_form       = FBARateBulkUploadForm()
    upload_result  = None
    fba_result     = None

    if request.method == 'POST':
        action = request.POST.get('action')

        # ── FBA rate upload ─────────────────────────────────────────────────
        if action == 'upload_fba_rates':
            fba_form = FBARateBulkUploadForm(request.POST, request.FILES)
            if fba_form.is_valid():
                fba_result = _process_fba_rates_file(
                    request.FILES['file'],
                    overwrite=fba_form.cleaned_data['overwrite'],
                    user=request.user,
                )
                if fba_result['errors']:
                    messages.warning(request, f"FBA rates: {len(fba_result['errors'])} row errors. "
                                              f"{fba_result['created']} created, {fba_result['updated']} updated.")
                else:
                    messages.success(request, f"✓ FBA rates: {fba_result['created']} created, "
                                              f"{fba_result['updated']} updated.")
                AuditLog.objects.create(user=request.user, action='create',
                    resource='fba_rates:csv', ip_address=request.META.get('REMOTE_ADDR'))

                # Auto-resync each affected (marketplace, window) — re-aggregates
                # only the days where the new rate applies.
                if fba_result['affected_windows']:
                    from .sync import sync_window
                    today = date.today()
                    for mp, win_start, win_end in sorted(fba_result['affected_windows']):
                        end = min(win_end, today)
                        if end < win_start:
                            continue
                        try:
                            res = sync_window(mp, win_start, end, max_wait_seconds=60)
                            if res.get('status') in ('OK', 'CACHED', 'FRESH'):
                                messages.info(request,
                                    f"↻ Recomputed {res.get('days_written', 0)} days "
                                    f"({mp.upper()} {win_start} → {end}) with the new FBA rate.")
                            else:
                                messages.warning(request,
                                    f"FBA rate saved for {mp.upper()} {win_start} → {end}, but resync "
                                    f"returned {res.get('status')}. Re-run "
                                    f"`python manage.py backfill_history --start {win_start} --end {end} "
                                    f"--marketplace {mp}` when Amazon's report is ready.")
                        except Exception as exc:
                            logger.warning('FBA resync failed for %s %s→%s: %s', mp, win_start, end, exc)

        elif action == 'upload_csv':
            upload_form = COGSBulkUploadForm(request.POST, request.FILES)
            if upload_form.is_valid():
                upload_result = _process_cogs_csv(
                    request.FILES['csv_file'],
                    overwrite=upload_form.cleaned_data['overwrite'],
                    user=request.user,
                )
                if upload_result['errors']:
                    messages.warning(request, f"{len(upload_result['errors'])} row errors.")
                else:
                    messages.success(request, f"✓ {upload_result['created']} created, {upload_result['updated']} updated.")
                AuditLog.objects.create(user=request.user, action='create',
                    resource='cogs:csv', ip_address=request.META.get('REMOTE_ADDR'))

                # Re-aggregate ONLY the months touched by this upload.
                # Other months (e.g. April) stay untouched.
                if upload_result['affected']:
                    summary = _resync_months_after_cogs(upload_result['affected'], user=request.user)
                    for mp, m, days, status in summary:
                        if status in ('OK', 'CACHED', 'FRESH'):
                            messages.info(request,
                                f"↻ Recomputed {days} days for {mp.upper()} {m:%b %Y} with the new COGS.")
                        else:
                            messages.warning(request,
                                f"COGS saved for {mp.upper()} {m:%b %Y}, but resync returned {status}. "
                                "Run `python manage.py backfill_history --start {0} --end today --marketplace {1}` "
                                "to recompute when Amazon's report is ready.".format(m, mp))

        elif action == 'manual_entry':
            manual_form = COGSEntryForm(request.POST)
            if manual_form.is_valid():
                e = manual_form.save(commit=False)
                e.uploaded_by = request.user
                e.save()
                # Re-aggregate just this (marketplace, month)
                summary = _resync_months_after_cogs(
                    {(e.product.marketplace, e.month)}, user=request.user
                )
                msg = 'COGS entry saved.'
                for mp, m, days, status in summary:
                    if status in ('OK', 'CACHED', 'FRESH'):
                        msg += f' ↻ Recomputed {days} days for {mp.upper()} {m:%b %Y}.'
                messages.success(request, msg)
                return redirect('dashboard:cogs')

    recent     = COGSEntry.objects.select_related('product').order_by('-month', 'product__asin')[:100]
    recent_fba = FBAFeeRate.objects.select_related('product').order_by('-effective_from', 'product__asin')[:50]
    ctx = {
        'upload_form':   upload_form,
        'manual_form':   manual_form,
        'upload_result': upload_result,
        'recent':        recent,
        'fba_form':      fba_form,
        'fba_result':    fba_result,
        'recent_fba':    recent_fba,
    }
    return render(request, 'dashboard/cogs.html', ctx)


def _process_cogs_csv(f, overwrite=False, user=None):
    result = {'created': 0, 'updated': 0, 'errors': [], 'affected': set()}
    content = f.read().decode('utf-8-sig')
    reader  = csv.DictReader(io.StringIO(content))
    raw_headers = reader.fieldnames or []
    normalized = {h.strip().lower(): h for h in raw_headers if h}

    # Accept both legacy and unified business format:
    # SKU, ASIN, Region, Month, Cogs, FBA, ProductType, PackSize, Variant
    has_new_format = {'sku', 'asin', 'region', 'month', 'cogs', 'fba', 'producttype', 'packsize', 'variant'}.issubset(set(normalized.keys()))
    has_legacy_format = {'asin', 'marketplace', 'month', 'unit_cost'}.issubset(set(normalized.keys()))
    if not has_new_format and not has_legacy_format:
        result['errors'].append(
            'Missing columns. Required either '
            '[SKU, ASIN, Region, Month, Cogs, FBA, ProductType, PackSize, Variant] or '
            '[asin, marketplace, month, unit_cost].'
        )
        return result

    def cell(row, key, default=''):
        src = normalized.get(key.lower())
        return (row.get(src, default) if src else default)

    def normalize_marketplace(value):
        v = (value or '').strip().lower()
        aliases = {
            'us': 'usa', 'usa': 'usa', 'united states': 'usa',
            'ca': 'ca', 'canada': 'ca',
            'uk': 'uk', 'gb': 'uk', 'united kingdom': 'uk',
            'de': 'de', 'germany': 'de',
            'ae': 'ae', 'uae': 'ae', 'united arab emirates': 'ae',
            'sa': 'sa', 'ksa': 'sa', 'saudi arabia': 'sa',
        }
        return aliases.get(v, v)

    for i, row in enumerate(reader, start=2):
        try:
            asin = cell(row, 'asin').strip().upper()
            if not asin:
                raise ValueError('ASIN is required')
            mp_raw = cell(row, 'region') if has_new_format else cell(row, 'marketplace')
            mp = normalize_marketplace(mp_raw)
            ms = cell(row, 'month').strip()
            if len(ms) == 7:
                ms += '-01'
            month = datetime.strptime(ms, '%Y-%m-%d').date().replace(day=1)
            sku = cell(row, 'sku').strip()
            product_type = cell(row, 'producttype').strip() or cell(row, 'product_type').strip()
            pack_size = cell(row, 'packsize').strip() or cell(row, 'pack_size').strip()
            variant = cell(row, 'variant').strip()
            if has_new_format:
                missing = []
                if not sku:
                    missing.append('SKU')
                if not product_type:
                    missing.append('ProductType')
                if not pack_size:
                    missing.append('PackSize')
                if not variant:
                    missing.append('Variant')
                if missing:
                    raise ValueError(f"Missing required fields: {', '.join(missing)}")
            generated_title = ' - '.join([part for part in [product_type, pack_size, variant] if part]).strip()
            product, _ = Product.objects.get_or_create(
                asin=asin, marketplace=mp,
                defaults={'title': generated_title or asin, 'brand': 'Infinitee Xclusives'}
            )
            if sku and product.sku != sku:
                product.sku = sku
            if generated_title and (not product.title or product.title == product.asin):
                product.title = generated_title
            if product_type and not product.category:
                product.category = product_type
            if sku or generated_title or product_type:
                product.save(update_fields=['sku', 'title', 'category', 'updated_at'])
            defaults = {
                'unit_cost':     Decimal(str(cell(row, 'cogs', cell(row, 'unit_cost', 0)) or 0)),
                'shipping_cost': Decimal(str(cell(row, 'fba', cell(row, 'shipping_cost', 0)) or 0)),
                'duties_cost':   Decimal(str(row.get('duties_cost', 0) or 0)),
                'prep_cost':     Decimal(str(row.get('prep_cost', 0) or 0)),
                'other_cost':    Decimal(str(row.get('other_cost', 0) or 0)),
                'uploaded_by':   user,
            }
            if overwrite:
                _, created = COGSEntry.objects.update_or_create(
                    product=product, month=month, defaults=defaults)
            else:
                _, created = COGSEntry.objects.get_or_create(
                    product=product, month=month, defaults=defaults)
            if created:
                result['created'] += 1
            else:
                result['updated'] += 1
            result['affected'].add((mp, month))
        except Exception as e:
            result['errors'].append(f'Row {i}: {e}')
    return result


_GR_CONDITION = {'LN': 'Like New', 'VG': 'Very Good', 'GD': 'Good',
                 'AC': 'Acceptable', 'FR': 'Fair'}


@login_required
@permission_required('can_manage_cogs')
def cogs_missing_csv(request):
    """
    Download the SKUs that sold units in a month but have no COGS mapped,
    as a CSV pre-filled in the exact format the COGS bulk upload accepts
    (SKU, ASIN, Region, Month, Cogs, FBA, ProductType, PackSize, Variant).
    Fill in / adjust the Cogs column and upload it back on this page.

    Grade-and-resell SKUs (AMZN.GR.<parent>-<hash>-<cond>) are matched to
    their parent product: the suggested Cogs is the parent's per-unit cost
    and the ASIN is a unique pseudo-ASIN ('GR' + hash) — deliberately NOT
    the parent's real ASIN, which would re-key the parent product's SKU
    and break its COGS mapping.

    GET params: mp (marketplace), month (YYYY-MM)
    """
    import csv as _csv
    from django.http import HttpResponse
    from .cogs_recalc import month_cogs_unit_map
    from .models import UnifiedSkuUnits

    mp = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(mp):
        return JsonResponse({'error': 'forbidden'}, status=403)
    try:
        y, m = request.GET.get('month', '').split('-')[:2]
        month = date(int(y), int(m), 1)
    except (ValueError, IndexError):
        return JsonResponse({'error': 'month must be YYYY-MM'}, status=400)

    cost = month_cogs_unit_map(mp, month)
    missing = [u for u in UnifiedSkuUnits.objects.filter(marketplace=mp, month=month)
               if (u.order_units or 0) > 0 and u.sku.upper() not in cost]
    missing.sort(key=lambda u: -(u.order_units or 0))

    products = {(p.sku or '').upper(): p
                for p in Product.objects.filter(marketplace=mp) if p.sku}

    def match_parent(sku: str):
        """AMZN.GR.<parent-sku>-<hash>-<cond> → (parent Product, hash, cond)."""
        s = sku.upper()
        if not s.startswith('AMZN.GR.'):
            return None, '', ''
        parts = s[len('AMZN.GR.'):].split('-')
        for i in range(len(parts), 0, -1):
            cand = '-'.join(parts[:i])
            if cand in products:
                rest = parts[i:]
                cond = rest[-1] if rest else ''
                hsh = '-'.join(rest[:-1]) if len(rest) > 1 else ''
                return products[cand], hsh, cond
        return None, '', ''

    resp = HttpResponse(content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = (
        f'attachment; filename="missing_cogs_{mp}_{month:%Y-%m}.csv"')
    resp.write('\ufeff')     # BOM so Excel opens UTF-8 correctly
    w = _csv.writer(resp)
    w.writerow(['SKU', 'ASIN', 'Region', 'Month', 'Cogs', 'FBA',
                'ProductType', 'PackSize', 'Variant',
                'UnitsSold', 'ParentSKU', 'Note'])
    for u in missing:
        parent, hsh, cond = match_parent(u.sku)
        if parent:
            # Deterministic, collision-free pseudo-ASIN from the full SKU
            # (the raw hash segment can exceed 14 chars and collide on truncate)
            import hashlib
            pseudo_asin = 'GR' + hashlib.md5(
                u.sku.upper().encode()).hexdigest()[:14].upper()
            suggested = cost.get((parent.sku or '').upper(), '')
            w.writerow([u.sku, pseudo_asin, mp, f'{month:%Y-%m}',
                        suggested, 0,
                        parent.category or 'GradeResell', '1',
                        _GR_CONDITION.get(cond, cond or 'GR'),
                        u.order_units, parent.sku,
                        'Graded return — suggested Cogs = parent cost; adjust if needed'])
        else:
            w.writerow([u.sku, '', mp, f'{month:%Y-%m}',
                        '', 0, '', '', '',
                        u.order_units, '',
                        'No parent match — fill ASIN/ProductType/PackSize/Variant'])
    AuditLog.objects.create(user=request.user, action='export',
        resource=f'cogs:missing:{mp}:{month:%Y-%m}',
        ip_address=request.META.get('REMOTE_ADDR'))
    return resp


def _process_fba_rates_file(f, overwrite=True, user=None):
    """
    Parse a CSV or .xlsx upload of per-SKU FBA fees with effective dates.
    Required columns (case-insensitive): SKU, EffectiveFrom, FBAFee.
    Optional: ASIN, Region (defaults to product's marketplace).
    Returns {'created', 'updated', 'errors', 'affected_windows'}.
    `affected_windows` is a set of (marketplace, window_start, window_end) tuples
    so the caller can resync exactly the affected days.
    """
    result = {'created': 0, 'updated': 0, 'errors': [], 'affected_windows': set()}
    name = (getattr(f, 'name', '') or '').lower()

    rows = []
    try:
        if name.endswith('.xlsx'):
            from openpyxl import load_workbook
            wb = load_workbook(f, data_only=True)
            ws = wb.active
            sheet_rows = list(ws.iter_rows(values_only=True))
            if not sheet_rows:
                result['errors'].append('Sheet is empty.')
                return result
            headers = [str(h).strip() if h is not None else '' for h in sheet_rows[0]]
            for r in sheet_rows[1:]:
                rows.append({headers[i]: r[i] for i in range(len(headers)) if headers[i]})
        else:
            content = f.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(content))
            rows = list(reader)
    except Exception as exc:
        result['errors'].append(f'Could not read file: {exc}')
        return result

    if not rows:
        result['errors'].append('No data rows found.')
        return result

    # Build flexible header lookup
    sample_headers = {(k or '').strip().lower(): k for k in rows[0].keys() if k}

    def col(row, *aliases):
        for a in aliases:
            src = sample_headers.get(a.lower())
            if src is not None:
                v = row.get(src)
                if v is not None and str(v).strip() != '':
                    return v
        return None

    mp_aliases = {
        'us': 'usa', 'usa': 'usa', 'united states': 'usa',
        'ca': 'ca', 'canada': 'ca',
        'uk': 'uk', 'gb': 'uk', 'united kingdom': 'uk',
        'de': 'de', 'germany': 'de',
        'ae': 'ae', 'uae': 'ae',
        'sa': 'sa', 'ksa': 'sa',
    }

    # Track per-(marketplace, product) the set of effective_from dates so we
    # can compute the resync windows after we've inserted everything.
    from collections import defaultdict
    new_dates = defaultdict(set)  # (mp, product_id) → {effective_from, ...}

    for i, row in enumerate(rows, start=2):
        try:
            sku  = (str(col(row, 'sku')  or '')).strip()
            asin = (str(col(row, 'asin') or '')).strip().upper()
            mp_raw = col(row, 'region', 'marketplace', 'mp')
            mp_in  = (str(mp_raw or '')).strip().lower()
            mp_in  = mp_aliases.get(mp_in, mp_in)
            eff_raw = col(row, 'effectivefrom', 'effective_from', 'effective from', 'date')
            fee_raw = col(row, 'fbafee', 'fba_fee', 'fba fee', 'fee')

            if not sku and not asin:
                raise ValueError('Need SKU or ASIN')
            if not eff_raw or fee_raw in (None, ''):
                raise ValueError('Missing EffectiveFrom or FBAFee')

            # Parse effective_from (accept date or datetime or ISO string)
            if hasattr(eff_raw, 'date'):
                eff_dt = eff_raw.date() if hasattr(eff_raw, 'hour') else eff_raw
            else:
                eff_dt = datetime.strptime(str(eff_raw).strip()[:10], '%Y-%m-%d').date()

            try:
                fee = Decimal(str(fee_raw).replace(',', '').replace('$', ''))
            except Exception:
                raise ValueError(f'Invalid FBAFee: {fee_raw!r}')

            # Find product
            qs = Product.objects
            if mp_in:
                qs = qs.filter(marketplace=mp_in)
            product = (qs.filter(sku=sku).first() if sku
                       else qs.filter(asin=asin).first())
            if not product and asin:
                product = qs.filter(asin=asin).first()
            if not product:
                raise ValueError(f'Product not found (sku={sku} asin={asin} region={mp_in})')

            defaults = {'fba_fee_per_unit': fee, 'uploaded_by': user}
            if overwrite:
                _, created = FBAFeeRate.objects.update_or_create(
                    product=product, effective_from=eff_dt, defaults=defaults,
                )
            else:
                _, created = FBAFeeRate.objects.get_or_create(
                    product=product, effective_from=eff_dt, defaults=defaults,
                )
            if created:
                result['created'] += 1
            else:
                result['updated'] += 1
            new_dates[(product.marketplace, product.id)].add(eff_dt)
        except Exception as exc:
            result['errors'].append(f'Row {i}: {exc}')

    # Compute resync windows: for each (mp, product), the affected window for
    # a new effective_from = [that date, day_before_next_rate_for_same_product].
    # We collapse into per-marketplace windows so the resync is one report each.
    today_d = date.today()
    per_mp_window = {}   # mp → (min_start, max_end)
    for (mp, product_id), dates in new_dates.items():
        all_dates = sorted({
            d for d in FBAFeeRate.objects
                          .filter(product_id=product_id)
                          .values_list('effective_from', flat=True)
        })
        for d in dates:
            # Window for this effective_from: [d, day before next, or today]
            try:
                next_d = next(x for x in all_dates if x > d)
                end = next_d - timedelta(days=1)
            except StopIteration:
                end = today_d
            cur = per_mp_window.get(mp)
            if cur is None:
                per_mp_window[mp] = (d, end)
            else:
                per_mp_window[mp] = (min(cur[0], d), max(cur[1], end))
    result['affected_windows'] = {(mp, s, e) for mp, (s, e) in per_mp_window.items()}
    return result


def _resync_months_after_cogs(affected, user=None):
    """
    After COGS rows are uploaded for one or more (marketplace, month) pairs,
    rebuild only those months' DailyMetric rows. Other months stay untouched
    — uploading May COGS does NOT change April's report.
    Returns a list of (mp, month, days_written) tuples for messaging.
    """
    from .sync import sync_window
    today = date.today()
    summary = []
    for mp, month_first in sorted(affected):
        # Last day of that month, capped at today (don't fetch future days)
        if month_first.month == 12:
            month_last = date(month_first.year + 1, 1, 1) - timedelta(days=1)
        else:
            month_last = date(month_first.year, month_first.month + 1, 1) - timedelta(days=1)
        end = min(month_last, today)
        if end < month_first:
            continue
        try:
            res = sync_window(mp, month_first, end, max_wait_seconds=60)
            summary.append((mp, month_first, res.get('days_written', 0), res.get('status', '?')))
        except Exception as exc:
            logger.warning('COGS resync failed for %s %s: %s', mp, month_first, exc)
            summary.append((mp, month_first, 0, f'ERROR: {exc}'))
    return summary


@login_required
@permission_required('can_manage_targets')
def targets(request):
    today = date.today()
    start_month = date(today.year, 1, 1)
    planning_months = [date(today.year, m, 1) for m in range(1, 13)]

    def make_row_key(product_type: str, pack_size: str) -> str:
        key = f'{product_type}__{pack_size}'.lower()
        return re.sub(r'[^a-z0-9_]+', '_', key)

    def split_title_parts(title: str):
        parts = [p.strip() for p in (title or '').split('-') if p.strip()]
        product_type = parts[0] if parts else (title or '').strip() or 'Unknown'
        if len(parts) > 1:
            pack_size = parts[1]
        else:
            m = re.search(r'(\d+\s*-\s*pack|\d+\s*pack)', (title or '').lower())
            pack_size = m.group(1).replace(' ', '') if m else '-'
        return product_type, (pack_size or '-')

    # ── Handle POST ─────────────────────────────────────────────────────────
    if request.method == 'POST':
        action = request.POST.get('action', 'save')
        pk     = request.POST.get('pk')

        if action == 'delete' and pk:
            MonthlyTarget.objects.filter(pk=pk).delete()
            messages.success(request, 'Target deleted.')
            return redirect('dashboard:targets')

        # ── Bulk save: 12 months for one marketplace at once ─────────────────
        if action == 'bulk_save':
            mp = request.POST.get('bulk_marketplace', 'usa')
            saved = 0
            for month_date in planning_months:
                key = month_date.strftime('%Y-%m')
                rev  = request.POST.get(f'rev_{key}')
                ppc  = request.POST.get(f'ppc_{key}')
                tacos = request.POST.get(f'tacos_{key}')
                gm   = request.POST.get(f'gm_{key}')
                units = request.POST.get(f'units_{key}')
                if not rev:
                    continue
                MonthlyTarget.objects.update_or_create(
                    marketplace=mp, month=month_date,
                    defaults={
                        'revenue_target': rev,
                        'ppc_budget':     ppc or 0,
                        'tacos_target':   tacos or 15,
                        'gm_target':      gm or 25,
                        'units_target':   units or 0,
                        'created_by':     request.user,
                    }
                )
                saved += 1
            messages.success(request, f'✓ {saved} monthly targets saved for {mp.upper()}.')
            AuditLog.objects.create(user=request.user, action='update',
                resource=f'targets:bulk:{mp}:{today.year}',
                ip_address=request.META.get('REMOTE_ADDR'))
            return redirect(f'/dashboard/targets/?view=annual&mp={mp}')

        if action == 'bulk_save_products':
            mp = request.POST.get('bulk_marketplace', 'usa')
            products = Product.objects.filter(marketplace=mp).order_by('title', 'asin')
            group_pairs = {}
            for p in products:
                pt, ps = split_title_parts(p.title or p.asin)
                group_pairs[make_row_key(pt, ps)] = (pt, ps)
            saved = 0
            for row_key, (product_type, pack_size) in group_pairs.items():
                for month_date in planning_months:
                    key = month_date.strftime('%Y-%m')
                    raw_val = request.POST.get(f'rev_{row_key}_{key}', '').strip()
                    if raw_val == '':
                        continue
                    ProductTypePackMonthlyTarget.objects.update_or_create(
                        marketplace=mp,
                        product_type=product_type,
                        pack_size=pack_size,
                        month=month_date,
                        defaults={'revenue_target': raw_val, 'created_by': request.user},
                    )
                    saved += 1

            messages.success(request, f'✓ {saved} product targets saved for {mp.upper()} ({today:%b}–Dec).')
            return redirect(f'/dashboard/targets/?view=annual&mp={mp}')

        # ── Excel upload: annual product-level targets (12 months × N products) ──
        if action == 'upload_targets_xlsx' and request.FILES.get('targets_xlsx'):
            try:
                from openpyxl import load_workbook
            except ImportError:
                messages.error(request, 'openpyxl is not installed on the server. Run: pip install openpyxl')
                return redirect(f'/dashboard/targets/?view=annual&mp={request.POST.get("bulk_marketplace", "usa")}')

            mp_default = request.POST.get('bulk_marketplace', 'usa')
            f = request.FILES['targets_xlsx']
            try:
                wb = load_workbook(f, data_only=True)
            except Exception as e:
                messages.error(request, f'Could not read Excel file: {e}')
                return redirect(f'/dashboard/targets/?view=annual&mp={mp_default}')

            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                messages.error(request, 'Sheet has no data rows.')
                return redirect(f'/dashboard/targets/?view=annual&mp={mp_default}')

            # Header parsing (case-insensitive). Required: ProductType, PackSize.
            # Optional: Marketplace. Month columns: YYYY-MM, "Jan", "January", "Jan 2026", etc.
            raw_headers = [str(h).strip() if h is not None else '' for h in rows[0]]
            header_idx = {h.lower(): i for i, h in enumerate(raw_headers) if h}

            def col(*aliases):
                for a in aliases:
                    if a.lower() in header_idx:
                        return header_idx[a.lower()]
                return None

            i_pt   = col('producttype', 'product type', 'product_type')
            i_ps   = col('packsize',    'pack size',    'pack_size')
            i_mp   = col('marketplace', 'region', 'mp')
            if i_pt is None or i_ps is None:
                messages.error(request, 'Missing required columns: ProductType and PackSize.')
                return redirect(f'/dashboard/targets/?view=annual&mp={mp_default}')

            # Map remaining columns to months. Accept 2026-01, "Jan", "January", "Jan-2026", "Jan 2026"
            month_aliases = {
                'jan':1,'january':1,'feb':2,'february':2,'mar':3,'march':3,
                'apr':4,'april':4,'may':5,'jun':6,'june':6,'jul':7,'july':7,
                'aug':8,'august':8,'sep':9,'sept':9,'september':9,
                'oct':10,'october':10,'nov':11,'november':11,'dec':12,'december':12,
            }
            month_columns = {}  # column index → date(year, month, 1)
            for idx, h in enumerate(raw_headers):
                if not h or idx in (i_pt, i_ps, i_mp):
                    continue
                hl = h.strip().lower().replace(' ', '-').replace('_', '-')
                # YYYY-MM
                m = re.match(r'^(\d{4})-(\d{1,2})(?:-\d{1,2})?$', hl)
                if m:
                    y, mo = int(m.group(1)), int(m.group(2))
                    if 1 <= mo <= 12:
                        month_columns[idx] = date(y, mo, 1)
                        continue
                # Month name (with optional year)
                parts = re.split(r'[-/]', hl)
                month_name = parts[0]
                year_part  = parts[1] if len(parts) > 1 else None
                if month_name in month_aliases:
                    mo = month_aliases[month_name]
                    try:
                        y = int(year_part) if year_part and len(year_part) == 4 else today.year
                    except ValueError:
                        y = today.year
                    month_columns[idx] = date(y, mo, 1)

            if not month_columns:
                messages.error(request, 'No month columns detected. Use headers like "2026-01" or "Jan".')
                return redirect(f'/dashboard/targets/?view=annual&mp={mp_default}')

            mp_aliases = {
                'us':'usa','usa':'usa','united states':'usa',
                'ca':'ca','canada':'ca',
                'uk':'uk','gb':'uk','united kingdom':'uk',
                'de':'de','germany':'de',
                'ae':'ae','uae':'ae','united arab emirates':'ae',
                'sa':'sa','ksa':'sa','saudi arabia':'sa',
            }

            saved = 0
            row_errors = []
            for r_idx, row in enumerate(rows[1:], start=2):
                product_type = (str(row[i_pt]).strip() if row[i_pt] is not None else '')
                pack_size    = (str(row[i_ps]).strip() if row[i_ps] is not None else '')
                if not product_type or not pack_size:
                    continue
                if i_mp is not None and row[i_mp] not in (None, ''):
                    raw_mp = str(row[i_mp]).strip().lower()
                    mp = mp_aliases.get(raw_mp, raw_mp)
                else:
                    mp = mp_default

                for c_idx, month_date in month_columns.items():
                    val = row[c_idx]
                    if val is None or val == '':
                        continue
                    try:
                        amount = Decimal(str(val).replace(',', '').replace('$', ''))
                    except Exception:
                        row_errors.append(f'Row {r_idx} col {raw_headers[c_idx]}: invalid value "{val}"')
                        continue
                    ProductTypePackMonthlyTarget.objects.update_or_create(
                        marketplace=mp,
                        product_type=product_type,
                        pack_size=pack_size,
                        month=month_date,
                        defaults={'revenue_target': amount, 'created_by': request.user},
                    )
                    saved += 1

            if row_errors:
                messages.warning(request, f'Saved {saved} cells. {len(row_errors)} errors: ' + '; '.join(row_errors[:5]))
            else:
                messages.success(request, f'✓ Uploaded {saved} target cells from spreadsheet.')
            AuditLog.objects.create(user=request.user, action='upload',
                resource=f'targets:xlsx:{mp_default}',
                ip_address=request.META.get('REMOTE_ADDR'))
            return redirect(f'/dashboard/targets/?view=annual&mp={mp_default}')

        # ── Single save ───────────────────────────────────────────────────────
        instance = MonthlyTarget.objects.filter(pk=pk).first() if pk else None
        form = MonthlyTargetForm(request.POST, instance=instance)
        if form.is_valid():
            t = form.save(commit=False)
            if not instance:
                t.created_by = request.user
            t.save()
            messages.success(request, f'Target saved for {t.get_marketplace_display()} — {t.month:%B %Y}.')
            AuditLog.objects.create(user=request.user, action='update',
                resource=f'target:{t.marketplace}:{t.month}',
                ip_address=request.META.get('REMOTE_ADDR'))
            return redirect('dashboard:targets')
        # form invalid — fall through to render with errors
    else:
        form = MonthlyTargetForm()

    # ── View mode ────────────────────────────────────────────────────────────
    view_mode     = request.GET.get('view', 'annual')   # 'monthly' | 'annual'
    active_mp     = request.GET.get('mp', 'usa')

    # ── All existing targets ─────────────────────────────────────────────────
    all_targets = MonthlyTarget.objects.order_by('-month', 'marketplace')
    grouped = {}
    for t in all_targets:
        key = str(t.month)[:7]
        grouped.setdefault(key, []).append(t)

    # ── Annual planning grid (12 months × this marketplace) ──────────────────
    year = today.year
    annual_months = planning_months
    annual_targets_map = {}  # key: 'YYYY-MM' → MonthlyTarget or None

    for month_date in annual_months:
        existing = MonthlyTarget.objects.filter(
            marketplace=active_mp, month=month_date
        ).first()
        annual_targets_map[month_date.strftime('%Y-%m')] = existing

    products = Product.objects.filter(marketplace=active_mp).order_by('title', 'asin')
    p_targets = ProductTypePackMonthlyTarget.objects.filter(
        marketplace=active_mp,
        month__gte=start_month,
        month__year=today.year,
    )
    p_target_map = {}
    for t in p_targets:
        p_target_map.setdefault((t.product_type, t.pack_size), {})[t.month.strftime('%Y-%m')] = t

    def infer_pack_size(title: str) -> str:
        m = re.search(r'(\d+\s*-\s*pack|\d+\s*pack)', title.lower())
        if not m:
            return '-'
        return m.group(1).replace(' ', '')

    grouped_pairs = {}
    for p in products:
        pt, ps = split_title_parts(p.title or p.asin)
        grouped_pairs[(pt, ps)] = True
    for t in p_targets:
        grouped_pairs[(t.product_type, t.pack_size)] = True

    product_rows = []
    for product_type, pack_size in sorted(grouped_pairs.keys()):
        per_month_targets = p_target_map.get((product_type, pack_size), {})
        yearly_total = sum(
            float(t.revenue_target) for t in per_month_targets.values()
            if t and t.revenue_target is not None
        )
        row_key = make_row_key(product_type, pack_size)
        product_rows.append({
            'row_key': row_key,
            'product_type': product_type,
            'pack_size': pack_size,
            'targets_by_month': per_month_targets,
            'yearly_total': yearly_total,
        })

    month_totals = {}
    for month_date in annual_months:
        key = month_date.strftime('%Y-%m')
        month_totals[key] = float(
            ProductTypePackMonthlyTarget.objects.filter(
                marketplace=active_mp,
                month=month_date,
            ).aggregate(total=Sum('revenue_target'))['total'] or 0
        )
    grand_total = sum(month_totals.values())

    next_month = date(today.year + (1 if today.month == 12 else 0), 1 if today.month == 12 else today.month + 1, 1)
    this_month_total = month_totals.get(today.strftime('%Y-%m'), 0)
    next_month_total = month_totals.get(next_month.strftime('%Y-%m'), 0)
    tracking = {
        'this_month': today.replace(day=1),
        'next_month': next_month,
        'this_month_total': this_month_total,
        'next_month_total': next_month_total,
        'delta_to_next': next_month_total - this_month_total,
        'grand_total': grand_total,
    }

    marketplace_list = [
        ('usa', 'United States', '🇺🇸'),
        ('ca',  'Canada',        '🇨🇦'),
        ('uk',  'United Kingdom','🇬🇧'),
        ('de',  'Germany',       '🇩🇪'),
        ('ae',  'UAE',           '🇦🇪'),
        ('sa',  'Saudi Arabia',  '🇸🇦'),
    ]

    ctx = {
        'form':              form,
        'grouped':           grouped,
        'marketplace_list':  marketplace_list,
        'view_mode':         view_mode,
        'active_mp':         active_mp,
        'annual_months':     annual_months,
        'annual_targets_map': annual_targets_map,
        'product_rows':      product_rows,
        'month_totals':      month_totals,
        'tracking':          tracking,
        'year':              year,
        'today':             today,
    }
    return render(request, 'dashboard/targets.html', ctx)


@login_required
@permission_required('can_manage_catalog')
def catalog(request):
    mp = request.GET.get('mp', 'all')
    qs = Product.objects.order_by('marketplace', 'asin')
    if mp != 'all':
        qs = qs.filter(marketplace=mp)
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        data  = json.loads(request.body)
        pk    = data.get('pk')
        field = data.get('field')
        val   = data.get('value')
        allowed = ['status','sku','category','title']
        if field in allowed and pk:
            Product.objects.filter(pk=pk).update(**{field: val})
            return JsonResponse({'ok': True})
        return JsonResponse({'error': 'Invalid'}, status=400)

    grouped_catalog = {}
    for p in qs:
        title = p.title or ''
        parts = [s.strip() for s in title.split('-') if s.strip()]
        product_type = parts[0] if parts else (p.category or 'Misc')
        pack_size = parts[1] if len(parts) > 1 else 'Unspecified'
        variant = parts[2] if len(parts) > 2 else ''
        grouped_catalog.setdefault(product_type, {}).setdefault(pack_size, []).append({
            'sku': p.sku or p.asin,
            'asin': p.asin,
            'variant': variant or title or p.asin,
            'pk': p.pk,
        })

    grouped_rows = []
    for product_type in sorted(grouped_catalog.keys()):
        packs = grouped_catalog[product_type]
        pack_rows = []
        sku_count = 0
        for pack in sorted(packs.keys()):
            items = packs[pack]
            sku_count += len(items)
            pack_rows.append({'pack': pack, 'items': items})
        grouped_rows.append({
            'product_type': product_type,
            'pack_count': len(pack_rows),
            'sku_count': sku_count,
            'packs': pack_rows,
        })

    ctx = {
        'products': qs, 'mp': mp,
        'grouped_rows': grouped_rows,
        'allowed_marketplaces': _allowed_marketplaces(request.user),
    }
    return render(request, 'dashboard/catalog.html', ctx)


@login_required
@permission_required('can_manage_catalog')
def product_form(request, pk=None):
    instance = get_object_or_404(Product, pk=pk) if pk else None
    form = ProductForm(request.POST or None, instance=instance)
    if request.method == 'POST' and form.is_valid():
        p = form.save(commit=False)
        p.updated_by = request.user
        p.save()
        messages.success(request, f'Product {p.asin} saved.')
        return redirect('dashboard:catalog')
    return render(request, 'dashboard/product_form.html', {'form': form, 'instance': instance})


@login_required
@permission_required('can_generate_ai_summary')
def executive_summary(request):
    return render(request, 'dashboard/summary.html')


@login_required
@permission_required('can_generate_ai_summary')
def summary_stream(request):
    import requests as http

    mp     = request.GET.get('marketplace', 'usa')
    rev    = request.GET.get('revenue', 'N/A')
    units  = request.GET.get('units', 'N/A')
    ppc    = request.GET.get('ppc_spend', 'N/A')
    tacos  = request.GET.get('tacos', 'N/A')
    gm_pct = request.GET.get('gm_pct', 'N/A')
    cm_pct = request.GET.get('cm_pct', 'N/A')
    vs_tgt = request.GET.get('vs_target', 'N/A')
    acos   = request.GET.get('acos_avg', 'N/A')

    today  = date.today()
    target = MonthlyTarget.objects.filter(marketplace=mp, month=today.replace(day=1)).first()
    target_info = (
        f"Monthly revenue target ${target.revenue_target:,.0f}, TACoS target {target.tacos_target}%, PPC budget ${target.ppc_budget:,.0f}."
        if target else "No monthly targets set for this marketplace."
    )

    metrics_7d = DailyMetric.objects.filter(marketplace=mp, date__gte=today-timedelta(days=7)).order_by('date')
    trend_info = ""
    if metrics_7d.exists():
        rev_7d = [f"${float(m.revenue):,.0f}" for m in metrics_7d]
        trend_info = f"7-day revenue: {', '.join(rev_7d)}"

    # Priority: new AIProviderConfig(anthropic) → legacy AnthropicConfig → settings fallback
    from apps.amazon_api.models import AIProviderConfig as _AIProv
    ai_prov_cfg   = _AIProv.get_for('anthropic')
    anthropic_cfg = AnthropicConfig.get_active()

    if ai_prov_cfg:
        api_key = ai_prov_cfg.api_key
        model   = ai_prov_cfg.get_model() or settings.ANTHROPIC_MODEL
    elif anthropic_cfg:
        api_key = anthropic_cfg.api_key or settings.ANTHROPIC_API_KEY
        model   = anthropic_cfg.model or settings.ANTHROPIC_MODEL
    else:
        api_key = settings.ANTHROPIC_API_KEY
        model   = settings.ANTHROPIC_MODEL

    if not api_key:
        def _err():
            yield 'data: {"error": "Anthropic API key not configured. Go to API Config → Anthropic to add your key."}\n\n'
        return StreamingHttpResponse(_err(), content_type='text/event-stream')

    system = """You are a Senior Amazon E-Commerce Analyst for Infinitee Xclusives, a private-label Home & Kitchen brand (towels, bedsheets) across 6 Amazon marketplaces. Manufacturing: Pakistan/India, 45-day lead time.

Provide CEO-level analysis. No filler. Use exact numbers.

Structure response with EXACTLY these markdown sections:
## 🔑 Key Insight
## 📊 Performance Interpretation
## ✅ Recommended Actions
## ⚠️ Risks & Watch Items"""

    prompt = f"""Executive summary for {mp.upper()} marketplace — {today}:

**KPIs:** Revenue: {rev} | Units: {units} | PPC: {ppc} | TACoS: {tacos} | GM%: {gm_pct} | CM%: {cm_pct} | vs Target: {vs_tgt} | ACoS: {acos}
**Targets:** {target_info}
**Trend:** {trend_info or 'No historical data yet.'}"""

    AuditLog.objects.create(user=request.user, action='ai_summary',
        resource=f'summary:{mp}:{today}', ip_address=request.META.get('REMOTE_ADDR'))

    def generate():
        try:
            resp = http.post(
                'https://api.anthropic.com/v1/messages',
                headers={'x-api-key': api_key, 'anthropic-version': '2023-06-01', 'Content-Type': 'application/json'},
                json={'model': model, 'max_tokens': 1024, 'stream': True, 'system': system,
                      'messages': [{'role': 'user', 'content': prompt}]},
                stream=True, timeout=60,
            )
            # Surface HTTP-level errors immediately (400 = bad key/model/credits,
            # 401 = invalid key, 429 = rate limit, 529 = overloaded).
            # Without this check the error body is never emitted because it
            # doesn't contain SSE `data:` lines, leaving the UI stuck forever.
            if not resp.ok:
                try:
                    err_body = resp.json()
                    err_msg = err_body.get('error', {}).get('message') or resp.text[:300]
                except Exception:
                    err_msg = resp.text[:300] or f'HTTP {resp.status_code}'
                logger.error('Anthropic API error %s: %s', resp.status_code, err_msg)
                yield f'data: {json.dumps({"error": f"[{resp.status_code}] {err_msg}"})}\n\n'
                return

            for raw in resp.iter_lines():
                if not raw:
                    continue
                line = raw.decode('utf-8') if isinstance(raw, bytes) else raw
                if line.startswith('data:'):
                    payload = line[5:].strip()
                    if payload == '[DONE]':
                        yield 'data: [DONE]\n\n'
                        break
                    try:
                        evt = json.loads(payload)
                        if evt.get('type') == 'content_block_delta':
                            delta = evt.get('delta', {}).get('text', '')
                            if delta:
                                yield f'data: {json.dumps({"text": delta})}\n\n'
                        elif evt.get('type') == 'error':
                            # Streaming error event from Anthropic
                            err_msg = evt.get('error', {}).get('message', str(evt))
                            yield f'data: {json.dumps({"error": err_msg})}\n\n'
                            return
                    except json.JSONDecodeError:
                        pass
        except Exception as e:
            logger.error(f'Summary stream error: {e}')
            yield f'data: {json.dumps({"error": str(e)})}\n\n'

    r = StreamingHttpResponse(generate(), content_type='text/event-stream')
    r['Cache-Control'] = 'no-cache'
    r['X-Accel-Buffering'] = 'no'
    return r


@login_required
def export_csv(request):
    mp    = request.GET.get('mp', 'usa')
    start = request.GET.get('start', str(date.today() - timedelta(days=30)))
    end   = request.GET.get('end',   str(date.today()))
    if not request.user.can_access_marketplace(mp):
        return HttpResponse('Access denied', status=403)
    AuditLog.objects.create(user=request.user, action='export',
        resource=f'historical:{mp}', ip_address=request.META.get('REMOTE_ADDR'))
    qs = DailyMetric.objects.filter(marketplace=mp, date__gte=start, date__lte=end).order_by('date')
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="ix_{mp}_{start}_{end}.csv"'
    w = csv.writer(response)
    w.writerow(['Date','Marketplace','Revenue','Units','Orders','Sessions','CVR%',
                'PPC Spend','TACoS%','ACoS%','Gross Margin','GM%','CM','CM%'])
    for m in qs:
        rev_f  = float(m.revenue or 0)
        cm_f   = float(m.contribution_margin or 0)
        ppc_f  = float(m.ppc_spend or 0)
        gm_f   = cm_f - ppc_f        # GM = CM − PPC
        gm_pct_f = (gm_f / rev_f * 100) if rev_f else 0.0
        w.writerow([m.date,m.marketplace,m.revenue,m.units,m.orders,m.sessions,
                    f'{float(m.conversion_rate)*100:.2f}',m.ppc_spend,
                    f'{float(m.tacos)*100:.2f}',f'{float(m.acos)*100:.2f}',
                    f'{gm_f:.2f}',f'{gm_pct_f:.2f}',
                    m.contribution_margin,f'{float(m.cm_pct)*100:.2f}'])
    return response


@login_required
@login_required
@permission_required('can_view_dashboard')
def fba_fee_drift(request):
    """Page shell — JS calls /api/fba-fee-drift/ to populate the table + cards."""
    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        marketplace = _allowed_marketplaces(request.user)[0]

    ctx = {
        'marketplace':          marketplace,
        'allowed_marketplaces': _allowed_marketplaces(request.user),
        'can_export_cogs':      request.user.has_perm_flag('can_manage_cogs'),
    }
    return render(request, 'dashboard/fba_fee_drift.html', ctx)


@login_required
@permission_required('can_view_dashboard')
def api_fba_fee_drift(request):
    """
    JSON for the FBA Fee Drift page.

    Query params:
        mp            (str)   — marketplace, default 'usa'
        status        (csv)   — filter to a subset: critical,warn,ok,no_upload
        brand         (str)   — exact-match brand filter
        family        (str)   — exact-match product family filter
        min_impact    (float) — hide rows with dollar_impact below this
    """
    from .fba_drift import compute_drift, summarize
    from .models import SettlementReport

    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        return JsonResponse({'error': 'forbidden'}, status=403)

    # Window and volume floor are user-controllable: settlement data arrives in
    # arrears and varies by marketplace, so a fixed 14-day / 10-unit view hides
    # most of the catalogue. `show_all` drops the volume floor so SKUs with
    # little or no settlement volume are still listed (as 'no_actuals').
    try:
        window_days = max(1, min(int(request.GET.get('window_days') or 14), 365))
    except ValueError:
        window_days = 14
    show_all = (request.GET.get('show_all') or '').lower() in ('1', 'true', 'yes')

    rows = compute_drift(marketplace, window_days=window_days,
                         include_zero_volume=show_all)
    summary = summarize(rows)

    # Apply optional filters from query string
    statuses_csv = (request.GET.get('status') or '').strip().lower()
    statuses = [s for s in statuses_csv.split(',') if s] if statuses_csv else []
    brand_filter = (request.GET.get('brand') or '').strip()
    family_filter = (request.GET.get('family') or '').strip()
    try:
        min_impact = float(request.GET.get('min_impact') or 0)
    except ValueError:
        min_impact = 0.0

    filtered = []
    for r in rows:
        if statuses and r.status not in statuses:
            continue
        if brand_filter and r.brand != brand_filter:
            continue
        if family_filter and r.product_family != family_filter:
            continue
        if r.dollar_impact < min_impact:
            continue
        filtered.append(r.as_dict())

    # Distinct brand/family lists for the filter dropdowns
    brands   = sorted({r.brand          for r in rows if r.brand})
    families = sorted({r.product_family for r in rows if r.product_family})

    last_settle = (SettlementReport.objects
                    .filter(marketplace=marketplace, status='ok')
                    .order_by('-end_date', '-synced_at').first())

    # Coverage — so the page can distinguish "no drift" from "no data".
    from .models import SkuFeeActual
    cov = SkuFeeActual.objects.filter(marketplace=marketplace)
    coverage = {
        'has_settlement_data': cov.exists(),
        'latest_actual': (cov.order_by('-date')
                          .values_list('date', flat=True).first().isoformat()
                          if cov.exists() else None),
        'window_days': window_days,
        'show_all': show_all,
    }

    return JsonResponse({
        'marketplace':         marketplace,
        'rows':                filtered,
        'summary':             summary,
        'coverage':            coverage,
        'brand_options':       brands,
        'family_options':      families,
        'last_settlement':     {
            'end_date': last_settle.end_date.isoformat()
                        if last_settle and last_settle.end_date else None,
            'synced_at': last_settle.synced_at.isoformat()
                         if last_settle else None,
            'report_id': last_settle.report_id if last_settle else None,
        } if last_settle else None,
    })


@login_required
@permission_required('can_view_dashboard')
def fba_drift_export_xlsx(request):
    """Export the drift table exactly as it is on screen.

    Uses the same filters as api_fba_fee_drift so the file always matches the
    view — including the default "All SKUs", which shows uploaded vs actual for
    every SKU, not only the ones that drifted. Distinct from
    `fba_drift_corrected_xlsx`, which produces a re-upload template of
    suggested fees for drifting SKUs only.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    from .fba_drift import compute_drift

    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        return JsonResponse({'error': 'forbidden'}, status=403)

    try:
        window_days = max(1, min(int(request.GET.get('window_days') or 30), 365))
    except ValueError:
        window_days = 30

    # The export is deliberately COMPLETE: every SKU, every marketplace the
    # user can see, one sheet per region. Screen filters are for finding
    # something; a download is for working the whole list offline, so status /
    # brand / min-impact filters are NOT applied here and the volume floor is
    # dropped. (`fba_drift_corrected_xlsx` remains the narrow, drifting-only
    # re-upload template.)
    marketplaces = _allowed_marketplaces(request.user)
    if request.GET.get('mp') and request.GET.get('scope') == 'current':
        marketplaces = [marketplace]

    _LABEL = {'critical': 'Action needed', 'warn': 'Drifting',
              'ok': 'In line', 'no_upload': 'No uploaded fee',
              'no_actuals': 'No settlement data in window'}
    _MP_NAME = {'usa': 'USA', 'ca': 'Canada', 'uk': 'UK', 'de': 'Germany',
                'ae': 'UAE', 'sa': 'KSA'}
    headers = ['SKU', 'ASIN', 'Product', 'Brand', 'Family',
               'Uploaded fee', f'Actual avg ({window_days}d)', 'Actual latest',
               'Delta', 'Delta %', f'Units ({window_days}d)', '$ impact',
               'Status', 'Latest actual date']
    fill = PatternFill('solid', fgColor='232F3E')

    wb = Workbook()
    wb.remove(wb.active)
    totals = []
    for mp in marketplaces:
        rows = compute_drift(mp, window_days=window_days,
                             include_zero_volume=True)
        rows.sort(key=lambda r: -r.dollar_impact)
        ws = wb.create_sheet(title=_MP_NAME.get(mp, mp.upper())[:31])
        ws.append(headers)
        for r in rows:
            d = r.as_dict()
            ws.append([
                d.get('sku', ''), d.get('asin', ''),
                (d.get('product_name') or '')[:80],
                d.get('brand', ''), d.get('product_family', ''),
                round(float(d.get('uploaded_fee') or 0), 2),
                round(float(d.get('actual_fee_avg') or 0), 2),
                round(float(d.get('actual_fee_latest') or 0), 2),
                round(float(d.get('delta') or 0), 2),
                round(float(d.get('pct') or 0), 1),
                int(d.get('actual_units') or 0),
                round(float(d.get('dollar_impact') or 0), 2),
                _LABEL.get(d.get('status'), d.get('status', '')),
                d.get('actual_latest_date') or '',
            ])
        for c in ws[1]:
            c.font = Font(bold=True, color='FFFFFF')
            c.fill = fill
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:N{max(ws.max_row, 1)}'
        for col, w in zip('ABCDEFGHIJKLMN',
                          [22, 14, 42, 18, 18, 13, 15, 13, 10, 10, 12, 12, 22, 16]):
            ws.column_dimensions[col].width = w
        totals.append((mp, len(rows),
                       sum(1 for r in rows if r.status in ('warn', 'critical'))))

    # Leading summary sheet so the workbook explains itself.
    s = wb.create_sheet(title='Summary', index=0)
    s.append(['FBA fee drift — all SKUs'])
    s.append([f'Generated {date.today().isoformat()} · '
              f'{window_days}-day settlement window · every SKU included'])
    s.append([])
    s.append(['Region', 'SKUs', 'Drifting (🟡+🔴)'])
    for mp, n, drift in totals:
        s.append([_MP_NAME.get(mp, mp.upper()), n, drift])
    s.append([])
    s.append(['A region with 0 SKUs, or every row showing "No settlement data",',
              'has no settlement reports ingested yet — drift cannot be',
              'computed there until it does.'])
    s['A1'].font = Font(bold=True, size=13)
    for c in s[4]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = fill
    for col, w in zip('ABC', [34, 12, 18]):
        s.column_dimensions[col].width = w

    resp = HttpResponse(content_type=(
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))
    resp['Content-Disposition'] = (
        f'attachment; filename="fba-fee-drift-all-regions-'
        f'{date.today().isoformat()}.xlsx"')
    wb.save(resp)
    return resp


@login_required
@permission_required('can_manage_cogs')
def fba_drift_corrected_xlsx(request):
    """
    Generate an XLSX matching the existing COGS / FBA-rate upload format,
    pre-filled with one row per currently-drifting SKU (the "suggested fee"
    column is the actual 14-day weighted avg from settlements). The team
    reviews the file and re-uploads via the existing FBA upload page.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse('openpyxl is not installed. Run: pip install openpyxl', status=500)

    from .fba_drift import compute_drift
    from datetime import date as _d

    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        return HttpResponse('forbidden', status=403)

    rows = compute_drift(marketplace)
    # Only emit drifting SKUs — no point exporting rows that are already correct
    drift = [r for r in rows if r.status in ('warn', 'critical')]

    wb = Workbook()
    ws = wb.active
    ws.title = 'FBA Fee Corrections'

    headers = [
        'SKU', 'ASIN', 'Product name',
        'Current uploaded fee', 'Suggested fee (14d actual avg)',
        '∆ per unit', '∆ %', 'Units (14d)', '$ impact', 'Status',
        'New effective_from (edit before upload)',
    ]
    ws.append(headers)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='232F3E')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    today = _d.today().isoformat()
    for r in drift:
        ws.append([
            r.sku, r.asin, r.product_name,
            r.uploaded_fee, r.actual_fee_avg,
            r.delta, r.pct, r.actual_units, r.dollar_impact, r.status,
            today,
        ])

    # Auto-width-ish: widen the columns proportionally
    widths = [22, 14, 40, 12, 14, 10, 8, 10, 12, 11, 22]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = (
        f'attachment; filename=fba_fee_corrections_{marketplace}_{today}.xlsx')
    wb.save(resp)
    return resp


# ══════════════════════════════════════════════════════════════════════════════
# MANAGEMENT P&L
# ══════════════════════════════════════════════════════════════════════════════
from django.views.decorators.http import require_POST as _require_POST
from django.core.exceptions import PermissionDenied as _PermissionDenied


@login_required
@permission_required('can_view_dashboard')
def pnl_statement(request):
    """Page shell — JS calls /api/pnl-statement/ to populate."""
    from datetime import date as _d
    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        marketplace = _allowed_marketplaces(request.user)[0]
    ctx = {
        'marketplace':          marketplace,
        'allowed_marketplaces': _allowed_marketplaces(request.user),
        'today':                _d.today(),
        'can_edit':             request.user.has_perm_flag('can_manage_cogs'),
    }
    return render(request, 'dashboard/pnl_statement.html', ctx)


def _pnl_month_arg(request):
    """Parse ?month=YYYY-MM (default: current month)."""
    from datetime import date as _d
    raw = (request.GET.get('month') or '').strip()
    if raw:
        try:
            y, m = raw.split('-')[:2]
            return _d(int(y), int(m), 1)
        except (ValueError, IndexError):
            pass
    t = _d.today()
    return t.replace(day=1)


def _prev_month(m):
    from datetime import date as _d
    return (m.replace(day=1) - timedelta(days=1)).replace(day=1)


def _same_month_last_year(m):
    try:
        return m.replace(year=m.year - 1)
    except ValueError:
        return m


@login_required
@permission_required('can_view_dashboard')
def api_pnl_statement(request):
    """
    JSON for the Management P&L. Returns the current month + prior month +
    same-month-last-year columns, in the region's native currency, OR the
    global USD consolidation when ?scope=global.

    Query params:
        mp     (str)  — marketplace (ignored when scope=global)
        month  (str)  — YYYY-MM (default: current month)
        scope  (str)  — 'region' (default) | 'global'
    """
    from .pnl_engine import build_statement, build_consolidated, currency_for

    month   = _pnl_month_arg(request)
    scope   = (request.GET.get('scope') or 'region').lower()
    prev    = _prev_month(month)
    yoy     = _same_month_last_year(month)

    if scope == 'global':
        mps = _allowed_marketplaces(request.user)
        cur_st  = build_consolidated(month, mps)
        prev_st = build_consolidated(prev,  mps)
        yoy_st  = build_consolidated(yoy,   mps)
        currency = 'USD'
        marketplace = None
    else:
        marketplace = request.GET.get('mp', 'usa')
        if not request.user.can_access_marketplace(marketplace):
            return JsonResponse({'error': 'forbidden'}, status=403)
        cur_st  = build_statement(marketplace, month)
        prev_st = build_statement(marketplace, prev)
        yoy_st  = build_statement(marketplace, yoy)
        currency = currency_for(marketplace)

    # Index prior/yoy totals by line key for fast column lookup
    prev_by_key = {r['key']: r for r in prev_st['rows']}
    yoy_by_key  = {r['key']: r for r in yoy_st['rows']}

    rows = []
    for r in cur_st['rows']:
        p = prev_by_key.get(r['key'], {})
        y = yoy_by_key.get(r['key'], {})
        cur_total  = r['total']
        prev_total = p.get('total', 0)
        yoy_total  = y.get('total', 0)
        # MoM delta % (skip for pct/unit lines where it's not meaningful)
        mom = None
        if not r['is_pct'] and prev_total:
            mom = (cur_total - prev_total) / abs(prev_total) * 100
        rows.append({
            **r,
            'prev':     prev_total,
            'yoy':      yoy_total,
            'mom_pct':  round(mom, 1) if mom is not None else None,
        })

    # Signed per-section totals over the INPUT lines (auto+manual, using the
    # line's sign; computed subtotals excluded to avoid double counting).
    from .pnl_lines import LINE_BY_KEY as _LBK
    section_totals: dict = {}
    for r in rows:
        ln = _LBK.get(r['key'])
        if not ln or ln['source'] not in ('auto', 'manual') \
                or r['is_pct'] or r['is_unit'] or r['section'] == 'metrics':
            continue
        sgn = 1 if ln.get('sign') == '+' else -1
        section_totals[r['section']] = (
            section_totals.get(r['section'], 0.0) + sgn * r['total'])
    section_totals = {k: round(v, 2) for k, v in section_totals.items()}

    # Manual-entry metadata (note + invoice) so the edit modal can prefill.
    manual_meta: dict = {}
    if scope != 'global' and marketplace:
        from .models import MonthlyPnLEntry
        for e in MonthlyPnLEntry.objects.filter(
                marketplace=marketplace, month=month):
            manual_meta.setdefault(e.line_key, {})[e.channel] = {
                'note':        e.note or '',
                'invoice_url': e.invoice.url if e.invoice else None,
            }

    return JsonResponse({
        'scope':          scope,
        'marketplace':    marketplace,
        'currency':       currency,
        'month':          month.isoformat(),
        'prev_month':     prev.isoformat(),
        'yoy_month':      yoy.isoformat(),
        'has_settlement': cur_st.get('has_settlement', False),
        'missing_fx':     cur_st.get('missing_fx', []),
        'per_region':     cur_st.get('per_region', []),
        'rows':           rows,
        'section_totals': section_totals,
        'manual_meta':    manual_meta,
        'section_order':  ['revenue', 'cogs', 'amazon_fees', 'marketing',
                            'other_income', 'gross_margin', 'storage',
                            'operating_expenses', 'human_resource', 'net',
                            'metrics'],
    })


@login_required
@permission_required('can_manage_cogs')
@_require_POST
def save_pnl_entry(request):
    """
    Upsert one manual P&L line value. Accepts JSON body OR multipart form
    (multipart allows the optional invoice file):
        marketplace, month: 'YYYY-MM', channel: 'amazon'|'retail',
        line_key, amount (signed — negatives allowed), note?, invoice? (file)
    """
    from .models import MonthlyPnLEntry
    from .pnl_lines import LINE_BY_KEY

    invoice_file = None
    if request.content_type and request.content_type.startswith('multipart'):
        data = request.POST
        invoice_file = request.FILES.get('invoice')
    else:
        try:
            data = json.loads(request.body)
        except ValueError:
            return JsonResponse({'error': 'bad json'}, status=400)

    mp = data.get('marketplace', 'usa')
    if not request.user.can_access_marketplace(mp):
        raise _PermissionDenied

    line_key = data.get('line_key', '')
    ln = LINE_BY_KEY.get(line_key)
    if not ln or ln['source'] != 'manual':
        return JsonResponse({'error': f'not a manual line: {line_key}'}, status=400)

    try:
        y, m = str(data['month']).split('-')[:2]
        from datetime import date as _d
        month = _d(int(y), int(m), 1)
        amount = float(data.get('amount') or 0)   # signed — credits allowed
    except (ValueError, KeyError, IndexError):
        return JsonResponse({'error': 'bad month/amount'}, status=400)

    channel = data.get('channel', 'amazon')
    if channel not in ('amazon', 'retail'):
        channel = 'amazon'

    if invoice_file is not None and invoice_file.size > 15 * 1024 * 1024:
        return JsonResponse({'error': 'invoice too large (max 15 MiB)'}, status=400)

    defaults = {'amount': amount, 'updated_by': request.user,
                'note': str(data.get('note') or '')[:512]}
    obj, _created = MonthlyPnLEntry.objects.update_or_create(
        marketplace=mp, month=month, channel=channel, line_key=line_key,
        defaults=defaults,
    )
    if invoice_file is not None:
        obj.invoice = invoice_file
        obj.save(update_fields=['invoice', 'updated_at'])
    return JsonResponse({'status': 'ok',
                          'invoice_url': obj.invoice.url if obj.invoice else None})


@login_required
@permission_required('can_manage_cogs')
@_require_POST
def save_fx_rate(request):
    """Upsert a monthly FX rate. JSON: {month:'YYYY-MM', currency, rate_to_usd}."""
    from .models import MonthlyFXRate
    try:
        data = json.loads(request.body)
        y, m = str(data['month']).split('-')[:2]
        from datetime import date as _d
        month = _d(int(y), int(m), 1)
        currency = str(data['currency']).upper()[:4]
        rate = float(data['rate_to_usd'])
    except (ValueError, KeyError, IndexError):
        return JsonResponse({'error': 'bad payload'}, status=400)

    MonthlyFXRate.objects.update_or_create(
        month=month, currency=currency,
        defaults={'rate_to_usd': rate, 'updated_by': request.user},
    )
    return JsonResponse({'status': 'ok'})


@login_required
@permission_required('can_manage_cogs')
@_require_POST
def import_pnl_xlsx(request):
    """
    Import manual P&L lines from an uploaded Excel that matches the client's
    'P&L Summary' layout (column A = labels, one value column). Maps labels →
    line keys via pnl_lines.label_to_key and writes manual lines only.

    POST multipart: marketplace, month (YYYY-MM), channel, file
    """
    from .pnl_importer import import_pnl_excel_bytes

    mp = request.POST.get('marketplace', 'usa')
    if not request.user.can_access_marketplace(mp):
        raise _PermissionDenied
    f = request.FILES.get('file')
    if f is None:
        return JsonResponse({'status': 'failed', 'message': 'No file.'}, status=400)
    month_raw = request.POST.get('month', '')
    channel   = request.POST.get('channel', 'amazon')
    try:
        from datetime import date as _d
        y, m = month_raw.split('-')[:2]
        month = _d(int(y), int(m), 1)
    except (ValueError, IndexError):
        return JsonResponse({'status': 'failed', 'message': 'Bad month.'}, status=400)

    try:
        result = import_pnl_excel_bytes(
            file_bytes=f.read(), original_filename=f.name,
            marketplace=mp, month=month, channel=channel, user=request.user)
    except Exception as exc:
        return JsonResponse({'status': 'failed', 'message': str(exc)}, status=500)
    return JsonResponse(result)


@login_required
@permission_required('can_view_dashboard')
def cash_flow(request):
    """Cash Flow & Balance page shell — JS calls /api/cash-flow/."""
    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        marketplace = _allowed_marketplaces(request.user)[0]
    return render(request, 'dashboard/cash_flow.html', {
        'marketplace':          marketplace,
        'allowed_marketplaces': _allowed_marketplaces(request.user),
    })


@login_required
@permission_required('can_view_dashboard')
def api_cash_flow(request):
    """
    Cash flow from the Unified Transaction uploads.

    Per uploaded month: net proceeds (earned into the Amazon balance),
    payouts made (Transfers to bank), balance movement (earned − paid out),
    deferred (earned but not yet released by Amazon = Amazon still owes it),
    plus the individual payout events and a cumulative running movement.
    """
    from .models import SettlementLineActual, AmazonPayout

    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        return JsonResponse({'error': 'forbidden'}, status=403)

    cash_rows = SettlementLineActual.objects.filter(
        marketplace=marketplace, line_key__startswith='cash_',
    ).order_by('month')
    by_month: dict = {}
    for r in cash_rows:
        m = by_month.setdefault(r.month.isoformat(), {})
        m[r.line_key] = float(r.amount)

    months = []
    cumulative = 0.0
    for miso in sorted(by_month):
        c = by_month[miso]
        proceeds = c.get('cash_net_proceeds', 0.0)
        paid     = c.get('cash_payouts', 0.0)
        movement = proceeds - paid
        cumulative += movement
        months.append({
            'month':        miso,
            'net_proceeds': round(proceeds, 2),
            'payouts':      round(paid, 2),
            'movement':     round(movement, 2),
            'cumulative':   round(cumulative, 2),
            'deferred':     round(c.get('cash_deferred', 0.0), 2),
            'released':     round(c.get('cash_released', 0.0), 2),
        })

    payouts = [{
        'date':        p.payout_date.isoformat(),
        'month':       p.month.isoformat(),
        'amount':      float(p.amount),
        'description': p.description,
    } for p in AmazonPayout.objects.filter(
        marketplace=marketplace).order_by('-payout_date')[:200]]

    latest = months[-1] if months else None
    return JsonResponse({
        'marketplace': marketplace,
        'months':      months,
        'payouts':     payouts,
        'latest':      latest,
    })


@login_required
@permission_required('can_view_dashboard')
def mcf_orders(request):
    """MCF Orders page shell — JS calls /api/mcf/."""
    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        marketplace = _allowed_marketplaces(request.user)[0]
    return render(request, 'dashboard/mcf_orders.html', {
        'marketplace':          marketplace,
        'allowed_marketplaces': _allowed_marketplaces(request.user),
    })


def _mcf_filtered_qs(request):
    from .models import McfOrder
    marketplace = request.GET.get('mp', 'usa')
    qs = McfOrder.objects.filter(marketplace=marketplace)
    status = (request.GET.get('status') or '').strip()
    if status:
        qs = qs.filter(status=status)
    q = (request.GET.get('q') or '').strip()
    if q:
        qs = (qs.filter(seller_order_id__icontains=q)
              | qs.filter(displayable_order_id__icontains=q)
              | qs.filter(recipient_name__icontains=q))
    try:
        days = int(request.GET.get('days') or 90)
    except ValueError:
        days = 90
    from django.utils import timezone as _tz
    qs = qs.filter(received_date__gte=_tz.now() - timedelta(days=days))
    return marketplace, qs.order_by('-received_date')


@login_required
@permission_required('can_view_dashboard')
def api_mcf_orders(request):
    """MCF orders + tracking, filtered by ?status=&q=&days=."""
    marketplace, qs = _mcf_filtered_qs(request)
    if not request.user.can_access_marketplace(marketplace):
        return JsonResponse({'error': 'forbidden'}, status=403)
    from .models import McfOrder
    statuses = sorted(set(McfOrder.objects.filter(
        marketplace=marketplace).values_list('status', flat=True)))
    rows = [{
        'seller_order_id':      o.seller_order_id,
        'displayable_order_id': o.displayable_order_id,
        'status':               o.status,
        'received_date':        o.received_date.isoformat()[:16].replace('T', ' ')
                                 if o.received_date else '',
        'recipient':            o.recipient_name,
        'city':                 o.city, 'state': o.state,
        'units':                o.units,
        'items':                o.items,
        'packages':             o.packages,
        'synced_at':            o.synced_at.isoformat()[:16].replace('T', ' '),
    } for o in qs[:1000]]
    return JsonResponse({'marketplace': marketplace, 'rows': rows,
                          'statuses': statuses, 'count': len(rows)})


@login_required
@permission_required('can_view_dashboard')
@_require_POST
def api_mcf_sync(request):
    """Pull fresh MCF orders + tracking from Amazon. JSON: {marketplace, days}."""
    from apps.dashboard.management.commands.sync_mcf_orders import sync_mcf
    try:
        data = json.loads(request.body)
    except ValueError:
        data = {}
    mp = data.get('marketplace', 'usa')
    if not request.user.can_access_marketplace(mp):
        raise _PermissionDenied
    try:
        days = min(int(data.get('days') or 30), 120)
    except (TypeError, ValueError):
        days = 30
    try:
        res = sync_mcf(mp, days)
    except Exception as exc:
        return JsonResponse({'status': 'failed',
                              'error': f'{type(exc).__name__}: {exc}'}, status=500)
    return JsonResponse(res)


@login_required
@permission_required('can_view_dashboard')
def mcf_export_csv(request):
    """Download the filtered MCF orders as CSV — one row per package/tracking."""
    import csv as _csv
    from django.http import HttpResponse
    marketplace, qs = _mcf_filtered_qs(request)
    if not request.user.can_access_marketplace(marketplace):
        return JsonResponse({'error': 'forbidden'}, status=403)

    resp = HttpResponse(content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = (
        f'attachment; filename="mcf_tracking_{marketplace}.csv"')
    resp.write('\ufeff')     # BOM so Excel opens UTF-8 correctly
    w = _csv.writer(resp)
    w.writerow(['Order ID', 'Displayable Order ID', 'Status', 'Received',
                 'Recipient', 'City', 'State', 'SKUs', 'Units',
                 'Carrier', 'Tracking Number', 'Ship Date', 'ETA'])
    for o in qs:
        skus = '; '.join(f'{i["sku"]} x{i["qty"]}' for i in (o.items or []))
        base = [o.seller_order_id, o.displayable_order_id, o.status,
                o.received_date.strftime('%Y-%m-%d %H:%M') if o.received_date else '',
                o.recipient_name, o.city, o.state, skus, o.units]
        pkgs = o.packages or []
        if pkgs:
            for p in pkgs:
                w.writerow(base + [p.get('carrier', ''), p.get('tracking', ''),
                                    p.get('ship_date', ''), p.get('eta', '')])
        else:
            w.writerow(base + ['', '', '', ''])
    return resp


@login_required
@permission_required('can_manage_cogs')
@_require_POST
def api_recalc_cogs(request):
    """
    Recalculate stored COGS for one month from the currently-uploaded COGS
    rates. JSON body: {marketplace, month:'YYYY-MM'}. Refreshes the
    Management P&L, daily metrics/SKU snapshots, hourly snapshots and
    Campaign P&L for that month.
    """
    from .cogs_recalc import recalc_cogs
    try:
        data = json.loads(request.body)
        mp = data.get('marketplace', 'usa')
        y, m = str(data['month']).split('-')[:2]
        from datetime import date as _d
        month = _d(int(y), int(m), 1)
    except (ValueError, KeyError, IndexError):
        return JsonResponse({'error': 'bad payload'}, status=400)
    if not request.user.can_access_marketplace(mp):
        raise _PermissionDenied
    try:
        summary = recalc_cogs(mp, month)
    except Exception as exc:
        return JsonResponse({'status': 'failed',
                              'error': f'{type(exc).__name__}: {exc}'}, status=500)
    return JsonResponse({'status': 'ok', 'summary': summary})


@login_required
@permission_required('can_manage_cogs')
@_require_POST
def api_sync_pnl_month(request):
    """
    Rebuild one month's P&L from Settlement Flat File V2 — the single source
    of truth for the Management P&L.
    JSON: {marketplace, month:'YYYY-MM'}.

    WHAT CHANGED, AND WHY
        This used to request a Payments Date-Range Transaction report and fall
        back to the Finances API. Both wrote SettlementLineActual under their
        own source_note, alongside ingest_settlement_reports writing the same
        line keys — so one statement could blend three sources, and
        ingest_settlement_reports would then ADD its settlement amounts on top
        of whatever row it found. USA 2026-07 gross_sales reached $2,471,529
        against $1,361,984 rebuilt from the settlements themselves.

        It now calls rebuild_settlement_month, which recomputes the month from
        every settlement covering it and REPLACES the stored lines. Pressing
        the button twice gives the same answer as pressing it once.

    THE CURRENT MONTH
        The old Date-Range path had to request a whole calendar month, so
        asking for 1–31 Aug on the 25th returned nothing and the sync failed.
        There is no such request here: settlements are generated by Amazon on
        its own ~14-day cycle, and rows are attributed by posted-date. An open
        month therefore returns what has actually settled so far — correct by
        definition, and it fills in as later settlements arrive.
    """
    from datetime import date as _d
    from django.core.management import call_command
    from io import StringIO

    try:
        data = json.loads(request.body)
        mp = data.get('marketplace', 'usa')
        y, m = str(data['month']).split('-')[:2]
        month = _d(int(y), int(m), 1)
    except (ValueError, KeyError, IndexError):
        return JsonResponse({'error': 'bad payload'}, status=400)
    if not request.user.can_access_marketplace(mp):
        raise _PermissionDenied

    buf = StringIO()
    try:
        call_command('rebuild_settlement_month',
                     marketplace=mp, month=f'{month:%Y-%m}',
                     stdout=buf, stderr=buf)
    except Exception as exc:
        return JsonResponse(
            {'status': 'failed',
             'error': f'{type(exc).__name__}: {exc}',
             'detail': buf.getvalue()[-2000:]}, status=200)

    from .models import SettlementLineActual
    row = SettlementLineActual.objects.filter(
        marketplace=mp, month=month, line_key='gross_sales').first()
    return JsonResponse({
        'status': 'ok',
        'message': (
            f'{month:%B %Y} rebuilt from Settlement V2 — '
            f'gross sales {float(row.amount):,.2f} across {row.units:,} units.'
            if row else
            f'{month:%B %Y} rebuilt from Settlement V2 — no settled sales yet.'),
        'detail': buf.getvalue()[-2000:],
    })


@login_required
@permission_required('can_manage_cogs')
@_require_POST
def import_unified_txn(request):
    """
    Import the Seller Central Unified Transaction (Date-Range Transaction)
    report — the single authoritative posted-date source for the P&L. Parses
    revenue + all fees + COGS for the month and stores monthly line actuals.

    POST multipart: marketplace, month (YYYY-MM), file
    """
    from .unified_txn_importer import import_unified_csv_bytes

    mp = request.POST.get('marketplace', 'usa')
    if not request.user.can_access_marketplace(mp):
        raise _PermissionDenied
    f = request.FILES.get('file')
    if f is None:
        return JsonResponse({'status': 'failed', 'message': 'No file.'}, status=400)
    try:
        from datetime import date as _d
        y, m = request.POST.get('month', '').split('-')[:2]
        month = _d(int(y), int(m), 1)
    except (ValueError, IndexError):
        return JsonResponse({'status': 'failed', 'message': 'Bad month.'}, status=400)
    if f.size > 60 * 1024 * 1024:
        return JsonResponse({'status': 'failed', 'message': 'File too large (max 60 MiB).'}, status=400)

    try:
        result = import_unified_csv_bytes(
            file_bytes=f.read(), original_filename=f.name,
            marketplace=mp, month=month, user=request.user)
    except Exception as exc:
        return JsonResponse({'status': 'failed', 'message': f'{type(exc).__name__}: {exc}'}, status=500)
    return JsonResponse(result)


@permission_required('can_manage_cogs')
def fba_rates_template_xlsx(request):
    """Download an Excel template pre-filled with the user's products
    and example peak/off-peak effective dates. If FBAFeeRate rows already
    exist for a product, they're included so the user can edit + re-upload.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        return HttpResponse('openpyxl is not installed. Run: pip install openpyxl', status=500)

    mp = request.GET.get('mp', 'usa')
    today = date.today()

    # Typical Amazon US peak surcharge cycle: Oct 15 → Jan 14 (US/CA/MX).
    # Pick the next peak start at/after today, and the matching peak end.
    peak_start_year = today.year if today < date(today.year, 10, 15) else today.year
    peak_start = date(peak_start_year, 10, 15)
    peak_end_next = date(peak_start_year + 1, 1, 15)

    # Existing rates pre-fill (keyed by product so we don't duplicate rows)
    existing = {}
    for r in FBAFeeRate.objects.filter(product__marketplace=mp).select_related('product'):
        existing.setdefault(r.product_id, []).append(
            (r.effective_from, float(r.fba_fee_per_unit))
        )

    products = list(Product.objects.filter(marketplace=mp).order_by('sku', 'asin'))

    wb = Workbook()
    ws = wb.active
    ws.title = f'FBA Rates {mp.upper()}'

    headers = ['SKU', 'ASIN', 'Region', 'EffectiveFrom', 'FBAFee']
    ws.append(headers)
    bold = Font(bold=True)
    fill = PatternFill('solid', fgColor='F3F4F5')
    for c_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c_idx)
        cell.font = bold
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center')

    for p in products:
        existing_for_p = sorted(existing.get(p.id, []))
        if existing_for_p:
            # Use real history — user edits / re-uploads
            for eff, fee in existing_for_p:
                ws.append([p.sku or '', p.asin, mp, eff.isoformat(), fee])
        else:
            # Two example rows so the user sees the peak / off-peak pattern
            ws.append([p.sku or '', p.asin, mp, peak_start.isoformat(),    ''])  # peak start
            ws.append([p.sku or '', p.asin, mp, peak_end_next.isoformat(), ''])  # peak end

    # Column widths
    widths = [22, 14, 8, 14, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    # A small instructions sheet
    info = wb.create_sheet('How to fill')
    info.append(['How to use this template'])
    info.append([])
    info.append(['1. Each row = one FBA fulfilment rate that takes effect on EffectiveFrom.'])
    info.append(['2. The rate stays in effect until the next EffectiveFrom for the same SKU.'])
    info.append(['3. Two rows per SKU per year is typical:'])
    info.append(['     - Oct 15 (peak surcharge begins)'])
    info.append(['     - Jan 15 (peak ends, off-peak rate resumes)'])
    info.append([])
    info.append(['Columns:'])
    info.append(['  SKU            - your seller SKU'])
    info.append(['  ASIN           - Amazon ASIN (used if SKU is blank)'])
    info.append(['  Region         - usa / ca / uk / de / ae / sa'])
    info.append(['  EffectiveFrom  - YYYY-MM-DD (date the rate begins)'])
    info.append(['  FBAFee         - USD per unit (Amazon’s published rate)'])
    info.append([])
    info.append(['Tip: if you don’t upload any FBA rates, the dashboard falls back to'])
    info.append(['     the FBA column from your COGS upload (one rate per month).'])
    info.column_dimensions['A'].width = 75
    info['A1'].font = Font(bold=True, size=13)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="fba_rates_template_{mp}.xlsx"'
    )
    wb.save(response)
    return response


@login_required
@permission_required('can_manage_targets')
def targets_template_xlsx(request):
    """Download an Excel template pre-filled with the user's product groups
    for the active marketplace. Header row: ProductType, PackSize, Marketplace,
    then the 12 months of the current year (YYYY-MM)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        return HttpResponse('openpyxl is not installed. Run: pip install openpyxl', status=500)

    mp = request.GET.get('mp', 'usa')
    year = date.today().year

    months = [date(year, m, 1) for m in range(1, 13)]
    months_iso = [m.strftime('%Y-%m') for m in months]

    # Discover existing (product_type, pack_size) pairs from the catalog
    def split_title_parts(title: str):
        parts = [p.strip() for p in (title or '').split('-') if p.strip()]
        product_type = parts[0] if parts else (title or '').strip() or 'Unknown'
        pack_size    = parts[1] if len(parts) > 1 else '-'
        return product_type, pack_size

    pairs = set()
    for p in Product.objects.filter(marketplace=mp):
        pairs.add(split_title_parts(p.title or p.asin))
    for t in ProductTypePackMonthlyTarget.objects.filter(marketplace=mp):
        pairs.add((t.product_type, t.pack_size))

    # Load existing target values to pre-fill the cells
    existing = {}
    for t in ProductTypePackMonthlyTarget.objects.filter(
        marketplace=mp, month__year=year,
    ):
        existing[(t.product_type, t.pack_size, t.month.strftime('%Y-%m'))] = float(t.revenue_target or 0)

    wb = Workbook()
    ws = wb.active
    ws.title = f'Targets {mp.upper()} {year}'

    headers = ['ProductType', 'PackSize', 'Marketplace'] + months_iso
    ws.append(headers)
    bold = Font(bold=True)
    fill = PatternFill('solid', fgColor='F3F4F5')
    for c_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c_idx)
        cell.font = bold
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center')

    for product_type, pack_size in sorted(pairs):
        row = [product_type, pack_size, mp]
        for mi in months_iso:
            row.append(existing.get((product_type, pack_size, mi), ''))
        ws.append(row)

    # Column widths
    widths = [22, 14, 12] + [11] * 12
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="annual_targets_template_{mp}_{year}.xlsx"'
    )
    wb.save(response)
    return response


# ═══════════════════════════════════════════════════════════════════════════
# HOURLY PATTERNS — page + JSON API
# Reads from HourlyMetricSnapshot / HourlySkuSnapshot (populated by the cron).
# ═══════════════════════════════════════════════════════════════════════════
# Hourly Patterns metric registry. PPC is INTENTIONALLY excluded — Amazon
# publishes PPC at daily granularity only. Showing it here as $0 or as a
# revenue-share allocation would violate the page's "real hourly data" rule.
# Future "Estimated Ads Model" view would surface modelled PPC, clearly labelled.
_HOURLY_METRIC_FIELDS = {
    # metric_id → (cell_attr, label, format, requires_full_ads)
    # When requires_full_ads is True, days missing SB or SD are excluded from
    # the metric's aggregates (avoids "misleading aggregated values" per contract).
    'cm':        ('cm',        'Contribution Margin', 'currency', False),
    'revenue':   ('revenue',   'Revenue',             'currency', False),
    'units':     ('units',     'Units Sold',          'integer',  False),
    'orders':    ('orders',    'Orders',              'integer',  False),
    'cm_pct':    ('cm_pct',    'CM %',                'percent',  False),
    'gm':        ('gm',        'Gross Margin',        'currency', False),
    'gm_pct':    ('gm_pct',    'GM %',                'percent',  False),
    'ppc_sp':    ('ppc_sp',    'SP PPC Spend (real)', 'currency', False),
    'ppc_total': ('ppc_total', 'PPC Spend (SP+SB+SD)','currency', True),
    'ppc_pct':   ('ppc_pct',   'TACoS (PPC ÷ Rev)',   'percent',  True),
}

_WEEKDAY_NAMES = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']


@login_required
@permission_required('can_view_dashboard')
def hourly_patterns(request):
    """Page shell — JS calls /api/hourly-patterns/ to populate."""
    from datetime import date as _d
    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        marketplace = _allowed_marketplaces(request.user)[0]

    from .hourly_aggregator import list_product_groups
    ctx = {
        'marketplace':           marketplace,
        'allowed_marketplaces':  _allowed_marketplaces(request.user),
        'today':                 _d.today(),
        'metric_options':        [
            {'id': k, 'label': v[1], 'requires_full_ads': v[3]}
            for k, v in _HOURLY_METRIC_FIELDS.items()
        ],
        'group_options':         list_product_groups(marketplace),
    }
    return render(request, 'dashboard/hourly_patterns.html', ctx)


@login_required
@permission_required('can_view_dashboard')
def api_hourly_patterns(request):
    """
    JSON for the Hourly Patterns page — strict completeness contract.

    Display window:    [requested_end - days + 1, T-2]
    Ingestion window:  ends at T-1 (yesterday is OK to store but not show).

    Per the Hourly Analytics System contract:
      • Days that fail CORE completeness (SP-hourly + Orders) are excluded.
      • SB/SD allocation is uniform (daily ÷ 24), ONLY for days where
        sb_daily / sd_daily was successfully synced. Otherwise that source is
        None on every cell, and ppc_total / ppc_pct are None too.
      • PPC-dependent metrics (ppc_total, ppc_pct) are aggregated ONLY over days
        where all ad sources are complete — never partially.

    Query params:
      mp       — marketplace (default usa)
      days     — 7 | 14 | 30 (default 7)
      metric   — cm | revenue | units | orders | cm_pct | gm | gm_pct
                 | ppc_sp | ppc_total | ppc_pct  (default cm)
    """
    from datetime import date as _d, timedelta as _td
    from collections import defaultdict
    from .hourly_aggregator import build_hourly_cells, clamp_to_t_minus_2

    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        return JsonResponse({'error': 'forbidden'}, status=403)

    try:
        days = max(1, min(int(request.GET.get('days', 7)), 30))
    except ValueError:
        days = 7

    metric_id = request.GET.get('metric', 'cm')
    if metric_id not in _HOURLY_METRIC_FIELDS:
        metric_id = 'cm'
    cell_attr, label, fmt, requires_full_ads = _HOURLY_METRIC_FIELDS[metric_id]

    # Optional product-group filter: 'all' (default) | 'unallocated' | '<pt>|<pack>'
    group_id = request.GET.get('group', 'all') or 'all'

    today = _d.today()
    end   = clamp_to_t_minus_2(today - _td(days=1), today)   # T-2 cutoff
    start = end - _td(days=days - 1)

    # ── Build gated cells via aggregator ─────────────────────────────────────
    agg = build_hourly_cells(marketplace, start, end, group_id=group_id)

    # Empty-state: CORE missing for every day in the window
    if not agg.renderable_dates:
        return JsonResponse({
            'marketplace':      marketplace,
            'days':             days,
            'metric':           metric_id,
            'metric_label':     label,
            'metric_format':    fmt,
            'period_start':     start.isoformat(),
            'period_end':       end.isoformat(),
            'display_cutoff':   end.isoformat(),
            'heatmap':          [],
            'kpi':              {'best_hour': None, 'worst_hour': None,
                                 'best_dow':  None, 'worst_dow':  None},
            'hour_detail':      [],
            'weekday_pattern':  [],
            'completeness':     agg.completeness,
            'sb_sd_estimated':  False,
            'empty_state': {
                'reason':  'core_missing',
                'message': ('No days in the selected window have BOTH SP-hourly '
                            'and Orders successfully synced. Hourly Patterns '
                            'cannot render without Core completeness.'),
            },
        })

    # Some convenience flags
    renderable_set = set(agg.renderable_dates)

    # Helper: per-cell value for the selected metric
    def _cell_value(cell):
        v = getattr(cell, cell_attr)
        return v   # may be None for ppc_total / ppc_pct

    # When the metric requires full ads, only consider cells with a non-None
    # value AND days where both SB & SD were successful.
    sb_set = set(agg.sb_complete_days)
    sd_set = set(agg.sd_complete_days)
    full_ads_days = sb_set & sd_set
    def _eligible(cell) -> bool:
        if requires_full_ads:
            if cell.date not in full_ads_days:
                return False
            v = _cell_value(cell)
            return v is not None
        return True

    # ── HEATMAP (period-scoped, renderable days only) ────────────────────────
    heatmap = []
    cur = start
    while cur <= end:
        if cur in renderable_set:
            day_status = agg.ads_status[cur]
            cells_out = []
            for c in (agg.cells_by_date[cur] or [None] * 24):
                if c is None:
                    cells_out.append(None)
                    continue
                v = _cell_value(c)
                cell_dict = c.as_dict()
                # `value` is the metric the heatmap colours by. None when the
                # metric is unavailable for this cell (ppc_total without SB/SD).
                cell_dict['value'] = v
                cells_out.append(cell_dict)
            heatmap.append({
                'date':          cur.isoformat(),
                'weekday':       _WEEKDAY_NAMES[cur.weekday()],
                'status':        'rendered',
                'ads_status':    day_status,
                'sb_estimated':  day_status['sb_daily'],   # daily ÷ 24
                'sd_estimated':  day_status['sd_daily'],
                'hours':         cells_out,
            })
        else:
            heatmap.append({
                'date':       cur.isoformat(),
                'weekday':    _WEEKDAY_NAMES[cur.weekday()],
                'status':     'core_missing',
                'ads_status': day_ads_complete_lookup(marketplace, cur),
                'hours':      None,
            })
        cur += _td(days=1)

    # ── KPI: best/worst hour by metric, best/worst day-of-week ───────────────
    hour_totals = defaultdict(list)   # hour → [value, ...]
    day_totals  = defaultdict(float)  # date → total of metric
    dow_totals  = defaultdict(list)   # weekday → [day total]

    for cell in agg.iter_cells():
        if not _eligible(cell):
            continue
        v = _cell_value(cell)
        if v is None:
            continue
        v = float(v)
        hour_totals[cell.hour].append(v)
        day_totals[cell.date] += v

    for d, total in day_totals.items():
        dow_totals[d.weekday()].append(total)

    def _avg(values):
        return (sum(values) / len(values)) if values else 0

    hour_avgs = {h: _avg(v) for h, v in hour_totals.items()}
    dow_avgs  = {dow: _avg(v) for dow, v in dow_totals.items()}

    if hour_avgs:
        best_h, worst_h = max(hour_avgs, key=hour_avgs.get), min(hour_avgs, key=hour_avgs.get)
        best_hour  = {'hour': best_h,  'avg': round(hour_avgs[best_h], 2)}
        worst_hour = {'hour': worst_h, 'avg': round(hour_avgs[worst_h], 2)}
    else:
        best_hour = worst_hour = None

    if dow_avgs:
        best_d, worst_d = max(dow_avgs, key=dow_avgs.get), min(dow_avgs, key=dow_avgs.get)
        best_dow  = {'weekday': _WEEKDAY_NAMES[best_d],  'avg': round(dow_avgs[best_d], 2)}
        worst_dow = {'weekday': _WEEKDAY_NAMES[worst_d], 'avg': round(dow_avgs[worst_d], 2)}
    else:
        best_dow = worst_dow = None

    # ── HOUR DETAIL TABLE — 24 rows ─────────────────────────────────────────
    # All averages computed only over eligible cells (per-metric ad-completeness).
    # For each hour, we also report the number of days that contributed
    # ("n_days_used / n_days_renderable") so the UI can footnote partial coverage.
    by_hour_rev = defaultdict(list)
    by_hour_units = defaultdict(list)
    by_hour_orders = defaultdict(list)
    by_hour_cm = defaultdict(list)
    by_hour_cm_pct = defaultdict(list)
    by_hour_gm = defaultdict(list)
    by_hour_gm_pct = defaultdict(list)
    by_hour_ppc_sp = defaultdict(list)
    by_hour_ppc_total = defaultdict(list)   # only days with full ads
    by_hour_ppc_pct  = defaultdict(list)

    for cell in agg.iter_cells():
        by_hour_rev[cell.hour].append(cell.revenue)
        by_hour_units[cell.hour].append(cell.units)
        by_hour_orders[cell.hour].append(cell.orders)
        by_hour_cm[cell.hour].append(cell.cm)
        by_hour_cm_pct[cell.hour].append(cell.cm_pct)
        by_hour_gm[cell.hour].append(cell.gm)
        by_hour_gm_pct[cell.hour].append(cell.gm_pct)
        by_hour_ppc_sp[cell.hour].append(cell.ppc_sp)
        if cell.date in full_ads_days and cell.ppc_total is not None:
            by_hour_ppc_total[cell.hour].append(cell.ppc_total)
            if cell.ppc_pct is not None:
                by_hour_ppc_pct[cell.hour].append(cell.ppc_pct)

    n_full_ads = len(full_ads_days)
    n_render   = len(agg.renderable_dates)

    hour_detail = []
    for h in range(24):
        rev_list = by_hour_rev.get(h, [])
        hour_detail.append({
            'hour':              h,
            'avg_revenue':       round(_avg(rev_list), 2),
            'avg_units':         round(_avg(by_hour_units.get(h, [])), 2),
            'avg_orders':        round(_avg(by_hour_orders.get(h, [])), 2),
            'avg_cm':            round(_avg(by_hour_cm.get(h, [])), 2),
            'avg_cm_pct':        round(_avg(by_hour_cm_pct.get(h, [])), 2),
            'avg_gm':            round(_avg(by_hour_gm.get(h, [])), 2),
            'avg_gm_pct':        round(_avg(by_hour_gm_pct.get(h, [])), 2),
            'avg_ppc_sp':        round(_avg(by_hour_ppc_sp.get(h, [])), 2),
            'avg_ppc_total':     round(_avg(by_hour_ppc_total.get(h, [])), 2)
                                  if by_hour_ppc_total.get(h) else None,
            'avg_ppc_pct':       round(_avg(by_hour_ppc_pct.get(h, [])), 2)
                                  if by_hour_ppc_pct.get(h)  else None,
            'best_day_value':    round(max(rev_list), 2) if rev_list else 0,
            'worst_day_value':   round(min(rev_list), 2) if rev_list else 0,
            'n_days_used':       len(rev_list),
            'n_full_ads_days':   len(by_hour_ppc_total.get(h, [])),
        })

    # ── WEEKDAY × HOUR PATTERN (30-day rolling, gated by completeness) ───────
    wp_end   = clamp_to_t_minus_2(today - _td(days=1), today)
    wp_start = wp_end - _td(days=29)
    wp_agg   = build_hourly_cells(marketplace, wp_start, wp_end)
    wp_full_ads = set(wp_agg.sb_complete_days) & set(wp_agg.sd_complete_days)

    weekday_buckets = [[[] for _ in range(24)] for _ in range(7)]
    for cell in wp_agg.iter_cells():
        v = getattr(cell, cell_attr)
        if v is None:
            continue
        if requires_full_ads and cell.date not in wp_full_ads:
            continue
        weekday_buckets[cell.date.weekday()][cell.hour].append(float(v))

    weekday_pattern = []
    for dow in range(7):
        weekday_pattern.append({
            'weekday':      _WEEKDAY_NAMES[dow],
            'avg_per_hour': [round(_avg(weekday_buckets[dow][h]), 2) for h in range(24)],
        })

    return JsonResponse({
        'marketplace':       marketplace,
        'days':              days,
        'metric':            metric_id,
        'metric_label':      label,
        'metric_format':     fmt,
        'requires_full_ads': requires_full_ads,
        'period_start':      start.isoformat(),
        'period_end':        end.isoformat(),
        'display_cutoff':    end.isoformat(),   # always T-2
        'heatmap':           heatmap,
        'kpi': {
            'best_hour':  best_hour,
            'worst_hour': worst_hour,
            'best_dow':   best_dow,
            'worst_dow':  worst_dow,
        },
        'hour_detail':       hour_detail,
        'weekday_pattern':   weekday_pattern,
        'completeness':      agg.completeness,
        # quick flags for the UI
        'sb_sd_estimated':   bool(agg.sb_complete_days or agg.sd_complete_days),
        'n_days_renderable': n_render,
        'n_full_ads_days':   n_full_ads,
        # ── Group filter context ───────────────────────────────────────────
        'group':             group_id,
        # 4-card window-total strip — Revenue / Total PPC / TACoS / GM%
        'summary': (lambda: {
            'revenue':   round(sum(c.revenue   for c in agg.iter_cells()), 2),
            'ppc_total': round(sum(((c.ppc_total if c.ppc_total is not None else c.ppc_sp))
                                    for c in agg.iter_cells()), 2),
            'cm':        round(sum(c.cm        for c in agg.iter_cells()), 2),
            'gm':        round(sum((c.cm
                                     - (c.ppc_total if c.ppc_total is not None else c.ppc_sp))
                                    for c in agg.iter_cells()), 2),
        })(),
    })


def day_ads_complete_lookup(marketplace, d):
    """Wrapper so the heatmap can report ads_status even for hidden days."""
    from .completeness import day_ads_complete as _f
    return _f(marketplace, d)


@login_required
@permission_required('can_view_dashboard')
def api_hourly_patterns_sku(request):
    """
    Per-SKU drill-down for a single (marketplace, date, hour) cell.

    Obeys the same completeness contract as the page:
      • Day must pass CORE completeness (SP-hourly + Orders).
      • Day must be ≤ T-2 (display cutoff).
      • Otherwise return status='unavailable' so the UI can show
        "data not available" instead of an empty list.

    GET /dashboard/api/hourly-patterns/sku/?mp=usa&date=2026-06-22&hour=14
    """
    from datetime import date as _d, datetime as _dt, timedelta as _td
    from .models import HourlySkuSnapshot
    from .completeness import day_core_complete, day_ads_complete

    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        return JsonResponse({'error': 'forbidden'}, status=403)

    try:
        day  = _dt.strptime(request.GET.get('date', ''), '%Y-%m-%d').date()
        hour = int(request.GET.get('hour', 0))
        assert 0 <= hour <= 23
    except (ValueError, AssertionError):
        return JsonResponse({'error': 'bad date/hour'}, status=400)

    today   = _d.today()
    cutoff  = today - _td(days=2)

    # ── Completeness gates ──────────────────────────────────────────────────
    if day > cutoff:
        return JsonResponse({
            'marketplace': marketplace,
            'date':        day.isoformat(),
            'hour':        hour,
            'status':      'unavailable',
            'reason':      'after_display_cutoff',
            'message':     f'Data for {day} is past the T-2 display cutoff ({cutoff}).',
            'rows':        [],
        })

    if not day_core_complete(marketplace, day):
        return JsonResponse({
            'marketplace': marketplace,
            'date':        day.isoformat(),
            'hour':        hour,
            'status':      'unavailable',
            'reason':      'core_incomplete',
            'ads_status':  day_ads_complete(marketplace, day),
            'message':     'Day failed CORE completeness (SP-hourly + Orders).',
            'rows':        [],
        })

    # Optional group filter — drill-down inherits the page's product-group
    # selection so the user sees only that group's SKUs in the panel.
    from .hourly_aggregator import parse_group_id, _group_sku_set
    group_id = request.GET.get('group', 'all') or 'all'
    group_mode, group_payload = parse_group_id(group_id)

    qs = (HourlySkuSnapshot.objects
          .filter(marketplace=marketplace, date=day, hour=hour))
    if group_mode == 'group':
        qs = qs.filter(sku__in=_group_sku_set(marketplace, group_payload))
    elif group_mode == 'unallocated':
        # No SKU side for unallocated PPC — return empty list with a notice.
        return JsonResponse({
            'marketplace': marketplace, 'date': day.isoformat(), 'hour': hour,
            'status': 'ok', 'group': group_id,
            'ads_status': day_ads_complete(marketplace, day),
            'rows': [],
            'message': 'Unallocated PPC has no associated SKUs.',
        })
    rows = list(qs.order_by('-revenue')
                  .values('sku', 'asin', 'qty', 'revenue', 'contribution_margin'))

    out = []
    for r in rows:
        rev = float(r['revenue'])
        cm  = float(r['contribution_margin'])
        out.append({
            'sku':     r['sku'],
            'asin':    r['asin'],
            'qty':     int(r['qty']),
            'revenue': round(rev, 2),
            'cm':      round(cm, 2),
            'cm_pct':  round((cm / rev * 100) if rev else 0, 2),
        })
    return JsonResponse({
        'marketplace': marketplace,
        'date':        day.isoformat(),
        'hour':        hour,
        'status':      'ok',
        'ads_status':  day_ads_complete(marketplace, day),
        'rows':        out,
    })


# ═══════════════════════════════════════════════════════════════════════════
# MANUAL HOURLY CSV UPLOAD (Seller Central GUI exports — 14d window per file)
# ═══════════════════════════════════════════════════════════════════════════
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt   # see comment in view
from django.core.exceptions import PermissionDenied


@login_required
@permission_required('can_view_dashboard')
@require_POST
def upload_manual_hourly(request):
    """
    POST multipart/form-data with:
        marketplace : 'usa' (...)
        ad_type     : 'sp' | 'sb' | 'sd'
        file        : .csv from Seller Central's hourly export (≤ 14 days)

    Returns JSON:
        {status: 'ok'|'failed', message: str, rows_imported, days_covered,
         date_range_start, date_range_end, audit_id, parse_diagnostics:{found,missing}}
    """
    from .manual_hourly_importer import import_hourly_csv_bytes

    mp      = request.POST.get('marketplace', 'usa')
    ad_type = request.POST.get('ad_type', 'sp')
    f       = request.FILES.get('file')

    if not request.user.can_access_marketplace(mp):
        raise PermissionDenied

    if f is None:
        return JsonResponse({'status': 'failed',
                             'message': 'No file in request.'}, status=400)
    if ad_type not in ('sp', 'sb', 'sd'):
        return JsonResponse({'status': 'failed',
                             'message': f'Bad ad_type: {ad_type}'}, status=400)
    if f.size > 25 * 1024 * 1024:    # 25 MiB hard cap
        return JsonResponse({'status': 'failed',
                             'message': 'File too large (max 25 MiB).'}, status=400)

    try:
        result = import_hourly_csv_bytes(
            marketplace       = mp,
            ad_type           = ad_type,
            file_bytes        = f.read(),
            original_filename = f.name,
            uploaded_by_user  = request.user,
        )
    except Exception as e:
        return JsonResponse({'status': 'failed', 'message': str(e)}, status=400)

    pr = result['parse_result']
    drange = result['date_range']
    return JsonResponse({
        'status':           result['status'],
        'message':          result.get('error') or 'Upload succeeded.',
        'rows_in_file':     pr.rows_in_file,
        'rows_imported':    result['rows_imported'],
        'days_covered':     result['days_covered'],
        'date_range_start': drange[0].isoformat() if drange and drange[0] else None,
        'date_range_end':   drange[1].isoformat() if drange and drange[1] else None,
        'audit_id':         result['upload_audit_id'],
        'parse_diagnostics': {
            'columns_found':   pr.columns_found,
            'columns_missing': pr.columns_missing,
            'rows_skipped':    pr.rows_skipped,
            'errors':          pr.errors,
        },
    })


@login_required
@permission_required('can_view_dashboard')
def list_manual_hourly_uploads(request):
    """List recent manual uploads for the audit panel on the Hourly Patterns page."""
    from .models import AdsManualHourlyUpload
    mp = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(mp):
        return JsonResponse({'error': 'forbidden'}, status=403)

    qs = AdsManualHourlyUpload.objects.filter(marketplace=mp).order_by('-uploaded_at')[:30]
    out = [{
        'id':                u.id,
        'ad_type':           u.ad_type,
        'filename':          u.original_filename,
        'uploaded_by':       getattr(u.uploaded_by, 'email', '?') if u.uploaded_by else '?',
        'uploaded_at':       u.uploaded_at.isoformat(),
        'date_range_start':  u.date_range_start.isoformat() if u.date_range_start else None,
        'date_range_end':    u.date_range_end.isoformat()   if u.date_range_end   else None,
        'days_covered':      u.days_covered,
        'rows_in_file':      u.rows_in_file,
        'rows_imported':     u.rows_imported,
        'status':            u.status,
        'error_message':     u.error_message,
    } for u in qs]
    return JsonResponse({'marketplace': mp, 'uploads': out})


# ═════════════════════════════════════════════════════════════════════════════
# PHASE 1 — CAMPAIGN PERFORMANCE CENTER
#
# Page    : /dashboard/campaigns/        (campaigns_list — server shell)
# API     : /dashboard/api/campaigns/    (api_campaigns_list — JSON for table)
#
# Data sources:
#   • CampaignProfitDaily  → pre-aggregated daily P&L per campaign (hot path).
#   • PPCCampaignSnapshot  → "Today" view (when CampaignProfitDaily isn't built
#                             for today — search-term/advertised-product reports
#                             are T-1 minimum).
#   • Campaign (dim table) → campaign name + type + brand + product family;
#                             joined on the fly so the parsed dim doesn't fall
#                             behind. If a campaign_id has no Campaign row,
#                             fall back to PPCCampaignSnapshot.campaign_name.
# ═════════════════════════════════════════════════════════════════════════════

# Window-period selector — matches Daily Dashboard convention.
_CAMPAIGN_PERIODS = {
    # id      label           is_today    days_back   anchor_offset_days
    'today':       ('Today',       True,    1,   0),
    'yesterday':   ('Yesterday',   False,   1,   1),
    '7d':          ('Last 7D',     False,   7,   1),
    '30d':         ('Last 30D',    False,   30,  1),
    'mtd':         ('Month-to-date', False, None, 1),  # special-cased
}


def _resolve_campaign_period(period_id: str, today: date) -> tuple[date, date, bool]:
    """
    period_id → (start_date, end_date, is_today_view).

    Today view ends at today; non-today periods cap at T-1.
    """
    if period_id not in _CAMPAIGN_PERIODS:
        period_id = '7d'
    label, is_today, days_back, anchor_offset = _CAMPAIGN_PERIODS[period_id]

    if is_today:
        return today, today, True
    if period_id == 'mtd':
        end = today - timedelta(days=anchor_offset)
        start = end.replace(day=1)
        return start, end, False
    end = today - timedelta(days=anchor_offset)
    start = end - timedelta(days=days_back - 1)
    return start, end, False


@login_required
@permission_required('can_view_dashboard')
def campaigns_list(request):
    """Page shell — JS calls /api/campaigns/ to populate the table."""
    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        marketplace = _allowed_marketplaces(request.user)[0]

    ctx = {
        'marketplace':          marketplace,
        'allowed_marketplaces': _allowed_marketplaces(request.user),
        'period_options': [
            {'id': pid, 'label': v[0]} for pid, v in _CAMPAIGN_PERIODS.items()
        ],
        'today':                date.today(),
    }
    return render(request, 'dashboard/campaigns.html', ctx)


@login_required
@permission_required('can_view_dashboard')
def api_campaigns_list(request):
    """
    JSON for the Campaign Performance Center table.

    Query params:
        mp           — marketplace (default usa)
        period       — today | yesterday | 7d | 30d | mtd (default 7d)
        campaign_type — sp | sb | sd | all (default all)
        status       — enabled | paused | all (default all)
        brand        — exact-match brand filter; '' means no filter

    Response envelope:
        {
          'marketplace':   str,
          'period':        {'id', 'label', 'start', 'end'},
          'is_today':      bool,
          'data_source':   'live_summary' | 'profit_daily',
          'today_warning': str | null,   # banner text for Today view
          'kpi': {
            'spend', 'ad_revenue', 'gross_profit', 'margin_pct',
            'tacos', 'acos', 'roas',
            'campaign_count',
          },
          'rows': [{
            'campaign_id', 'campaign_name', 'campaign_type', 'brand',
            'spend', 'ad_revenue', 'orders', 'units', 'impressions', 'clicks',
            'profit', 'margin_pct', 'tacos', 'contribution_to_profit_pct',
            'acos', 'roas', 'ctr', 'cpc', 'cvr',
            'attribution_coverage_pct',
            'sku_count_attributed',
          }, ...],
        }
    """
    from .models import (
        CampaignProfitDaily, PPCCampaignSnapshot, Campaign, DailyMetric,
    )

    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        return JsonResponse({'error': 'forbidden'}, status=403)

    period_id = (request.GET.get('period') or '7d').lower()
    if period_id not in _CAMPAIGN_PERIODS:
        period_id = '7d'

    campaign_type = (request.GET.get('campaign_type') or 'all').lower()
    if campaign_type not in ('sp', 'sb', 'sd', 'all'):
        campaign_type = 'all'

    today = date.today()
    start, end, is_today = _resolve_campaign_period(period_id, today)

    # ── Window-level account revenue (drives TACoS) ─────────────────────────
    window_revenue = (DailyMetric.objects.filter(
        marketplace=marketplace, date__gte=start, date__lte=end
    ).aggregate(total=Sum('revenue'))['total'] or Decimal('0'))

    # ── Dim lookup: campaign_id → {name, type, brand, product_family} ───────
    dim_qs = Campaign.objects.filter(marketplace=marketplace)
    if campaign_type != 'all':
        dim_qs = dim_qs.filter(campaign_type=campaign_type)
    dim_map = {c.campaign_id: {
        'campaign_name':  c.campaign_name,
        'campaign_type':  c.campaign_type,
        'brand':          c.brand,
        'product_family': c.product_family,
        'portfolio_id':   c.portfolio_id,
    } for c in dim_qs.only('campaign_id', 'campaign_name', 'campaign_type',
                            'brand', 'product_family', 'portfolio_id')}

    # ── Aggregate per-campaign rows ─────────────────────────────────────────
    if is_today:
        # Today: CampaignProfitDaily isn't computed yet; fall back to
        # PPCCampaignSnapshot (live SUMMARY). Profit columns are unavailable
        # because advertised-product report is T-1 minimum — clearly flag.
        rows_by_campaign = _aggregate_today_campaigns(
            marketplace, today, campaign_type, dim_map)
        data_source = 'live_summary'
        today_warning = ('Detailed advertising analytics are available through T-1. '
                         "Today's view contains only currently available PPC metrics; "
                         'profit / margin / TACoS will appear once yesterday\'s data lands.')
    else:
        rows_by_campaign = _aggregate_window_campaigns(
            marketplace, start, end, campaign_type, dim_map)
        data_source = 'profit_daily'
        today_warning = None

    # ── KPI strip (window totals) ───────────────────────────────────────────
    # Coerce None → 0 for sum-safety; today rows have profit=None which
    # otherwise breaks sum(). Display layer still surfaces None as "—".
    _z = lambda v: v if v is not None else Decimal('0')
    sum_spend  = sum(_z(r['spend'])      for r in rows_by_campaign)
    sum_rev    = sum(_z(r['ad_revenue']) for r in rows_by_campaign)
    sum_profit = sum(_z(r['profit'])     for r in rows_by_campaign)
    sum_imp    = sum((r['impressions'] or 0) for r in rows_by_campaign)
    sum_clicks = sum((r['clicks']      or 0) for r in rows_by_campaign)
    sum_orders = sum((r['orders']      or 0) for r in rows_by_campaign)

    kpi = {
        'spend':           round(float(sum_spend), 2),
        'ad_revenue':      round(float(sum_rev),   2),
        'gross_profit':    round(float(sum_profit), 2) if not is_today else None,
        'margin_pct':      round(float(sum_profit / sum_rev * 100), 2)
                            if (sum_rev > 0 and not is_today) else None,
        'tacos':           round(float(sum_spend / window_revenue * 100), 2)
                            if window_revenue > 0 else None,
        'acos':            round(float(sum_spend / sum_rev * 100), 2)
                            if sum_rev > 0 else None,
        'roas':            round(float(sum_rev / sum_spend), 2)
                            if sum_spend > 0 else None,
        'impressions':     int(sum_imp),
        'clicks':          int(sum_clicks),
        'orders':          int(sum_orders),
        'campaign_count':  len(rows_by_campaign),
        'window_revenue':  round(float(window_revenue), 2),
    }

    # ── Per-row derivations: TACoS, contribution-to-profit %, etc. ──────────
    # Done in a second pass so the totals (denominators) are already known.
    total_window_profit = sum_profit if not is_today else Decimal('0')
    for r in rows_by_campaign:
        r['tacos'] = (round(float(r['spend'] / window_revenue * 100), 2)
                      if window_revenue > 0 else None)
        if not is_today and total_window_profit > 0:
            r['contribution_to_profit_pct'] = round(
                float(r['profit'] / total_window_profit * 100), 2)
        else:
            r['contribution_to_profit_pct'] = None
        # Round Decimal → float at the boundary (None passes through unchanged).
        for k in ('spend', 'ad_revenue', 'profit', 'cogs', 'fees'):
            if k in r and r[k] is not None:
                r[k] = round(float(r[k]), 2)
        for k in ('margin_pct', 'acos', 'roas', 'ctr', 'cpc', 'cvr',
                  'attribution_coverage_pct'):
            if k in r and r[k] is not None:
                r[k] = round(float(r[k]), 4)

    # Sort by profit DESC (today: by spend DESC since profit is null)
    sort_key = 'spend' if is_today else 'profit'
    rows_by_campaign.sort(key=lambda r: r[sort_key] or 0, reverse=True)

    return JsonResponse({
        'marketplace':   marketplace,
        'period': {
            'id':    period_id,
            'label': _CAMPAIGN_PERIODS[period_id][0],
            'start': start.isoformat(),
            'end':   end.isoformat(),
        },
        'is_today':      is_today,
        'data_source':   data_source,
        'today_warning': today_warning,
        'kpi':           kpi,
        'rows':          rows_by_campaign,
    })


def _aggregate_today_campaigns(marketplace, today_d, campaign_type, dim_map):
    """
    "Today" view: SUM(PPCCampaignSnapshot) for today (refreshed every 30 min by
    sync_today_ppc). Profit / margin / TACoS are not computed — return None.
    """
    from .models import PPCCampaignSnapshot

    qs = PPCCampaignSnapshot.objects.filter(marketplace=marketplace, date=today_d)
    if campaign_type != 'all':
        qs = qs.filter(campaign_type=campaign_type)

    rows = []
    for s in qs.values(
        'campaign_id', 'campaign_name', 'campaign_type', 'state',
        'impressions', 'clicks', 'spend',
        'sales_7d', 'orders_7d', 'units_7d',
        'acos', 'roas', 'ctr', 'cvr', 'cpc',
    ):
        cid = s['campaign_id']
        dim = dim_map.get(cid, {})
        impr   = int(s['impressions'] or 0)
        clicks = int(s['clicks']      or 0)
        spend  = Decimal(s['spend']    or 0)
        sales  = Decimal(s['sales_7d'] or 0)
        orders = int(s['orders_7d']    or 0)

        rows.append({
            'campaign_id':    cid,
            'campaign_name':  dim.get('campaign_name') or s['campaign_name'],
            'campaign_type':  dim.get('campaign_type') or s['campaign_type'],
            'brand':          dim.get('brand', ''),
            'state':          s['state'],
            'impressions':    impr,
            'clicks':         clicks,
            'spend':          spend,
            'ad_revenue':     sales,
            'orders':         orders,
            'units':          int(s['units_7d'] or 0),
            'profit':         None,   # T-1 only
            'margin_pct':     None,
            'attribution_coverage_pct': None,
            'sku_count_attributed':     None,
            'acos':  float(s['acos']) * 100 if s['acos'] else None,
            'roas':  float(s['roas']) if s['roas'] else None,
            'ctr':   float(s['ctr'])  * 100 if s['ctr']  else None,
            'cvr':   float(s['cvr'])  * 100 if s['cvr']  else None,
            'cpc':   float(s['cpc'])         if s['cpc']  else None,
        })
    return rows


def _aggregate_window_campaigns(marketplace, start, end, campaign_type, dim_map):
    """
    Non-today period: aggregate CampaignProfitDaily over [start, end].
    All derived ratios are recomputed from the SUMS so they don't drift on
    weighted averages.
    """
    from .models import CampaignProfitDaily, PPCCampaignSnapshot

    qs = CampaignProfitDaily.objects.filter(
        marketplace=marketplace, date__gte=start, date__lte=end)
    if campaign_type != 'all':
        qs = qs.filter(source_ad_type=campaign_type)

    agg = {}
    for r in qs.values(
        'campaign_id', 'source_ad_type',
        'spend', 'ad_revenue', 'attributed_units', 'attributed_orders',
        'cogs_attributed', 'referral_fee_attributed', 'fba_fee_attributed',
        'other_fees_attributed', 'gross_profit',
        'sku_count_attributed', 'attribution_coverage_pct',
    ):
        cid = r['campaign_id']
        b = agg.setdefault(cid, {
            'campaign_id': cid, 'source_ad_type': r['source_ad_type'],
            'spend': Decimal('0'), 'ad_revenue': Decimal('0'),
            'cogs': Decimal('0'), 'fees': Decimal('0'),
            'units': 0, 'orders': 0,
            'profit': Decimal('0'),
            'sku_set_size_sum': 0, 'days_count': 0,
            'coverage_weighted_sum': Decimal('0'),
        })
        b['spend']      += Decimal(r['spend']      or 0)
        b['ad_revenue'] += Decimal(r['ad_revenue'] or 0)
        b['cogs']       += Decimal(r['cogs_attributed'] or 0)
        b['fees']       += (Decimal(r['referral_fee_attributed'] or 0)
                          + Decimal(r['fba_fee_attributed']      or 0)
                          + Decimal(r['other_fees_attributed']   or 0))
        b['units']      += int(r['attributed_units']  or 0)
        b['orders']     += int(r['attributed_orders'] or 0)
        b['profit']     += Decimal(r['gross_profit']  or 0)
        b['sku_set_size_sum'] += int(r['sku_count_attributed'] or 0)
        b['days_count']       += 1
        # weighted coverage: weight by ad_revenue
        b['coverage_weighted_sum'] += (Decimal(r['attribution_coverage_pct'] or 0)
                                       * Decimal(r['ad_revenue'] or 0))

    # Pull window impressions/clicks from PPCCampaignSnapshot (those aren't on
    # CampaignProfitDaily — keep that table P&L-focused).
    ps_qs = PPCCampaignSnapshot.objects.filter(
        marketplace=marketplace, date__gte=start, date__lte=end)
    if campaign_type != 'all':
        ps_qs = ps_qs.filter(campaign_type=campaign_type)

    impr_clicks = {}
    for r in ps_qs.values('campaign_id').annotate(
        impressions=Sum('impressions'), clicks=Sum('clicks')
    ):
        impr_clicks[r['campaign_id']] = {
            'impressions': int(r['impressions'] or 0),
            'clicks':      int(r['clicks']      or 0),
        }

    rows = []
    for cid, b in agg.items():
        dim = dim_map.get(cid, {})
        revenue = b['ad_revenue']
        spend   = b['spend']
        profit  = b['profit']
        coverage = (b['coverage_weighted_sum'] / revenue
                    if revenue > 0 else Decimal('0'))
        impr  = impr_clicks.get(cid, {}).get('impressions', 0)
        clicks = impr_clicks.get(cid, {}).get('clicks', 0)

        rows.append({
            'campaign_id':    cid,
            'campaign_name':  dim.get('campaign_name', cid),
            'campaign_type':  dim.get('campaign_type', b['source_ad_type']),
            'brand':          dim.get('brand', ''),
            'state':          'enabled',  # range-level — state varies day-to-day
            'impressions':    impr,
            'clicks':         clicks,
            'spend':          spend,
            'ad_revenue':     revenue,
            'profit':         profit,
            'cogs':           b['cogs'],
            'fees':           b['fees'],
            'orders':         b['orders'],
            'units':          b['units'],
            'margin_pct':     (profit / revenue * 100) if revenue > 0 else None,
            'acos':           (spend  / revenue * 100) if revenue > 0 else None,
            'roas':           (revenue / spend)         if spend   > 0 else None,
            'ctr':            (clicks / impr  * 100) if impr   > 0 else None,
            'cvr':            (b['orders'] / clicks * 100) if clicks > 0 else None,
            'cpc':            (spend  / clicks)        if clicks > 0 else None,
            'attribution_coverage_pct': float(coverage),
            'sku_count_attributed':     b['sku_set_size_sum'] // max(b['days_count'], 1),
            # tacos + contribution_to_profit_pct filled in caller (needs totals)
        })
    return rows


# ═════════════════════════════════════════════════════════════════════════════
# PHASE 1 — CAMPAIGN DETAIL PAGE
#
# /dashboard/campaigns/<cid>/             campaign_detail (page shell)
# /dashboard/api/campaigns/<cid>/         api_campaign_detail (KPI + trend)
# /dashboard/api/campaigns/<cid>/skus/    api_campaign_top_skus  ← PRIMARY tab
# /dashboard/api/campaigns/<cid>/daily/   api_campaign_daily
#
# Top Contributing SKUs is the keystone tab — it answers "what did this
# campaign actually sell, and was it profitable?" by joining
# AdsAdvertisedProductDailySnapshot (per-ASIN attributed revenue/units)
# with DailySkuSnapshot (per-SKU COGS + fees) and the campaign's spend share.
# ═════════════════════════════════════════════════════════════════════════════


def _resolve_campaign_dim(marketplace: str, campaign_id: str) -> dict:
    """Lookup campaign dim row; fall back to PPCCampaignSnapshot if dim is empty."""
    from .models import Campaign, PPCCampaignSnapshot

    c = Campaign.objects.filter(
        marketplace=marketplace, campaign_id=campaign_id
    ).only('campaign_name', 'campaign_type', 'brand', 'product_family',
           'portfolio_name', 'state').first()
    if c:
        return {
            'campaign_id':    campaign_id,
            'campaign_name':  c.campaign_name,
            'campaign_type':  c.campaign_type,
            'brand':          c.brand,
            'product_family': c.product_family,
            'portfolio_name': c.portfolio_name,
            'state':          c.state or 'enabled',
        }
    # Fall back to PPCCampaignSnapshot latest
    ps = PPCCampaignSnapshot.objects.filter(
        marketplace=marketplace, campaign_id=campaign_id
    ).order_by('-date').only(
        'campaign_name', 'campaign_type', 'state').first()
    if ps:
        return {
            'campaign_id':    campaign_id,
            'campaign_name':  ps.campaign_name,
            'campaign_type':  ps.campaign_type,
            'brand':          '',
            'product_family': '',
            'portfolio_name': '',
            'state':          ps.state,
        }
    return {'campaign_id': campaign_id, 'campaign_name': campaign_id,
            'campaign_type': 'sp', 'brand': '', 'product_family': '',
            'portfolio_name': '', 'state': 'unknown'}


@login_required
@permission_required('can_view_dashboard')
def campaign_detail(request, campaign_id: str):
    """Campaign Detail page shell — JS fetches KPIs + tabs."""
    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        marketplace = _allowed_marketplaces(request.user)[0]

    dim = _resolve_campaign_dim(marketplace, campaign_id)
    ctx = {
        'marketplace':          marketplace,
        'allowed_marketplaces': _allowed_marketplaces(request.user),
        'campaign_id':          campaign_id,
        'campaign_name':        dim['campaign_name'],
        'campaign_type':        dim['campaign_type'],
        'campaign_brand':       dim['brand'],
        'campaign_state':       dim['state'],
        # SKU Intelligence P0 — context handoff: when the user arrives from a
        # SKU's campaign-drivers panel, show a breadcrumb back to that SKU.
        'from_sku':             (request.GET.get('from_sku') or '')[:64],
        'period_options': [
            {'id': pid, 'label': v[0]} for pid, v in _CAMPAIGN_PERIODS.items()
        ],
        'today':                date.today(),
    }
    return render(request, 'dashboard/campaign_detail.html', ctx)


@login_required
@permission_required('can_view_dashboard')
def api_campaign_detail(request, campaign_id: str):
    """KPI strip + daily trend series for the campaign over the selected period."""
    from .models import (
        CampaignProfitDaily, PPCCampaignSnapshot, DailyMetric,
    )

    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        return JsonResponse({'error': 'forbidden'}, status=403)

    period_id = (request.GET.get('period') or '7d').lower()
    if period_id not in _CAMPAIGN_PERIODS:
        period_id = '7d'
    today = date.today()
    start, end, is_today = _resolve_campaign_period(period_id, today)

    dim = _resolve_campaign_dim(marketplace, campaign_id)

    # ── Daily rows from CampaignProfitDaily ─────────────────────────────────
    profit_rows = list(CampaignProfitDaily.objects.filter(
        marketplace=marketplace, campaign_id=campaign_id,
        date__gte=start, date__lte=end,
    ).values('date', 'spend', 'ad_revenue', 'attributed_units',
             'attributed_orders', 'gross_profit', 'margin_pct',
             'acos', 'roas', 'tacos', 'attribution_coverage_pct'))

    # Impressions/clicks come from PPCCampaignSnapshot (campaign-day fact).
    ps_rows = {r['date']: r for r in PPCCampaignSnapshot.objects.filter(
        marketplace=marketplace, campaign_id=campaign_id,
        date__gte=start, date__lte=end,
    ).values('date', 'impressions', 'clicks',
             'spend', 'sales_7d', 'orders_7d', 'units_7d')}

    # ── Build per-day series for the trend chart ────────────────────────────
    trend = []
    profit_map = {r['date']: r for r in profit_rows}

    cur = start
    while cur <= end:
        prof = profit_map.get(cur)
        ps   = ps_rows.get(cur)
        if prof or ps:
            spend     = float(prof['spend'])     if prof else float(ps['spend'] or 0)
            revenue   = float(prof['ad_revenue']) if prof else float(ps['sales_7d'] or 0)
            profit    = float(prof['gross_profit']) if prof else None
            margin    = float(prof['margin_pct'])   if prof else None
            acos_v    = float(prof['acos']) * 100 if prof else (
                        float(ps['spend'] or 0) / float(ps['sales_7d']) * 100
                        if ps and ps['sales_7d'] else None)
            roas_v    = float(prof['roas'])     if prof else (
                        float(ps['sales_7d'] or 0) / float(ps['spend'])
                        if ps and ps['spend'] else None)
            orders    = int(prof['attributed_orders']) if prof else int(ps['orders_7d'] or 0)
            units     = int(prof['attributed_units'])  if prof else int(ps['units_7d']  or 0)
            impr      = int(ps['impressions'] or 0) if ps else 0
            clicks    = int(ps['clicks']      or 0) if ps else 0
            trend.append({
                'date':    cur.isoformat(),
                'spend':   round(spend,   2),
                'revenue': round(revenue, 2),
                'profit':  round(profit,  2) if profit is not None else None,
                'margin_pct': round(margin, 2) if margin is not None else None,
                'acos':    round(acos_v, 2) if acos_v is not None else None,
                'roas':    round(roas_v, 2) if roas_v is not None else None,
                'orders':  orders,
                'units':   units,
                'impressions': impr,
                'clicks':  clicks,
            })
        cur += timedelta(days=1)

    # ── Window aggregates (KPI strip) ───────────────────────────────────────
    sum_spend  = sum(t['spend']   for t in trend)
    sum_rev    = sum(t['revenue'] for t in trend)
    sum_profit = sum(t['profit']  for t in trend if t['profit'] is not None)
    sum_imp    = sum(t['impressions'] for t in trend)
    sum_clicks = sum(t['clicks']      for t in trend)
    sum_orders = sum(t['orders']      for t in trend)
    sum_units  = sum(t['units']       for t in trend)

    window_revenue = float(DailyMetric.objects.filter(
        marketplace=marketplace, date__gte=start, date__lte=end
    ).aggregate(total=Sum('revenue'))['total'] or 0)

    # Weighted attribution coverage across the window (weight by ad_revenue)
    weighted_cov_num = sum(float(r['attribution_coverage_pct'] or 0) *
                           float(r['ad_revenue'] or 0) for r in profit_rows)
    coverage = (weighted_cov_num / sum_rev) if sum_rev > 0 else None

    has_profit = any(t['profit'] is not None for t in trend)
    kpi = {
        'spend':         round(sum_spend, 2),
        'ad_revenue':    round(sum_rev,   2),
        'gross_profit':  round(sum_profit, 2) if has_profit else None,
        'margin_pct':    round(sum_profit / sum_rev * 100, 2)
                          if (sum_rev > 0 and has_profit) else None,
        'tacos':         round(sum_spend / window_revenue * 100, 2)
                          if window_revenue > 0 else None,
        'acos':          round(sum_spend / sum_rev * 100, 2) if sum_rev > 0 else None,
        'roas':          round(sum_rev / sum_spend, 2) if sum_spend > 0 else None,
        'orders':        sum_orders,
        'units':         sum_units,
        'impressions':   sum_imp,
        'clicks':        sum_clicks,
        'ctr':           round(sum_clicks / sum_imp * 100, 2) if sum_imp > 0 else None,
        'cvr':           round(sum_orders / sum_clicks * 100, 2) if sum_clicks > 0 else None,
        'cpc':           round(sum_spend / sum_clicks, 2)     if sum_clicks > 0 else None,
        'attribution_coverage_pct': round(coverage, 2) if coverage is not None else None,
        'window_revenue': round(window_revenue, 2),
    }

    return JsonResponse({
        'marketplace':  marketplace,
        'campaign':     dim,
        'period': {
            'id':    period_id,
            'label': _CAMPAIGN_PERIODS[period_id][0],
            'start': start.isoformat(),
            'end':   end.isoformat(),
        },
        'is_today':     is_today,
        'kpi':          kpi,
        'trend':        trend,
        'today_warning': ('Detailed advertising analytics are available through T-1. '
                          "Today's view contains only currently available PPC metrics."
                          if is_today else None),
    })


@login_required
@permission_required('can_view_dashboard')
def api_campaign_top_skus(request, campaign_id: str):
    """
    Per-SKU rollup for the campaign's selected period.

    Joins:
      • AdsAdvertisedProductDailySnapshot — per (campaign, ASIN, SKU)
        attributed revenue / units (truth-of-record for what the campaign sold).
      • DailySkuSnapshot — per-SKU per-day TOTAL cgs / amz_fee / fulfill;
        divided by qty to get per-unit costs.
      • PPCCampaignSnapshot — campaign-level daily spend (for the spend-
        allocation column, prorated by each SKU's revenue share).

    Returns:
        rows: [{
          sku, asin, product_name, revenue, units, orders,
          spend_allocation,           # campaign spend × (sku_revenue / campaign_revenue)
          cogs, fees, profit, margin_pct,
          contribution_pct,           # this SKU's profit / total SKU profit
        }, ...]
    """
    from .models import (
        AdsAdvertisedProductDailySnapshot, DailySkuSnapshot,
        PPCCampaignSnapshot, Product,
    )

    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        return JsonResponse({'error': 'forbidden'}, status=403)

    period_id = (request.GET.get('period') or '7d').lower()
    if period_id not in _CAMPAIGN_PERIODS:
        period_id = '7d'
    today = date.today()
    start, end, is_today = _resolve_campaign_period(period_id, today)

    if is_today:
        # Per-SKU campaign attribution is T-1; show empty with banner.
        return JsonResponse({
            'marketplace': marketplace,
            'campaign_id': campaign_id,
            'period': {'id': period_id, 'label': _CAMPAIGN_PERIODS[period_id][0],
                       'start': start.isoformat(), 'end': end.isoformat()},
            'is_today': True,
            'rows': [],
            'today_warning': ('Top Contributing SKUs are computed from the T-1 '
                              'advertised-product report. Switch to Yesterday or a '
                              'longer period to see SKU-level attribution.'),
        })

    # ── 1. Pull per-(asin, sku) advertised-product rows ─────────────────────
    ap_qs = AdsAdvertisedProductDailySnapshot.objects.filter(
        marketplace=marketplace, campaign_id=campaign_id,
        date__gte=start, date__lte=end,
    ).values('date', 'asin', 'advertised_sku',
             'sales_7d', 'units_7d', 'orders_7d')

    # ── 2. Pull per-day campaign-level spend (for spend allocation) ─────────
    spend_by_date = {r['date']: float(r['spend'] or 0)
                     for r in PPCCampaignSnapshot.objects.filter(
                         marketplace=marketplace, campaign_id=campaign_id,
                         date__gte=start, date__lte=end,
                     ).values('date', 'spend')}

    # ── 3. Pull DailySkuSnapshot rows for all SKUs in the AP window ─────────
    skus_in_play = {r['advertised_sku'] for r in ap_qs if r['advertised_sku']}
    # Per-unit cost lookup, keyed by (sku, date). Same logic as compute_campaign_profit.
    sku_cost_by_date: dict[tuple, dict] = {}
    if skus_in_play:
        for row in DailySkuSnapshot.objects.filter(
            marketplace=marketplace, date__gte=start, date__lte=end,
            sku__in=skus_in_play
        ).values('sku', 'date', 'qty', 'cgs', 'amz_fee', 'fulfill', 'revenue'):
            qty = max(int(row['qty'] or 0), 1)
            sku_cost_by_date[(row['sku'], row['date'])] = {
                'cogs_per_unit':     Decimal(row['cgs'])     / qty,
                'referral_per_unit': Decimal(row['amz_fee']) / qty,
                'fba_per_unit':      Decimal(row['fulfill']) / qty,
            }

    # Fallback: Product table for SKUs missing from DailySkuSnapshot
    missing_skus = skus_in_play - {k[0] for k in sku_cost_by_date}
    product_fallback: dict[str, dict] = {}
    if missing_skus:
        for p in Product.objects.filter(marketplace=marketplace, sku__in=missing_skus).only(
            'sku', 'asin', 'title', 'fba_fee', 'referral_fee_pct',
        ):
            product_fallback[p.sku] = {
                'fba_per_unit':   Decimal(p.fba_fee or 0),
                'referral_pct':   Decimal(p.referral_fee_pct or 15) / Decimal('100'),
                'title':          p.title,
                'asin':           p.asin,
            }

    # ── 4. Product title lookup for display ─────────────────────────────────
    product_meta = {p.sku: p for p in Product.objects.filter(
        marketplace=marketplace,
        sku__in=skus_in_play,
    ).only('sku', 'title', 'asin')}

    # ── 5. Compute per-day campaign revenue for spend-allocation denominator ─
    campaign_rev_by_date: dict[date, float] = {}
    for r in ap_qs:
        campaign_rev_by_date[r['date']] = (
            campaign_rev_by_date.get(r['date'], 0.0)
            + float(r['sales_7d'] or 0))

    # ── 6. Aggregate to (sku, asin) ─────────────────────────────────────────
    agg: dict[tuple, dict] = {}
    for r in ap_qs:
        sku = r['advertised_sku'] or ''
        asin = r['asin'] or ''
        key = (sku, asin)
        revenue = float(r['sales_7d'] or 0)
        units   = int(r['units_7d']   or 0)
        orders  = int(r['orders_7d']  or 0)
        d       = r['date']

        b = agg.setdefault(key, {
            'sku': sku, 'asin': asin,
            'revenue': 0.0, 'units': 0, 'orders': 0,
            'spend_allocation': 0.0,
            'cogs': Decimal('0'), 'fees': Decimal('0'),
        })
        b['revenue'] += revenue
        b['units']   += units
        b['orders']  += orders

        # Spend allocation: this row gets a share of the day's campaign spend
        # proportional to its share of the day's total ad-attributed revenue.
        # If campaign_rev_by_date[d] == 0 (rare: spend without sales), allocate
        # nothing — we don't have a meaningful denominator.
        denom = campaign_rev_by_date.get(d, 0)
        if denom > 0:
            b['spend_allocation'] += (
                spend_by_date.get(d, 0.0) * (revenue / denom))

        # Per-unit costs
        cm = sku_cost_by_date.get((sku, d))
        if cm:
            b['cogs'] += units * cm['cogs_per_unit']
            b['fees'] += units * (cm['referral_per_unit'] + cm['fba_per_unit'])
        elif sku in product_fallback:
            m = product_fallback[sku]
            b['fees'] += units * m['fba_per_unit'] + Decimal(str(revenue)) * m['referral_pct']
            # No COGS row → leave cogs at 0; visible as inflated margin

    # ── 7. Final shape ──────────────────────────────────────────────────────
    rows = []
    total_profit = Decimal('0')
    for (sku, asin), b in agg.items():
        profit = Decimal(str(b['revenue'])) - b['cogs'] - b['fees'] - Decimal(str(b['spend_allocation']))
        margin_pct = (float(profit) / b['revenue'] * 100) if b['revenue'] > 0 else None
        meta = product_meta.get(sku)
        rows.append({
            'sku':           sku,
            'asin':          asin or (meta.asin if meta else ''),
            'product_name':  (meta.title if meta else
                              product_fallback.get(sku, {}).get('title', ''))[:96],
            'revenue':       round(b['revenue'], 2),
            'units':         b['units'],
            'orders':        b['orders'],
            'spend_allocation': round(b['spend_allocation'], 2),
            'cogs':          round(float(b['cogs']), 2),
            'fees':          round(float(b['fees']), 2),
            'profit':        round(float(profit), 2),
            'margin_pct':    round(margin_pct, 2) if margin_pct is not None else None,
        })
        if profit > 0:
            total_profit += profit

    # Contribution % = sku.profit / total_positive_profit (negatives ignored
    # so the percentages sum to ≤ 100 and represent profit upside).
    for r in rows:
        if total_profit > 0 and r['profit'] > 0:
            r['contribution_pct'] = round(r['profit'] / float(total_profit) * 100, 2)
        else:
            r['contribution_pct'] = 0.0

    rows.sort(key=lambda r: r['profit'], reverse=True)

    return JsonResponse({
        'marketplace':  marketplace,
        'campaign_id':  campaign_id,
        'period': {'id': period_id, 'label': _CAMPAIGN_PERIODS[period_id][0],
                   'start': start.isoformat(), 'end': end.isoformat()},
        'is_today':     False,
        'rows':         rows,
        'sku_count':    len(rows),
    })


@login_required
@permission_required('can_view_dashboard')
def api_campaign_daily(request, campaign_id: str):
    """Day-by-day fact rows for the Daily Performance tab — reuses the same
    series the trend chart consumes but returned as a flat tabular form."""
    from .models import CampaignProfitDaily, PPCCampaignSnapshot

    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        return JsonResponse({'error': 'forbidden'}, status=403)

    period_id = (request.GET.get('period') or '7d').lower()
    if period_id not in _CAMPAIGN_PERIODS:
        period_id = '7d'
    today = date.today()
    start, end, is_today = _resolve_campaign_period(period_id, today)

    profit_map = {r['date']: r for r in CampaignProfitDaily.objects.filter(
        marketplace=marketplace, campaign_id=campaign_id,
        date__gte=start, date__lte=end,
    ).values('date', 'spend', 'ad_revenue', 'attributed_units',
             'attributed_orders', 'gross_profit', 'margin_pct',
             'acos', 'roas', 'attribution_coverage_pct', 'sku_count_attributed')}

    ps_map = {r['date']: r for r in PPCCampaignSnapshot.objects.filter(
        marketplace=marketplace, campaign_id=campaign_id,
        date__gte=start, date__lte=end,
    ).values('date', 'impressions', 'clicks', 'spend',
             'sales_7d', 'orders_7d', 'units_7d')}

    rows = []
    cur = start
    while cur <= end:
        p  = profit_map.get(cur)
        ps = ps_map.get(cur)
        if not p and not ps:
            cur += timedelta(days=1); continue

        spend   = float(p['spend'])     if p else float(ps['spend'] or 0)
        revenue = float(p['ad_revenue']) if p else float(ps['sales_7d'] or 0)
        orders  = int(p['attributed_orders']) if p else int(ps['orders_7d'] or 0)
        units   = int(p['attributed_units'])  if p else int(ps['units_7d']  or 0)
        impr    = int(ps['impressions'] or 0) if ps else 0
        clicks  = int(ps['clicks']      or 0) if ps else 0
        profit  = float(p['gross_profit']) if p else None
        margin  = float(p['margin_pct'])   if p else None
        cov     = float(p['attribution_coverage_pct']) if p else None

        rows.append({
            'date':        cur.isoformat(),
            'impressions': impr,
            'clicks':      clicks,
            'ctr':         round(clicks / impr * 100, 2) if impr > 0 else None,
            'spend':       round(spend, 2),
            'cpc':         round(spend / clicks, 2) if clicks > 0 else None,
            'revenue':     round(revenue, 2),
            'orders':      orders,
            'units':       units,
            'cvr':         round(orders / clicks * 100, 2) if clicks > 0 else None,
            'profit':      round(profit, 2) if profit is not None else None,
            'margin_pct':  round(margin, 2) if margin is not None else None,
            'acos':        round(spend / revenue * 100, 2) if revenue > 0 else None,
            'roas':        round(revenue / spend, 2) if spend > 0 else None,
            'attribution_coverage_pct': round(cov, 2) if cov is not None else None,
            'sku_count_attributed':     int(p['sku_count_attributed']) if p else None,
        })
        cur += timedelta(days=1)

    return JsonResponse({
        'marketplace':  marketplace,
        'campaign_id':  campaign_id,
        'period': {'id': period_id, 'label': _CAMPAIGN_PERIODS[period_id][0],
                   'start': start.isoformat(), 'end': end.isoformat()},
        'is_today':     is_today,
        'rows':         rows,
    })


# ═════════════════════════════════════════════════════════════════════════════
# PHASE 1 — SEARCH TERM INTELLIGENCE
#
# /dashboard/search-terms/                      search_terms (page shell)
# /dashboard/api/search-terms/                  api_search_terms (aggregated)
# /dashboard/api/search-terms/detail/?term=...  api_search_term_detail (per-campaign)
#
# Estimated Profit uses the campaign-level blended contribution-margin %
# proxy — per-SKU per-search-term attribution is impossible from Amazon's data
# (the search-term report doesn't tell us which ASIN was bought). The proxy is
# the standard Pacvue / Perpetua approach: for each (search_term, campaign,
# date) row, profit ≈ sales × (campaign_cm_pct) − spend, where campaign_cm_pct
# comes from CampaignProfitDaily.
# ═════════════════════════════════════════════════════════════════════════════

# Auto-tag thresholds — kept in lock-step with compute_campaign_profit.py so a
# row's tag here matches its tag in the CampaignSearchTermSummary counter.
_ST_HIGH_SPEND        = 5.0
_ST_HIGH_CTR          = 0.005
_ST_LOW_CVR           = 0.02
_ST_HIGH_ROAS         = 5.0
_ST_HIGH_PROFIT       = 50.0
_ST_LOSING_MONEY      = -20.0
_ST_HIGH_PROFIT_ACOS  = 0.30   # for scaling_opportunity

_ST_ALL_TAGS = (
    'high_spend_no_sales',
    'high_ctr_low_cvr',
    'high_profit',
    'scaling_opportunity',
    'losing_money',
)


def _tag_search_term(spend, sales, orders, clicks, impr, est_profit, acos) -> list[str]:
    tags = []
    if spend > _ST_HIGH_SPEND and orders == 0:
        tags.append('high_spend_no_sales')
    ctr = (clicks / impr) if impr else 0
    cvr = (orders / clicks) if clicks else 0
    if ctr > _ST_HIGH_CTR and cvr < _ST_LOW_CVR and clicks > 10:
        tags.append('high_ctr_low_cvr')
    roas = (sales / spend) if spend else 0
    if roas > _ST_HIGH_ROAS and spend > _ST_HIGH_SPEND:
        tags.append('high_profit')
    if est_profit > _ST_HIGH_PROFIT and acos is not None and acos < _ST_HIGH_PROFIT_ACOS:
        tags.append('scaling_opportunity')
    if est_profit < _ST_LOSING_MONEY:
        tags.append('losing_money')
    return tags


@login_required
@permission_required('can_view_dashboard')
def search_terms(request):
    """Page shell — JS calls /api/search-terms/ with filters.

    P2 — accepts an investigation context handed over from Campaign Detail
    (campaign_id / target_id / from_sku / from_campaign / period) so the user
    never re-selects what they already chose upstream.
    """
    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        marketplace = _allowed_marketplaces(request.user)[0]

    ctx = {
        'marketplace':          marketplace,
        'ctx_campaign_id':      (request.GET.get('campaign_id') or '')[:64],
        'ctx_campaign_name':    (request.GET.get('from_campaign') or '')[:256],
        'ctx_target_id':        (request.GET.get('target_id') or '')[:64],
        'ctx_from_sku':         (request.GET.get('from_sku') or '')[:64],
        'allowed_marketplaces': _allowed_marketplaces(request.user),
        'period_options': [
            {'id': pid, 'label': v[0]} for pid, v in _CAMPAIGN_PERIODS.items()
            if pid != 'today'  # search-term reports are T-1 minimum
        ],
        'tag_options': [
            {'id': 'high_spend_no_sales',   'label': 'High Spend, No Sales'},
            {'id': 'high_ctr_low_cvr',      'label': 'High CTR, Low CVR'},
            {'id': 'high_profit',           'label': 'High Profit'},
            {'id': 'scaling_opportunity',   'label': 'Scaling Opportunity'},
            {'id': 'losing_money',          'label': 'Losing Money'},
        ],
        'today':                date.today(),
    }
    return render(request, 'dashboard/search_terms.html', ctx)


@login_required
@permission_required('can_view_dashboard')
def api_search_terms(request):
    """
    Aggregated search-term rollup for the selected window.

    Query params:
        mp           — marketplace (default usa)
        period       — yesterday | 7d | 30d | mtd (default 7d)
        campaign_id  — optional; filter to one campaign
        ad_type      — sp | sb | all (default all). SD has no search-term report.
        tags         — comma-separated; filter rows that have AT LEAST ONE of these tags
        sort         — spend | sales | profit | impressions | clicks | orders (default spend)
        dir          — asc | desc (default desc)
        limit        — default 500 (max 2000)

    Response envelope:
        {
          marketplace, period, kpi, tag_counts, rows
        }

    Rows are server-side sorted + limited to keep payloads bounded — the
    underlying fact table can exceed 100k rows/day, so we never ship all of it.
    """
    from .models import (
        AdsSearchTermDailySnapshot, CampaignProfitDaily, Campaign,
    )
    from collections import defaultdict

    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        return JsonResponse({'error': 'forbidden'}, status=403)

    period_id = (request.GET.get('period') or '7d').lower()
    if period_id == 'today' or period_id not in _CAMPAIGN_PERIODS:
        period_id = '7d'
    today = date.today()
    start, end, _ = _resolve_campaign_period(period_id, today)

    campaign_id  = (request.GET.get('campaign_id') or '').strip()
    ad_type      = (request.GET.get('ad_type') or 'all').lower()
    if ad_type not in ('sp', 'sb', 'all'):
        ad_type = 'all'
    tag_filter   = [t.strip() for t in (request.GET.get('tags') or '').split(',')
                    if t.strip() in _ST_ALL_TAGS]
    sort_key     = (request.GET.get('sort') or 'spend').lower()
    if sort_key not in ('spend', 'sales', 'profit', 'impressions',
                        'clicks', 'orders', 'roas', 'acos'):
        sort_key = 'spend'
    direction    = (request.GET.get('dir') or 'desc').lower()
    desc         = (direction != 'asc')
    try:
        limit = max(1, min(int(request.GET.get('limit') or 500), 2000))
    except ValueError:
        limit = 500

    # ── 1. Pull campaign CM% map for the proxy profit calc ──────────────────
    cm_map: dict[tuple, float] = {}
    cp_qs = CampaignProfitDaily.objects.filter(
        marketplace=marketplace, date__gte=start, date__lte=end)
    if campaign_id:
        cp_qs = cp_qs.filter(campaign_id=campaign_id)
    if ad_type != 'all':
        cp_qs = cp_qs.filter(source_ad_type=ad_type)
    for r in cp_qs.values('campaign_id', 'date', 'ad_revenue',
                           'contribution_margin'):
        rev = float(r['ad_revenue'] or 0)
        if rev > 0:
            cm_map[(r['campaign_id'], r['date'])] = (
                float(r['contribution_margin'] or 0) / rev)
    cm_data_ready = bool(cm_map)

    # ── 2. Stream search-term rows ──────────────────────────────────────────
    st_qs = AdsSearchTermDailySnapshot.objects.filter(
        marketplace=marketplace, date__gte=start, date__lte=end)
    if campaign_id:
        st_qs = st_qs.filter(campaign_id=campaign_id)
    if ad_type != 'all':
        st_qs = st_qs.filter(source_ad_type=ad_type)
    # P2 — Target → Search Term drill. AdsSearchTermDailySnapshot already
    # records the target the term matched against, so narrowing to one target
    # needs no new table and no new endpoint. Absent = unchanged behaviour.
    target_id = (request.GET.get('target_id') or '').strip()
    if target_id:
        st_qs = st_qs.filter(target_id=target_id)

    # Bound the work: pull only the top N rows by spend at the SQL layer.
    # We still need to aggregate by search_term_hash across days/campaigns,
    # so this is a heuristic — pulling 10× the limit gives the aggregator
    # enough fuel to surface the genuine top-N after rollup.
    raw = st_qs.values(
        'campaign_id', 'date', 'search_term', 'search_term_hash',
        'spend', 'sales_7d', 'orders_7d', 'clicks', 'impressions',
        'match_type', 'target_id',
    ).order_by('-spend')[:max(limit * 10, 5000)]

    # ── 3. Aggregate by search_term_hash ────────────────────────────────────
    agg = defaultdict(lambda: {
        'search_term':       '',
        'search_term_hash':  '',
        'spend':             0.0,
        'sales':             0.0,
        'orders':            0,
        'clicks':            0,
        'impressions':       0,
        'est_profit':        0.0,
        'campaigns':         set(),
        'match_types':       set(),
        'target_ids':        set(),
    })
    for r in raw:
        h = r['search_term_hash']
        b = agg[h]
        b['search_term']      = r['search_term']
        b['search_term_hash'] = h
        if r.get('match_type'):
            b['match_types'].add(r['match_type'])
        if r.get('target_id'):
            b['target_ids'].add(r['target_id'])
        s = float(r['spend']    or 0)
        v = float(r['sales_7d'] or 0)
        o = int(r['orders_7d']  or 0)
        c = int(r['clicks']     or 0)
        i = int(r['impressions'] or 0)
        b['spend']       += s
        b['sales']       += v
        b['orders']      += o
        b['clicks']      += c
        b['impressions'] += i
        cm_pct = cm_map.get((r['campaign_id'], r['date']), 0.0)
        b['est_profit']  += (v * cm_pct - s)
        b['campaigns'].add(r['campaign_id'])

    # ── 4. Compute per-row derived metrics + auto-tags ──────────────────────
    rows = []
    tag_counts = {t: 0 for t in _ST_ALL_TAGS}
    for b in agg.values():
        spend = b['spend']
        sales = b['sales']
        ctr   = (b['clicks']  / b['impressions']) if b['impressions'] else None
        cvr   = (b['orders']  / b['clicks'])      if b['clicks']      else None
        acos  = (spend / sales)                   if sales else None
        roas  = (sales / spend)                   if spend else None
        est_margin_pct = (b['est_profit'] / sales * 100) if sales else None

        tags = _tag_search_term(
            spend=spend, sales=sales, orders=b['orders'],
            clicks=b['clicks'], impr=b['impressions'],
            est_profit=b['est_profit'], acos=acos,
        )
        for t in tags:
            tag_counts[t] += 1

        rows.append({
            'search_term':         b['search_term'],
            'spend':               round(spend, 2),
            'sales':               round(sales, 2),
            'orders':              b['orders'],
            'clicks':              b['clicks'],
            'impressions':         b['impressions'],
            'ctr':                 round(ctr * 100, 2)  if ctr  is not None else None,
            'cvr':                 round(cvr * 100, 2)  if cvr  is not None else None,
            'acos':                round(acos * 100, 2) if acos is not None else None,
            'roas':                round(roas, 2)       if roas is not None else None,
            'estimated_profit':    round(b['est_profit'], 2),
            'estimated_margin_pct': round(est_margin_pct, 2) if est_margin_pct is not None else None,
            'campaign_count':      len(b['campaigns']),
            # P2 — how the term matched (may be several across days/targets)
            'match_types':         sorted(b['match_types']),
            'target_count':        len(b['target_ids']),
            'tags':                tags,
        })

    # ── 5. Tag-filter: keep rows that have at least one of the requested tags
    if tag_filter:
        wanted = set(tag_filter)
        rows = [r for r in rows if any(t in wanted for t in r['tags'])]

    # ── 6. Sort + limit ─────────────────────────────────────────────────────
    sort_field_map = {
        'spend': 'spend', 'sales': 'sales', 'profit': 'estimated_profit',
        'impressions': 'impressions', 'clicks': 'clicks', 'orders': 'orders',
        'roas': 'roas', 'acos': 'acos',
    }
    sf = sort_field_map[sort_key]
    rows.sort(key=lambda r: (r[sf] if r[sf] is not None else (0 if desc else float('inf'))),
              reverse=desc)
    rows = rows[:limit]

    # ── 7. KPI strip ────────────────────────────────────────────────────────
    kpi = {
        'distinct_terms':       len(agg),
        'shown_terms':          len(rows),
        'spend':                round(sum(r['spend'] for r in rows), 2),
        'sales':                round(sum(r['sales'] for r in rows), 2),
        'estimated_profit':     round(sum(r['estimated_profit'] for r in rows), 2),
        'orders':               sum(r['orders'] for r in rows),
        'truncated':            len(agg) > limit,
    }

    return JsonResponse({
        'marketplace': marketplace,
        'period': {'id': period_id, 'label': _CAMPAIGN_PERIODS[period_id][0],
                   'start': start.isoformat(), 'end': end.isoformat()},
        'campaign_id':  campaign_id or None,
        'target_id':    target_id or None,
        'ad_type':      ad_type,
        'sort':         {'key': sort_key, 'dir': direction},
        'limit':        limit,
        'kpi':          kpi,
        'tag_counts':   tag_counts,
        'rows':         rows,
        'cm_data_ready': cm_data_ready,
        'profit_proxy_note': ('Estimated Profit is calculated using each campaign\'s '
                              'blended contribution-margin %. Per-SKU per-search-term '
                              'attribution is not provided by Amazon — this is the '
                              'standard proxy used in enterprise PPC tools.'),
        'cm_not_ready_warning': (None if cm_data_ready else
            'CampaignProfitDaily is not yet populated for this window — '
            'Estimated Profit currently shows −spend on every row. '
            'Run `compute_campaign_profit --backfill-window 30` (or wait for '
            'the orchestrator\'s Phase 3) to fill in real profit numbers.'),
    })


@login_required
@permission_required('can_view_dashboard')
def api_search_term_detail(request):
    """
    Per-(campaign, date) breakdown of a single search term — for the
    drill-down modal triggered by clicking a search-term row.

    Query params:
        mp, period, term (the search_term text, exact match)
    """
    from .models import AdsSearchTermDailySnapshot, CampaignProfitDaily, Campaign
    import hashlib
    from collections import defaultdict

    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        return JsonResponse({'error': 'forbidden'}, status=403)

    period_id = (request.GET.get('period') or '7d').lower()
    if period_id == 'today' or period_id not in _CAMPAIGN_PERIODS:
        period_id = '7d'
    today = date.today()
    start, end, _ = _resolve_campaign_period(period_id, today)

    term = (request.GET.get('term') or '').strip()
    if not term:
        return JsonResponse({'error': 'term parameter required'}, status=400)

    term_hash = hashlib.sha1(term.lower().encode('utf-8')).hexdigest()

    # CM% lookup
    cm_map: dict[tuple, float] = {}
    for r in CampaignProfitDaily.objects.filter(
        marketplace=marketplace, date__gte=start, date__lte=end,
    ).values('campaign_id', 'date', 'ad_revenue', 'contribution_margin'):
        rev = float(r['ad_revenue'] or 0)
        if rev > 0:
            cm_map[(r['campaign_id'], r['date'])] = (
                float(r['contribution_margin'] or 0) / rev)

    qs = list(AdsSearchTermDailySnapshot.objects.filter(
        marketplace=marketplace, date__gte=start, date__lte=end,
        search_term_hash=term_hash,
    ).values('campaign_id', 'date', 'spend', 'sales_7d', 'orders_7d',
             'clicks', 'impressions', 'match_type'))

    # Campaign-name lookup — prefer Campaign dim, fall back to PPCCampaignSnapshot
    # (the Campaign dim table is populated on demand; the snapshot is always there).
    from .models import PPCCampaignSnapshot
    cids_in_play = {r['campaign_id'] for r in qs}
    cnames = {c.campaign_id: c.campaign_name for c in Campaign.objects.filter(
        marketplace=marketplace, campaign_id__in=cids_in_play
    ).only('campaign_id', 'campaign_name')}
    missing_cids = cids_in_play - set(cnames)
    if missing_cids:
        for r in PPCCampaignSnapshot.objects.filter(
            marketplace=marketplace, campaign_id__in=missing_cids
        ).order_by('campaign_id', '-date').values('campaign_id', 'campaign_name'):
            cnames.setdefault(r['campaign_id'], r['campaign_name'])

    by_campaign: dict[str, dict] = defaultdict(lambda: {
        'campaign_id': '', 'campaign_name': '',
        'spend': 0.0, 'sales': 0.0, 'orders': 0,
        'clicks': 0, 'impressions': 0, 'est_profit': 0.0,
        'match_types': set(),
    })
    for r in qs:
        cid = r['campaign_id']
        b = by_campaign[cid]
        b['campaign_id']   = cid
        b['campaign_name'] = cnames.get(cid, cid)
        b['spend']       += float(r['spend']    or 0)
        b['sales']       += float(r['sales_7d'] or 0)
        b['orders']      += int(r['orders_7d']  or 0)
        b['clicks']      += int(r['clicks']     or 0)
        b['impressions'] += int(r['impressions'] or 0)
        if r['match_type']:
            b['match_types'].add(r['match_type'])
        cm = cm_map.get((cid, r['date']), 0.0)
        b['est_profit']  += float(r['sales_7d'] or 0) * cm - float(r['spend'] or 0)

    rows = []
    for b in by_campaign.values():
        spend = b['spend']; sales = b['sales']
        rows.append({
            'campaign_id':       b['campaign_id'],
            'campaign_name':     b['campaign_name'],
            'spend':             round(spend, 2),
            'sales':             round(sales, 2),
            'orders':            b['orders'],
            'clicks':            b['clicks'],
            'impressions':       b['impressions'],
            'estimated_profit':  round(b['est_profit'], 2),
            'acos':              round(spend / sales * 100, 2) if sales else None,
            'roas':              round(sales / spend, 2)       if spend else None,
            'match_types':       sorted(b['match_types']),
        })
    rows.sort(key=lambda r: r['spend'], reverse=True)

    return JsonResponse({
        'marketplace': marketplace,
        'period': {'id': period_id, 'label': _CAMPAIGN_PERIODS[period_id][0],
                   'start': start.isoformat(), 'end': end.isoformat()},
        'search_term': term,
        'rows': rows,
        'total_campaigns': len(rows),
    })


# ═════════════════════════════════════════════════════════════════════════════
# PHASE 1 — PLACEMENT ANALYTICS
#
# /dashboard/placements/              placements (page shell)
# /dashboard/api/placements/          api_placements (JSON)
#
# Sponsored Products only — SB and SD do NOT expose placement breakdowns in
# Ads API v3. The placements are normalized to:
#   top_of_search | rest_of_search | product_pages | other
#
# Estimated Profit uses the same campaign-margin proxy as Search Term
# Intelligence: placement_sales × campaign_cm_pct − placement_spend.
# ═════════════════════════════════════════════════════════════════════════════

_PLACEMENT_LABEL = {
    'top_of_search':   'Top of Search',
    'product_pages':   'Product Pages',
    'other_on_amazon': 'Other on-Amazon',
    'off_amazon':      'Off Amazon',
    # Legacy buckets — kept so historical rows still render
    'rest_of_search':  'Rest of Search (legacy)',
    'other':           'Other (legacy)',
}
# Display order for the bar visualization. Amazon's v3 spCampaigns placement
# report returns four real categories — Rest of Search is rolled into
# "Other on-Amazon" by Amazon's own grouping; we surface that with a tooltip.
_PLACEMENT_ORDER = (
    'top_of_search', 'product_pages',
    'other_on_amazon', 'off_amazon',
    'rest_of_search', 'other',   # legacy tail — invisible unless data exists
)


@login_required
@permission_required('can_view_dashboard')
def placements(request):
    """Page shell — JS calls /api/placements/ for the visualization data."""
    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        marketplace = _allowed_marketplaces(request.user)[0]

    ctx = {
        'marketplace':          marketplace,
        'allowed_marketplaces': _allowed_marketplaces(request.user),
        'period_options': [
            {'id': pid, 'label': v[0]} for pid, v in _CAMPAIGN_PERIODS.items()
            if pid != 'today'  # placement reports are T-1 minimum
        ],
        'today':                date.today(),
    }
    return render(request, 'dashboard/placements.html', ctx)


@login_required
@permission_required('can_view_dashboard')
def api_placements(request):
    """
    Account-level placement rollup for the selected window.

    Query params:
        mp           — marketplace (default usa)
        period       — yesterday | 7d | 30d | mtd (default 7d)
        campaign_id  — optional; scope to one campaign

    Returns:
        kpi:           totals across all placements
        placements:    [{placement, label, impressions, clicks, ctr, spend,
                          spend_pct, sales, sales_pct, orders, acos, roas,
                          estimated_profit, profit_pct, ...}, ...]
                       — sorted in fixed display order, includes 0-row placeholders
        top_campaigns: per-campaign placement mix table (top N by total spend)
    """
    from .models import (
        AdsPlacementDailySnapshot, CampaignProfitDaily, Campaign,
        PPCCampaignSnapshot,
    )
    from collections import defaultdict

    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        return JsonResponse({'error': 'forbidden'}, status=403)

    period_id = (request.GET.get('period') or '7d').lower()
    if period_id == 'today' or period_id not in _CAMPAIGN_PERIODS:
        period_id = '7d'
    today = date.today()
    start, end, _ = _resolve_campaign_period(period_id, today)

    campaign_id = (request.GET.get('campaign_id') or '').strip()

    # ── 1. Campaign CM% map for the proxy profit calc ───────────────────────
    cm_map: dict[tuple, float] = {}
    cp_qs = CampaignProfitDaily.objects.filter(
        marketplace=marketplace, date__gte=start, date__lte=end)
    if campaign_id:
        cp_qs = cp_qs.filter(campaign_id=campaign_id)
    for r in cp_qs.values('campaign_id', 'date', 'ad_revenue',
                           'contribution_margin'):
        rev = float(r['ad_revenue'] or 0)
        if rev > 0:
            cm_map[(r['campaign_id'], r['date'])] = (
                float(r['contribution_margin'] or 0) / rev)
    cm_data_ready = bool(cm_map)

    # ── 2. Pull placement fact rows ─────────────────────────────────────────
    qs = AdsPlacementDailySnapshot.objects.filter(
        marketplace=marketplace, date__gte=start, date__lte=end,
        source_ad_type='sp',  # only SP has placement data
    )
    if campaign_id:
        qs = qs.filter(campaign_id=campaign_id)

    # ── 3. Aggregate by placement ───────────────────────────────────────────
    placement_agg = defaultdict(lambda: {
        'impressions': 0, 'clicks': 0,
        'spend': 0.0, 'sales': 0.0, 'orders': 0,
        'est_profit': 0.0,
    })
    # Per-campaign × placement aggregation for the drill-down table
    campaign_placement_agg: dict[tuple, dict] = defaultdict(lambda: {
        'impressions': 0, 'clicks': 0,
        'spend': 0.0, 'sales': 0.0, 'orders': 0,
        'est_profit': 0.0,
    })
    campaigns_in_play = set()

    for r in qs.values('campaign_id', 'date', 'placement',
                        'impressions', 'clicks', 'spend',
                        'sales_7d', 'orders_7d'):
        plc = r['placement'] or 'other'
        cid = r['campaign_id']
        campaigns_in_play.add(cid)
        spend = float(r['spend'] or 0)
        sales = float(r['sales_7d'] or 0)
        cm_pct = cm_map.get((cid, r['date']), 0.0)
        est_profit = sales * cm_pct - spend

        for bucket in (placement_agg[plc], campaign_placement_agg[(cid, plc)]):
            bucket['impressions'] += int(r['impressions'] or 0)
            bucket['clicks']      += int(r['clicks']      or 0)
            bucket['spend']       += spend
            bucket['sales']       += sales
            bucket['orders']      += int(r['orders_7d']   or 0)
            bucket['est_profit']  += est_profit

    # ── 4. Window totals (for % denominators) ───────────────────────────────
    tot_spend  = sum(b['spend']      for b in placement_agg.values()) or 1e-9
    tot_sales  = sum(b['sales']      for b in placement_agg.values()) or 1e-9
    # For profit denominator, use sum of POSITIVE profit buckets only — so
    # negative-profit placements show a 0% share rather than a misleading
    # negative percentage. Cap at 100% in display layer.
    pos_profit_sum = sum(max(0, b['est_profit']) for b in placement_agg.values()) or 1e-9

    # ── 5. Build placement rows in fixed display order ──────────────────────
    placements_out = []
    for plc in _PLACEMENT_ORDER:
        b = placement_agg.get(plc) or {
            'impressions': 0, 'clicks': 0, 'spend': 0.0, 'sales': 0.0,
            'orders': 0, 'est_profit': 0.0,
        }
        spend = b['spend']; sales = b['sales']
        placements_out.append({
            'placement':         plc,
            'label':             _PLACEMENT_LABEL[plc],
            'impressions':       b['impressions'],
            'clicks':            b['clicks'],
            'ctr':               round(b['clicks'] / b['impressions'] * 100, 2)
                                   if b['impressions'] else None,
            'spend':             round(spend, 2),
            'spend_pct':         round(spend / tot_spend * 100, 2) if tot_spend > 0 else 0,
            'sales':             round(sales, 2),
            'sales_pct':         round(sales / tot_sales * 100, 2) if tot_sales > 0 else 0,
            'orders':            b['orders'],
            'acos':              round(spend / sales * 100, 2) if sales > 0 else None,
            'roas':              round(sales / spend, 2)       if spend > 0 else None,
            'estimated_profit':  round(b['est_profit'], 2),
            'profit_pct':        round(max(0, b['est_profit']) / pos_profit_sum * 100, 2)
                                   if pos_profit_sum > 0 else 0,
            'cpc':               round(spend / b['clicks'], 2) if b['clicks'] else None,
        })

    # Hide legacy buckets ('other', 'rest_of_search') from display when empty —
    # Amazon's current v3 placement report doesn't return these strings; they
    # remain in the schema only for backward-compat with historical data.
    _LEGACY = {'other', 'rest_of_search'}
    placements_out = [p for p in placements_out
                       if p['placement'] not in _LEGACY or p['impressions'] > 0]

    # ── 6. Per-campaign placement mix (top N by total spend) ────────────────
    # Campaign-name lookup with fallback (same pattern as search-term detail)
    cnames = {c.campaign_id: c.campaign_name for c in Campaign.objects.filter(
        marketplace=marketplace, campaign_id__in=campaigns_in_play
    ).only('campaign_id', 'campaign_name')}
    missing = campaigns_in_play - set(cnames)
    if missing:
        for r in PPCCampaignSnapshot.objects.filter(
            marketplace=marketplace, campaign_id__in=missing
        ).order_by('campaign_id', '-date').values('campaign_id', 'campaign_name'):
            cnames.setdefault(r['campaign_id'], r['campaign_name'])

    # Reorganize: campaign_id → {tos:{...}, ros:{...}, pp:{...}, total_spend}
    by_campaign: dict[str, dict] = defaultdict(lambda: {
        'campaign_id':       '',
        'campaign_name':     '',
        'total_spend':       0.0,
        'total_sales':       0.0,
        'total_profit':      0.0,
        'placements':        {p: None for p in _PLACEMENT_ORDER},
    })
    for (cid, plc), b in campaign_placement_agg.items():
        c = by_campaign[cid]
        c['campaign_id']   = cid
        c['campaign_name'] = cnames.get(cid, cid)
        c['total_spend']  += b['spend']
        c['total_sales']  += b['sales']
        c['total_profit'] += b['est_profit']
        # Per-placement row for this campaign
        c['placements'][plc] = {
            'spend':            round(b['spend'], 2),
            'sales':            round(b['sales'], 2),
            'orders':           b['orders'],
            'estimated_profit': round(b['est_profit'], 2),
            'spend_pct':        0,  # filled below
        }

    # Fill per-campaign spend_pct
    top_campaign_rows = []
    for c in sorted(by_campaign.values(), key=lambda x: x['total_spend'], reverse=True)[:30]:
        ts = c['total_spend'] or 1e-9
        for plc in _PLACEMENT_ORDER:
            p = c['placements'][plc]
            if p:
                p['spend_pct'] = round(p['spend'] / ts * 100, 2) if ts > 0 else 0
        top_campaign_rows.append({
            'campaign_id':    c['campaign_id'],
            'campaign_name':  c['campaign_name'],
            'total_spend':    round(c['total_spend'],  2),
            'total_sales':    round(c['total_sales'],  2),
            'total_profit':   round(c['total_profit'], 2),
            'placements':     c['placements'],
        })

    # ── 7. Top-level KPI ────────────────────────────────────────────────────
    total_profit = sum(p['estimated_profit'] for p in placements_out)
    kpi = {
        'total_spend':      round(tot_spend if tot_spend > 1e-9 else 0, 2),
        'total_sales':      round(tot_sales if tot_sales > 1e-9 else 0, 2),
        'total_profit':     round(total_profit, 2),
        'total_impressions': sum(p['impressions'] for p in placements_out),
        'total_clicks':      sum(p['clicks']      for p in placements_out),
        'total_orders':      sum(p['orders']      for p in placements_out),
        'campaigns':         len(campaigns_in_play),
    }

    return JsonResponse({
        'marketplace':   marketplace,
        'period': {'id': period_id, 'label': _CAMPAIGN_PERIODS[period_id][0],
                   'start': start.isoformat(), 'end': end.isoformat()},
        'campaign_id':   campaign_id or None,
        'kpi':           kpi,
        'placements':    placements_out,
        'top_campaigns': top_campaign_rows,
        'cm_data_ready': cm_data_ready,
        'cm_not_ready_warning': (None if cm_data_ready else
            'CampaignProfitDaily is not yet populated for this window — '
            'Estimated Profit currently shows −spend on every placement. '
            'Run `compute_campaign_profit --backfill-window 30` to fill in '
            'real profit numbers.'),
        'profit_proxy_note': ('Estimated Profit uses each campaign\'s blended '
                              'contribution-margin %. Per-SKU per-placement '
                              'attribution is not provided by Amazon.'),
    })


@login_required
@permission_required('can_view_dashboard')
def api_campaign_targeting(request, campaign_id: str):
    """
    Per-target rollup for one campaign over the selected window.

    Aggregates AdsTargetingDailySnapshot (keywords + product targets +
    audiences) by target_id. Returns one row per target with full metrics.
    """
    from .models import AdsTargetingDailySnapshot, CampaignProfitDaily
    from collections import defaultdict

    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        return JsonResponse({'error': 'forbidden'}, status=403)

    period_id = (request.GET.get('period') or '7d').lower()
    if period_id == 'today' or period_id not in _CAMPAIGN_PERIODS:
        period_id = '7d'
    today = date.today()
    start, end, _ = _resolve_campaign_period(period_id, today)

    # Campaign CM% map for proxy profit (same as search-term flow)
    cm_map: dict[date, float] = {}
    for r in CampaignProfitDaily.objects.filter(
        marketplace=marketplace, campaign_id=campaign_id,
        date__gte=start, date__lte=end,
    ).values('date', 'ad_revenue', 'contribution_margin'):
        rev = float(r['ad_revenue'] or 0)
        if rev > 0:
            cm_map[r['date']] = float(r['contribution_margin'] or 0) / rev
    cm_data_ready = bool(cm_map)

    qs = AdsTargetingDailySnapshot.objects.filter(
        marketplace=marketplace, campaign_id=campaign_id,
        date__gte=start, date__lte=end,
    ).values('date', 'target_id', 'target_type', 'expression', 'match_type',
             'impressions', 'clicks', 'spend', 'sales_7d', 'orders_7d',
             'units_7d')

    agg: dict[str, dict] = defaultdict(lambda: {
        'target_id': '', 'target_type': '', 'expression': '', 'match_types': set(),
        'impressions': 0, 'clicks': 0, 'spend': 0.0, 'sales': 0.0,
        'orders': 0, 'units': 0, 'est_profit': 0.0,
    })
    for r in qs:
        tid = r['target_id'] or '(unknown)'
        b = agg[tid]
        b['target_id']   = tid
        b['target_type'] = r['target_type'] or 'other'
        b['expression']  = r['expression']  or ''
        if r['match_type']:
            b['match_types'].add(r['match_type'])
        spend = float(r['spend']    or 0)
        sales = float(r['sales_7d'] or 0)
        b['impressions'] += int(r['impressions'] or 0)
        b['clicks']      += int(r['clicks']      or 0)
        b['spend']       += spend
        b['sales']       += sales
        b['orders']      += int(r['orders_7d']   or 0)
        b['units']       += int(r['units_7d']    or 0)
        b['est_profit']  += sales * cm_map.get(r['date'], 0.0) - spend

    rows = []
    for b in agg.values():
        spend = b['spend']; sales = b['sales']
        clicks = b['clicks']; impr = b['impressions']
        rows.append({
            'target_id':         b['target_id'],
            'target_type':       b['target_type'],
            'expression':        b['expression'],
            'match_types':       sorted(b['match_types']) or [''],
            'impressions':       impr,
            'clicks':            clicks,
            'ctr':               round(clicks / impr * 100, 2) if impr else None,
            'spend':             round(spend, 2),
            'cpc':               round(spend / clicks, 2) if clicks else None,
            'sales':             round(sales, 2),
            'orders':            b['orders'],
            'cvr':               round(b['orders'] / clicks * 100, 2) if clicks else None,
            'acos':              round(spend / sales * 100, 2) if sales else None,
            'roas':              round(sales / spend, 2)       if spend else None,
            'estimated_profit':  round(b['est_profit'], 2),
        })
    rows.sort(key=lambda r: r['spend'], reverse=True)

    return JsonResponse({
        'marketplace':   marketplace,
        'campaign_id':   campaign_id,
        'period': {'id': period_id, 'label': _CAMPAIGN_PERIODS[period_id][0],
                   'start': start.isoformat(), 'end': end.isoformat()},
        'rows':          rows,
        'target_count':  len(rows),
        'cm_data_ready': cm_data_ready,
        'profit_proxy_note': ('Estimated Profit uses the campaign\'s blended '
                              'contribution-margin %. Per-SKU per-target '
                              'attribution is not provided by Amazon.'),
    })


# ─── Campaign Hourly Performance (SP only) ──────────────────────────────────
#
# Powered by PPCCampaignHourlySnapshot — the AMS-sourced real hourly facts.
# SB and SD are not produced at hourly resolution by Amazon's APIs; the page
# surfaces an explanatory banner if a non-SP campaign is selected.

@login_required
@permission_required('can_view_dashboard')
def api_campaign_hourly(request, campaign_id: str):
    """
    Per-(date, hour) PPC facts for ONE campaign over a 7/14/30D window.

    Returns RAW underlying numbers (spend, sales, orders, clicks, impressions)
    so the JS client can switch the metric (ACoS, ROAS, CPC, CTR, CVR, …)
    without re-fetching.

    Response envelope:
      period:        {id, label, start, end}
      campaign_type: 'sp' | 'sb' | 'sd'  (banner gating)
      sp_only_msg:   string | null       (set if campaign isn't SP)
      days:          [{date, weekday, hours: [{hour, spend, sales, orders,
                                                clicks, impressions}, …24]}]
      dow_pattern:   weekday-averaged 7×24 grid (avg per (DOW, hour))
      window_totals: per-hour totals across the whole window  (used by best/worst)
      best/worst:    {by_hour: {acos, roas, cpc, …},  by_dow: {…}}
    """
    from .models import PPCCampaignHourlySnapshot, Campaign, PPCCampaignSnapshot
    from collections import defaultdict

    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        return JsonResponse({'error': 'forbidden'}, status=403)

    period_id = (request.GET.get('period') or '7d').lower()
    if period_id not in _CAMPAIGN_PERIODS or period_id == 'today':
        period_id = '7d'
    today = date.today()
    start, end, _ = _resolve_campaign_period(period_id, today)

    # ── Anchor the window to where the data actually ENDS ──────────────────
    # Manual CSV uploads and AMS S3 ingest can both fall behind. If the most
    # recent hourly row for THIS campaign is older than the calendar window's
    # end, we slide the window back so the user sees populated data instead
    # of a wall of empty rows. We also surface a freshness banner so they
    # know the data isn't up-to-date.
    from .models import PPCCampaignHourlySnapshot as _Hour
    latest_data_date = _Hour.objects.filter(
        marketplace=marketplace, campaign_id=campaign_id,
    ).aggregate(mx=models.Max('date'))['mx']

    freshness = None
    if latest_data_date and latest_data_date < end:
        # Slide window so it ENDS at the latest data row
        span_days = (end - start).days
        end   = latest_data_date
        start = end - timedelta(days=span_days)
        days_stale = (today - timedelta(days=1) - latest_data_date).days
        freshness = {
            'latest_data_date': latest_data_date.isoformat(),
            'days_stale':       days_stale,
            'window_anchored':  True,
            'message':          (f'Latest hourly data for this campaign: '
                                  f'{latest_data_date.isoformat()} ({days_stale} day(s) stale). '
                                  f'Window auto-anchored. Upload a fresh manual CSV '
                                  f'or check AMS S3 ingest if you expected newer data.'),
        }
    elif latest_data_date is None:
        freshness = {
            'latest_data_date': None, 'days_stale': None,
            'window_anchored':  False,
            'message':          'No hourly data found for this campaign. '
                                'PPCCampaignHourlySnapshot is empty for this campaign_id.',
        }

    # Campaign type — drives the SP-only banner. Look in the dim table first;
    # fall back to PPCCampaignSnapshot.
    ctype = (Campaign.objects.filter(marketplace=marketplace,
                                       campaign_id=campaign_id)
             .values_list('campaign_type', flat=True).first()
             or PPCCampaignSnapshot.objects.filter(
                 marketplace=marketplace, campaign_id=campaign_id
             ).order_by('-date').values_list('campaign_type', flat=True).first()
             or 'sp')

    sp_only_msg = None
    if ctype != 'sp':
        sp_only_msg = (f'This is a Sponsored {"Brands" if ctype=="sb" else "Display"} '
                        f'campaign. Amazon does not publish hourly facts for SB/SD — '
                        f'only SP exposes true hourly data via the Marketing Stream. '
                        f'Switch to an SP campaign to see hourly performance.')

    # Pre-fill days × hours grid with empty cells so the JS heatmap is dense.
    days_count = (end - start).days + 1
    by_day: dict[date, dict] = {}
    cur = start
    while cur <= end:
        by_day[cur] = {
            'date':    cur.isoformat(),
            'weekday': cur.weekday(),       # 0 = Mon, 6 = Sun
            'hours':   [{'hour': h, 'spend': 0.0, 'sales': 0.0,
                          'orders': 0, 'clicks': 0, 'impressions': 0}
                          for h in range(24)],
        }
        cur += timedelta(days=1)

    # Pull raw hourly rows. Empty result is FINE — heatmap stays gray-empty.
    rows = PPCCampaignHourlySnapshot.objects.filter(
        marketplace=marketplace, campaign_id=campaign_id,
        date__gte=start, date__lte=end,
    ).values('date', 'hour', 'spend', 'sales_7d', 'orders_7d',
             'clicks', 'impressions')

    for r in rows:
        d = r['date']
        h = int(r['hour'] or 0)
        if d not in by_day or not (0 <= h < 24):
            continue
        cell = by_day[d]['hours'][h]
        cell['spend']       += float(r['spend']    or 0)
        cell['sales']       += float(r['sales_7d'] or 0)
        cell['orders']      += int(r['orders_7d']  or 0)
        cell['clicks']      += int(r['clicks']     or 0)
        cell['impressions'] += int(r['impressions'] or 0)

    days = [by_day[d] for d in sorted(by_day)]

    # ── Day-of-week × hour pattern (averaged) ───────────────────────────────
    dow_sums = [[{'spend': 0.0, 'sales': 0.0, 'orders': 0,
                   'clicks': 0, 'impressions': 0, 'n_days': 0}
                  for _ in range(24)] for _ in range(7)]
    # n_days counted once per (dow, hour) per real day (the row exists in by_day)
    seen_days_per_dow = [set() for _ in range(7)]
    for d in by_day:
        seen_days_per_dow[d.weekday()].add(d)
    for d, day in by_day.items():
        dow = d.weekday()
        for h, cell in enumerate(day['hours']):
            s = dow_sums[dow][h]
            s['spend']       += cell['spend']
            s['sales']       += cell['sales']
            s['orders']      += cell['orders']
            s['clicks']      += cell['clicks']
            s['impressions'] += cell['impressions']
    dow_pattern = []
    for dow in range(7):
        n = max(len(seen_days_per_dow[dow]), 1)
        hours = []
        for h in range(24):
            s = dow_sums[dow][h]
            hours.append({
                'hour':        h,
                'spend':       round(s['spend'] / n, 2),
                'sales':       round(s['sales'] / n, 2),
                'orders':      round(s['orders'] / n, 2),
                'clicks':      round(s['clicks'] / n, 2),
                'impressions': round(s['impressions'] / n, 2),
            })
        dow_pattern.append({'weekday': dow, 'n_days': n, 'hours': hours})

    # ── Per-hour-of-day window totals + per-DOW window totals ──────────────
    # For "best/worst hour-of-day" + "best/worst day-of-week" callouts.
    hour_totals = [{'hour': h, 'spend': 0.0, 'sales': 0.0,
                     'orders': 0, 'clicks': 0, 'impressions': 0}
                    for h in range(24)]
    dow_totals  = [{'weekday': dow, 'spend': 0.0, 'sales': 0.0,
                     'orders': 0, 'clicks': 0, 'impressions': 0}
                    for dow in range(7)]
    for d, day in by_day.items():
        dow = d.weekday()
        for h, cell in enumerate(day['hours']):
            hour_totals[h]['spend']       += cell['spend']
            hour_totals[h]['sales']       += cell['sales']
            hour_totals[h]['orders']      += cell['orders']
            hour_totals[h]['clicks']      += cell['clicks']
            hour_totals[h]['impressions'] += cell['impressions']
            dow_totals[dow]['spend']       += cell['spend']
            dow_totals[dow]['sales']       += cell['sales']
            dow_totals[dow]['orders']      += cell['orders']
            dow_totals[dow]['clicks']      += cell['clicks']
            dow_totals[dow]['impressions'] += cell['impressions']

    def _derive(t):
        spend = t['spend']; sales = t['sales']
        clicks = t['clicks']; impr = t['impressions']; orders = t['orders']
        return {
            'acos': (spend / sales * 100) if sales > 0 else None,
            'roas': (sales / spend)        if spend > 0 else None,
            'cpc':  (spend / clicks)       if clicks > 0 else None,
            'ctr':  (clicks / impr * 100)  if impr   > 0 else None,
            'cvr':  (orders / clicks * 100) if clicks > 0 else None,
        }

    def _pick_best(rows, key, *, lower_is_better=False, min_spend=5):
        # Filter rows with enough spend to be meaningful — avoids "best hour
        # was 3am with $0.10 spend and 0% ACoS" outliers.
        elig = []
        for r in rows:
            if r['spend'] < min_spend: continue
            d = _derive(r)
            if d.get(key) is None: continue
            elig.append((d[key], r))
        if not elig:
            return None
        elig.sort(key=lambda kv: kv[0], reverse=not lower_is_better)
        best_val, best_row = elig[0]
        return {**best_row, key: round(float(best_val), 2)}

    best_worst = {
        'by_hour': {
            'best_roas':  _pick_best(hour_totals, 'roas', lower_is_better=False),
            'worst_roas': _pick_best(hour_totals, 'roas', lower_is_better=True),
            'best_acos':  _pick_best(hour_totals, 'acos', lower_is_better=True),
            'worst_acos': _pick_best(hour_totals, 'acos', lower_is_better=False),
        },
        'by_dow': {
            'best_roas':  _pick_best(dow_totals,  'roas', lower_is_better=False),
            'worst_roas': _pick_best(dow_totals,  'roas', lower_is_better=True),
            'best_acos':  _pick_best(dow_totals,  'acos', lower_is_better=True),
            'worst_acos': _pick_best(dow_totals,  'acos', lower_is_better=False),
        },
    }

    # Window-level totals + headline KPIs
    window = {
        'spend':       round(sum(t['spend']       for t in hour_totals), 2),
        'sales':       round(sum(t['sales']       for t in hour_totals), 2),
        'orders':      sum(t['orders']      for t in hour_totals),
        'clicks':      sum(t['clicks']      for t in hour_totals),
        'impressions': sum(t['impressions'] for t in hour_totals),
        'days_count':  days_count,
    }
    wd = _derive({'spend': window['spend'], 'sales': window['sales'],
                  'orders': window['orders'], 'clicks': window['clicks'],
                  'impressions': window['impressions']})
    window.update({
        'acos': round(wd['acos'], 2) if wd['acos'] is not None else None,
        'roas': round(wd['roas'], 2) if wd['roas'] is not None else None,
        'cpc':  round(wd['cpc'],  2) if wd['cpc']  is not None else None,
        'ctr':  round(wd['ctr'],  2) if wd['ctr']  is not None else None,
        'cvr':  round(wd['cvr'],  2) if wd['cvr']  is not None else None,
    })

    return JsonResponse({
        'marketplace':   marketplace,
        'campaign_id':   campaign_id,
        'campaign_type': ctype,
        'sp_only_msg':   sp_only_msg,
        'freshness':     freshness,
        'period': {'id': period_id, 'label': _CAMPAIGN_PERIODS[period_id][0],
                   'start': start.isoformat(), 'end': end.isoformat()},
        'days':           days,
        'dow_pattern':    dow_pattern,
        'hour_totals':    hour_totals,
        'dow_totals':     dow_totals,
        'best_worst':     best_worst,
        'window':         window,
        'metric_options': [
            {'id': 'spend',       'label': 'Spend',       'fmt': 'currency', 'higher_better': False},
            {'id': 'sales',       'label': 'Sales',       'fmt': 'currency', 'higher_better': True},
            {'id': 'orders',      'label': 'Orders',      'fmt': 'integer',  'higher_better': True},
            {'id': 'clicks',      'label': 'Clicks',      'fmt': 'integer',  'higher_better': True},
            {'id': 'impressions', 'label': 'Impressions', 'fmt': 'integer',  'higher_better': True},
            {'id': 'acos',        'label': 'ACoS',        'fmt': 'percent',  'higher_better': False, 'derived': True},
            {'id': 'roas',        'label': 'ROAS',        'fmt': 'ratio',    'higher_better': True,  'derived': True},
            {'id': 'cpc',         'label': 'CPC',         'fmt': 'currency', 'higher_better': False, 'derived': True},
            {'id': 'ctr',         'label': 'CTR',         'fmt': 'percent',  'higher_better': True,  'derived': True},
            {'id': 'cvr',         'label': 'CVR',         'fmt': 'percent',  'higher_better': True,  'derived': True},
        ],
    })


# ═════════════════════════════════════════════════════════════════════════════
# PHASE 1 — LEADERBOARDS
#
# /dashboard/leaderboards/      leaderboards (page shell)
# /dashboard/api/leaderboards/  api_leaderboards (all 6 boards in one call)
#
# 6 boards:
#   1. Most Profitable Campaigns        (top N by gross_profit)
#   2. Most Profitable Search Terms     (top N by est. profit; proxy method)
#   3. Highest Margin Campaigns         (top N by margin_pct, min-spend gated)
#   4. Highest Revenue Campaigns        (top N by ad_revenue)
#   5. Biggest Money Losers             (bottom N by gross_profit, i.e. most negative)
#   6. Highest Spend No Sales Terms     (top N by spend WHERE orders == 0)
# ═════════════════════════════════════════════════════════════════════════════

# Min spend gate for "Highest Margin Campaigns" — exclude tiny outlier
# campaigns ($1 of spend × 1 order = 9999% margin) from the ranking.
_MARGIN_BOARD_MIN_SPEND = Decimal('50.00')


@login_required
@permission_required('can_view_dashboard')
def leaderboards(request):
    """Page shell — JS fires a single /api/leaderboards/ call."""
    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        marketplace = _allowed_marketplaces(request.user)[0]

    ctx = {
        'marketplace':          marketplace,
        'allowed_marketplaces': _allowed_marketplaces(request.user),
        'period_options': [
            {'id': pid, 'label': v[0]} for pid, v in _CAMPAIGN_PERIODS.items()
            if pid != 'today'  # leaderboards need profit / margin which is T-1
        ],
        'today':                date.today(),
    }
    return render(request, 'dashboard/leaderboards.html', ctx)


@login_required
@permission_required('can_view_dashboard')
def api_leaderboards(request):
    """Returns all 6 boards in one envelope so the page renders in a single fetch."""
    from .models import (
        CampaignProfitDaily, AdsSearchTermDailySnapshot,
        Campaign, PPCCampaignSnapshot,
    )
    from collections import defaultdict

    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        return JsonResponse({'error': 'forbidden'}, status=403)

    period_id = (request.GET.get('period') or '7d').lower()
    if period_id == 'today' or period_id not in _CAMPAIGN_PERIODS:
        period_id = '7d'
    today = date.today()
    start, end, _ = _resolve_campaign_period(period_id, today)

    try:
        top_n = max(1, min(int(request.GET.get('limit') or 10), 50))
    except ValueError:
        top_n = 10

    # ── Campaign-level aggregation from CampaignProfitDaily ─────────────────
    camp_agg: dict[str, dict] = defaultdict(lambda: {
        'campaign_id':   '',
        'source_ad_type': 'sp',
        'spend':         Decimal('0'),
        'ad_revenue':    Decimal('0'),
        'gross_profit':  Decimal('0'),
        'orders':        0,
        'units':         0,
    })
    for r in CampaignProfitDaily.objects.filter(
        marketplace=marketplace, date__gte=start, date__lte=end,
    ).values('campaign_id', 'source_ad_type', 'spend', 'ad_revenue',
             'gross_profit', 'attributed_orders', 'attributed_units'):
        cid = r['campaign_id']
        b = camp_agg[cid]
        b['campaign_id']    = cid
        b['source_ad_type'] = r['source_ad_type']
        b['spend']         += Decimal(r['spend']        or 0)
        b['ad_revenue']    += Decimal(r['ad_revenue']   or 0)
        b['gross_profit']  += Decimal(r['gross_profit'] or 0)
        b['orders']        += int(r['attributed_orders'] or 0)
        b['units']         += int(r['attributed_units']  or 0)

    # Campaign-name lookup with fallback
    cnames = {c.campaign_id: c.campaign_name for c in Campaign.objects.filter(
        marketplace=marketplace, campaign_id__in=camp_agg
    ).only('campaign_id', 'campaign_name')}
    missing = set(camp_agg) - set(cnames)
    if missing:
        for r in PPCCampaignSnapshot.objects.filter(
            marketplace=marketplace, campaign_id__in=missing
        ).order_by('campaign_id', '-date').values('campaign_id', 'campaign_name'):
            cnames.setdefault(r['campaign_id'], r['campaign_name'])

    def camp_row(cid, b):
        spend = b['spend']; rev = b['ad_revenue']; prof = b['gross_profit']
        margin_pct = (prof / rev * 100) if rev > 0 else Decimal('0')
        return {
            'campaign_id':    cid,
            'campaign_name':  cnames.get(cid, cid),
            'campaign_type':  b['source_ad_type'],
            'spend':          round(float(spend), 2),
            'ad_revenue':     round(float(rev),   2),
            'gross_profit':   round(float(prof),  2),
            'margin_pct':     round(float(margin_pct), 2),
            'orders':         b['orders'],
            'units':          b['units'],
            'acos':           round(float(spend / rev * 100), 2) if rev > 0 else None,
            'roas':           round(float(rev / spend), 2) if spend > 0 else None,
        }

    all_camp_rows = [camp_row(cid, b) for cid, b in camp_agg.items()]

    # ── Boards 1, 3, 4, 5 are slices of the campaign aggregate ──────────────
    board_most_profit = sorted(
        all_camp_rows, key=lambda r: r['gross_profit'], reverse=True)[:top_n]
    board_money_losers = sorted(
        all_camp_rows, key=lambda r: r['gross_profit'])[:top_n]
    board_highest_rev = sorted(
        all_camp_rows, key=lambda r: r['ad_revenue'], reverse=True)[:top_n]
    # Margin board — only campaigns with meaningful spend, to avoid tiny outliers
    qual_for_margin = [r for r in all_camp_rows
                        if r['spend'] >= float(_MARGIN_BOARD_MIN_SPEND)
                        and r['ad_revenue'] > 0]
    board_highest_margin = sorted(
        qual_for_margin, key=lambda r: r['margin_pct'], reverse=True)[:top_n]

    # ── Boards 2 + 6 — search-term aggregation ──────────────────────────────
    # Pull a generous slice (top by spend or top by sales) to feed the aggregator.
    # This bounds the rows we have to process in Python while still surfacing
    # the genuine top-N after rollup.
    st_qs = AdsSearchTermDailySnapshot.objects.filter(
        marketplace=marketplace, date__gte=start, date__lte=end,
    ).values('campaign_id', 'date', 'search_term', 'search_term_hash',
             'spend', 'sales_7d', 'orders_7d', 'clicks', 'impressions'
    ).order_by('-spend')[:20000]

    # CM% map from CampaignProfitDaily
    cm_map: dict[tuple, float] = {}
    for r in CampaignProfitDaily.objects.filter(
        marketplace=marketplace, date__gte=start, date__lte=end,
    ).values('campaign_id', 'date', 'ad_revenue', 'contribution_margin'):
        rev = float(r['ad_revenue'] or 0)
        if rev > 0:
            cm_map[(r['campaign_id'], r['date'])] = (
                float(r['contribution_margin'] or 0) / rev)
    cm_data_ready = bool(cm_map)

    st_agg: dict[str, dict] = defaultdict(lambda: {
        'search_term': '',
        'spend': 0.0, 'sales': 0.0, 'orders': 0,
        'clicks': 0, 'impressions': 0, 'est_profit': 0.0,
        'campaigns': set(),
    })
    for r in st_qs:
        h = r['search_term_hash']
        b = st_agg[h]
        b['search_term'] = r['search_term']
        spend = float(r['spend']    or 0)
        sales = float(r['sales_7d'] or 0)
        b['spend']       += spend
        b['sales']       += sales
        b['orders']      += int(r['orders_7d']    or 0)
        b['clicks']      += int(r['clicks']       or 0)
        b['impressions'] += int(r['impressions']  or 0)
        b['est_profit']  += sales * cm_map.get((r['campaign_id'], r['date']), 0.0) - spend
        b['campaigns'].add(r['campaign_id'])

    def st_row(b):
        spend = b['spend']; sales = b['sales']
        return {
            'search_term':      b['search_term'],
            'spend':            round(spend, 2),
            'sales':            round(sales, 2),
            'orders':           b['orders'],
            'clicks':           b['clicks'],
            'impressions':      b['impressions'],
            'estimated_profit': round(b['est_profit'], 2),
            'acos':             round(spend / sales * 100, 2) if sales > 0 else None,
            'roas':             round(sales / spend, 2)       if spend > 0 else None,
            'cvr':              round(b['orders'] / b['clicks'] * 100, 2) if b['clicks'] else None,
            'campaign_count':   len(b['campaigns']),
        }

    all_st_rows = [st_row(b) for b in st_agg.values()]

    board_most_profit_terms = sorted(
        all_st_rows, key=lambda r: r['estimated_profit'], reverse=True)[:top_n]

    # Board 6 — High Spend No Sales: spend > $5, orders == 0
    no_sales_rows = [r for r in all_st_rows
                       if r['orders'] == 0 and r['spend'] > 5]
    board_no_sales = sorted(
        no_sales_rows, key=lambda r: r['spend'], reverse=True)[:top_n]

    return JsonResponse({
        'marketplace':   marketplace,
        'period': {'id': period_id, 'label': _CAMPAIGN_PERIODS[period_id][0],
                   'start': start.isoformat(), 'end': end.isoformat()},
        'top_n':         top_n,
        'cm_data_ready': cm_data_ready,
        'cm_not_ready_warning': (None if cm_data_ready else
            'CampaignProfitDaily is not yet populated for this window — '
            'profit-based rankings reflect −spend on every row. '
            'Run `compute_campaign_profit --backfill-window 30` to fill in '
            'real numbers.'),
        'boards': {
            'most_profitable_campaigns':    board_most_profit,
            'most_profitable_search_terms': board_most_profit_terms,
            'highest_margin_campaigns':     board_highest_margin,
            'highest_revenue_campaigns':    board_highest_rev,
            'biggest_money_losers':         board_money_losers,
            'highest_spend_no_sales_terms': board_no_sales,
        },
        'meta': {
            'campaigns_evaluated':  len(all_camp_rows),
            'search_terms_evaluated': len(all_st_rows),
            'margin_min_spend':     float(_MARGIN_BOARD_MIN_SPEND),
        },
    })


# ═════════════════════════════════════════════════════════════════════════════
# PHASE 2 — EXECUTIVE P&L CENTER (Sections 21A + 21B)
#
# /dashboard/pnl/                 pnl_daily (page shell)
# /dashboard/api/pnl/             api_pnl_daily (JSON: KPIs, waterfall, trend)
#
# Reads DailyMetric directly — no new ingestion needed. All P&L building
# blocks are already frozen at sync time:
#   revenue, ppc_spend, amazon_fee (referral), fba_fee, cgs (COGS),
#   contribution_margin, gross_margin, tacos
#
# Net Profit = revenue − ppc_spend − amazon_fee − fba_fee − cgs
# (matches gross_margin field; we recompute at read time for stability.)
# ═════════════════════════════════════════════════════════════════════════════

_PNL_PERIODS = {
    # id            label             days_back   anchor_offset_days
    'yesterday':    ('Yesterday',     1,    1),
    '7d':           ('Last 7D',       7,    1),
    '14d':          ('Last 14D',      14,   1),
    '30d':          ('Last 30D',      30,   1),
    'mtd':          ('Month-to-date', None, 1),  # special-cased
    'last_month':   ('Last Month',    None, 1),  # special-cased
}


def _resolve_pnl_period(period_id: str, today: date) -> tuple[date, date, date, date]:
    """
    Returns (start, end, prev_start, prev_end).
    prev_* covers the immediately previous window of equal length, for
    period-vs-period comparisons.
    """
    if period_id not in _PNL_PERIODS:
        period_id = '7d'
    label, days_back, anchor_offset = _PNL_PERIODS[period_id]

    if period_id == 'mtd':
        end = today - timedelta(days=anchor_offset)
        start = end.replace(day=1)
        span = (end - start).days + 1
        prev_end = start - timedelta(days=1)
        prev_start = prev_end - timedelta(days=span - 1)
        return start, end, prev_start, prev_end

    if period_id == 'last_month':
        # First day of current month minus one = last day of last month
        anchor = today - timedelta(days=anchor_offset)
        first_of_this_month = anchor.replace(day=1)
        end = first_of_this_month - timedelta(days=1)
        start = end.replace(day=1)
        span = (end - start).days + 1
        prev_end = start - timedelta(days=1)
        prev_start = prev_end - timedelta(days=span - 1)
        return start, end, prev_start, prev_end

    # Fixed-length windows (yesterday, 7d, 14d, 30d)
    end = today - timedelta(days=anchor_offset)
    start = end - timedelta(days=days_back - 1)
    prev_end = start - timedelta(days=1)
    prev_start = prev_end - timedelta(days=days_back - 1)
    return start, end, prev_start, prev_end


@login_required
@permission_required('can_view_dashboard')
def pnl_daily(request):
    """Page shell — JS calls /api/pnl/ for the data."""
    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        marketplace = _allowed_marketplaces(request.user)[0]

    ctx = {
        'marketplace':          marketplace,
        'allowed_marketplaces': _allowed_marketplaces(request.user),
        'period_options': [
            {'id': pid, 'label': v[0]} for pid, v in _PNL_PERIODS.items()
        ],
        'today':                date.today(),
    }
    return render(request, 'dashboard/pnl_daily.html', ctx)


@login_required
@permission_required('can_view_dashboard')
def api_pnl_daily(request):
    """
    Returns:
        period:        {id, label, start, end}
        prev_period:   {start, end}
        kpi:           {revenue, ppc_spend, referral_fee, fba_fee, cogs,
                          gross_profit, net_margin_pct, tacos, acos, roas,
                          orders, units}  — each with .current, .prev, .delta, .delta_pct
        waterfall:     ordered list of {label, type, value, running} for the
                          waterfall chart (Revenue → minus PPC → minus Referral → ...)
        trend:         per-day list for the trend chart
    """
    from .models import DailyMetric

    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        return JsonResponse({'error': 'forbidden'}, status=403)

    period_id = (request.GET.get('period') or '7d').lower()
    if period_id not in _PNL_PERIODS:
        period_id = '7d'
    today = date.today()
    start, end, prev_start, prev_end = _resolve_pnl_period(period_id, today)

    # ── Pull both windows in one query each ─────────────────────────────────
    cur_rows = list(DailyMetric.objects.filter(
        marketplace=marketplace, date__gte=start, date__lte=end
    ).values('date', 'revenue', 'units', 'orders',
             'ppc_spend', 'amazon_fee', 'fba_fee', 'cgs',
             'ppc_impressions', 'ppc_clicks'))

    prev_rows = list(DailyMetric.objects.filter(
        marketplace=marketplace, date__gte=prev_start, date__lte=prev_end
    ).values('date', 'revenue', 'units', 'orders',
             'ppc_spend', 'amazon_fee', 'fba_fee', 'cgs',
             'ppc_impressions', 'ppc_clicks'))

    def aggregate(rows):
        agg = {
            'revenue':      Decimal('0'),
            'ppc_spend':    Decimal('0'),
            'referral_fee': Decimal('0'),
            'fba_fee':      Decimal('0'),
            'cogs':         Decimal('0'),
            'orders':       0,
            'units':        0,
            'impressions':  0,
            'clicks':       0,
        }
        for r in rows:
            agg['revenue']      += Decimal(r['revenue']    or 0)
            agg['ppc_spend']    += Decimal(r['ppc_spend']  or 0)
            agg['referral_fee'] += Decimal(r['amazon_fee'] or 0)
            agg['fba_fee']      += Decimal(r['fba_fee']    or 0)
            agg['cogs']         += Decimal(r['cgs']        or 0)
            agg['orders']       += int(r['orders']         or 0)
            agg['units']        += int(r['units']          or 0)
            agg['impressions']  += int(r['ppc_impressions'] or 0)
            agg['clicks']       += int(r['ppc_clicks']      or 0)
        # Derived
        agg['gross_profit'] = (agg['revenue'] - agg['ppc_spend']
                                - agg['referral_fee'] - agg['fba_fee']
                                - agg['cogs'])
        agg['net_margin_pct'] = (agg['gross_profit'] / agg['revenue'] * 100
                                  if agg['revenue'] > 0 else Decimal('0'))
        agg['tacos'] = (agg['ppc_spend'] / agg['revenue'] * 100
                          if agg['revenue'] > 0 else Decimal('0'))
        agg['acos']  = (agg['ppc_spend'] / agg['revenue'] * 100
                          if agg['revenue'] > 0 else Decimal('0'))  # placeholder; needs ppc_sales for real ACoS
        return agg

    cur  = aggregate(cur_rows)
    prev = aggregate(prev_rows)

    def kpi(field, fmt='currency'):
        c = float(cur.get(field, 0) or 0)
        p = float(prev.get(field, 0) or 0)
        delta = c - p
        delta_pct = (delta / p * 100) if p else (100.0 if c else 0.0)
        return {'current': round(c, 2), 'previous': round(p, 2),
                'delta': round(delta, 2), 'delta_pct': round(delta_pct, 1),
                'fmt': fmt}

    kpis = {
        'revenue':         kpi('revenue'),
        'orders':          kpi('orders', fmt='int'),
        'units':           kpi('units',  fmt='int'),
        'ppc_spend':       kpi('ppc_spend'),
        'referral_fee':    kpi('referral_fee'),
        'fba_fee':         kpi('fba_fee'),
        'cogs':            kpi('cogs'),
        'gross_profit':    kpi('gross_profit'),
        'net_margin_pct':  kpi('net_margin_pct', fmt='percent'),
        'tacos':           kpi('tacos',          fmt='percent'),
    }

    # ── Waterfall data ──────────────────────────────────────────────────────
    rev   = float(cur['revenue'])
    ppc   = float(cur['ppc_spend'])
    ref   = float(cur['referral_fee'])
    fba   = float(cur['fba_fee'])
    cogs  = float(cur['cogs'])
    profit = float(cur['gross_profit'])

    waterfall = []
    running = 0.0
    waterfall.append({'label': 'Revenue',       'type': 'start', 'value': rev,    'running': rev})
    running = rev
    waterfall.append({'label': 'PPC Spend',     'type': 'minus', 'value': -ppc,   'running': running - ppc});  running -= ppc
    waterfall.append({'label': 'Referral Fees', 'type': 'minus', 'value': -ref,   'running': running - ref});  running -= ref
    waterfall.append({'label': 'FBA Fees',      'type': 'minus', 'value': -fba,   'running': running - fba});  running -= fba
    waterfall.append({'label': 'COGS',          'type': 'minus', 'value': -cogs,  'running': running - cogs}); running -= cogs
    waterfall.append({'label': 'Gross Profit',  'type': 'end',   'value': profit, 'running': profit})

    # ── Per-day trend ───────────────────────────────────────────────────────
    by_date = {r['date']: r for r in cur_rows}
    trend = []
    cur_date = start
    while cur_date <= end:
        r = by_date.get(cur_date)
        if r:
            d_rev = float(r['revenue']    or 0)
            d_ppc = float(r['ppc_spend']  or 0)
            d_ref = float(r['amazon_fee'] or 0)
            d_fba = float(r['fba_fee']    or 0)
            d_cogs= float(r['cgs']        or 0)
            d_profit = d_rev - d_ppc - d_ref - d_fba - d_cogs
            d_margin = (d_profit / d_rev * 100) if d_rev > 0 else 0
            d_tacos  = (d_ppc    / d_rev * 100) if d_rev > 0 else 0
        else:
            d_rev=d_ppc=d_ref=d_fba=d_cogs=d_profit=d_margin=d_tacos = 0
        trend.append({
            'date':         cur_date.isoformat(),
            'revenue':      round(d_rev, 2),
            'ppc_spend':    round(d_ppc, 2),
            'referral_fee': round(d_ref, 2),
            'fba_fee':      round(d_fba, 2),
            'cogs':         round(d_cogs, 2),
            'gross_profit': round(d_profit, 2),
            'net_margin_pct': round(d_margin, 2),
            'tacos':        round(d_tacos,  2),
            'orders':       int((r['orders'] if r else 0) or 0),
            'units':        int((r['units']  if r else 0) or 0),
        })
        cur_date += timedelta(days=1)

    return JsonResponse({
        'marketplace':  marketplace,
        'period': {'id': period_id, 'label': _PNL_PERIODS[period_id][0],
                   'start': start.isoformat(), 'end': end.isoformat(),
                   'days': (end - start).days + 1},
        'prev_period': {'start': prev_start.isoformat(), 'end': prev_end.isoformat()},
        'kpi':       kpis,
        'waterfall': waterfall,
        'trend':     trend,
    })


# ─── 21E — SKU Profitability ────────────────────────────────────────────────

# Auto-highlight thresholds. Conservative so the highlights actually mean
# something for the user.
_SKU_HIGH_REVENUE_MIN = Decimal('100')
_SKU_LOW_MARGIN_PCT   = Decimal('5')
_SKU_PROFITABLE_MIN   = Decimal('100')


"""
SKU Intelligence P1 — deterministic attention signals.

ONE implementation drives the row chips, the attention-summary counts and the
server-side signal filter (spec rule: never duplicate the business logic).
No composite score — every signal is an explicit threshold on existing metrics,
and the thresholds live in one dict so changing one changes it everywhere.
All comparisons are window vs the immediately-previous window of equal length
(_resolve_pnl_period already returns both).
"""
SKU_SIGNAL_THRESHOLDS = {
    'min_spend':        25.0,   # $ — below this a SKU is noise, not a signal
    'acos_up_rel':      0.20,   # ACOS worsened >20% relative vs prior window
    'ppc_dependent':    0.70,   # PPC share of revenue above 70%
    'organic_down_rel': 0.20,   # organic revenue down >20% while PPC flat/up
    'scaling_acos':     35.0,   # ACOS below this AND revenue growing →
    'scaling_rev_up':   0.15,   #   revenue up >15% = scaling candidate
    'low_conf':         0.50,   # spend-weighted allocation confidence below
    'capped_rate':      0.30,   # driver campaign budget-capped ≥30% of days
    'capped_min_share': 0.25,   # …and that campaign carries ≥25% of SKU spend
}

SKU_SIGNALS = [   # (id, label) — display order for the attention bar
    ('losing',        'Losing'),
    ('acos_up',       'ACOS ↑'),
    ('ppc_dependent', 'PPC-Dependent'),
    ('organic_down',  'Organic ↓'),
    ('scaling',       'Scaling'),
    ('capped',        'Capped'),
    ('low_conf',      'Low Confidence'),
]


def _sku_signals(*, profit, spend, revenue, prev_revenue, prev_spend,
                 ppc_sales, organic, prev_organic, prev_ppc_sales,
                 confidence, capped_driver) -> list[str]:
    """Deterministic signals for one SKU. All floats; None = unknown."""
    T = SKU_SIGNAL_THRESHOLDS
    out = []
    big = spend > T['min_spend']
    if profit < 0 and big:
        out.append('losing')
    # ACOS↑ — relative worsening vs prior window (both windows need revenue)
    if big and revenue > 0 and prev_revenue and prev_revenue > 0 and prev_spend:
        acos_now, acos_prev = spend / revenue, prev_spend / prev_revenue
        if acos_prev > 0 and (acos_now - acos_prev) / acos_prev > T['acos_up_rel']:
            out.append('acos_up')
    if revenue > 0 and ppc_sales is not None and ppc_sales / revenue > T['ppc_dependent']:
        out.append('ppc_dependent')
    if (organic is not None and prev_organic and prev_organic > 0
            and ppc_sales is not None and prev_ppc_sales is not None):
        org_drop = (prev_organic - organic) / prev_organic
        if org_drop > T['organic_down_rel'] and ppc_sales >= prev_ppc_sales * 0.95:
            out.append('organic_down')
    if (revenue > 0 and spend > 0 and prev_revenue and prev_revenue > 0
            and spend / revenue * 100 < T['scaling_acos']
            and (revenue - prev_revenue) / prev_revenue > T['scaling_rev_up']):
        out.append('scaling')
    if capped_driver:
        out.append('capped')
    if big and confidence is not None and confidence < T['low_conf']:
        out.append('low_conf')
    return out


@login_required
@permission_required('can_view_dashboard')
def pnl_skus(request):
    """Page shell — JS calls /api/pnl/skus/."""
    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        marketplace = _allowed_marketplaces(request.user)[0]
    ctx = {
        'marketplace':          marketplace,
        'allowed_marketplaces': _allowed_marketplaces(request.user),
        'period_options': [
            {'id': pid, 'label': v[0]} for pid, v in _PNL_PERIODS.items()
        ],
        'today':                date.today(),
    }
    return render(request, 'dashboard/pnl_skus.html', ctx)


@login_required
@permission_required('can_view_dashboard')
def api_pnl_skus(request):
    """
    Per-SKU P&L for the window.

    Inputs:
      • DailySkuSnapshot — per-SKU per-day revenue, qty, cgs, amz_fee, fulfill
      • SkuPpcAllocation — per-SKU per-day ad spend (already attributed)
      • Product           — brand + title for joining/display

    Returns rows with: sku, asin, product_name, brand, revenue, units,
    ppc_spend, referral_fee, fba_fee, cogs, gross_profit, margin_pct,
    acos, roas, tacos, contribution_pct, tags
    """
    from .models import DailySkuSnapshot, SkuPpcAllocation, Product
    from collections import defaultdict

    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        return JsonResponse({'error': 'forbidden'}, status=403)

    period_id = (request.GET.get('period') or '7d').lower()
    if period_id not in _PNL_PERIODS:
        period_id = '7d'
    today = date.today()
    start, end, prev_start, prev_end = _resolve_pnl_period(period_id, today)

    brand_filter = (request.GET.get('brand') or '').strip()
    tag_filter   = [t.strip() for t in (request.GET.get('tags') or '').split(',') if t.strip()]
    try:
        limit = max(1, min(int(request.GET.get('limit') or 500), 2000))
    except ValueError:
        limit = 500

    # ── 1. Aggregate DailySkuSnapshot ───────────────────────────────────────
    sku_agg: dict[str, dict] = defaultdict(lambda: {
        'sku': '', 'asin': '',
        'revenue': Decimal('0'), 'units': 0,
        'cogs': Decimal('0'), 'referral_fee': Decimal('0'), 'fba_fee': Decimal('0'),
    })
    for r in DailySkuSnapshot.objects.filter(
        marketplace=marketplace, date__gte=start, date__lte=end,
    ).values('sku', 'asin', 'revenue', 'qty', 'cgs', 'amz_fee', 'fulfill'):
        s = r['sku']
        b = sku_agg[s]
        b['sku']           = s
        b['asin']          = r['asin'] or b['asin']
        b['revenue']      += Decimal(r['revenue'] or 0)
        b['units']        += int(r['qty']         or 0)
        b['cogs']         += Decimal(r['cgs']     or 0)
        b['referral_fee'] += Decimal(r['amz_fee'] or 0)
        b['fba_fee']      += Decimal(r['fulfill'] or 0)

    # ── 2. PPC spend per SKU from SkuPpcAllocation ──────────────────────────
    for r in SkuPpcAllocation.objects.filter(
        marketplace=marketplace, date__gte=start, date__lte=end,
    ).values('sku').annotate(spend=Sum('sku_ppc_spend')):
        s = r['sku']
        if s in sku_agg:
            sku_agg[s]['ppc_spend'] = Decimal(r['spend'] or 0)
        else:
            sku_agg[s] = {**sku_agg[s], 'sku': s, 'ppc_spend': Decimal(r['spend'] or 0)}

    # Ensure every row has ppc_spend
    for b in sku_agg.values():
        b.setdefault('ppc_spend', Decimal('0'))

    # ── 2b. P1 signal inputs — all grouped queries over the two windows ─────
    from .models import AdsAdvertisedProductDailySnapshot, CampaignBudgetUsageDaily
    from django.db.models import F as _F

    # Prior window: revenue + spend per SKU (Δ columns, ACOS↑, SCALING).
    prev_rev = {r['sku']: (float(r['rev'] or 0), int(r['qty'] or 0))
                for r in DailySkuSnapshot.objects.filter(
                    marketplace=marketplace,
                    date__gte=prev_start, date__lte=prev_end)
                .values('sku').annotate(rev=Sum('revenue'), qty=Sum('qty'))}
    prev_spend = {r['sku']: float(r['sp'] or 0)
                  for r in SkuPpcAllocation.objects.filter(
                      marketplace=marketplace,
                      date__gte=prev_start, date__lte=prev_end)
                  .values('sku').annotate(sp=Sum('sku_ppc_spend'))}

    # Attributed PPC sales per SKU — SP/SD by advertised_sku (7d window);
    # SB is ASIN-level (14d) and is folded in via each SKU's ASIN, flagged
    # as an estimate. Same methodology as api_sku_campaigns; labelled, never
    # presented as exact organic attribution.
    def _ppc_sales_maps(s, e):
        by_sku = {r['advertised_sku']: float(r['sa'] or 0)
                  for r in AdsAdvertisedProductDailySnapshot.objects.filter(
                      marketplace=marketplace, date__gte=s, date__lte=e)
                  .exclude(source_ad_type='sb').exclude(advertised_sku='')
                  .values('advertised_sku').annotate(sa=Sum('sales_7d'))}
        by_asin = {r['asin']: float(r['sa'] or 0)
                   for r in AdsAdvertisedProductDailySnapshot.objects.filter(
                       marketplace=marketplace, source_ad_type='sb',
                       date__gte=s, date__lte=e)
                   .values('asin').annotate(sa=Sum('sales_7d'))}
        return by_sku, by_asin
    cur_ap_sku, cur_ap_asin = _ppc_sales_maps(start, end)
    prv_ap_sku, prv_ap_asin = _ppc_sales_maps(prev_start, prev_end)

    # Spend-weighted allocation confidence per SKU (LOW-CONF chip).
    conf_by_sku = {}
    for r in (SkuPpcAllocation.objects.filter(
                  marketplace=marketplace, date__gte=start, date__lte=end)
              .values('sku').annotate(sp=Sum('sku_ppc_spend'),
                                      wc=Sum(_F('confidence_score')
                                             * _F('sku_ppc_spend')))):
        sp = float(r['sp'] or 0)
        conf_by_sku[r['sku']] = (float(r['wc'] or 0) / sp) if sp > 0 else None

    # CAPPED chip: campaigns budget-capped ≥30% of active days, joined to
    # SKUs through their allocation spend (driver = ≥25% of the SKU's spend).
    T = SKU_SIGNAL_THRESHOLDS
    cap_agg = {}
    for r in (CampaignBudgetUsageDaily.objects.filter(
                  marketplace=marketplace, date__gte=start, date__lte=end)
              .values('campaign_id', 'usage_pct')):
        a = cap_agg.setdefault(str(r['campaign_id']), [0, 0])
        a[0] += 1
        if float(r['usage_pct'] or 0) >= CampaignBudgetUsageDaily.OUT_OF_BUDGET_PCT:
            a[1] += 1
    capped_campaigns = {cid for cid, (n, c) in cap_agg.items()
                        if n > 0 and c / n >= T['capped_rate']}
    capped_skus = set()
    if capped_campaigns:
        for r in (SkuPpcAllocation.objects.filter(
                      marketplace=marketplace, date__gte=start, date__lte=end)
                  .values('sku', 'campaign_id').annotate(sp=Sum('sku_ppc_spend'))):
            if str(r['campaign_id']) in capped_campaigns:
                tot = float(sku_agg.get(r['sku'], {}).get('ppc_spend', 0) or 0)
                if tot > 0 and float(r['sp'] or 0) / tot >= T['capped_min_share']:
                    capped_skus.add(r['sku'])

    # ── 3. Product meta lookup (brand + title) — with AMZN.GR.* fallback ────
    # Direct Product rows first, then resolve AMZN.GR.* variants to their parent
    # SKU's metadata so the Brand column shows the correct brand instead of '—'.
    skus_in_play = set(sku_agg.keys())
    pmeta = {p.sku: p for p in Product.objects.filter(
        marketplace=marketplace, sku__in=skus_in_play
    ).only('sku', 'title', 'brand', 'asin')}
    # Backfill from parent for the unmatched (variant) SKUs
    for sku in skus_in_play - set(pmeta):
        if sku.startswith('AMZN.GR.'):
            parent = _resolve_amzn_parent_product(marketplace, sku)
            if parent:
                pmeta[sku] = parent

    if brand_filter:
        sku_agg = {s: b for s, b in sku_agg.items()
                   if pmeta.get(s) and pmeta[s].brand == brand_filter}

    # ── 4. Compute derived columns + tags ───────────────────────────────────
    # Need account-level revenue + total profit for contribution % later
    total_profit = Decimal('0')
    rows = []
    for s, b in sku_agg.items():
        revenue = b['revenue']
        ppc     = b['ppc_spend']
        ref     = b['referral_fee']
        fba     = b['fba_fee']
        cogs    = b['cogs']
        profit  = revenue - ppc - ref - fba - cogs
        margin_pct = (profit / revenue * 100) if revenue > 0 else Decimal('0')
        acos       = (ppc    / revenue * 100) if revenue > 0 else None
        roas       = (revenue / ppc)           if ppc     > 0 else None
        tacos      = acos  # SKU-level TACoS = SKU-level ACoS when ad rev ≈ total SKU rev

        # Auto-highlight tags
        tags = []
        if profit < 0:
            tags.append('negative_margin')
        if revenue >= _SKU_HIGH_REVENUE_MIN and margin_pct < _SKU_LOW_MARGIN_PCT and margin_pct > 0:
            tags.append('high_revenue_low_profit')
        if profit >= _SKU_PROFITABLE_MIN:
            tags.append('profitable')
        if revenue > 0 and ppc / revenue > Decimal('0.30'):
            tags.append('ad_dependent')   # >30% TACoS

        # ── P1: PPC/organic estimate, deltas, deterministic signals ─────────
        rev_f, ppc_f = float(revenue), float(ppc)
        row_asin = (b['asin'] or '').upper()
        ppc_sales = cur_ap_sku.get(s, 0.0) + cur_ap_asin.get(row_asin, 0.0)
        # Organic = total − attributed ad sales. Attribution windows (7d/14d)
        # can overshoot a short revenue window — clamp and flag, never hide.
        organic, organic_flag = None, 'unavailable'
        if rev_f > 0 or ppc_sales > 0:
            raw = rev_f - ppc_sales
            organic, organic_flag = max(raw, 0.0), ('overshoot' if raw < 0 else 'est')
        p_rev = prev_rev.get(s, (0.0, 0))[0]
        p_sp = prev_spend.get(s, 0.0)
        p_ppc_sales = prv_ap_sku.get(s, 0.0) + prv_ap_asin.get(row_asin, 0.0)
        p_organic = max(p_rev - p_ppc_sales, 0.0) if p_rev > 0 else None
        rev_d = ((rev_f - p_rev) / p_rev * 100) if p_rev > 0 else None
        sp_d = ((ppc_f - p_sp) / p_sp * 100) if p_sp > 0 else None
        acos_d = None
        if rev_f > 0 and p_rev > 0 and p_sp > 0:
            a_now, a_prev = ppc_f / rev_f, p_sp / p_rev
            acos_d = (a_now - a_prev) / a_prev * 100 if a_prev > 0 else None
        confidence = conf_by_sku.get(s)
        signals = _sku_signals(
            profit=float(profit), spend=ppc_f, revenue=rev_f,
            prev_revenue=p_rev, prev_spend=p_sp, ppc_sales=ppc_sales,
            organic=organic, prev_organic=p_organic,
            prev_ppc_sales=p_ppc_sales, confidence=confidence,
            capped_driver=(s in capped_skus))

        meta = pmeta.get(s)
        rows.append({
            'sku':           s,
            'asin':          (meta.asin if meta else b['asin']) or '',
            'product_name':  (meta.title if meta else '')[:96],
            'brand':         (meta.brand if meta else '') or '',
            'revenue':       round(float(revenue), 2),
            'units':         b['units'],
            'ppc_spend':     round(float(ppc),  2),
            'referral_fee':  round(float(ref),  2),
            'fba_fee':       round(float(fba),  2),
            'cogs':          round(float(cogs), 2),
            'gross_profit':  round(float(profit), 2),
            'margin_pct':    round(float(margin_pct), 2),
            'acos':          round(float(acos),  2) if acos is not None else None,
            'roas':          round(float(roas),  2) if roas is not None else None,
            'tacos':         round(float(tacos), 2) if tacos is not None else None,
            'tags':          tags,
            # P1 — PPC/organic split (estimate), deltas, signals, confidence
            'ppc_sales':     round(ppc_sales, 2),
            'organic':       round(organic, 2) if organic is not None else None,
            'organic_flag':  organic_flag,
            'ppc_share_pct': round(ppc_sales / rev_f * 100, 1) if rev_f > 0 else None,
            'revenue_delta_pct': round(rev_d, 1) if rev_d is not None else None,
            'spend_delta_pct':   round(sp_d, 1) if sp_d is not None else None,
            'acos_delta_pct':    round(acos_d, 1) if acos_d is not None else None,
            'confidence':    round(confidence, 2) if confidence is not None else None,
            'signals':       signals,
        })
        if profit > 0:
            total_profit += profit

    # ── 5. Apply tag filter + contribution % + sort + limit ─────────────────
    if tag_filter:
        wanted = set(tag_filter)
        rows = [r for r in rows if any(t in wanted for t in r['tags'])]

    # P1 — attention counts are computed over the population BEFORE the signal
    # filter is applied (so the bar always shows the full picture), and the
    # same signal list drives the filter — one implementation.
    signal_counts = {sid: sum(1 for r in rows if sid in r['signals'])
                     for sid, _lbl in SKU_SIGNALS}

    # Signal filter — AND-combined (Losing + High ACOS + PPC-dependent).
    signal_filter = [x.strip() for x in
                     (request.GET.get('signals') or '').split(',') if x.strip()]
    if signal_filter:
        rows = [r for r in rows
                if all(sid in r['signals'] for sid in signal_filter)]

    # Compound numeric conditions — same syntax + helpers as the Marketing
    # Optimizer (metric:op:value;…), server-side.
    from .views_marketing import _parse_conds, _num_ok
    for metric, op, val in _parse_conds(request.GET.get('cond')):
        rows = [r for r in rows if _num_ok(r.get(metric), op, val)]

    for r in rows:
        if total_profit > 0 and r['gross_profit'] > 0:
            r['contribution_pct'] = round(r['gross_profit'] / float(total_profit) * 100, 2)
        else:
            r['contribution_pct'] = 0.0

    rows.sort(key=lambda r: r['gross_profit'], reverse=True)
    truncated = len(rows) > limit
    rows = rows[:limit]

    # ── KPI strip ────────────────────────────────────────────────────────────
    kpi = {
        'sku_count':      len(sku_agg),
        'shown':          len(rows),
        'truncated':      truncated,
        'revenue':        round(float(sum(Decimal(str(r['revenue']))      for r in rows)), 2),
        'gross_profit':   round(float(sum(Decimal(str(r['gross_profit'])) for r in rows)), 2),
        'ppc_spend':      round(float(sum(Decimal(str(r['ppc_spend']))    for r in rows)), 2),
        'units':          sum(r['units'] for r in rows),
        'tag_counts': {
            'negative_margin':         sum(1 for r in rows if 'negative_margin' in r['tags']),
            'high_revenue_low_profit': sum(1 for r in rows if 'high_revenue_low_profit' in r['tags']),
            'profitable':              sum(1 for r in rows if 'profitable' in r['tags']),
            'ad_dependent':            sum(1 for r in rows if 'ad_dependent' in r['tags']),
        },
        'signal_counts': signal_counts,
    }

    # ── Brand list (for the brand-filter dropdown) ──────────────────────────
    brands = sorted({(meta.brand or '') for meta in pmeta.values()
                      if meta.brand} | {''}, key=lambda x: (x == '', x))

    return JsonResponse({
        'marketplace': marketplace,
        'period': {'id': period_id, 'label': _PNL_PERIODS[period_id][0],
                   'start': start.isoformat(), 'end': end.isoformat()},
        'brand_filter': brand_filter or None,
        'brands':       brands,
        'tag_filter':   tag_filter,
        'kpi':          kpi,
        'rows':         rows,
    })


# ─── SKU Intelligence P0 — SKU → Campaign drivers ───────────────────────────

def _campaign_budget_action(marketplace, campaign_id, start, end):
    """P4 — propose a budget increase ONLY where the evidence supports it.

    Conditions (all reused, none invented):
      • budget-capped on ≥30% of active days   — Budget & Pacing's cap rate
      • ACOS below 35%                          — the P1 'scaling' efficiency bar
      • a CURRENT budget is actually stored     — else no honest before/after

    Returns None when any condition fails: the opportunity stays diagnostic and
    no executable recommendation is offered.
    """
    from decimal import Decimal

    from . import ad_actions as ACT
    from .models import CampaignBudgetUsageDaily, CampaignProfitDaily

    rows = list(CampaignBudgetUsageDaily.objects.filter(
        marketplace=marketplace, campaign_id=str(campaign_id),
        date__gte=start, date__lte=end).values('usage_pct'))
    if not rows:
        return None
    capped = sum(1 for r in rows
                 if float(r['usage_pct'] or 0)
                 >= CampaignBudgetUsageDaily.OUT_OF_BUDGET_PCT)
    cap_rate = capped / len(rows)
    if cap_rate < SKU_SIGNAL_THRESHOLDS['capped_rate']:
        return None

    agg = CampaignProfitDaily.objects.filter(
        marketplace=marketplace, campaign_id=str(campaign_id),
        date__gte=start, date__lte=end).aggregate(sp=Sum('spend'),
                                                  rv=Sum('ad_revenue'))
    spend, rev = float(agg['sp'] or 0), float(agg['rv'] or 0)
    acos = (spend / rev * 100) if rev > 0 else None
    if acos is None or acos >= SKU_SIGNAL_THRESHOLDS['scaling_acos']:
        return None      # constrained but not efficient — budget is not the fix

    current, as_of, source = ACT.current_campaign_budget(marketplace, campaign_id)
    if current is None or current <= 0:
        return None      # no current value → no honest "current → proposed"

    proposed = (current * Decimal('1.30')).quantize(Decimal('0.01'))
    return {
        'action_type': 'campaign_budget',
        'campaign_id': str(campaign_id),
        'current_value': float(current),
        'proposed_value': float(proposed),
        'change_pct': 30.0,
        'current_source': source,
        'current_as_of': as_of.isoformat() if as_of else None,
        'reason': (f'Hit its daily budget on {capped} of {len(rows)} active days '
                   f'({cap_rate * 100:.0f}%) while holding ACOS at {acos:.1f}% — '
                   f'constrained and efficient, so the cap is limiting sales.'),
        'evidence': [
            {'label': 'Capped days',   'value': f'{capped}/{len(rows)} ({cap_rate * 100:.0f}%)'},
            {'label': 'ACOS',          'value': f'{acos:.1f}%'},
            {'label': 'Spend',         'value': f'${spend:,.0f}'},
            {'label': 'Ad revenue',    'value': f'${rev:,.0f}'},
            {'label': 'Current budget','value': f'${float(current):,.2f}/day'},
        ],
        'period': {'start': start.isoformat(), 'end': end.isoformat()},
        'capability': ACT.write_capability(marketplace),
        'note': ('Diagnosis with a proposed step. Nothing is sent to Amazon '
                 'until a person reviews and approves it.'),
    }


@login_required
@permission_required('can_view_dashboard')
def api_campaign_opportunities(request, campaign_id: str):
    """
    P3 — contextual opportunities for ONE campaign. Lazy: only runs when the
    Opportunities tab is opened (P3.13).

    Reuses, never rebuilds:
      • AdsTargetingDailySnapshot  — target rows (same source as the Targeting tab)
      • AdsSearchTermDailySnapshot — term rows  (same source as the Search Terms tab)
      • _tag_search_term           — the EXISTING deterministic rules that already
                                     drive the Search Terms page and the Optimizer
      • StiOpportunity / AIRecommendation — surfaced, not regenerated

    Diagnosis only — nothing here changes a bid, budget, target or anything on
    Amazon.
    """
    from collections import defaultdict

    from . import opportunities as OPP
    from .models import (AdsSearchTermDailySnapshot, AdsTargetingDailySnapshot,
                         AIRecommendation, StiOpportunity)

    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        return JsonResponse({'error': 'forbidden'}, status=403)

    period_id = (request.GET.get('period') or '7d').lower()
    if period_id == 'today' or period_id not in _CAMPAIGN_PERIODS:
        period_id = '7d'
    today = date.today()
    start, end, _ = _resolve_campaign_period(period_id, today)
    from_sku = (request.GET.get('from_sku') or '')[:64]

    # ── Targets — one grouped query, same fact table as the Targeting tab ───
    tgt = defaultdict(lambda: {'spend': 0.0, 'sales': 0.0, 'orders': 0,
                               'clicks': 0, 'impressions': 0,
                               'expression': '', 'target_id': ''})
    for r in (AdsTargetingDailySnapshot.objects
              .filter(marketplace=marketplace, campaign_id=campaign_id,
                      date__gte=start, date__lte=end)
              .values('target_id', 'expression')
              .annotate(sp=Sum('spend'), sa=Sum('sales_7d'), o=Sum('orders_7d'),
                        c=Sum('clicks'), i=Sum('impressions'))):
        b = tgt[r['target_id']]
        b['target_id'] = r['target_id']
        b['expression'] = r['expression'] or r['target_id']
        b['spend'] += float(r['sp'] or 0); b['sales'] += float(r['sa'] or 0)
        b['orders'] += int(r['o'] or 0); b['clicks'] += int(r['c'] or 0)
        b['impressions'] += int(r['i'] or 0)

    # ── Search terms — one grouped query, same fact table as the ST tab ─────
    st = defaultdict(lambda: {'spend': 0.0, 'sales': 0.0, 'orders': 0,
                              'clicks': 0, 'impressions': 0, 'search_term': ''})
    for r in (AdsSearchTermDailySnapshot.objects
              .filter(marketplace=marketplace, campaign_id=campaign_id,
                      date__gte=start, date__lte=end)
              .values('search_term_hash', 'search_term')
              .annotate(sp=Sum('spend'), sa=Sum('sales_7d'), o=Sum('orders_7d'),
                        c=Sum('clicks'), i=Sum('impressions'))
              .order_by('-sp')[:800]):
        b = st[r['search_term_hash']]
        b['search_term'] = r['search_term']
        b['spend'] += float(r['sp'] or 0); b['sales'] += float(r['sa'] or 0)
        b['orders'] += int(r['o'] or 0); b['clicks'] += int(r['c'] or 0)
        b['impressions'] += int(r['i'] or 0)

    def _derive(d):
        d['ctr'] = (d['clicks'] / d['impressions'] * 100) if d['impressions'] else None
        d['cvr'] = (d['orders'] / d['clicks'] * 100) if d['clicks'] else None
        d['estimated_profit'] = 0.0     # profit proxy belongs to the ST tab
        return d

    cards = OPP.entity_opportunities(
        rows=[_derive(v) for v in tgt.values()], level='target',
        campaign_id=campaign_id, tagger=_tag_search_term,
        drill_builder=lambda r: {'label': 'Search terms for this target',
                                 'target_id': r['target_id']})
    cards += OPP.entity_opportunities(
        rows=[_derive(v) for v in st.values()], level='search_term',
        campaign_id=campaign_id, tagger=_tag_search_term)

    # ── Existing systems, surfaced (never regenerated) ──────────────────────
    ai = list(AIRecommendation.objects.filter(
        marketplace=marketplace, scope_type='campaign', scope_id=str(campaign_id),
        status__in=['new', 'acknowledged']).order_by('-rank_score')[:3])
    # STI opportunities whose subject is a term this campaign actually ran —
    # that is the only defensible campaign linkage STI's schema supports.
    terms_lc = {v['search_term'].lower() for v in st.values() if v['search_term']}
    sti = []
    if terms_lc:
        sti = [o for o in StiOpportunity.objects.filter(
                   marketplace=marketplace, status='open').order_by('-score')[:200]
               if (o.subject or '').lower() in terms_lc][:4]
    subjects = {c['subject'].lower() for c in cards if c.get('subject')}
    cards += OPP.sti_cards(sti, level='campaign', dedupe_subjects=subjects)
    cards += OPP.ai_cards(ai, level='campaign')
    cards = OPP.dedupe(cards)

    # ── P4 — the ONE actionable recommendation this data supports ───────────
    # A budget increase is proposable only when the campaign is demonstrably
    # constrained AND efficient, using the SAME thresholds the Budget & Pacing
    # page already applies (cap ≥30% of active days; ACOS below target). Every
    # other opportunity above stays diagnostic (P4.5): organic decline does not
    # imply "spend more", and PPC dependency does not imply "spend less".
    action = _campaign_budget_action(marketplace, campaign_id, start, end)

    return JsonResponse({
        'action': action,
        'marketplace': marketplace, 'campaign_id': campaign_id,
        'period': {'id': period_id, 'label': _CAMPAIGN_PERIODS[period_id][0],
                   'start': start.isoformat(), 'end': end.isoformat()},
        'from_sku': from_sku or None,
        'counts': {'targets': len(tgt), 'search_terms': len(st),
                   'opportunities': len(cards)},
        'opportunities': cards,
        'note': ('Diagnosis only — Pulse never changes bids, budgets or targets. '
                 'Target and search-term rules are the same ones used by the '
                 'Search Terms page.'),
    })


def _sku_opportunity_cards(*, marketplace, sku, asin, rows, context, start, end,
                           prev_revenue, prev_spend, prev_ppc_sales, prev_organic):
    """P3 — contextual opportunities for one SKU.

    Deterministic cards come from numbers already computed in this request
    (driver shares + P1 context); STI and AIRecommendation records are
    SURFACED, never regenerated. Two small indexed queries are added — both
    only run when a SKU row is expanded (P3.13).
    """
    from . import opportunities as OPP
    from .models import AIRecommendation, Product, ProductGroup, StiOpportunity

    # P1 signals, recomputed from the SAME engine (no second rule set).
    signals = _sku_signals(
        profit=0.0,   # profit-based signals belong to the table, not this panel
        spend=float(context.get('spend') or 0),
        revenue=float(context.get('revenue') or 0),
        prev_revenue=prev_revenue, prev_spend=prev_spend,
        ppc_sales=float(context.get('ppc_sales') or 0),
        organic=context.get('organic'), prev_organic=prev_organic,
        prev_ppc_sales=prev_ppc_sales, confidence=None, capped_driver=False)

    def _campaign_url(cid):
        return (f'/dashboard/campaigns/{cid}/?mp={marketplace}'
                f'&from_sku={sku}')

    cards = OPP.sku_opportunities(sku=sku, driver_rows=rows, context=context,
                                  signals=signals, campaign_url=_campaign_url)

    # ── Existing STI opportunities for the group this SKU belongs to ────────
    # StiOpportunity is scoped to a ProductGroup; a SKU joins through its
    # catalogue category. Open items only — status lives on the STI record.
    sti = []
    try:
        cat = (Product.objects.filter(marketplace=marketplace, sku=sku)
               .values_list('category', flat=True).first())
        if cat:
            gids = [g.pk for g in ProductGroup.objects.filter(active=True)
                    if cat in (g.categories or [])]
            if gids:
                sti = list(StiOpportunity.objects
                           .filter(product_group_id__in=gids,
                                   marketplace=marketplace, status='open')
                           .order_by('-score')[:3])
    except Exception:
        sti = []      # STI is optional context — never break the panel over it

    # ── Existing AI recommendations scoped to this SKU ──────────────────────
    ai = list(AIRecommendation.objects.filter(
        marketplace=marketplace, scope_type='sku', scope_id=sku,
        status__in=['new', 'acknowledged']).order_by('-rank_score')[:3])

    # Pulse deterministic first, then STI, then AI (P3.6: measured data wins).
    subjects = {c['subject'].lower() for c in cards if c.get('subject')}
    cards += OPP.sti_cards(sti, level='sku', dedupe_subjects=subjects)
    cards += OPP.ai_cards(ai, level='sku')
    return OPP.dedupe(cards)


@login_required
@permission_required('can_view_dashboard')
def api_sku_campaigns(request):
    """
    Campaign drivers for ONE SKU over the selected period — the reverse of
    api_campaign_top_skus. Read-only exposure layer over the existing
    attribution engine; changes no methodology.

    Sources (each one grouped query — no N+1):
      • SkuPpcAllocation   — authoritative per-SKU PPC spend per campaign,
                             with attribution_source / confidence / settlement.
      • AdsAdvertisedProductDailySnapshot — attributed sales/orders/units per
                             campaign. SP+SD join by SKU; SB joins by ASIN
                             (purchasedProduct has no SKU and uses a 14-day
                             window) and is flagged, never silently merged.
      • Campaign           — dim: name / type / state.

    Query params: mp, period (same _PNL_PERIODS vocabulary as api_pnl_skus),
                  sku (required), asin (optional; resolved from the latest
                  DailySkuSnapshot row when absent — needed for the SB join).

    Response rows sorted by spend desc ("where is the money going?"); the UI
    renders spend-share vs sales-share as paired bars so the mismatch reads
    at a glance.
    """
    from .models import (AdsAdvertisedProductDailySnapshot, Campaign,
                         DailySkuSnapshot, SkuPpcAllocation)
    from django.db.models import F, Q

    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        return JsonResponse({'error': 'forbidden'}, status=403)

    sku = (request.GET.get('sku') or '').strip()
    if not sku:
        return JsonResponse({'error': 'sku is required'}, status=400)

    period_id = (request.GET.get('period') or '7d').lower()
    if period_id not in _PNL_PERIODS:
        period_id = '7d'
    today = date.today()
    start, end, _ps, _pe = _resolve_pnl_period(period_id, today)

    asin = (request.GET.get('asin') or '').strip()
    if not asin:
        snap = (DailySkuSnapshot.objects
                .filter(marketplace=marketplace, sku=sku)
                .exclude(asin='').order_by('-date')
                .values_list('asin', flat=True).first())
        asin = snap or ''

    # ── 1. Per-campaign spend AND sales — BOTH from the advertised-product
    #      report, so the two halves of every comparison share one source.
    #
    # Why not SkuPpcAllocation for the per-campaign split: the allocation
    # engine splits each SP campaign's spend across ASINs using ACCOUNT-WIDE
    # daily proportions (ppc_allocator._pass1_sp — its own comment notes that
    # per-campaign/per-ASIN spend "isn't stored"). That spreads a SKU's spend
    # across every SP campaign in the account, so campaigns that never
    # advertised this SKU appear as spend-with-no-sales. Amazon's
    # advertised-product report DOES carry per-campaign per-SKU spend, so the
    # driver table uses it. SkuPpcAllocation remains authoritative for the
    # SKU's TOTAL spend (below) and for SKU-level P&L, which is untouched.
    ap_filter = Q(advertised_sku__iexact=sku)
    if asin:
        ap_filter |= Q(asin__iexact=asin, source_ad_type='sb')
    ap = (AdsAdvertisedProductDailySnapshot.objects
          .filter(marketplace=marketplace, date__gte=start, date__lte=end)
          .filter(ap_filter)
          .values('campaign_id', 'source_ad_type')
          .annotate(spend=Sum('spend'), sales=Sum('sales_7d'),
                    orders=Sum('orders_7d'), units=Sum('units_7d')))
    camp: dict[str, dict] = {}
    for r in ap:
        cid = str(r['campaign_id'])
        c = camp.setdefault(cid, {
            'campaign_id': cid, 'campaign_type': '', 'spend': 0.0,
            'sales': 0.0, 'orders': 0, 'units': 0,
            'sb_asin_level': False, 'wconf': 0.0, 'sources': {},
            'settlement': 'locked'})
        c['spend'] += float(r['spend'] or 0)
        c['sales'] += float(r['sales'] or 0)
        c['orders'] += int(r['orders'] or 0)
        c['units'] += int(r['units'] or 0)
        c['campaign_type'] = c['campaign_type'] or (r['source_ad_type'] or '')
        if r['source_ad_type'] == 'sb':
            c['sb_asin_level'] = True

    # ── 2. Allocation: SKU total (reconciliation) + attribution quality ─────
    # Still read, for two things only: the SKU's authoritative total spend
    # (which reconciles to campaign totals) and the confidence / settlement
    # badges. Never for the per-campaign split.
    _SETTLE_RANK = {'provisional': 0, 'settling': 1, 'locked': 2}
    alloc_total = 0.0
    worst_settlement = 'locked'
    for r in (SkuPpcAllocation.objects
              .filter(marketplace=marketplace, sku__iexact=sku,
                      date__gte=start, date__lte=end)
              .values('campaign_id', 'attribution_source', 'settlement_state')
              .annotate(spend=Sum('sku_ppc_spend'),
                        wconf=Sum(F('confidence_score') * F('sku_ppc_spend')))):
        sp = float(r['spend'] or 0)
        alloc_total += sp
        if (_SETTLE_RANK.get(r['settlement_state'], 0)
                < _SETTLE_RANK.get(worst_settlement, 2)):
            worst_settlement = r['settlement_state']
        c = camp.get(str(r['campaign_id']))
        if c is not None:
            c['wconf'] += float(r['wconf'] or 0)
            c['sources'][r['attribution_source']] = (
                c['sources'].get(r['attribution_source'], 0.0) + sp)
            if (_SETTLE_RANK.get(r['settlement_state'], 0)
                    < _SETTLE_RANK.get(c['settlement'], 2)):
                c['settlement'] = r['settlement_state']
    for c in camp.values():
        c['settlement'] = c['settlement'] if c['sources'] else worst_settlement

    # ── 3. Campaign dim ─────────────────────────────────────────────────────
    dim = {str(d['campaign_id']): d for d in
           Campaign.objects.filter(marketplace=marketplace,
                                   campaign_id__in=list(camp))
           .values('campaign_id', 'campaign_name', 'campaign_type', 'state')}

    total_spend = sum(c['spend'] for c in camp.values())
    total_sales = sum(c['sales'] for c in camp.values())
    # Spend the allocator assigned to this SKU that the advertised-product
    # report does not tie to any specific campaign. Shown, never hidden.
    residual = round(alloc_total - total_spend, 2)

    rows = []
    for cid, c in camp.items():
        d = dim.get(cid, {})
        # Every row now originates in Amazon's own per-campaign per-SKU report,
        # so spend and sales are always from the same source and comparable.
        linked = True
        s = c
        spend, sales = c['spend'], c['sales']
        # Dominant attribution source = the one carrying the most spend.
        dominant = max(c['sources'], key=c['sources'].get) if c['sources'] else ''
        conf = (c['wconf'] / spend) if spend > 0 else None
        rows.append({
            'campaign_id':    cid,
            'campaign_name':  d.get('campaign_name') or cid,
            'campaign_type':  d.get('campaign_type') or c['campaign_type'] or '',
            'state':          d.get('state') or '',
            'linked':         linked,
            'spend':          round(spend, 2),
            'spend_share':    round(spend / total_spend * 100, 1) if total_spend > 0 else 0.0,
            'ppc_sales':      round(sales, 2) if linked else None,
            'sales_share':    (round(sales / total_sales * 100, 1)
                               if (linked and total_sales > 0) else (0.0 if linked else None)),
            'orders':         s['orders'] if linked else None,
            'units':          s['units'] if linked else None,
            'acos':           round(spend / sales * 100, 1) if (linked and sales > 0) else None,
            'roas':           (round(sales / spend, 2) if (linked and spend > 0) else None),
            'attribution_source': dominant,
            'source_mix':     {k: round(v / spend * 100)
                               for k, v in c['sources'].items()} if spend > 0 else {},
            'confidence':     round(conf, 2) if conf is not None else None,
            'settlement':     c['settlement'],
            'sb_asin_level':  s['sb_asin_level'],
        })
    rows.sort(key=lambda r: r['spend'], reverse=True)

    # ── P1: SKU context + what-changed + daily trend ────────────────────────
    # Same sources the table already trusts — no second calculation engine.
    def _sku_window(s0, e0):
        d0 = (DailySkuSnapshot.objects
              .filter(marketplace=marketplace, sku=sku,
                      date__gte=s0, date__lte=e0)
              .aggregate(rev=Sum('revenue'), qty=Sum('qty')))
        sp0 = (SkuPpcAllocation.objects
               .filter(marketplace=marketplace, sku__iexact=sku,
                       date__gte=s0, date__lte=e0)
               .aggregate(s=Sum('sku_ppc_spend'))['s'])
        ap0 = (AdsAdvertisedProductDailySnapshot.objects
               .filter(marketplace=marketplace, date__gte=s0, date__lte=e0)
               .filter(ap_filter)
               .aggregate(s=Sum('sales_7d'))['s'])
        return (float(d0['rev'] or 0), int(d0['qty'] or 0),
                float(sp0 or 0), float(ap0 or 0))

    rev_c, units_c, spend_c, psales_c = _sku_window(start, end)
    rev_p, units_p, spend_p, psales_p = _sku_window(_ps, _pe)
    raw_org = rev_c - psales_c
    organic_c = max(raw_org, 0.0) if (rev_c > 0 or psales_c > 0) else None
    organic_p = max(rev_p - psales_p, 0.0) if rev_p > 0 else None

    def _pct(cur, prev):
        return round((cur - prev) / prev * 100, 1) if prev and prev > 0 else None

    context = {
        'revenue': round(rev_c, 2), 'units': units_c,
        'ppc_sales': round(psales_c, 2),
        'organic': round(organic_c, 2) if organic_c is not None else None,
        'organic_flag': ('overshoot' if raw_org < 0 else 'est')
                        if organic_c is not None else 'unavailable',
        'ppc_share_pct': round(psales_c / rev_c * 100, 1) if rev_c > 0 else None,
        'spend': round(spend_c, 2),
        'acos': round(spend_c / psales_c * 100, 1) if psales_c > 0 else None,
        'tacos': round(spend_c / rev_c * 100, 1) if rev_c > 0 else None,
        'roas': round(psales_c / spend_c, 2) if spend_c > 0 else None,
        'deltas': {'revenue': _pct(rev_c, rev_p),
                   'ppc_sales': _pct(psales_c, psales_p),
                   'organic': _pct(organic_c or 0, organic_p),
                   'spend': _pct(spend_c, spend_p)},
    }

    # Daily trend — the selected window, widened to ≥30 days so a 1-day view
    # still shows a readable line. Organic/day derived client-side (rev−ppc).
    t_start = min(start, end - timedelta(days=29))
    t_rev = {r['date']: float(r['rev'] or 0)
             for r in DailySkuSnapshot.objects
             .filter(marketplace=marketplace, sku=sku,
                     date__gte=t_start, date__lte=end)
             .values('date').annotate(rev=Sum('revenue'))}
    t_sp = {r['date']: float(r['s'] or 0)
            for r in SkuPpcAllocation.objects
            .filter(marketplace=marketplace, sku__iexact=sku,
                    date__gte=t_start, date__lte=end)
            .values('date').annotate(s=Sum('sku_ppc_spend'))}
    t_ps = {r['date']: float(r['s'] or 0)
            for r in AdsAdvertisedProductDailySnapshot.objects
            .filter(marketplace=marketplace, date__gte=t_start, date__lte=end)
            .filter(ap_filter)
            .values('date').annotate(s=Sum('sales_7d'))}
    trend = []
    dcur = t_start
    while dcur <= end:
        trend.append({'date': dcur.isoformat(),
                      'revenue': round(t_rev.get(dcur, 0.0), 2),
                      'ppc_sales': round(t_ps.get(dcur, 0.0), 2),
                      'spend': round(t_sp.get(dcur, 0.0), 2)})
        dcur += timedelta(days=1)

    # Group-fallback allocation can spread residual cents across hundreds of
    # campaigns; those rows are noise, not drivers. Drop sub-cent rows, cap the
    # list, and report how much of the spend the shown rows cover so nothing
    # is hidden silently. Shares stay computed against the FULL totals.
    total_campaigns = len(rows)
    rows = [r for r in rows if r['spend'] >= 0.01 or (r['ppc_sales'] or 0) >= 0.01]
    try:
        limit = max(1, min(int(request.GET.get('limit') or 25), 200))
    except ValueError:
        limit = 25
    truncated = len(rows) > limit
    shown_spend = sum(r['spend'] for r in rows[:limit])
    rows = rows[:limit]

    return JsonResponse({
        'marketplace': marketplace,
        'sku': sku, 'asin': asin,
        'period': {'id': period_id, 'label': _PNL_PERIODS[period_id][0],
                   'start': start.isoformat(), 'end': end.isoformat()},
        'totals': {'spend': round(total_spend, 2),
                   'ppc_sales': round(total_sales, 2),
                   'campaigns': total_campaigns,
                   'shown': len(rows), 'truncated': truncated,
                   'shown_spend_pct': round(shown_spend / total_spend * 100, 1)
                                      if total_spend > 0 else None,
                   'acos': round(total_spend / total_sales * 100, 1)
                           if total_sales > 0 else None,
                   # Reconciliation, shown rather than hidden.
                   'allocated_total': round(alloc_total, 2),
                   'residual': residual,
                   'residual_pct': round(residual / alloc_total * 100, 1)
                                   if alloc_total > 0 else None},
        'context': context,
        'trend': trend,
        'opportunities': _sku_opportunity_cards(
            marketplace=marketplace, sku=sku, asin=asin, rows=rows,
            context=context, start=start, end=end,
            prev_revenue=rev_p, prev_spend=spend_p,
            prev_ppc_sales=psales_p, prev_organic=organic_p),
        'rows': rows,
        'notes': {
            'spend_source': ("Amazon's advertised-product report — per-campaign, "
                             "per-SKU spend"),
            'sales_source': ('Same report — SP/SD by SKU (7-day attribution); '
                             'SB by ASIN (14-day window, may span sibling SKUs '
                             'of the same ASIN)'),
            'residual': (
                f'The allocation engine assigns ${alloc_total:,.2f} of ad spend '
                f'to this SKU in total; ${total_spend:,.2f} of that is tied to a '
                f'specific campaign by the advertised-product report. The '
                f'${residual:,.2f} difference is spend the engine spread across '
                f'campaigns using account-wide proportions and is not shown '
                f'against individual campaigns here.') if residual > 0.01 else '',
        },
    })


# ─── 21C + 21D — Marketplace + Brand P&L Breakdowns ─────────────────────────

@login_required
@permission_required('can_view_dashboard')
def pnl_breakdown(request):
    """Page shell — JS calls /api/pnl/breakdown/ for both Marketplace and Brand boards."""
    ctx = {
        'allowed_marketplaces': _allowed_marketplaces(request.user),
        'period_options': [
            {'id': pid, 'label': v[0]} for pid, v in _PNL_PERIODS.items()
        ],
        'today':                date.today(),
    }
    return render(request, 'dashboard/pnl_breakdown.html', ctx)


@login_required
@permission_required('can_view_dashboard')
def api_pnl_breakdown(request):
    """Returns both Marketplace and Brand P&L boards in one envelope.

    Marketplace board: aggregate DailyMetric per marketplace (only marketplaces
        the user has access to).
    Brand board:        aggregate DailySkuSnapshot joined with Product.brand;
        scoped to a single marketplace (passed as `mp`, default 'usa').
    """
    from .models import DailyMetric, DailySkuSnapshot, Product, SkuPpcAllocation
    from collections import defaultdict

    period_id = (request.GET.get('period') or '7d').lower()
    if period_id not in _PNL_PERIODS:
        period_id = '7d'
    today = date.today()
    start, end, _, _ = _resolve_pnl_period(period_id, today)

    # ── 1. Marketplace board — over all marketplaces the user can see ────────
    user_mps = _allowed_marketplaces(request.user)
    mp_agg: dict[str, dict] = {mp: {
        'marketplace':    mp,
        'revenue':        Decimal('0'),
        'ppc_spend':      Decimal('0'),
        'referral_fee':   Decimal('0'),
        'fba_fee':        Decimal('0'),
        'cogs':           Decimal('0'),
        'orders':         0,
        'units':          0,
    } for mp in user_mps}

    for r in DailyMetric.objects.filter(
        marketplace__in=user_mps, date__gte=start, date__lte=end,
    ).values('marketplace', 'revenue', 'ppc_spend', 'amazon_fee', 'fba_fee',
             'cgs', 'orders', 'units'):
        b = mp_agg[r['marketplace']]
        b['revenue']      += Decimal(r['revenue']    or 0)
        b['ppc_spend']    += Decimal(r['ppc_spend']  or 0)
        b['referral_fee'] += Decimal(r['amazon_fee'] or 0)
        b['fba_fee']      += Decimal(r['fba_fee']    or 0)
        b['cogs']         += Decimal(r['cgs']        or 0)
        b['orders']       += int(r['orders']         or 0)
        b['units']        += int(r['units']          or 0)

    total_profit_mp = Decimal('0')
    mp_rows_temp = []
    for mp, b in mp_agg.items():
        profit = b['revenue'] - b['ppc_spend'] - b['referral_fee'] - b['fba_fee'] - b['cogs']
        margin_pct = (profit / b['revenue'] * 100) if b['revenue'] > 0 else Decimal('0')
        tacos      = (b['ppc_spend'] / b['revenue'] * 100) if b['revenue'] > 0 else Decimal('0')
        mp_rows_temp.append({
            'marketplace':   mp,
            'revenue':       round(float(b['revenue']), 2),
            'ppc_spend':     round(float(b['ppc_spend']), 2),
            'referral_fee':  round(float(b['referral_fee']), 2),
            'fba_fee':       round(float(b['fba_fee']), 2),
            'cogs':          round(float(b['cogs']), 2),
            'gross_profit':  round(float(profit), 2),
            'margin_pct':    round(float(margin_pct), 2),
            'tacos':         round(float(tacos), 2),
            'orders':        b['orders'],
            'units':         b['units'],
        })
        if profit > 0:
            total_profit_mp += profit

    for r in mp_rows_temp:
        if total_profit_mp > 0 and r['gross_profit'] > 0:
            r['contribution_pct'] = round(r['gross_profit'] / float(total_profit_mp) * 100, 2)
        else:
            r['contribution_pct'] = 0.0
    mp_rows = sorted(mp_rows_temp, key=lambda r: r['gross_profit'], reverse=True)

    # ── 2. Brand board — scoped to one marketplace (default usa) ────────────
    mp = request.GET.get('mp', 'usa')
    if mp not in user_mps:
        mp = user_mps[0]

    # Per-SKU rollup, then group by brand
    sku_pl: dict[str, dict] = defaultdict(lambda: {
        'revenue': Decimal('0'), 'units': 0,
        'cogs': Decimal('0'), 'referral_fee': Decimal('0'), 'fba_fee': Decimal('0'),
        'ppc_spend': Decimal('0'),
    })
    for r in DailySkuSnapshot.objects.filter(
        marketplace=mp, date__gte=start, date__lte=end,
    ).values('sku', 'revenue', 'qty', 'cgs', 'amz_fee', 'fulfill'):
        b = sku_pl[r['sku']]
        b['revenue']      += Decimal(r['revenue'] or 0)
        b['units']        += int(r['qty'] or 0)
        b['cogs']         += Decimal(r['cgs']     or 0)
        b['referral_fee'] += Decimal(r['amz_fee'] or 0)
        b['fba_fee']      += Decimal(r['fulfill'] or 0)

    for r in SkuPpcAllocation.objects.filter(
        marketplace=mp, date__gte=start, date__lte=end,
    ).values('sku').annotate(spend=Sum('sku_ppc_spend')):
        sku_pl[r['sku']]['ppc_spend'] = Decimal(r['spend'] or 0)

    # Map sku → brand via Product. Falls back to AMZN.GR.* parent-SKU lookup
    # so Vine/Launchpad/gift variants get rolled into the parent's brand
    # instead of polluting the "(unbranded)" bucket.
    brand_by_sku: dict[str, str] = {}
    for sku in sku_pl.keys():
        p = _resolve_amzn_parent_product(mp, sku)
        brand_by_sku[sku] = (p.brand if (p and p.brand) else '(unbranded)')

    brand_agg: dict[str, dict] = defaultdict(lambda: {
        'brand': '',
        'sku_count': 0,
        'revenue': Decimal('0'), 'units': 0,
        'cogs': Decimal('0'), 'referral_fee': Decimal('0'), 'fba_fee': Decimal('0'),
        'ppc_spend': Decimal('0'),
    })
    for s, b in sku_pl.items():
        br = brand_by_sku.get(s, '(unbranded)')
        bb = brand_agg[br]
        bb['brand']        = br
        bb['sku_count']   += 1
        bb['revenue']     += b['revenue']
        bb['units']       += b['units']
        bb['cogs']        += b['cogs']
        bb['referral_fee']+= b['referral_fee']
        bb['fba_fee']     += b['fba_fee']
        bb['ppc_spend']   += b['ppc_spend']

    total_profit_br = Decimal('0')
    br_rows_temp = []
    for br, b in brand_agg.items():
        profit = b['revenue'] - b['ppc_spend'] - b['referral_fee'] - b['fba_fee'] - b['cogs']
        margin_pct = (profit / b['revenue'] * 100) if b['revenue'] > 0 else Decimal('0')
        tacos      = (b['ppc_spend'] / b['revenue'] * 100) if b['revenue'] > 0 else Decimal('0')
        br_rows_temp.append({
            'brand':         br,
            'sku_count':     b['sku_count'],
            'revenue':       round(float(b['revenue']), 2),
            'ppc_spend':     round(float(b['ppc_spend']), 2),
            'referral_fee':  round(float(b['referral_fee']), 2),
            'fba_fee':       round(float(b['fba_fee']), 2),
            'cogs':          round(float(b['cogs']), 2),
            'gross_profit':  round(float(profit), 2),
            'margin_pct':    round(float(margin_pct), 2),
            'tacos':         round(float(tacos), 2),
            'units':         b['units'],
        })
        if profit > 0:
            total_profit_br += profit

    for r in br_rows_temp:
        if total_profit_br > 0 and r['gross_profit'] > 0:
            r['contribution_pct'] = round(r['gross_profit'] / float(total_profit_br) * 100, 2)
        else:
            r['contribution_pct'] = 0.0
    br_rows = sorted(br_rows_temp, key=lambda r: r['gross_profit'], reverse=True)

    return JsonResponse({
        'period': {'id': period_id, 'label': _PNL_PERIODS[period_id][0],
                   'start': start.isoformat(), 'end': end.isoformat()},
        'marketplace_for_brand_board': mp,
        'marketplaces': mp_rows,
        'brands':       br_rows,
    })


# ─── 21I — Executive Morning Report ─────────────────────────────────────────

@login_required
@permission_required('can_view_dashboard')
def morning_report(request):
    """Page shell — JS calls /api/morning-report/."""
    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        marketplace = _allowed_marketplaces(request.user)[0]
    ctx = {
        'marketplace':          marketplace,
        'allowed_marketplaces': _allowed_marketplaces(request.user),
        'today':                date.today(),
    }
    return render(request, 'dashboard/morning_report.html', ctx)


@login_required
@permission_required('can_view_dashboard')
def api_morning_report(request):
    """
    Single-shot 30-second briefing on yesterday's business.

    Returns a fully composed summary:
      • Yesterday's headline KPIs vs the day before
      • Best campaign (most profit) and most profitable SKU
      • Biggest concern (high-spend-no-sales term, or biggest losing campaign,
        or margin collapse, whichever is most actionable)
      • Top recommendation (highest-ROAS scaling opportunity)
      • Profit movement attribution — what drove yesterday's profit change
    """
    from .models import (
        DailyMetric, DailySkuSnapshot, SkuPpcAllocation, Product,
        CampaignProfitDaily, Campaign, AdsSearchTermDailySnapshot,
        PPCCampaignSnapshot,
    )

    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        return JsonResponse({'error': 'forbidden'}, status=403)

    today = date.today()
    yday = today - timedelta(days=1)
    dby  = today - timedelta(days=2)

    # ── 1. Yesterday's headline P&L vs day-before ───────────────────────────
    def day_pl(d):
        r = DailyMetric.objects.filter(marketplace=marketplace, date=d).first()
        if not r:
            return None
        rev = float(r.revenue);    ppc = float(r.ppc_spend)
        ref = float(r.amazon_fee); fba = float(r.fba_fee)
        cogs = float(r.cgs)
        profit = rev - ppc - ref - fba - cogs
        return {
            'date':     d.isoformat(),
            'revenue':  round(rev, 2),
            'ppc':      round(ppc, 2),
            'referral': round(ref, 2),
            'fba':      round(fba, 2),
            'cogs':     round(cogs, 2),
            'profit':   round(profit, 2),
            'margin_pct': round((profit / rev * 100) if rev > 0 else 0, 2),
            'tacos':      round((ppc    / rev * 100) if rev > 0 else 0, 2),
            'acos':       round((ppc    / rev * 100) if rev > 0 else 0, 2),
            'orders':   r.orders,
            'units':    r.units,
        }
    yday_pl = day_pl(yday)
    dby_pl  = day_pl(dby)

    def delta(a, b):
        if a is None or b is None:
            return None
        d = a - b
        return {'delta': round(d, 2),
                'pct':   round((d / b * 100), 1) if b else (100.0 if a else 0.0)}

    headline_deltas = {}
    if yday_pl and dby_pl:
        for k in ('revenue', 'profit', 'margin_pct', 'tacos', 'ppc', 'orders', 'units'):
            headline_deltas[k] = delta(yday_pl[k], dby_pl[k])

    # ── 2. Best campaign yesterday (most profit) ────────────────────────────
    best_camp = CampaignProfitDaily.objects.filter(
        marketplace=marketplace, date=yday, gross_profit__gt=0,
    ).order_by('-gross_profit').values(
        'campaign_id', 'gross_profit', 'spend', 'ad_revenue', 'margin_pct'
    ).first()
    if best_camp:
        cname = (Campaign.objects.filter(
            marketplace=marketplace, campaign_id=best_camp['campaign_id']
        ).values_list('campaign_name', flat=True).first()
                  or PPCCampaignSnapshot.objects.filter(
            marketplace=marketplace, campaign_id=best_camp['campaign_id']
        ).order_by('-date').values_list('campaign_name', flat=True).first()
                  or best_camp['campaign_id'])
        best_camp = {
            'campaign_id':  best_camp['campaign_id'],
            'campaign_name': cname,
            'profit':       round(float(best_camp['gross_profit']), 2),
            'revenue':      round(float(best_camp['ad_revenue']),   2),
            'spend':        round(float(best_camp['spend']),        2),
            'margin_pct':   round(float(best_camp['margin_pct']),   2),
        }

    # ── 3. Most profitable SKU yesterday ────────────────────────────────────
    sku_rows = list(DailySkuSnapshot.objects.filter(
        marketplace=marketplace, date=yday
    ).values('sku', 'revenue', 'qty', 'cgs', 'amz_fee', 'fulfill'))
    sku_ppc = dict(SkuPpcAllocation.objects.filter(
        marketplace=marketplace, date=yday
    ).values_list('sku', 'sku_ppc_spend'))

    best_sku = None
    for r in sku_rows:
        s = r['sku']
        rev = float(r['revenue'] or 0)
        if rev <= 0:
            continue
        ppc  = float(sku_ppc.get(s, 0) or 0)
        cogs = float(r['cgs']     or 0)
        ref  = float(r['amz_fee'] or 0)
        fba  = float(r['fulfill'] or 0)
        profit = rev - ppc - ref - fba - cogs
        margin = (profit / rev * 100) if rev > 0 else 0
        if best_sku is None or profit > best_sku['profit']:
            pmeta = Product.objects.filter(marketplace=marketplace, sku=s).only('title').first()
            best_sku = {
                'sku': s,
                'product_name': pmeta.title if pmeta else '',
                'revenue': round(rev, 2),
                'profit':  round(profit, 2),
                'margin_pct': round(margin, 2),
            }

    # ── 4. Biggest concern — pick the most actionable ───────────────────────
    concern = None

    # 4a. Margin collapse (>20% drop day-over-day)
    if (yday_pl and dby_pl and dby_pl['margin_pct'] > 0
        and yday_pl['margin_pct'] < dby_pl['margin_pct'] * 0.8):
        concern = {
            'severity': 'critical',
            'type':     'margin_collapse',
            'title':    'Margin collapse',
            'message':  (f"Yesterday's margin dropped to {yday_pl['margin_pct']:.1f}% "
                          f"from {dby_pl['margin_pct']:.1f}% the day before — "
                          f"a {(yday_pl['margin_pct'] - dby_pl['margin_pct']):.1f} pt drop."),
        }
    # 4b. Highest-spend campaign with negative profit yesterday
    if not concern:
        worst_camp = CampaignProfitDaily.objects.filter(
            marketplace=marketplace, date=yday, spend__gt=10
        ).order_by('gross_profit').values(
            'campaign_id', 'gross_profit', 'spend'
        ).first()
        if worst_camp and float(worst_camp['gross_profit'] or 0) < -10:
            cname = (Campaign.objects.filter(
                marketplace=marketplace, campaign_id=worst_camp['campaign_id']
            ).values_list('campaign_name', flat=True).first()
                      or PPCCampaignSnapshot.objects.filter(
                marketplace=marketplace, campaign_id=worst_camp['campaign_id']
            ).order_by('-date').values_list('campaign_name', flat=True).first()
                      or worst_camp['campaign_id'])
            concern = {
                'severity': 'warning',
                'type':     'losing_campaign',
                'title':    'Biggest money-losing campaign yesterday',
                'message':  (f"Campaign \"{cname}\" lost "
                              f"${abs(float(worst_camp['gross_profit'])):,.0f} on "
                              f"${float(worst_camp['spend']):,.0f} of spend yesterday."),
                'campaign_id': worst_camp['campaign_id'],
            }
    # 4c. Top high-spend-no-sales search term yesterday
    if not concern:
        from collections import defaultdict
        st_no_sales: dict[str, dict] = defaultdict(lambda: {'spend': 0.0, 'clicks': 0, 'campaigns': 0})
        for r in AdsSearchTermDailySnapshot.objects.filter(
            marketplace=marketplace, date=yday, orders_7d=0,
        ).values('search_term', 'spend', 'clicks').order_by('-spend')[:50]:
            b = st_no_sales[r['search_term']]
            b['spend']  += float(r['spend']  or 0)
            b['clicks'] += int(r['clicks']   or 0)
        if st_no_sales:
            term, b = max(st_no_sales.items(), key=lambda kv: kv[1]['spend'])
            if b['spend'] > 5:
                concern = {
                    'severity': 'warning',
                    'type':     'wasted_term',
                    'title':    'Wasted-spend search term yesterday',
                    'message':  (f"Search term \"{term}\" spent ${b['spend']:.0f} on "
                                  f"{b['clicks']} clicks with zero orders yesterday."),
                }

    # ── 5. Top recommendation — highest-ROAS profitable campaign ────────────
    recommend = None
    scale_camp = CampaignProfitDaily.objects.filter(
        marketplace=marketplace, date=yday,
        gross_profit__gt=50, spend__gt=10, margin_pct__gt=0.10,  # >10% margin
    ).order_by('-roas').values(
        'campaign_id', 'gross_profit', 'margin_pct', 'roas', 'spend', 'ad_revenue'
    ).first()
    if scale_camp:
        cname = (Campaign.objects.filter(
            marketplace=marketplace, campaign_id=scale_camp['campaign_id']
        ).values_list('campaign_name', flat=True).first()
                  or PPCCampaignSnapshot.objects.filter(
            marketplace=marketplace, campaign_id=scale_camp['campaign_id']
        ).order_by('-date').values_list('campaign_name', flat=True).first()
                  or scale_camp['campaign_id'])
        recommend = {
            'severity': 'opportunity',
            'type':     'scale_campaign',
            'title':    'Scaling opportunity',
            'message':  (f"Campaign \"{cname}\" returned a {float(scale_camp['roas']):.2f}x ROAS "
                          f"with a {float(scale_camp['margin_pct']):.1f}% net margin yesterday "
                          f"(${float(scale_camp['gross_profit']):,.0f} profit on "
                          f"${float(scale_camp['spend']):,.0f} spend). "
                          f"Consider increasing budget."),
            'campaign_id': scale_camp['campaign_id'],
        }

    # ── 6. Profit movement attribution ──────────────────────────────────────
    # Simple: decompose the change in profit into its cost-bucket components
    movement = None
    if yday_pl and dby_pl:
        revenue_change = yday_pl['revenue'] - dby_pl['revenue']
        ppc_change     = yday_pl['ppc']     - dby_pl['ppc']
        ref_change     = yday_pl['referral']- dby_pl['referral']
        fba_change     = yday_pl['fba']     - dby_pl['fba']
        cogs_change    = yday_pl['cogs']    - dby_pl['cogs']
        net_change     = yday_pl['profit']  - dby_pl['profit']

        # Sorted by magnitude — surface the dominant driver(s)
        contributors = [
            {'name': 'Revenue',       'sign': '+', 'value': round(revenue_change, 2)},
            {'name': 'PPC Spend',     'sign': '−', 'value': round(-ppc_change,    2)},
            {'name': 'Referral Fees', 'sign': '−', 'value': round(-ref_change,    2)},
            {'name': 'FBA Fees',      'sign': '−', 'value': round(-fba_change,    2)},
            {'name': 'COGS',          'sign': '−', 'value': round(-cogs_change,   2)},
        ]
        contributors.sort(key=lambda x: abs(x['value']), reverse=True)
        movement = {
            'net_change':   round(net_change, 2),
            'contributors': contributors,
        }

    return JsonResponse({
        'marketplace': marketplace,
        'yesterday':   yday_pl,
        'day_before':  dby_pl,
        'deltas':      headline_deltas,
        'best_campaign': best_camp,
        'best_sku':      best_sku,
        'concern':       concern,
        'recommendation': recommend,
        'profit_movement': movement,
        'as_of_date':    yday.isoformat(),
    })


# ═════════════════════════════════════════════════════════════════════════════
# PHASE 3 — BRAND ANALYTICS
#
# /dashboard/ba/queries/             ba_queries (page shell)
# /dashboard/api/ba/queries/         api_ba_queries (JSON)
#
# Search Query Performance is the headline BA page — for each customer search
# query Amazon saw last week, we show OUR ASIN's click/conversion/purchase
# share. Auto-tags surface the most actionable findings:
#   • Dominant      — we own >50% of clicks on a query
#   • Weak          — high impressions but low click share (listing problem)
#   • Losing PDP    — high click share but low purchase share (PDP problem)
#   • Scaling Opp   — purchase share growing AND positive
#   • New           — query we didn't appear on previous week
# ═════════════════════════════════════════════════════════════════════════════

_BA_DOMINANT_CLICK    = 50.0  # %
_BA_WEAK_IMPR_MIN     = 100   # absolute count
_BA_WEAK_CLICK_MAX    = 5.0
_BA_PDP_LOSS_CLICK    = 10.0  # we click well…
_BA_PDP_LOSS_GAP      = 5.0   # …but purchase share lags click share by ≥5 pts
_BA_TAG_ALL = ('dominant', 'weak', 'losing_pdp', 'scaling_opp')


def _tag_query(row) -> list[str]:
    tags = []
    if float(row['brand_click_share']) >= _BA_DOMINANT_CLICK:
        tags.append('dominant')
    if (int(row['asin_impression_count']) >= _BA_WEAK_IMPR_MIN
            and float(row['brand_click_share']) <= _BA_WEAK_CLICK_MAX):
        tags.append('weak')
    if (float(row['brand_click_share']) >= _BA_PDP_LOSS_CLICK
            and float(row['brand_click_share']) - float(row['brand_purchase_share']) >= _BA_PDP_LOSS_GAP):
        tags.append('losing_pdp')
    # scaling_opp filled in caller after we know prior-week share
    return tags


@login_required
@permission_required('can_view_dashboard')
def ba_queries(request):
    """Page shell — JS calls /api/ba/queries/."""
    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        marketplace = _allowed_marketplaces(request.user)[0]
    ctx = {
        'marketplace':          marketplace,
        'allowed_marketplaces': _allowed_marketplaces(request.user),
        'today':                date.today(),
    }
    return render(request, 'dashboard/ba_queries.html', ctx)


@login_required
@permission_required('can_view_dashboard')
def api_ba_queries(request):
    """
    Query-level BA Search Query Performance data for the selected week + ASIN.

    Query params:
        mp           — marketplace (default usa)
        week_start   — Sunday (YYYY-MM-DD). Default: most recent week we have data for.
        asin         — scope to one ASIN. Default: top-revenue ASIN.
        tags         — comma-separated tag filter
        sort         — query_volume | impressions | click_share | purchase_share | (default query_volume)
        limit        — default 200, max 1000
    """
    from .models import BASearchQueryWeekly, DailySkuSnapshot, Product
    from django.db.models import Sum, Max

    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        return JsonResponse({'error': 'forbidden'}, status=403)

    # ── Resolve ASIN ────────────────────────────────────────────────────────
    asin = (request.GET.get('asin') or '').strip()
    if not asin:
        # Pick the ASIN with the most BA rows in the latest week
        latest = BASearchQueryWeekly.objects.filter(
            marketplace=marketplace).aggregate(mx=Max('week_start'))['mx']
        if not latest:
            return JsonResponse({
                'marketplace': marketplace,
                'kpi': {'queries_total': 0},
                'rows': [], 'asins': [], 'weeks': [], 'asin': None,
                'no_data': True,
                'no_data_msg': ('No Brand Analytics data ingested yet. Run '
                                '`python manage.py ingest_brand_analytics` '
                                'to start the weekly pipeline.'),
            })
        # Pick ASIN with most rows for the latest week
        top = BASearchQueryWeekly.objects.filter(
            marketplace=marketplace, week_start=latest,
        ).values('asin').annotate(n=Sum('asin_impression_count')).order_by('-n').first()
        asin = top['asin'] if top else None

    # ── Resolve week ────────────────────────────────────────────────────────
    week_start_str = (request.GET.get('week_start') or '').strip()
    if week_start_str:
        try:
            week_start = date.fromisoformat(week_start_str)
        except ValueError:
            week_start = BASearchQueryWeekly.objects.filter(
                marketplace=marketplace, asin=asin
            ).aggregate(mx=Max('week_start'))['mx']
    else:
        week_start = BASearchQueryWeekly.objects.filter(
            marketplace=marketplace, asin=asin
        ).aggregate(mx=Max('week_start'))['mx']

    if not week_start:
        return JsonResponse({
            'marketplace': marketplace, 'asin': asin,
            'kpi': {'queries_total': 0},
            'rows': [], 'asins': [], 'weeks': [],
            'no_data': True,
            'no_data_msg': f'No BA rows yet for ASIN {asin}.',
        })

    # ── Sibling lists for the toolbar ──────────────────────────────────────
    # NB: .order_by() with no args CLEARS the model's default Meta.ordering.
    # Without that clear, Django adds the ordering columns to the SELECT and
    # .distinct() ends up operating on (asin, week_start, score) tuples
    # instead of just asin — giving us one row per query (~thousands) instead
    # of one row per distinct ASIN.
    asins = sorted(set(BASearchQueryWeekly.objects.filter(marketplace=marketplace)
                       .order_by().values_list('asin', flat=True).distinct()))
    weeks = sorted(set(BASearchQueryWeekly.objects.filter(
                       marketplace=marketplace, asin=asin)
                       .order_by().values_list('week_start', flat=True).distinct()),
                   reverse=True)

    # Human-readable labels for every ASIN we have data for — drives the
    # dropdown so the user sees "B0982ZVLJF · Kitchen Towel 6-Pack Grey"
    # instead of the bare ASIN.
    asin_labels: dict[str, dict] = {}
    for p in Product.objects.filter(marketplace=marketplace, asin__in=asins).only(
        'asin', 'title', 'brand', 'sku',
    ):
        asin_labels[p.asin] = {
            'title': p.title or '',
            'brand': p.brand or '',
            'sku':   p.sku   or '',
        }

    # ── Pull current + prior-week rows ──────────────────────────────────────
    cur_qs = BASearchQueryWeekly.objects.filter(
        marketplace=marketplace, asin=asin, week_start=week_start,
    ).values('search_query', 'search_query_hash', 'search_query_score',
             'search_query_volume', 'impressions_total',
             'asin_impression_count', 'clicks_total',
             'asin_click_count', 'cart_adds_total', 'asin_cart_add_count',
             'purchases_total', 'asin_purchase_count',
             'brand_impressions_share', 'brand_click_share',
             'brand_cart_add_share', 'brand_purchase_share')

    prior_week = None
    if len(weeks) > 1:
        prior_week = weeks[1] if weeks[0] == week_start else weeks[0]
    prior_share_by_hash: dict[str, dict] = {}
    if prior_week:
        for r in BASearchQueryWeekly.objects.filter(
            marketplace=marketplace, asin=asin, week_start=prior_week,
        ).values('search_query_hash', 'brand_click_share', 'brand_purchase_share'):
            prior_share_by_hash[r['search_query_hash']] = {
                'click':    float(r['brand_click_share']),
                'purchase': float(r['brand_purchase_share']),
            }

    # ── Build rows + apply tags ─────────────────────────────────────────────
    rows = []
    tag_counts = {t: 0 for t in _BA_TAG_ALL}
    for r in cur_qs:
        tags = _tag_query(r)
        prior = prior_share_by_hash.get(r['search_query_hash'])
        cur_purchase = float(r['brand_purchase_share'])
        # scaling_opp: positive purchase share that GREW vs prior
        if prior and cur_purchase > 5 and cur_purchase > prior['purchase']:
            tags.append('scaling_opp')
        for t in tags:
            tag_counts[t] += 1

        rows.append({
            'search_query':         r['search_query'],
            'search_query_score':   r['search_query_score'],
            'search_query_volume':  r['search_query_volume'],
            'impressions_total':    r['impressions_total'],
            'asin_impression_count': r['asin_impression_count'],
            'clicks_total':         r['clicks_total'],
            'asin_click_count':     r['asin_click_count'],
            'purchases_total':      r['purchases_total'],
            'asin_purchase_count':  r['asin_purchase_count'],
            'impressions_share':    float(r['brand_impressions_share']),
            'click_share':          float(r['brand_click_share']),
            'cart_add_share':       float(r['brand_cart_add_share']),
            'purchase_share':       float(r['brand_purchase_share']),
            'prior_click_share':    prior['click']    if prior else None,
            'prior_purchase_share': prior['purchase'] if prior else None,
            'delta_click_share':    (float(r['brand_click_share']) - prior['click'])
                                      if prior else None,
            'delta_purchase_share': (float(r['brand_purchase_share']) - prior['purchase'])
                                      if prior else None,
            'tags':                 tags,
        })

    # ── Tag filter + sort + limit ──────────────────────────────────────────
    tag_filter = [t.strip() for t in (request.GET.get('tags') or '').split(',')
                   if t.strip() in _BA_TAG_ALL]
    if tag_filter:
        wanted = set(tag_filter)
        rows = [r for r in rows if any(t in wanted for t in r['tags'])]

    sort_key = (request.GET.get('sort') or 'search_query_volume').lower()
    if sort_key not in ('search_query_volume', 'search_query_score',
                        'impressions_total', 'asin_impression_count',
                        'click_share', 'purchase_share',
                        'delta_purchase_share'):
        sort_key = 'search_query_volume'
    # For search_query_score, smaller is better (rank 1 > rank 100)
    reverse = (sort_key != 'search_query_score')
    rows.sort(key=lambda r: (r[sort_key] if r[sort_key] is not None
                              else (-1 if reverse else 99999)),
              reverse=reverse)

    try:
        limit = max(1, min(int(request.GET.get('limit') or 200), 1000))
    except ValueError:
        limit = 200
    rows = rows[:limit]

    # ── KPI strip ──────────────────────────────────────────────────────────
    total_impr     = sum(r['asin_impression_count'] for r in rows)
    total_clicks   = sum(r['asin_click_count']      for r in rows)
    total_purchase = sum(r['asin_purchase_count']   for r in rows)
    # Volume-weighted avg share
    total_vol      = sum(r['search_query_volume']   for r in rows)
    weighted_click_share = (sum(r['click_share'] * r['search_query_volume']
                                  for r in rows) / total_vol) if total_vol else 0
    weighted_purchase_share = (sum(r['purchase_share'] * r['search_query_volume']
                                     for r in rows) / total_vol) if total_vol else 0

    # Product display name
    pmeta = Product.objects.filter(marketplace=marketplace, asin=asin
                                    ).only('title', 'brand', 'sku').first()
    asin_label = {
        'asin':  asin,
        'title': (pmeta.title if pmeta else '')[:96],
        'brand': (pmeta.brand if pmeta else ''),
        'sku':   (pmeta.sku   if pmeta else ''),
    }

    # ── Signal whether Δ vs prior is even possible ──────────────────────────
    # Δ columns need a SECOND week ingested for THIS ASIN. Without it we
    # surface a clear banner so the user doesn't wonder why the column is empty.
    delta_state = {
        'has_prior_week':  bool(prior_week),
        'prior_week':      prior_week.isoformat() if prior_week else None,
        'message':         (None if prior_week else
            'Δ vs prior is empty for this ASIN — only one week of SQP data '
            'is ingested. Run `python manage.py ingest_brand_analytics '
            f'--week-start {(week_start - timedelta(days=7)).isoformat()} '
            f'--top-asins 10 --kinds ba_search_query` to pull the prior week.'),
    }

    return JsonResponse({
        'marketplace':  marketplace,
        'asin':         asin,
        'asin_label':   asin_label,
        'asins':        asins,  # already sorted at the top
        'asin_labels':  asin_labels,
        'weeks':        [w.isoformat() for w in weeks],
        'week_start':   week_start.isoformat(),
        'week_end':     (week_start + timedelta(days=6)).isoformat(),
        'delta_state':  delta_state,
        'kpi': {
            'queries_total':           BASearchQueryWeekly.objects.filter(
                marketplace=marketplace, asin=asin, week_start=week_start).count(),
            'queries_shown':           len(rows),
            'queries_we_appear_in':    len(rows),
            'asin_impressions':        total_impr,
            'asin_clicks':             total_clicks,
            'asin_purchases':          total_purchase,
            'weighted_click_share':    round(weighted_click_share,    2),
            'weighted_purchase_share': round(weighted_purchase_share, 2),
        },
        'tag_counts':   tag_counts,
        'rows':         rows,
    })


# ─── 12H — Market Basket + Repeat Purchase ─────────────────────────────────

@login_required
@permission_required('can_view_dashboard')
def ba_baskets(request):
    """Page shell — JS calls /api/ba/baskets/."""
    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        marketplace = _allowed_marketplaces(request.user)[0]
    ctx = {
        'marketplace':          marketplace,
        'allowed_marketplaces': _allowed_marketplaces(request.user),
        'today':                date.today(),
    }
    return render(request, 'dashboard/ba_baskets.html', ctx)


@login_required
@permission_required('can_view_dashboard')
def api_ba_baskets(request):
    """
    Two boards in one response:
      • Market Basket    — most-frequently co-purchased pairs (cross-sell signal)
      • Repeat Purchase  — per-ASIN retention metrics (LTV signal)

    Query params:
        mp           — marketplace
        week_start   — Sunday (default: most recent available)
    """
    from .models import BAMarketBasketWeekly, BARepeatPurchaseWeekly, Product
    from django.db.models import Max, Count, Sum

    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        return JsonResponse({'error': 'forbidden'}, status=403)

    week_str = (request.GET.get('week_start') or '').strip()
    if week_str:
        try:
            week_start = date.fromisoformat(week_str)
        except ValueError:
            week_start = None
    else:
        week_start = None
    if not week_start:
        mb_week = BAMarketBasketWeekly.objects.filter(
            marketplace=marketplace).aggregate(mx=Max('week_start'))['mx']
        rp_week = BARepeatPurchaseWeekly.objects.filter(
            marketplace=marketplace).aggregate(mx=Max('week_start'))['mx']
        week_start = mb_week or rp_week

    if not week_start:
        return JsonResponse({
            'marketplace': marketplace,
            'no_data': True,
            'no_data_msg': ('No Market Basket or Repeat Purchase data ingested yet. '
                            'Run `python manage.py ingest_brand_analytics '
                            '--kinds ba_market_basket,ba_repeat_purchase`.'),
            'week_start': None, 'weeks': [],
            'market_basket': {}, 'repeat_purchase': {},
        })

    # Available weeks (intersection across both tables for cleaner UX)
    weeks = sorted(
        set(BAMarketBasketWeekly.objects.filter(marketplace=marketplace)
            .values_list('week_start', flat=True).distinct())
        | set(BARepeatPurchaseWeekly.objects.filter(marketplace=marketplace)
              .values_list('week_start', flat=True).distinct()),
        reverse=True,
    )

    # ── Market Basket ──────────────────────────────────────────────────────
    # Top co-purchase pairs by combination_pct, with rank #1 prioritised.
    # Also surface per-our-ASIN summary: how many partners we have.
    mb_rows_raw = list(BAMarketBasketWeekly.objects.filter(
        marketplace=marketplace, week_start=week_start,
    ).values('asin', 'purchased_asin', 'purchased_title',
             'purchased_frequency_rank', 'combination_pct'))

    # Brand check: figure out which of the purchased_asin values are also OUR
    # ASINs. That tells us "internal cross-sell" (bundle within our line) vs
    # "external cross-sell" (people pair us with competitor/complement).
    our_asins = set(Product.objects.filter(
        marketplace=marketplace).values_list('asin', flat=True))

    mb_pairs = []
    for r in mb_rows_raw:
        is_internal = r['purchased_asin'] in our_asins
        mb_pairs.append({
            'asin':            r['asin'],
            'purchased_asin':  r['purchased_asin'],
            'purchased_title': r['purchased_title'],
            'rank':            r['purchased_frequency_rank'],
            'combination_pct': round(float(r['combination_pct']) * 100, 2),
            'is_internal':     is_internal,
        })
    # Sort: strongest combination first, with internal cross-sell prioritised
    mb_pairs.sort(key=lambda r: (r['combination_pct'], r['is_internal']), reverse=True)

    # Per-our-ASIN summary
    per_asin: dict[str, dict] = {}
    for p in mb_pairs:
        b = per_asin.setdefault(p['asin'], {
            'asin': p['asin'], 'partner_count': 0,
            'avg_combo_pct': 0.0, 'top_partner': None,
            'internal_count': 0, 'external_count': 0,
        })
        b['partner_count'] += 1
        b['avg_combo_pct'] += p['combination_pct']
        if p['is_internal']: b['internal_count'] += 1
        else:                b['external_count'] += 1
        if b['top_partner'] is None or p['combination_pct'] > b['top_partner']['combo']:
            b['top_partner'] = {
                'asin':  p['purchased_asin'],
                'title': p['purchased_title'][:60],
                'combo': p['combination_pct'],
                'internal': p['is_internal'],
            }
    for b in per_asin.values():
        if b['partner_count']:
            b['avg_combo_pct'] = round(b['avg_combo_pct'] / b['partner_count'], 2)

    # ── Repeat Purchase ────────────────────────────────────────────────────
    rp_rows = list(BARepeatPurchaseWeekly.objects.filter(
        marketplace=marketplace, week_start=week_start,
    ).values('asin', 'orders', 'unique_customers', 'repeat_customers_pct',
             'repeat_purchase_revenue', 'repeat_purchase_revenue_pct'))

    # Product metadata for titles
    asins_for_meta = {r['asin'] for r in rp_rows} | {p['asin'] for p in mb_pairs[:200]}
    pmeta = {p.sku: p for p in Product.objects.filter(
        marketplace=marketplace, asin__in=asins_for_meta).only('asin', 'title')}
    asin_title = {p.asin: p.title for p in pmeta.values()}

    rp_out = []
    for r in rp_rows:
        rp_out.append({
            'asin':                 r['asin'],
            'title':                (asin_title.get(r['asin']) or '')[:96],
            'orders':               r['orders'],
            'unique_customers':     r['unique_customers'],
            'repeat_customers_pct': round(float(r['repeat_customers_pct']) * 100, 2),
            'repeat_revenue':       round(float(r['repeat_purchase_revenue']), 2),
            'repeat_revenue_pct':   round(float(r['repeat_purchase_revenue_pct']) * 100, 2),
        })
    rp_out.sort(key=lambda r: r['orders'], reverse=True)

    rp_totals = {
        'orders':           sum(r['orders']           for r in rp_out),
        'unique_customers': sum(r['unique_customers'] for r in rp_out),
        'repeat_revenue':   round(sum(r['repeat_revenue'] for r in rp_out), 2),
        'asin_count':       len(rp_out),
    }
    rp_totals['avg_repeat_pct'] = (
        round(sum(r['repeat_customers_pct'] * r['orders'] for r in rp_out)
               / max(rp_totals['orders'], 1), 2)
    )

    return JsonResponse({
        'marketplace':  marketplace,
        'week_start':   week_start.isoformat(),
        'week_end':     (week_start + timedelta(days=6)).isoformat(),
        'weeks':        [w.isoformat() for w in weeks],
        'market_basket': {
            'pair_count':  len(mb_pairs),
            'pairs':       mb_pairs[:200],
            'per_asin':    sorted(per_asin.values(),
                                   key=lambda r: r['partner_count'], reverse=True)[:50],
            'internal_count': sum(1 for p in mb_pairs if p['is_internal']),
            'external_count': sum(1 for p in mb_pairs if not p['is_internal']),
        },
        'repeat_purchase': {
            'rows':   rp_out,
            'totals': rp_totals,
        },
    })


# ─── 12F — Market Share Dashboard ───────────────────────────────────────────

@login_required
@permission_required('can_view_dashboard')
def ba_market_share(request):
    """Page shell — JS calls /api/ba/market-share/."""
    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        marketplace = _allowed_marketplaces(request.user)[0]
    ctx = {
        'marketplace':          marketplace,
        'allowed_marketplaces': _allowed_marketplaces(request.user),
    }
    return render(request, 'dashboard/ba_market_share.html', ctx)


@login_required
@permission_required('can_view_dashboard')
def api_ba_market_share(request):
    """
    Brand-level Market Share Dashboard built from BASearchQueryWeekly.

    The "market" is defined as the universe of customer searches where ANY of
    our ASINs appeared in the top-3 (clicked, converted, or purchased). Within
    that universe we measure:

      • Weighted click share / purchase share — share of clicks (or purchases)
        on our ASINs vs the other top-3 ASINs, weighted by query volume.
      • Win / Loss query buckets — where we hold ≥X% click share vs where we
        appear but lose to a competitor at the top of the page.
      • Competitor leaderboard — which non-our-brand ASINs most frequently
        appear in the top-3 alongside ours.
      • Per-our-ASIN contribution — which of our SKUs are pulling the share.

    Query params:
        mp           — marketplace
        week_start   — Sunday (default: most recent available)
    """
    from .models import BASearchQueryWeekly, Product
    from django.db.models import Max, Sum
    from collections import defaultdict

    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        return JsonResponse({'error': 'forbidden'}, status=403)

    week_str = (request.GET.get('week_start') or '').strip()
    week_start: date | None = None
    if week_str:
        try:
            week_start = date.fromisoformat(week_str)
        except ValueError:
            week_start = None
    if not week_start:
        week_start = BASearchQueryWeekly.objects.filter(
            marketplace=marketplace
        ).aggregate(mx=Max('week_start'))['mx']

    if not week_start:
        return JsonResponse({
            'marketplace':  marketplace,
            'no_data':      True,
            'no_data_msg':  ('No Search Query Performance data ingested yet. '
                              'Run `python manage.py ingest_brand_analytics '
                              '--kinds ba_search_query --top-asins 10`.'),
            'week_start':   None, 'weeks': [],
        })

    # All weeks we have any SQP data for (sorted descending) — drives the picker
    weeks = sorted(set(BASearchQueryWeekly.objects.filter(
        marketplace=marketplace
    ).order_by().values_list('week_start', flat=True).distinct()), reverse=True)

    # Our ASIN set (for "competitor" classification)
    our_asins = set(Product.objects.filter(marketplace=marketplace).values_list(
        'asin', flat=True))

    # Brand label for each of our ASINs (drives per-ASIN contribution display)
    asin_meta = {p.asin: {'title': p.title or '', 'brand': p.brand or '',
                          'sku': p.sku or ''}
                  for p in Product.objects.filter(marketplace=marketplace).only(
                      'asin', 'title', 'brand', 'sku')}

    # ── Pull every row for the week ────────────────────────────────────────
    rows = list(BASearchQueryWeekly.objects.filter(
        marketplace=marketplace, week_start=week_start,
    ).values('asin', 'search_query', 'search_query_hash', 'search_query_score',
             'search_query_volume',
             'brand_click_share', 'brand_purchase_share',
             'brand_impressions_share',
             'top_clicked_asins', 'top_converted_asins', 'top_purchased_asins'))

    if not rows:
        return JsonResponse({
            'marketplace': marketplace,
            'no_data':     True,
            'no_data_msg': f'No SQP data for week {week_start}.',
            'week_start':  week_start.isoformat(),
            'weeks':       [w.isoformat() for w in weeks],
        })

    # ── 1. Per-query aggregation ────────────────────────────────────────────
    # IMPORTANT: brand_click_share / brand_purchase_share are stored as
    # PERCENTAGES (0-100), not fractions. SQP returns them that way and the
    # normalizer keeps them verbatim. All math below treats them as percentages.
    #
    # A single query can appear multiple times in `rows` (once per OUR ASIN
    # that ranked in its top-3). For brand-level share we SUM each of our
    # ASIN's click/purchase share on that query — that's our COMBINED brand
    # share for the query.
    per_query: dict[str, dict] = defaultdict(lambda: {
        'search_query': '', 'search_query_hash': '',
        'volume': 0, 'score': 0,
        'our_click_share': 0.0, 'our_purchase_share': 0.0,
        'our_asins': [],
        'our_asin_count': 0,
    })

    for r in rows:
        h = r['search_query_hash']
        b = per_query[h]
        b['search_query']      = r['search_query']
        b['search_query_hash'] = h
        b['volume']            = int(r['search_query_volume'] or 0)
        b['score']             = int(r['search_query_score']  or 0)
        b['our_click_share']    += float(r['brand_click_share']    or 0)
        b['our_purchase_share'] += float(r['brand_purchase_share'] or 0)
        b['our_asins'].append(r['asin'])
        b['our_asin_count'] = len(b['our_asins'])

    # Cap shares at 100% — when many of our ASINs rank in the top-3 for the
    # same query, the summed shares can over-count due to rounding.
    for b in per_query.values():
        b['our_click_share']    = min(100.0, b['our_click_share'])
        b['our_purchase_share'] = min(100.0, b['our_purchase_share'])

    # ── 2. Account-level KPI (volume-weighted shares, units = %) ────────────
    total_volume = sum(b['volume'] for b in per_query.values()) or 1
    weighted_click_num    = sum(b['volume'] * b['our_click_share']    for b in per_query.values())
    weighted_purchase_num = sum(b['volume'] * b['our_purchase_share'] for b in per_query.values())

    queries_we_dominate = sum(1 for b in per_query.values() if b['our_click_share'] > 50)
    queries_in_top3     = len(per_query)

    # ── 3. Win / Loss buckets ───────────────────────────────────────────────
    # Win  = our_click_share >= 30%
    # Loss = our_click_share <  10% AND volume in the top quartile (we appear
    #         in a popular query but get nearly none of the clicks).
    volumes = sorted((b['volume'] for b in per_query.values()), reverse=True)
    p75_volume = volumes[len(volumes) // 4] if volumes else 0

    def row_for_query(b) -> dict:
        return {
            'search_query':       b['search_query'],
            'volume':             b['volume'],
            'score':              b['score'],
            'our_click_share':    round(b['our_click_share'],    2),
            'our_purchase_share': round(b['our_purchase_share'], 2),
            'our_asin_count':     b['our_asin_count'],
        }

    win_rows = sorted(
        (row_for_query(b) for b in per_query.values() if b['our_click_share'] >= 30),
        key=lambda r: r['volume'], reverse=True,
    )[:30]
    loss_rows = sorted(
        (row_for_query(b) for b in per_query.values()
          if b['our_click_share'] < 10 and b['volume'] >= p75_volume),
        key=lambda r: r['volume'], reverse=True,
    )[:30]

    # ── 4. Competitor leaderboard — pulled from BAMarketBasketWeekly ────────
    # The per-ASIN SQP report does NOT expose competitor top-3 ASINs (it only
    # tells us OUR share). For a real competitor signal we use the Market
    # Basket report's external pairs — ASINs that customers co-purchase with
    # ours but that aren't in our brand catalogue.
    from .models import BAMarketBasketWeekly
    competitor_agg: dict[str, dict] = defaultdict(lambda: {
        'asin': '', 'overlap_count': 0, 'total_combo_pct': 0.0,
        'top_our_asin': '', 'top_combo_pct': 0.0,
        'title': '',
    })
    for mb in BAMarketBasketWeekly.objects.filter(
        marketplace=marketplace, week_start=week_start,
    ).values('asin', 'purchased_asin', 'purchased_title',
             'purchased_frequency_rank', 'combination_pct'):
        if mb['purchased_asin'] in our_asins:
            continue  # internal cross-sell, not a competitor
        c = competitor_agg[mb['purchased_asin']]
        c['asin']             = mb['purchased_asin']
        c['title']            = c['title'] or (mb['purchased_title'] or '')[:80]
        c['overlap_count']   += 1
        combo_pct = float(mb['combination_pct'] or 0) * 100   # stored as 0-1
        c['total_combo_pct'] += combo_pct
        if combo_pct > c['top_combo_pct']:
            c['top_combo_pct'] = combo_pct
            c['top_our_asin']  = mb['asin']

    competitor_rows = sorted(competitor_agg.values(),
                              key=lambda r: r['overlap_count'], reverse=True)[:30]
    for c in competitor_rows:
        c['avg_combo_pct']   = round(c['total_combo_pct'] / max(c['overlap_count'], 1), 2)
        c['top_combo_pct']   = round(c['top_combo_pct'], 2)
        c['total_combo_pct'] = round(c['total_combo_pct'], 2)

    # ── 5. Per-our-ASIN contribution to brand share ─────────────────────────
    # Same percentage-units convention.
    per_asin_contrib: dict[str, dict] = defaultdict(lambda: {
        'asin': '', 'title': '', 'brand': '', 'sku': '',
        'query_count': 0,
        'volume_weighted_click_share':    0.0,
        'volume_weighted_purchase_share': 0.0,
        'total_query_volume':             0,
    })
    for r in rows:
        asin = r['asin']
        c = per_asin_contrib[asin]
        c['asin']        = asin
        meta = asin_meta.get(asin, {})
        c['title']       = (meta.get('title') or '')[:64]
        c['brand']       = meta.get('brand') or ''
        c['sku']         = meta.get('sku') or ''
        v = int(r['search_query_volume'] or 0)
        c['query_count']        += 1
        c['total_query_volume'] += v
        c['volume_weighted_click_share']    += v * float(r['brand_click_share']    or 0)
        c['volume_weighted_purchase_share'] += v * float(r['brand_purchase_share'] or 0)

    per_asin_rows = []
    for c in per_asin_contrib.values():
        vol = c['total_query_volume'] or 1
        per_asin_rows.append({
            **c,
            'avg_click_share':    round(c['volume_weighted_click_share']    / vol, 2),
            'avg_purchase_share': round(c['volume_weighted_purchase_share'] / vol, 2),
        })
    per_asin_rows.sort(key=lambda r: r['total_query_volume'], reverse=True)

    return JsonResponse({
        'marketplace':  marketplace,
        'week_start':   week_start.isoformat(),
        'week_end':     (week_start + timedelta(days=6)).isoformat(),
        'weeks':        [w.isoformat() for w in weeks],
        'kpi': {
            'queries_in_market':       queries_in_top3,
            'queries_we_dominate':     queries_we_dominate,
            'total_query_volume':      total_volume,
            # Shares are already in percentage units (0-100). Divide by total
            # volume gives the weighted-average percentage directly.
            'weighted_click_share':    round(weighted_click_num    / total_volume, 2),
            'weighted_purchase_share': round(weighted_purchase_num / total_volume, 2),
            'win_query_count':         len(win_rows),
            'loss_query_count':        len(loss_rows),
            'unique_competitors':      len(competitor_agg),
        },
        'wins':            win_rows,
        'losses':          loss_rows,
        'competitors':     competitor_rows,
        'per_asin':        per_asin_rows[:50],
    })


# ─── 12C + 12D — Click + Purchase Share Trend ───────────────────────────────

@login_required
@permission_required('can_view_dashboard')
def ba_share_trend(request):
    """Page shell — JS calls /api/ba/share-trend/."""
    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        marketplace = _allowed_marketplaces(request.user)[0]
    ctx = {
        'marketplace':          marketplace,
        'allowed_marketplaces': _allowed_marketplaces(request.user),
    }
    return render(request, 'dashboard/ba_share_trend.html', ctx)


@login_required
@permission_required('can_view_dashboard')
def api_ba_share_trend(request):
    """
    Click + Purchase Share trend over the weeks we have SQP data for.

    Returns three boards:
      • account_trend   — weighted brand-level share per week (drives the
                          line chart at the top)
      • movers_up/down  — queries with the biggest Δ click share between the
                          two most recent weeks
      • per_asin_trend  — per-our-ASIN weighted share per week (small lines)

    Shares are stored as percentages (0-100) in BASearchQueryWeekly — we
    keep that convention end-to-end.
    """
    from .models import BASearchQueryWeekly, Product
    from collections import defaultdict

    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        return JsonResponse({'error': 'forbidden'}, status=403)

    weeks = sorted(set(BASearchQueryWeekly.objects.filter(
        marketplace=marketplace
    ).order_by().values_list('week_start', flat=True).distinct()))

    if not weeks:
        return JsonResponse({
            'marketplace': marketplace,
            'no_data':     True,
            'no_data_msg': ('No Search Query Performance data ingested yet. '
                            'Run `python manage.py ingest_brand_analytics '
                            '--kinds ba_search_query --top-asins 10`.'),
            'weeks':       [], 'account_trend': [],
            'movers_up':   [], 'movers_down': [], 'per_asin_trend': [],
        })

    # ── Per-week account-level weighted-share aggregate ─────────────────────
    # For each week: sum(volume × share) / sum(volume) across all our queries.
    # If multiple of our ASINs rank for the same query in the same week, sum
    # their share (cap at 100%) — this is our combined brand share for that
    # query.
    account_trend = []
    per_query_per_week: dict[tuple, dict] = defaultdict(lambda: {
        'click': 0.0, 'purchase': 0.0, 'volume': 0,
    })
    rows = list(BASearchQueryWeekly.objects.filter(marketplace=marketplace).values(
        'asin', 'week_start', 'search_query', 'search_query_hash',
        'search_query_volume', 'brand_click_share', 'brand_purchase_share',
    ))

    for r in rows:
        key = (r['week_start'], r['search_query_hash'])
        bucket = per_query_per_week[key]
        bucket['click']    += float(r['brand_click_share']    or 0)
        bucket['purchase'] += float(r['brand_purchase_share'] or 0)
        bucket['volume']    = int(r['search_query_volume']    or 0)
        bucket['search_query'] = r['search_query']

    # Cap shares at 100% per query (defensive — single-ASIN rows can't go over,
    # but combined rows can over-count by a hair due to Amazon's rounding).
    for b in per_query_per_week.values():
        b['click']    = min(100.0, b['click'])
        b['purchase'] = min(100.0, b['purchase'])

    # Per-week rollup
    per_week_agg: dict[date, dict] = defaultdict(lambda: {
        'volume': 0, 'weighted_click_num': 0.0, 'weighted_purchase_num': 0.0,
        'queries_in_market': 0,
    })
    for (wk, _h), b in per_query_per_week.items():
        a = per_week_agg[wk]
        a['volume']               += b['volume']
        a['weighted_click_num']   += b['volume'] * b['click']
        a['weighted_purchase_num']+= b['volume'] * b['purchase']
        a['queries_in_market']    += 1

    for wk in weeks:
        a = per_week_agg[wk]
        vol = a['volume'] or 1
        account_trend.append({
            'week_start':              wk.isoformat(),
            'weighted_click_share':    round(a['weighted_click_num']    / vol, 2),
            'weighted_purchase_share': round(a['weighted_purchase_num'] / vol, 2),
            'total_query_volume':      a['volume'],
            'queries_in_market':       a['queries_in_market'],
        })

    # ── Movers — queries with biggest Δ between two most recent weeks ───────
    movers_up: list = []
    movers_down: list = []
    if len(weeks) >= 2:
        latest, prior = weeks[-1], weeks[-2]
        # Build {hash: {latest_click, prior_click, volume, query}}
        comparison: dict[str, dict] = {}
        for (wk, h), b in per_query_per_week.items():
            comp = comparison.setdefault(h, {
                'search_query': b.get('search_query') or '',
                'volume':       b['volume'],
                'latest_click': 0.0, 'prior_click': 0.0,
                'latest_purchase': 0.0, 'prior_purchase': 0.0,
            })
            if wk == latest:
                comp['latest_click']    = b['click']
                comp['latest_purchase'] = b['purchase']
                comp['volume']          = b['volume']
                comp['search_query']    = b.get('search_query') or comp['search_query']
            elif wk == prior:
                comp['prior_click']    = b['click']
                comp['prior_purchase'] = b['purchase']

        # Only consider queries with meaningful volume (top 25% by volume)
        all_vols = sorted((c['volume'] for c in comparison.values()), reverse=True)
        vol_floor = all_vols[len(all_vols) // 4] if all_vols else 0

        scored = []
        for h, c in comparison.items():
            if c['volume'] < vol_floor:
                continue
            delta_click    = c['latest_click']    - c['prior_click']
            delta_purchase = c['latest_purchase'] - c['prior_purchase']
            scored.append({
                'search_query':       c['search_query'],
                'volume':             c['volume'],
                'latest_click':       round(c['latest_click'],    2),
                'prior_click':        round(c['prior_click'],     2),
                'delta_click':        round(delta_click,          2),
                'latest_purchase':    round(c['latest_purchase'], 2),
                'prior_purchase':     round(c['prior_purchase'],  2),
                'delta_purchase':     round(delta_purchase,       2),
            })
        movers_up   = sorted(scored, key=lambda r: r['delta_click'], reverse=True)[:20]
        movers_down = sorted(scored, key=lambda r: r['delta_click'])[:20]
        movers_down = [m for m in movers_down if m['delta_click'] < 0]
        movers_up   = [m for m in movers_up   if m['delta_click'] > 0]

    # ── Per-our-ASIN weekly share trend ─────────────────────────────────────
    # weighted by query volume across each ASIN's queries that week
    per_asin_per_week: dict[tuple, dict] = defaultdict(lambda: {
        'click_num': 0.0, 'purchase_num': 0.0, 'volume': 0,
    })
    for r in rows:
        key = (r['asin'], r['week_start'])
        a = per_asin_per_week[key]
        a['click_num']    += int(r['search_query_volume'] or 0) * float(r['brand_click_share']    or 0)
        a['purchase_num'] += int(r['search_query_volume'] or 0) * float(r['brand_purchase_share'] or 0)
        a['volume']       += int(r['search_query_volume'] or 0)

    asins_seen = sorted({asin for asin, _wk in per_asin_per_week})
    pmeta = {p.asin: p for p in Product.objects.filter(
        marketplace=marketplace, asin__in=asins_seen).only('asin', 'title', 'sku')}

    per_asin_trend = []
    for asin in asins_seen:
        series = []
        for wk in weeks:
            a = per_asin_per_week.get((asin, wk))
            if not a or a['volume'] == 0:
                series.append({'week': wk.isoformat(),
                                'click_share': None, 'purchase_share': None,
                                'volume': 0})
                continue
            series.append({
                'week':           wk.isoformat(),
                'click_share':    round(a['click_num']    / a['volume'], 2),
                'purchase_share': round(a['purchase_num'] / a['volume'], 2),
                'volume':         a['volume'],
            })
        meta = pmeta.get(asin)
        per_asin_trend.append({
            'asin':  asin,
            'title': (meta.title if meta else '')[:60],
            'sku':   (meta.sku   if meta else ''),
            'series': series,
        })

    # ── State flags for UI hints ────────────────────────────────────────────
    state = {
        'weeks_count':       len(weeks),
        'has_movers':        len(weeks) >= 2,
        'movers_explanation': (None if len(weeks) >= 2 else
            'Movers (queries with biggest week-over-week Δ) need at least '
            'TWO weeks of SQP data. Run `python manage.py '
            'ingest_brand_analytics --week-start '
            f'{(weeks[0] - timedelta(days=7)).isoformat()} '
            '--top-asins 10 --kinds ba_search_query` to add a prior week.'),
    }

    return JsonResponse({
        'marketplace':     marketplace,
        'weeks':           [w.isoformat() for w in weeks],
        'account_trend':   account_trend,
        'movers_up':       movers_up,
        'movers_down':     movers_down,
        'per_asin_trend':  per_asin_trend,
        'state':           state,
    })


# ═════════════════════════════════════════════════════════════════════════════
# PHASE 4 — AI RECOMMENDATIONS PAGE
#
# /dashboard/ai/recommendations/                  ai_recommendations
# /dashboard/api/ai/recommendations/              api_ai_recommendations
# /dashboard/api/ai/recommendations/<pk>/status/  api_ai_rec_status (POST)
# /dashboard/api/ai/recommendations/regenerate/   api_ai_recs_regenerate (POST)
# ═════════════════════════════════════════════════════════════════════════════


@login_required
@permission_required('can_view_dashboard')
def ai_recommendations(request):
    """Page shell — JS fetches the recs from the API."""
    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        marketplace = _allowed_marketplaces(request.user)[0]
    ctx = {
        'marketplace':          marketplace,
        'allowed_marketplaces': _allowed_marketplaces(request.user),
    }
    return render(request, 'dashboard/ai_recommendations.html', ctx)


@login_required
@permission_required('can_view_dashboard')
def api_ai_recommendations(request):
    """
    Returns ranked AIRecommendation rows for the marketplace.

    Query params:
        mp        — marketplace
        status    — comma-list of statuses to include (default: new,acknowledged,snoozed)
        severity  — comma-list filter on severity
        category  — comma-list filter on category
    """
    from .models import AIRecommendation
    from django.db.models import Max

    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        return JsonResponse({'error': 'forbidden'}, status=403)

    qs = AIRecommendation.objects.filter(marketplace=marketplace)

    statuses = [s.strip() for s in (request.GET.get('status') or
                                      'new,acknowledged,snoozed').split(',')
                if s.strip()]
    if statuses:
        qs = qs.filter(status__in=statuses)

    sevs = [s.strip() for s in (request.GET.get('severity') or '').split(',') if s.strip()]
    if sevs:
        qs = qs.filter(severity__in=sevs)

    cats = [c.strip() for c in (request.GET.get('category') or '').split(',') if c.strip()]
    if cats:
        qs = qs.filter(category__in=cats)

    qs = qs.order_by('-rank_score', '-generated_at')[:200]

    rows = [{
        'id':              r.id,
        'severity':        r.severity,
        'category':        r.category,
        'scope_type':      r.scope_type,
        'scope_id':        r.scope_id,
        'scope_name':      r.scope_name,
        'headline':        r.headline,
        'evidence':        r.evidence,
        'suggested_action':r.suggested_action,
        'projected_impact':r.projected_impact,
        'confidence':      float(r.confidence),
        'rank_score':      float(r.rank_score),
        'status':          r.status,
        'snoozed_until':   r.snoozed_until.isoformat() if r.snoozed_until else None,
        'generated_at':    r.generated_at.isoformat() if r.generated_at else None,
        'reference_date':  r.reference_date.isoformat() if r.reference_date else None,
        'ai_model':        r.ai_model,
        'acknowledged_by': r.acknowledged_by.username if r.acknowledged_by else None,
        # Quick drill-down link the UI uses to navigate to scope
        'drill_url':       _build_scope_url(marketplace, r.scope_type, r.scope_id),
    } for r in qs]

    last_run = AIRecommendation.objects.filter(
        marketplace=marketplace).aggregate(mx=Max('generated_at'))['mx']

    # Severity / category tag counts (across ALL statuses for the marketplace,
    # so the filter pills always have a useful count)
    full = AIRecommendation.objects.filter(marketplace=marketplace)
    sev_counts = {k: full.filter(severity=k).count()
                   for k in ('critical', 'warning', 'opportunity', 'info')}
    cat_counts = {k: full.filter(category=k).count()
                   for k in ('ppc_scale', 'ppc_cut', 'ppc_negate', 'ppc_bid',
                             'sku_scale', 'sku_pause', 'margin_fix', 'inventory',
                             'listing', 'cross_sell', 'competitive', 'other')}
    status_counts = {k: full.filter(status=k).count()
                      for k, _ in __import__('apps.dashboard.models',
                                              fromlist=['AIRecommendation']
                                              ).AIRecommendation.STATUS}

    return JsonResponse({
        'marketplace':  marketplace,
        'rows':         rows,
        'kpi': {
            'total_active':  sum(1 for r in rows if r['status'] in ('new', 'acknowledged')),
            'critical_open': sum(1 for r in rows
                                  if r['severity'] == 'critical' and r['status'] != 'done'),
            'opportunity_open': sum(1 for r in rows
                                     if r['severity'] == 'opportunity' and r['status'] != 'done'),
        },
        'severity_counts': sev_counts,
        'category_counts': cat_counts,
        'status_counts':   status_counts,
        'last_generated':  last_run.isoformat() if last_run else None,
    })


def _build_scope_url(mp: str, scope_type: str, scope_id: str) -> str | None:
    if not scope_id:
        return None
    if scope_type == 'campaign':
        return f'/dashboard/campaigns/{scope_id}/?mp={mp}'
    if scope_type == 'sku':
        return f'/dashboard/pnl/skus/?mp={mp}'   # no per-sku detail page yet
    if scope_type == 'search_term':
        return f'/dashboard/search-terms/?mp={mp}'
    if scope_type == 'placement':
        return f'/dashboard/placements/?mp={mp}'
    return None


@login_required
@permission_required('can_view_dashboard')
def api_ai_rec_status(request, pk: int):
    """POST {status: ...} — update the status of a single recommendation.

    Allowed transitions: any → new/acknowledged/done/snoozed/dismissed.
    Status 'snoozed' may carry a `snoozed_until` ISO date.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'method not allowed'}, status=405)
    from .models import AIRecommendation
    try:
        body = json.loads(request.body or b'{}')
    except json.JSONDecodeError:
        return JsonResponse({'error': 'invalid json body'}, status=400)

    new_status = (body.get('status') or '').strip()
    if new_status not in {k for k, _ in AIRecommendation.STATUS}:
        return JsonResponse({'error': f'unknown status {new_status!r}'}, status=400)

    rec = AIRecommendation.objects.filter(pk=pk).first()
    if not rec:
        return JsonResponse({'error': 'not found'}, status=404)
    if not request.user.can_access_marketplace(rec.marketplace):
        return JsonResponse({'error': 'forbidden'}, status=403)

    rec.status = new_status
    if new_status in ('acknowledged', 'done'):
        rec.acknowledged_by = request.user
        rec.acknowledged_at = timezone.now()
    elif new_status == 'snoozed':
        try:
            rec.snoozed_until = (date.fromisoformat(body['snoozed_until'])
                                  if body.get('snoozed_until') else None)
        except ValueError:
            return JsonResponse({'error': 'invalid snoozed_until'}, status=400)
    if body.get('user_notes'):
        rec.user_notes = body['user_notes'][:2000]
    rec.save()
    return JsonResponse({'ok': True, 'id': rec.id, 'status': rec.status})


@login_required
@permission_required('can_view_dashboard')
def api_ai_recs_regenerate(request):
    """POST — trigger generate_ai_recommendations synchronously, return summary."""
    if request.method != 'POST':
        return JsonResponse({'error': 'method not allowed'}, status=405)
    mp = request.GET.get('mp') or 'usa'
    if not request.user.can_access_marketplace(mp):
        return JsonResponse({'error': 'forbidden'}, status=403)

    # We run the management command inline so error reporting stays simple.
    from django.core.management import call_command
    import io
    out = io.StringIO()
    err = io.StringIO()
    try:
        call_command('generate_ai_recommendations',
                      marketplace=mp, stdout=out, stderr=err)
        return JsonResponse({
            'ok': True,
            'stdout': out.getvalue()[-1500:],
            'stderr': err.getvalue()[-1500:],
        })
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)[:300]}, status=500)


# ─── Phase 4 — Morning Report AI Commentary (SSE-streamed) ──────────────────

@login_required
@permission_required('can_view_dashboard')
def api_morning_ai_commentary(request):
    """
    Streams a 2-3 paragraph CFO commentary on yesterday's data via Server-Sent
    Events. The frontend appends each delta to a panel at the top of the
    Morning Report.

    The briefing payload is the same gather_briefing_data() blob the
    Recommendations engine uses — so the commentary and the recommendations
    are reasoning about the same numbers.
    """
    import json as _json
    import requests as _http
    from .ai_insights import gather_briefing_data, get_anthropic_credentials

    marketplace = request.GET.get('mp', 'usa')
    if not request.user.can_access_marketplace(marketplace):
        return JsonResponse({'error': 'forbidden'}, status=403)

    api_key, model = get_anthropic_credentials()
    if not api_key:
        def _err():
            yield ('data: ' + _json.dumps({'error':
                'Anthropic API key not configured. Add one in API Config → Anthropic.'
            }) + '\n\n')
        return StreamingHttpResponse(_err(), content_type='text/event-stream')

    briefing = gather_briefing_data(marketplace)
    if not briefing.get('day_pl'):
        def _empty():
            yield ('data: ' + _json.dumps({'error':
                "No DailyMetric for yesterday yet — wait for the 00:45 "
                'finalize_yesterday cron, then refresh.'
            }) + '\n\n')
        return StreamingHttpResponse(_empty(), content_type='text/event-stream')

    AuditLog.objects.create(
        user=request.user, action='ai_morning_commentary',
        resource=f'morning:{marketplace}:{briefing["reference_date"]}',
        ip_address=request.META.get('REMOTE_ADDR'))

    system = ("You are a private-label Amazon e-commerce CFO writing a brief "
              "morning briefing to the operator. Read the JSON below and write "
              "EXACTLY 2-3 short paragraphs (no headings, no bullets, no markdown "
              "fences). Tone: calm, precise, opinionated. Always lead with the "
              "single most important fact about yesterday's profit. Reference "
              "actual numbers and named campaigns/SKUs when relevant. End with "
              "the one most useful action to take today.")
    user_msg = (f'Yesterday briefing for {marketplace.upper()}:\n\n'
                f'```json\n{_json.dumps(briefing, default=str, indent=2)}\n```\n\n'
                'Write the 2-3 paragraph briefing now.')

    def _stream():
        # Use Claude's native streaming so the UI fills word by word.
        try:
            with _http.post(
                'https://api.anthropic.com/v1/messages',
                headers={'x-api-key': api_key,
                          'anthropic-version': '2023-06-01',
                          'Content-Type': 'application/json'},
                json={'model': model, 'max_tokens': 1024, 'stream': True,
                       'system': system,
                       'messages': [{'role':'user', 'content': user_msg}]},
                stream=True, timeout=90,
            ) as resp:
                if not resp.ok:
                    try:
                        detail = resp.json()
                        err = (detail.get('error', {}) or {}).get('message') or str(detail)[:300]
                    except Exception:
                        err = resp.text[:300]
                    yield 'data: ' + _json.dumps({'error': f'http {resp.status_code}: {err}'}) + '\n\n'
                    return
                # Each SSE line from Anthropic looks like
                #   event: content_block_delta
                #   data: {"type":"content_block_delta","delta":{"type":"text_delta","text":"…"}}
                for raw_line in resp.iter_lines(decode_unicode=True):
                    if not raw_line: continue
                    if not raw_line.startswith('data:'): continue
                    payload = raw_line[5:].strip()
                    if payload == '[DONE]':
                        yield 'data: ' + _json.dumps({'done': True}) + '\n\n'
                        return
                    try:
                        evt = _json.loads(payload)
                    except _json.JSONDecodeError:
                        continue
                    if evt.get('type') == 'content_block_delta':
                        delta = (evt.get('delta') or {}).get('text') or ''
                        if delta:
                            yield 'data: ' + _json.dumps({'delta': delta}) + '\n\n'
                    elif evt.get('type') == 'message_stop':
                        yield 'data: ' + _json.dumps({'done': True}) + '\n\n'
                        return
                    elif evt.get('type') == 'error':
                        yield 'data: ' + _json.dumps({'error': str(evt.get('error'))[:300]}) + '\n\n'
                        return
        except Exception as e:
            yield 'data: ' + _json.dumps({'error': f'network: {e}'}) + '\n\n'

    return StreamingHttpResponse(_stream(), content_type='text/event-stream')


# ═══════════════════════════════════════════════════════════════════════════
#  P&L — overhead template download + Settlement V2 manual upload
# ═══════════════════════════════════════════════════════════════════════════

@login_required
@permission_required('can_manage_cogs')
def pnl_overhead_template_xlsx(request):
    """
    Download a blank overhead template matching the Import P&L Excel parser.

    Column A carries the exact labels PNL_LINES defines for source='manual'
    lines, so the importer's label matching cannot miss. Column B is empty for
    the user to fill. Amazon lines are deliberately absent — they come from
    Settlement V2 and the overhead importer has no path to them.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        return HttpResponse('openpyxl is not installed. Run: pip install openpyxl',
                            status=500)

    from .pnl_lines import PNL_LINES

    mp    = request.GET.get('mp', 'usa')
    month = request.GET.get('month', date.today().strftime('%Y-%m'))
    ccy   = (getattr(settings, 'AMAZON_MARKETPLACES', {})
             .get(mp, {}).get('currency', 'USD'))

    wb = Workbook()
    ws = wb.active
    ws.title = 'P&L Summary'

    head_fill = PatternFill('solid', fgColor='1F4E5F')
    head_font = Font(bold=True, color='FFFFFF', size=11)
    sec_fill  = PatternFill('solid', fgColor='E8EEF2')
    sec_font  = Font(bold=True, size=10)

    ws['A1'] = f'Overhead — {mp.upper()} — {month}'
    ws['B1'] = f'Amount ({ccy})'
    for c in ('A1', 'B1'):
        ws[c].fill = head_fill
        ws[c].font = head_font
        ws[c].alignment = Alignment(horizontal='left')

    row = 2
    last_section = None
    for ln in PNL_LINES:
        if ln.get('source') != 'manual':
            continue
        sec = ln.get('section') or ''
        if sec != last_section:
            ws.cell(row=row, column=1, value=str(sec)).fill = sec_fill
            ws.cell(row=row, column=1).font = sec_font
            ws.cell(row=row, column=2).fill = sec_fill
            row += 1
            last_section = sec
        ws.cell(row=row, column=1, value=ln['label'])
        ws.cell(row=row, column=2, value=None)
        row += 1

    ws.column_dimensions['A'].width = 46
    ws.column_dimensions['B'].width = 18

    note = ws.cell(row=row + 1, column=1,
                   value='Leave a row blank to skip it. Negative values are '
                         'allowed for credits/rebates. Do not rename column A '
                         '— the importer matches on these labels.')
    note.font = Font(italic=True, size=9, color='666666')

    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = (
        f'attachment; filename="pnl_overhead_template_{mp}_{month}.xlsx"')
    wb.save(resp)
    return resp


@login_required
@permission_required('can_manage_cogs')
@_require_POST
def import_settlement_v2(request):
    """
    Upload Settlement Flat File V2 report(s) fetched by hand from
    Seller Central → Payments → Reports Repository.

    Amazon expires report documents after a few months, so the API can no
    longer download older settlements — USA April and May 2026 already fail
    with a 400. Seller Central still serves them. Uploading here caches the
    file against its SettlementReport so rebuild_settlement_month can read it
    exactly as if it had downloaded it.

    This does NOT import figures directly. It only makes the source file
    available; the same rebuild does the parsing, dedup and arithmetic, so an
    uploaded month is computed identically to a downloaded one.

    POST multipart: marketplace, file(s)
    """
    from .settlement_cache import parse_bytes, describe, write_cached
    from .models import SettlementReport

    mp = request.POST.get('marketplace', 'usa')
    if not request.user.can_access_marketplace(mp):
        raise _PermissionDenied

    files = request.FILES.getlist('files') or request.FILES.getlist('file')
    if not files:
        return JsonResponse({'status': 'failed', 'message': 'No file.'}, status=400)

    results, months = [], set()
    for f in files:
        if f.size > 250 * 1024 * 1024:
            results.append({'file': f.name, 'ok': False,
                            'message': 'Too large (max 250 MiB).'})
            continue
        try:
            raw = f.read()
            rows = parse_bytes(raw)
        except Exception as exc:
            results.append({'file': f.name, 'ok': False,
                            'message': f'{type(exc).__name__}: {exc}'})
            continue

        info = describe(rows)
        if not info['rows'] or not info['first_posted']:
            results.append({'file': f.name, 'ok': False,
                            'message': 'No settlement rows with a posted-date — '
                                       'is this the Flat File V2 settlement report?'})
            continue

        # Attach to the SettlementReport this file belongs to, so the rebuild
        # knows to look for it. Match on settlement-id where we can; fall back
        # to the posted-date span; create a row when Amazon never listed it.
        rep = None
        for sid in info['settlement_ids']:
            rep = SettlementReport.objects.filter(
                marketplace=mp, report_id=f'upload-{sid}').first()
            if rep:
                break
        if rep is None:
            rep = SettlementReport.objects.filter(
                marketplace=mp,
                start_date__lte=info['first_posted'],
                end_date__gte=info['last_posted'],
            ).order_by('start_date').first()
        if rep is None:
            sid = info['settlement_ids'][0] if info['settlement_ids'] else f.name[:32]
            rep, _ = SettlementReport.objects.update_or_create(
                marketplace=mp, report_id=f'upload-{sid}',
                defaults={'document_id': '', 'status': 'ok',
                          'start_date': info['first_posted'],
                          'end_date': info['last_posted'],
                          'rows_processed': info['rows'],
                          'error_message': 'uploaded from Seller Central'},
            )

        write_cached(mp, rep.report_id, raw)
        m = info['first_posted'].replace(day=1)
        months.add(f'{m:%Y-%m}')
        if info['last_posted'].month != info['first_posted'].month:
            months.add(f'{info["last_posted"].replace(day=1):%Y-%m}')

        results.append({
            'file': f.name, 'ok': True,
            'report_id': rep.report_id,
            'settlement_ids': info['settlement_ids'],
            'rows': info['rows'],
            'posted': f"{info['first_posted']} → {info['last_posted']}",
        })

    ok = [r for r in results if r['ok']]
    return JsonResponse({
        'status': 'ok' if ok else 'failed',
        'message': (
            f'Cached {len(ok)} settlement file(s). Now press Sync from Amazon '
            f'for: {", ".join(sorted(months))} — the rebuild will read these '
            f'instead of downloading.'
            if ok else 'No file could be read.'),
        'months': sorted(months),
        'results': results,
    })
