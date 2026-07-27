"""
apps/dashboard/pnl_importer.py — import manual P&L lines from an Excel sheet.

Accepts the client's "P&L Summary" layout: column A holds the line label,
and a value column holds the amount. We match labels → line keys via
pnl_lines.label_to_key and write ONLY source='manual' lines (auto/computed
lines are ignored — those come from settlement/engine).

The value column is auto-detected: the first numeric column to the right of
column A on the row whose label matches a known manual line. Caller can pin
it via `value_col` (1-based) if auto-detect picks wrong.
"""
from __future__ import annotations

import io
from datetime import date

from .pnl_lines import label_to_key, LINE_BY_KEY


def import_pnl_excel_bytes(
    *,
    file_bytes:        bytes,
    original_filename: str,
    marketplace:       str,
    month:             date,
    channel:           str = 'amazon',
    user=None,
    sheet_name:        str | None = None,
    value_col:         int | None = None,
) -> dict:
    """
    Parse + upsert manual P&L lines. Returns:
        {status, message, rows_imported, lines_matched, lines_unmatched,
         matched: [...], unmatched: [...]}
    """
    from openpyxl import load_workbook
    from .models import MonthlyPnLEntry, ManualPnLUpload

    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        ManualPnLUpload.objects.create(
            marketplace=marketplace, month=month,
            original_filename=original_filename[:256],
            status='failed', error_message=f'open failed: {exc}'[:500],
            uploaded_by=user)
        return {'status': 'failed', 'message': f'Could not open workbook: {exc}'}

    # Pick the sheet — prefer one literally named like a P&L summary
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = None
        for cand in wb.sheetnames:
            if 'p&l' in cand.lower() or 'pnl' in cand.lower() or 'summary' in cand.lower():
                ws = wb[cand]
                break
        if ws is None:
            ws = wb.active

    matched   = []
    unmatched = []
    seen_keys = set()

    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if not label or not str(label).strip():
            continue
        key = label_to_key(str(label))
        if not key:
            continue
        ln = LINE_BY_KEY.get(key)
        if not ln or ln['source'] != 'manual':
            continue          # only import manual lines
        if key in seen_keys:
            continue          # first occurrence wins

        # Find the value: pinned column, else first numeric to the right
        amount = None
        if value_col:
            v = ws.cell(r, value_col).value
            amount = _as_float(v)
        else:
            for c in range(2, min(ws.max_column + 1, 30)):
                v = ws.cell(r, c).value
                fv = _as_float(v)
                if fv is not None:
                    amount = fv
                    break
        if amount is None:
            unmatched.append({'label': str(label), 'key': key, 'reason': 'no value'})
            continue

        MonthlyPnLEntry.objects.update_or_create(
            marketplace=marketplace, month=month, channel=channel, line_key=key,
            defaults={'amount': amount, 'updated_by': user,   # signed — allow credits/rebates
                       'note': f'imported from {original_filename[:80]}'},
        )
        seen_keys.add(key)
        matched.append({'label': str(label), 'key': key, 'amount': amount})

    audit = ManualPnLUpload.objects.create(
        marketplace=marketplace, month=month,
        original_filename=original_filename[:256],
        rows_imported=len(matched), lines_matched=len(matched),
        lines_unmatched=len(unmatched), status='ok', uploaded_by=user)

    return {
        'status':          'ok',
        'message':         f'Imported {len(matched)} manual line(s) for '
                            f'{month:%Y-%m} ({channel}).',
        'rows_imported':   len(matched),
        'lines_matched':   len(matched),
        'lines_unmatched': len(unmatched),
        'matched':         matched,
        'unmatched':       unmatched,
        'audit_id':        audit.id,
    }


def _as_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(',', '').replace('$', '')
    if s.startswith('(') and s.endswith(')'):   # accounting negatives
        s = '-' + s[1:-1]
    try:
        return float(s)
    except ValueError:
        return None
