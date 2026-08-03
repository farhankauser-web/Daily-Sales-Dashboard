"""
Link containers to the Amazon shipment ID they were dispatched under.

The ops workbook ("Containers Summary") carries the mapping ops already
maintain by hand:

    Container # | Region | Moved to | Shipment ID
    CCLU7914144 | USA    | AWD      | STAR-VYUQYVAQTT7T4

That ID is what makes per-container receipt tracking possible: asking Amazon
about STAR-VYUQYVAQTT7T4 returns declared vs received quantities for THAT
container, instead of the SKU-level aggregate we have to guess from today.

Reports by default; --apply writes. Existing IDs are never silently replaced
— a conflict is reported and skipped unless --overwrite is given.
"""
from __future__ import annotations

from django.core.management.base import BaseCommand, CommandError


class Command(BaseCommand):
    help = ('Backfill InTransitShipment.shipment_id (the Amazon shipment ID) '
            'from the Containers Summary workbook.')

    def add_arguments(self, parser):
        parser.add_argument('xlsx', help='Path to "Containers Summary" .xlsx')
        parser.add_argument('--sheet', default='Container Details')
        parser.add_argument('--apply', action='store_true',
                            help='Write the IDs. Without this, report only.')
        parser.add_argument('--overwrite', action='store_true',
                            help='Replace an existing, different shipment ID.')

    def handle(self, *args, **opts):
        try:
            import openpyxl
        except ImportError:
            raise CommandError('openpyxl not installed: pip install openpyxl')
        from apps.inventory_planning.models import InTransitShipment

        try:
            wb = openpyxl.load_workbook(opts['xlsx'], data_only=True)
        except Exception as exc:
            raise CommandError(f'cannot open workbook: {exc}')
        if opts['sheet'] not in wb.sheetnames:
            raise CommandError(f'sheet "{opts["sheet"]}" not found; '
                               f'have {wb.sheetnames}')
        ws = wb[opts['sheet']]

        # Find the header row rather than assuming row 1 — the ops file has a
        # blank spacer row above it, and that will drift as people edit.
        header, header_row = None, None
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15,
                                             values_only=True), start=1):
            cells = [str(c or '').strip().lower() for c in row]
            if 'container #' in cells or 'container' in cells:
                header, header_row = cells, i
                break
        if header is None:
            raise CommandError('could not find a header row containing '
                               '"Container #" in the first 15 rows')

        def col(*names):
            for n in names:
                if n in header:
                    return header.index(n)
            return None

        c_cont, c_ship = col('container #', 'container'), col('shipment id', 'shipment')
        c_dest = col('moved to', 'destination')
        if c_cont is None or c_ship is None:
            raise CommandError(f'need "Container #" and "Shipment ID" columns; '
                               f'found {header}')
        self.stdout.write(f'header found on row {header_row}: {header}')

        pairs, skipped_rows = {}, 0
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            cont = str(row[c_cont] or '').strip().upper()
            ship = str(row[c_ship] or '').strip()
            if not cont or not ship or cont == 'CONTAINER #':
                skipped_rows += 1
                continue
            pairs[cont] = (ship, str(row[c_dest] or '').strip() if c_dest is not None else '')

        self.stdout.write(f'workbook rows usable: {len(pairs)} '
                          f'(skipped {skipped_rows} blank/header)')

        matched, already, conflict, missing = [], [], [], []
        for cont, (ship, dest) in sorted(pairs.items()):
            sh = InTransitShipment.objects.filter(container_no__iexact=cont).first()
            if sh is None:
                missing.append((cont, ship, dest))
            elif not (sh.shipment_id or '').strip():
                matched.append((sh, cont, ship, dest))
            elif sh.shipment_id.strip() == ship:
                already.append(cont)
            else:
                conflict.append((sh, cont, sh.shipment_id, ship))

        self.stdout.write(
            f'\n  to link      : {len(matched)}'
            f'\n  already set  : {len(already)}'
            f'\n  conflicting  : {len(conflict)}'
            f'\n  no container : {len(missing)}   (in workbook, not in Pulse)')

        if matched:
            self.stdout.write('\nWILL LINK:')
            for sh, cont, ship, dest in matched[:60]:
                self.stdout.write(f'   {cont:<16} → {ship:<22} '
                                  f'[{dest or "?"}] status={sh.status}')
        if conflict:
            self.stdout.write(self.style.WARNING('\nCONFLICTS (kept unless --overwrite):'))
            for sh, cont, old, new in conflict:
                self.stdout.write(f'   {cont:<16} has {old!r}, workbook says {new!r}')
        if missing:
            self.stdout.write(self.style.NOTICE(
                f'\nIn workbook but no matching container in Pulse '
                f'({len(missing)}) — likely older than the data we imported:'))
            for cont, ship, dest in missing[:15]:
                self.stdout.write(f'   {cont:<16} {ship}')
            if len(missing) > 15:
                self.stdout.write(f'   … and {len(missing)-15} more')

        if not opts['apply']:
            self.stdout.write(self.style.WARNING(
                '\nReport only — nothing written. Re-run with --apply to link.'))
            return

        n = 0
        for sh, cont, ship, dest in matched:
            sh.shipment_id = ship[:64]
            sh.save(update_fields=['shipment_id'])
            n += 1
        if opts['overwrite']:
            for sh, cont, old, new in conflict:
                sh.shipment_id = new[:64]
                sh.save(update_fields=['shipment_id'])
                n += 1
        self.stdout.write(self.style.SUCCESS(f'\n✅ linked {n} container(s).'))
