"""Inventory Planner (Phase 1): projection page + API + imports/exports."""
from __future__ import annotations

import json
from collections import defaultdict
from datetime import date

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.views.decorators.http import require_POST

from apps.core.decorators import permission_required

from .models import DemandInput, InTransitShipment, PlanningSku, Warehouse
from .planning import build_projection

# Statuses that must NOT appear in the Containers — In Transit list.
# 'received'/'cancelled' are finished; 'receiving' has moved on to the
# Receiving tab, where the shipped-vs-counted picture lives.
_NOT_IN_TRANSIT = ['received', 'cancelled', 'receiving']


@login_required
@permission_required('can_view_inventory')
def planner(request):
    return render(request, 'inventory_planning/planner.html', {
        'can_edit': request.user.has_perm_flag('can_manage_cogs'),
    })


@login_required
@permission_required('can_view_inventory')
def containers(request):
    return render(request, 'inventory_planning/containers.html', {
        'can_edit': request.user.has_perm_flag('can_manage_cogs'),
    })


@login_required
@permission_required('can_view_inventory')
def api_containers(request):
    from .models import InTransitShipment
    region = request.GET.get('region', 'usa')
    show = request.GET.get('show', 'active')
    qs = (InTransitShipment.objects.filter(region=region)
          .select_related('destination').prefetch_related('lines')
          .order_by('eta_destination', 'eta_port'))
    if show == 'active':
        # 'receiving' is deliberately excluded here as well as the terminal
        # states: once Amazon starts counting a container in it belongs to the
        # Receiving tab, not In Transit. Leaving it in both would show the same
        # container twice, which is the duplication this stage exists to end.
        # The planner is unaffected — it keys off its own exclude() of
        # received/cancelled and still counts the un-received remainder.
        qs = qs.exclude(status__in=_NOT_IN_TRANSIT)
    rows = []
    for sh in qs:
        rows.append({
            'id': sh.pk,
            'container_no': sh.container_no,
            'shipment_id': sh.shipment_id,
            'vendor': sh.vendor,
            'destination': sh.destination.name if sh.destination else '—',
            'departure': sh.departure_date.isoformat() if sh.departure_date else None,
            'eta_port': sh.eta_port.isoformat() if sh.eta_port else None,
            'eta_dest': sh.eta_destination.isoformat() if sh.eta_destination else None,
            'status': sh.status,
            'units': sh.total_units,
            'lines': [{'sku': l.sku, 'units': l.units}
                      for l in sh.lines.all()],
        })
    kpi = {
        'active': (InTransitShipment.objects.filter(region=region)
                   .exclude(status__in=_NOT_IN_TRANSIT).count()),
        'units': sum(r['units'] for r in rows
                     if r['status'] not in _NOT_IN_TRANSIT),
    }
    # SKU metadata (name/type/category) for the matrix view
    from .models import PlanningSku
    skus = {s for r in rows for s in (ln['sku'] for ln in r['lines'])}
    meta = {p.sku.upper(): {'name': p.name, 'type': p.sku_type,
                            'category': p.category}
            for p in PlanningSku.objects.filter(region=region,
                                                sku__in=skus)}
    sku_meta = {s: meta.get(s, {'name': '', 'type': '', 'category': ''})
                for s in skus}
    return JsonResponse({'rows': rows, 'kpi': kpi, 'region': region,
                         'sku_meta': sku_meta})


@login_required
@permission_required('can_view_inventory')
def container_history(request):
    return render(request, 'inventory_planning/container_history.html', {
        'can_edit': request.user.has_perm_flag('can_manage_cogs'),
    })


@login_required
@permission_required('can_view_inventory')
def receiving(request):
    """Containers Amazon is currently counting in — the stage between In
    Transit and Container History."""
    return render(request, 'inventory_planning/receiving.html', {
        'can_edit': request.user.has_perm_flag('can_manage_cogs'),
    })


@login_required
@permission_required('can_view_inventory')
def api_receiving(request):
    """
    Containers linked to an Amazon shipment, with the shipped-vs-received
    picture per SKU.

    Three quantities per line, and they are not interchangeable:
      packed (B)   our packing list — the truth of what left the factory
      declared (A) what we told Amazon when generating labels; A ≥ B always
      received (C) what Amazon has counted in, converted from CASES to eaches

    Variance is B − C. It is never A − C: because we always declare at least
    as much as we pack, an A-based figure invents a shortage out of our own
    over-declaration.
    """
    from .models import InTransitShipment
    region = request.GET.get('region', 'usa')
    qs = (InTransitShipment.objects
          .filter(region=region)
          .exclude(shipment_id='')
          .exclude(status='cancelled')
          .select_related('destination')
          .prefetch_related('lines')
          .order_by('-amazon_synced_at', 'eta_destination'))
    show = (request.GET.get('show') or 'active')
    if show == 'active':
        qs = qs.exclude(status='received')

    rows = []
    for sh in qs[:300]:
        lines, packed, declared, received = [], 0, 0, 0
        for l in sh.lines.all():
            b = int(l.units or 0)
            a = int(l.amazon_expected_units or 0)
            c = int(l.amazon_received_units or 0)
            packed += b; declared += a; received += c
            lines.append({
                'sku': l.sku, 'packed': b, 'declared': a, 'received': c,
                'remaining': max(0, b - c),
                'short': max(0, b - c) if c else 0,
                'over': max(0, c - b),
                'per_case': int(l.units_per_case or 0),
                'pct': round(min(100, c / b * 100), 1) if b else 0,
            })
        lines.sort(key=lambda x: (-x['short'], -x['packed']))
        started = received > 0
        rows.append({
            'id': sh.pk,
            'container': sh.container_no or f'#{sh.pk}',
            'shipment_id': sh.shipment_id,
            'vendor': sh.vendor, 'status': sh.status,
            'amazon_status': sh.amazon_status,
            'destination': sh.destination.name if sh.destination else '',
            'eta': sh.eta_destination.isoformat() if sh.eta_destination else None,
            'synced': sh.amazon_synced_at.isoformat() if sh.amazon_synced_at else None,
            'packed': packed, 'declared': declared, 'received': received,
            'remaining': max(0, packed - received),
            # Only meaningful once Amazon has started; before that the whole
            # container would read as missing.
            'variance': (packed - received) if started else None,
            'over_declared': declared - packed,
            'started': started,
            'pct': round(min(100, received / packed * 100), 1) if packed else 0,
            'lines': lines,
        })
    return JsonResponse({'rows': rows, 'region': region})


@login_required
@permission_required('can_view_inventory')
def container_view(request, pk):
    """Standalone page (opens in a new tab) showing one container's contents
    in the Type/Name/SKU/Units matrix format."""
    from .models import InTransitShipment, PlanningSku
    sh = (InTransitShipment.objects.filter(pk=pk)
          .select_related('destination', 'received_by')
          .prefetch_related('lines').first())
    if not sh:
        from django.http import Http404
        raise Http404
    meta = {p.sku.upper(): p for p in PlanningSku.objects.filter(
        region=sh.region, sku__in=[l.sku for l in sh.lines.all()])}
    lines = []
    for l in sh.lines.all():
        p = meta.get(l.sku.upper())
        lines.append({'type': p.sku_type if p else '', 'name': p.name if p else '',
                      'sku': l.sku, 'units': l.units,
                      'received': l.received_units,
                      'disc': l.received_units - l.units})
    lines.sort(key=lambda x: (x['name'] or x['sku']))
    return render(request, 'inventory_planning/container_view.html',
                  {'sh': sh, 'lines': lines, 'archived': sh.is_archived})


@login_required
@permission_required('can_view_inventory')
def api_container_history(request):
    """Archived (received/cancelled) shipments — searchable/filterable."""
    from django.db.models import Q
    from .models import InTransitShipment
    region = request.GET.get('region', 'usa')
    qs = (InTransitShipment.objects.filter(
            region=region, status__in=['received', 'cancelled'])
          .select_related('destination', 'received_by')
          .prefetch_related('lines').order_by('-received_date', '-updated_at'))
    q = (request.GET.get('q') or '').strip()
    if q:
        qs = qs.filter(
            Q(container_no__icontains=q) | Q(shipment_id__icontains=q)
            | Q(po_number__icontains=q) | Q(vendor__icontains=q)
            | Q(destination__name__icontains=q))
    st = request.GET.get('status') or ''
    if st:
        qs = qs.filter(status=st)
    frm, to = request.GET.get('from'), request.GET.get('to')
    if frm:
        qs = qs.filter(received_date__gte=frm)
    if to:
        qs = qs.filter(received_date__lte=to)
    rows = []
    for sh in qs[:500]:
        disc = sum(l.received_units - l.units for l in sh.lines.all())
        rows.append({
            'id': sh.pk, 'container_no': sh.container_no,
            'shipment_id': sh.shipment_id, 'po_number': sh.po_number,
            'vendor': sh.vendor,
            'destination': sh.destination.name if sh.destination else '—',
            'departure': sh.departure_date.isoformat() if sh.departure_date else None,
            'eta_dest': sh.eta_destination.isoformat() if sh.eta_destination else None,
            'arrival': sh.received_date.isoformat() if sh.received_date else None,
            'status': sh.status,
            'shipped': sh.total_units, 'received': sh.total_received,
            'discrepancy': disc,
            'received_by': (getattr(sh.received_by, 'full_name', '')
                            or sh.received_by.email) if sh.received_by else '',
            'received_at': sh.received_at.strftime('%Y-%m-%d %H:%M')
                           if sh.received_at else '',
            'lines': [{'sku': l.sku, 'shipped': l.units,
                       'received': l.received_units} for l in sh.lines.all()],
        })
    return JsonResponse({'rows': rows, 'region': region})


@login_required
@permission_required('can_manage_cogs')
@require_POST
def api_receive_container(request):
    """Confirm receipt: archive the container, add its units to the
    destination warehouse (AWD/3PL), and record the audit trail.
    JSON: {id, region, warehouse_code?, arrival_date?, lines:[{sku,received}]}"""
    from datetime import datetime
    from django.db import transaction
    from django.utils import timezone
    from .models import InTransitShipment, Warehouse, WarehouseStock
    try:
        d = json.loads(request.body)
        sh = InTransitShipment.objects.prefetch_related('lines').get(
            pk=d['id'])
    except (ValueError, KeyError, InTransitShipment.DoesNotExist):
        return JsonResponse({'error': 'container not found'}, status=404)
    if sh.is_archived:
        return JsonResponse({'status': 'failed',
                             'error': 'Container already received/cancelled.'})

    wh = sh.destination
    if d.get('warehouse_code'):
        wh = Warehouse.objects.filter(code=d['warehouse_code']).first() or wh
    try:
        arrival = (datetime.strptime(d['arrival_date'], '%Y-%m-%d').date()
                   if d.get('arrival_date') else timezone.now().date())
    except (ValueError, TypeError):
        arrival = timezone.now().date()
    recv = {str(x.get('sku', '')).upper(): int(x.get('received') or 0)
            for x in (d.get('lines') or [])}
    now = timezone.now()
    added = 0
    with transaction.atomic():
        for line in sh.lines.all():
            got = recv.get(line.sku.upper(), line.units)   # default = shipped
            line.received_units = got
            line.save(update_fields=['received_units'])
            # add received units to AWD/3PL warehouse (FBA is API-fed)
            if wh and wh.kind in ('awd', '3pl') and got > 0:
                ws, _ = WarehouseStock.objects.get_or_create(
                    warehouse=wh, sku=line.sku.upper(),
                    defaults={'units': 0, 'as_of': now, 'source': 'received'})
                ws.units = (ws.units or 0) + got
                ws.as_of = now
                if ws.source not in ('api', 'import'):
                    ws.source = 'received'
                ws.save()
                added += got
        sh.status = 'received'
        sh.received_date = arrival
        sh.received_at = now
        sh.received_by = request.user
        if wh:
            sh.destination = wh
        sh.save()
    return JsonResponse({'status': 'ok', 'units_added': added,
                         'warehouse': wh.name if wh else None})


@login_required
@permission_required('can_view_inventory')
def runway(request):
    return render(request, 'inventory_planning/runway.html', {
        'can_edit': request.user.has_perm_flag('can_manage_cogs'),
    })


@login_required
@permission_required('can_view_inventory')
def api_planner(request):
    region = request.GET.get('region', 'usa')
    res = build_projection(region)
    # active shipments for the transit panel
    ships = []
    for sh in (InTransitShipment.objects
               .filter(region=region)
               .exclude(status__in=['received', 'cancelled'])
               .select_related('destination')
               .prefetch_related('lines')):
        ships.append({
            'id': sh.pk,
            'container': sh.container_no or sh.shipment_id or '—',
            'vendor': sh.vendor,
            'destination': sh.destination.name if sh.destination else '—',
            'departure': sh.departure_date.isoformat() if sh.departure_date else None,
            'eta_port': sh.eta_port.isoformat() if sh.eta_port else None,
            'eta_dest': sh.eta_destination.isoformat() if sh.eta_destination else None,
            'status': sh.status,
            'units': sh.total_units,
            'skus': sh.lines.count(),
        })
    res['shipments'] = ships
    return JsonResponse(res)


@login_required
@permission_required('can_manage_cogs')
@require_POST
def api_sync_stock(request):
    """Button: refresh FBA + AWD stock from Amazon now."""
    from .management.commands.sync_planning_inventory import sync_region
    region = json.loads(request.body or '{}').get('region', 'usa')
    try:
        return JsonResponse({'status': 'ok', **sync_region(region)})
    except Exception as exc:
        return JsonResponse({'status': 'failed',
                             'error': f'{type(exc).__name__}: {exc}'},
                            status=500)


@login_required
@permission_required('can_manage_cogs')
@require_POST
def import_workbook(request):
    """Upload ops' status-report workbook (SKU master + 3PL stock + transit)."""
    from .importer import import_status_workbook
    f = request.FILES.get('file')
    region = request.POST.get('region', 'usa')
    if f is None:
        return JsonResponse({'status': 'failed', 'message': 'No file.'},
                            status=400)
    try:
        res = import_status_workbook(f.read(), region=region,
                                     user=request.user)
    except Exception as exc:
        return JsonResponse({'status': 'failed',
                             'message': f'{type(exc).__name__}: {exc}'},
                            status=400)
    return JsonResponse({'status': 'ok', **res,
                         'message': (f'{res["skus"]} SKUs, {res["stock_rows"]} '
                                     f'stock rows, {res["shipments"]} shipments '
                                     f'imported.')})


@login_required
@permission_required('can_manage_cogs')
@require_POST
def save_container(request):
    """Ops adds/edits a container when it leaves port.
    JSON: {id?, region, container_no, vendor, destination_code, departure,
           eta_port, status, lines:[{sku, units}]}"""
    from datetime import datetime
    from .models import InTransitLine, InTransitShipment, Warehouse
    try:
        d = json.loads(request.body)
    except ValueError:
        return JsonResponse({'error': 'bad payload'}, status=400)

    def _d(v):
        try:
            return datetime.strptime(v, '%Y-%m-%d').date() if v else None
        except (ValueError, TypeError):
            return None

    region = d.get('region', 'usa')
    dest = None
    if d.get('destination_code'):
        dest = Warehouse.objects.filter(code=d['destination_code']).first()
    eta_port = _d(d.get('eta_port'))
    from datetime import timedelta
    eta_dest = eta_port + timedelta(days=10) if eta_port else None
    fields = dict(
        region=region,
        container_no=(d.get('container_no') or '')[:32],
        shipment_id=(d.get('shipment_id') or '')[:64],
        po_number=(d.get('po_number') or '')[:64],
        vendor=(d.get('vendor') or '')[:64],
        destination=dest,
        departure_date=_d(d.get('departure')),
        eta_port=eta_port, eta_destination=eta_dest,
        status=d.get('status') or 'in_transit',
        notes=(d.get('notes') or '')[:256])
    if d.get('id'):
        sh = InTransitShipment.objects.filter(pk=d['id'], region=region).first()
        if not sh:
            return JsonResponse({'error': 'container not found'}, status=404)
        for k, v in fields.items():
            setattr(sh, k, v)
        sh.save()
        sh.lines.all().delete()
    else:
        sh = InTransitShipment.objects.create(**fields)
    for ln in (d.get('lines') or []):
        sku = str(ln.get('sku') or '').strip().upper()
        try:
            units = int(ln.get('units') or 0)
        except (TypeError, ValueError):
            units = 0
        if sku and units > 0:
            InTransitLine.objects.create(shipment=sh, sku=sku, units=units)
    return JsonResponse({'status': 'ok', 'id': sh.pk})


@login_required
@permission_required('can_manage_cogs')
@require_POST
def delete_container(request):
    from .models import InTransitShipment
    pk = json.loads(request.body or '{}').get('id')
    InTransitShipment.objects.filter(pk=pk).delete()
    return JsonResponse({'status': 'ok'})


@login_required
@permission_required('can_view_inventory')
def container_detail(request, pk):
    """Lines of one container (for the edit modal)."""
    from .models import InTransitShipment
    sh = InTransitShipment.objects.filter(pk=pk).prefetch_related('lines').first()
    if not sh:
        return JsonResponse({'error': 'not found'}, status=404)
    return JsonResponse({
        'id': sh.pk, 'container_no': sh.container_no,
        'shipment_id': sh.shipment_id, 'po_number': sh.po_number,
        'vendor': sh.vendor,
        'destination_code': sh.destination.code if sh.destination else '',
        'departure': sh.departure_date.isoformat() if sh.departure_date else '',
        'eta_port': sh.eta_port.isoformat() if sh.eta_port else '',
        'status': sh.status,
        'lines': [{'sku': l.sku, 'units': l.units} for l in sh.lines.all()],
    })


@login_required
@permission_required('can_view_inventory')
def warehouses(request):
    """3PL/AWD destinations for the container form dropdown."""
    from .models import Warehouse
    region = request.GET.get('region', 'usa')
    ws = (Warehouse.objects.filter(region=region, is_active=True)
          .exclude(kind='fba').order_by('kind', 'name'))
    return JsonResponse({'warehouses': [
        {'id': w.pk, 'code': w.code, 'name': w.name, 'kind': w.kind}
        for w in ws]})


@login_required
@permission_required('can_manage_cogs')
@require_POST
def import_master(request):
    """File 1: SKU + PDS + 3PL inventory."""
    from .importer import import_master_file
    f = request.FILES.get('file')
    region = request.POST.get('region', 'usa')
    if f is None:
        return JsonResponse({'status': 'failed', 'message': 'No file.'}, status=400)
    try:
        r = import_master_file(f.read(), region=region, user=request.user)
    except Exception as exc:
        return JsonResponse({'status': 'failed',
                             'message': f'{type(exc).__name__}: {exc}'}, status=400)
    msg = (f'{r["skus"]} SKUs, {r["pds_rows"]} PDS updates, {r["stock_rows"]} '
           f'3PL stock rows. Warehouses: {", ".join(r["warehouses"]) or "—"}.'
           + (f' {r["discontinued"]} discontinued dropped.' if r['discontinued'] else ''))
    return JsonResponse({'status': 'ok', **r, 'message': msg})


@login_required
@permission_required('can_manage_cogs')
@require_POST
def import_containers_view(request):
    """File 2: container manifest → transit."""
    from .importer import import_containers
    f = request.FILES.get('file')
    region = request.POST.get('region', 'usa')
    if f is None:
        return JsonResponse({'status': 'failed', 'message': 'No file.'}, status=400)
    try:
        r = import_containers(f.read(), region=region, user=request.user)
    except Exception as exc:
        return JsonResponse({'status': 'failed',
                             'message': f'{type(exc).__name__}: {exc}'}, status=400)
    return JsonResponse({'status': 'ok', **r,
                         'message': (f'{r["containers"]} containers, '
                                     f'{r["units"]:,} units in transit.')})


@login_required
@permission_required('can_view_inventory')
def master_template(request):
    """Template for File 1 — pre-filled with active SKUs + current PDS/3PL."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from .models import DemandInput, PlanningSku, Warehouse, WarehouseStock
    region = request.GET.get('region', 'usa')
    tpl_whs = list(Warehouse.objects.filter(region=region, is_active=True,
                                            kind='3pl').order_by('name'))
    wb = Workbook(); ws = wb.active; ws.title = 'Master'
    hdr = ['SKU', 'Name', 'Category', 'SKU Type', 'Status', 'PDS',
           'In Hand Pakistan', 'In Production'] + [w.name for w in tpl_whs]
    ws.append(hdr)
    for c in ws[1]:
        c.font = Font(bold=True)
    pds = {d.sku: d.pds for d in DemandInput.objects.filter(region=region)
           .order_by('effective_from')}
    stock = defaultdict(dict)
    for s in WarehouseStock.objects.select_related('warehouse').filter(
            warehouse__in=tpl_whs):
        stock[s.sku.upper()][s.warehouse.name] = s.units
    for ps in PlanningSku.objects.filter(region=region, is_active=True).order_by('sku'):
        row = [ps.sku, ps.name, ps.category, ps.sku_type, ps.product_status,
               pds.get(ps.sku.upper(), ''), ps.factory_stock,
               ps.factory_production]
        row += [stock[ps.sku.upper()].get(w.name, 0) for w in tpl_whs]
        ws.append(row)
    for i in range(1, len(hdr) + 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = 16
    resp = HttpResponse(content_type='application/vnd.openxmlformats-'
                                     'officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="inventory_master.xlsx"'
    wb.save(resp)
    return resp


@login_required
@permission_required('can_view_inventory')
def container_template(request):
    """Template for File 2 — container manifest."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook(); ws = wb.active; ws.title = 'Containers'
    hdr = ['Container No', 'Vendor', 'Destination', 'Departure Date',
           'ETA Port', 'Status', 'SKU', 'Units']
    ws.append(hdr)
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.append(['ONEU1234567', 'AKT', 'AWD-USA', '2026-07-20', '2026-09-01',
               'departed', 'BTH-MT-GRN-900', 500])
    for i, w in enumerate([16, 12, 16, 14, 14, 12, 22, 10], start=1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w
    resp = HttpResponse(content_type='application/vnd.openxmlformats-'
                                     'officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="container_template.xlsx"'
    wb.save(resp)
    return resp


@login_required
@permission_required('can_manage_cogs')
@require_POST
def save_pds(request):
    """Quick PDS set from the grid: {sku, pds, region}."""
    try:
        d = json.loads(request.body)
        sku = str(d['sku']).strip().upper()
        pds = float(d['pds'])
    except (ValueError, KeyError, TypeError):
        return JsonResponse({'error': 'bad payload'}, status=400)
    DemandInput.objects.create(
        sku=sku, region=d.get('region', 'usa'), pds=pds,
        effective_from=date.today(), entered_by=request.user,
        note='set from planner grid')
    return JsonResponse({'status': 'ok'})


@login_required
@permission_required('can_manage_cogs')
@require_POST
def import_pds(request):
    """PDS bulk upload: xlsx with columns SKU | PDS (optional From, To)."""
    import openpyxl
    import io
    f = request.FILES.get('file')
    if f is None:
        return JsonResponse({'status': 'failed', 'message': 'No file.'},
                            status=400)
    region = request.POST.get('region', 'usa')
    wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
    ws = wb.active
    n = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        sku = str(row[0]).strip().upper()
        try:
            pds = float(row[1])
        except (TypeError, ValueError, IndexError):
            continue
        eff_from = row[2] if len(row) > 2 and row[2] else date.today()
        eff_to = row[3] if len(row) > 3 and row[3] else None
        if hasattr(eff_from, 'date'):
            eff_from = eff_from.date()
        if eff_to is not None and hasattr(eff_to, 'date'):
            eff_to = eff_to.date()
        DemandInput.objects.create(sku=sku, region=region, pds=pds,
                                   effective_from=eff_from, effective_to=eff_to,
                                   entered_by=request.user, note='bulk upload')
        n += 1
    return JsonResponse({'status': 'ok', 'message': f'{n} PDS rows imported.'})


@login_required
@permission_required('can_view_inventory')
def pds_template(request):
    """Download the PDS upload template pre-filled with current SKUs."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    region = request.GET.get('region', 'usa')
    wb = Workbook()
    ws = wb.active
    ws.title = 'PDS'
    ws.append(['SKU', 'PDS (units/day)', 'From (optional)', 'To (optional)'])
    for c in ws[1]:
        c.font = Font(bold=True)
    cur = {d.sku: d.pds for d in DemandInput.objects.filter(region=region)
           .order_by('effective_from')}
    for ps in PlanningSku.objects.filter(region=region, is_active=True):
        ws.append([ps.sku, cur.get(ps.sku.upper(), '')])
    for col, w in zip('ABCD', [26, 16, 14, 14]):
        ws.column_dimensions[col].width = w
    resp = HttpResponse(content_type='application/vnd.openxmlformats-'
                                     'officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="pds_template.xlsx"'
    wb.save(resp)
    return resp


@login_required
@permission_required('can_view_inventory')
def export_planner(request):
    """Export the current projection grid to xlsx."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    region = request.GET.get('region', 'usa')
    res = build_projection(region)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventory Planner'
    hdr = ['SKU Type', 'Category', 'Name', 'SKU', 'Average Daily Sale',
           'Average 30 days Sale', 'Average 90 days Sale',
           'Potential Daily Sale', 'Coverage Days (30 Day Avg Sale)',
           'Coverage Days (90 Day Avg Sale)', 'No. of Days(Available/PDS)',
           'Available', 'Inbound', 'Reserved', 'Total Inventory Amazon',
           'Coverage Days Amazon', 'Stock Required Days',
           'Units Required at Amazon FC', '3PL WH Inventory',
           'AWD Inventory', 'Total Warehouse Inventory',
           'Coverage Days WH Inventory', 'In Transit', 'Coverage Days',
           'Total coverage days USA', 'In Hand Stock Pakistan',
           'Coverage Days ', 'In Production', 'Total Coverage Days']
    ws.append(hdr)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in res['rows']:
        ws.append([r['sku_type'], r['category'], r['name'], r['sku'],
                   r['ads'], r['avg30'], r['avg90'], r['pds'],
                   r['cov_30'], r['cov_90'], r['days_avail_pds'],
                   r['available'], r['inbound'], r['reserved'],
                   r['total_amazon'], r['cov_amazon'], r['stock_req_days'],
                   r['units_req_fc'], r['jarrett'], r['awd'], r['total_wh'],
                   r['cov_wh'], r['transit'], r['cov_transit'],
                   r['total_cov_usa'], r['pak_stock'], r['cov_pak'],
                   r['in_production'], r['total_cov']])
    resp = HttpResponse(content_type='application/vnd.openxmlformats-'
                                     'officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = (f'attachment; filename='
                                   f'"inventory_planner_{region}.xlsx"')
    wb.save(resp)
    return resp


# ── Procurement (Phase 2): Suppliers → PO → Lines → Production Plans ────────

@login_required
@permission_required('can_view_inventory')
def suppliers(request):
    return render(request, 'inventory_planning/suppliers.html', {
        'can_edit': request.user.has_perm_flag('can_manage_cogs'),
    })


@login_required
@permission_required('can_view_inventory')
def purchase_orders(request):
    return render(request, 'inventory_planning/purchase_orders.html', {
        'can_edit': request.user.has_perm_flag('can_manage_cogs'),
    })


def _line_rows(lines):
    """Shared ledger shape for a set of PO Lines (SKU grain)."""
    return [{
        'id': l.pk, 'sku': l.sku, 'name': l.name,
        'category': l.group.category, 'pp': getattr(
            getattr(l.group, 'plan', None), 'pp_number', ''),
        'po': l.po.po_number, 'po_id': l.po_id,
        'fob': float(l.group.fob_rate),
        'ordered': l.ordered_units, 'wastage': l.wastage_units,
        'allocated': l.allocated_units, 'loaded': l.loaded_units,
        'received': l.received_units, 'transit': l.in_transit_units,
        'remaining': l.remaining_units, 'status': l.status,
        'ready': (l.expected_ready_date.isoformat()
                  if l.expected_ready_date else None),
    } for l in lines]


def _all_lines():
    from .models import POLine
    return (POLine.objects
            .select_related('group__plan', 'po__supplier')
            .prefetch_related('allocations__shipment'))


@login_required
@permission_required('can_view_inventory')
def api_suppliers(request):
    """Supplier dashboard — one row per supplier with the full ledger."""
    from .models import Supplier
    from .procurement import opening_by_supplier

    opening = opening_by_supplier()
    agg = {}
    for l in _all_lines():
        s = l.po.supplier
        a = agg.setdefault(s.pk, {'ordered': 0, 'wastage': 0, 'allocated': 0,
                                  'loaded': 0, 'received': 0, 'remaining': 0,
                                  'value': 0.0, 'pos': set()})
        a['ordered']   += l.ordered_units
        a['wastage']   += l.wastage_units
        a['allocated'] += l.allocated_units
        a['loaded']    += l.loaded_units
        a['received']  += l.received_units
        a['remaining'] += l.remaining_units
        a['value']     += float(l.group.fob_rate) * l.remaining_units
        if l.po.status not in ('closed', 'cancelled'):
            a['pos'].add(l.po_id)

    rows = []
    for s in Supplier.objects.filter(is_active=True):
        a = agg.get(s.pk, {'ordered': 0, 'wastage': 0, 'allocated': 0,
                           'loaded': 0, 'received': 0, 'remaining': 0,
                           'value': 0.0, 'pos': set()})
        op = opening.get(s.pk, 0)
        rows.append({
            'id': s.pk, 'code': s.code, 'name': s.name,
            'country': s.country, 'currency': s.currency,
            'open_pos': len(a['pos']),
            'opening': op,
            'ordered': a['ordered'], 'wastage': a['wastage'],
            'allocated': a['allocated'], 'loaded': a['loaded'],
            'in_transit': a['allocated'] - a['received'],
            'received': a['received'],
            # opening backlog carries into what is still owed to us
            'remaining': op + a['remaining'],
            'outstanding_value': round(a['value'], 2),
            'lead_days': s.production_lead_days + s.sea_lead_days
                         + s.port_to_wh_days,
            'capacity': s.monthly_capacity_units,
        })
    rows.sort(key=lambda r: -r['remaining'])
    tot = {k: sum(r[k] for r in rows) for k in
           ('opening', 'ordered', 'wastage', 'allocated', 'received',
            'remaining', 'in_transit')}
    tot['outstanding_value'] = round(sum(r['outstanding_value'] for r in rows), 2)
    tot['open_pos'] = sum(r['open_pos'] for r in rows)
    return JsonResponse({'rows': rows, 'totals': tot})


@login_required
@permission_required('can_view_inventory')
def api_supplier_detail(request, pk):
    """Drill 1 — a supplier's ledger broken down BY PRODUCT CATEGORY."""
    from .models import Supplier
    from .procurement import opening_by_category

    s = Supplier.objects.get(pk=pk)
    opening = opening_by_category(pk)
    cats = {}
    for l in _all_lines().filter(po__supplier_id=pk):
        c = l.group.category or 'Uncategorised'
        a = cats.setdefault(c, {'category': c, 'opening': opening.get(c, 0),
                                'ordered': 0, 'wastage': 0, 'allocated': 0,
                                'loaded': 0, 'received': 0, 'remaining': 0,
                                'value': 0.0, 'skus': set()})
        a['ordered']   += l.ordered_units
        a['wastage']   += l.wastage_units
        a['allocated'] += l.allocated_units
        a['loaded']    += l.loaded_units
        a['received']  += l.received_units
        a['remaining'] += l.remaining_units
        a['value']     += float(l.group.fob_rate) * l.remaining_units
        a['skus'].add(l.sku)
    # categories that exist only as opening backlog (no PO yet)
    for c, units in opening.items():
        if c not in cats:
            cats[c] = {'category': c, 'opening': units, 'ordered': 0,
                       'wastage': 0, 'allocated': 0, 'loaded': 0,
                       'received': 0, 'remaining': 0, 'value': 0.0,
                       'skus': set()}
    rows = []
    for a in cats.values():
        a['skus'] = len(a['skus'])
        a['value'] = round(a['value'], 2)
        a['remaining'] = a['opening'] + a['remaining']
        rows.append(a)
    rows.sort(key=lambda r: -r['remaining'])

    pos = [{'id': po.pk, 'po_number': po.po_number,
            'order_date': po.order_date.isoformat() if po.order_date else None,
            'status': po.status, 'ordered': po.ordered_units,
            'allocated': po.allocated_units, 'received': po.received_units,
            'remaining': po.remaining_units,
            'fob_value': float(po.fob_value)}
           for po in s.purchase_orders.prefetch_related('groups__lines', 'lines')
                      .order_by('-order_date')]
    return JsonResponse({'supplier': {'id': s.pk, 'name': s.name,
                                      'code': s.code, 'country': s.country},
                         'categories': rows, 'purchase_orders': pos})


@login_required
@permission_required('can_view_inventory')
def api_purchase_orders(request):
    from .models import PurchaseOrder
    qs = (PurchaseOrder.objects.select_related('supplier')
          .prefetch_related('groups', 'lines__allocations__shipment',
                            'lines__group'))
    sup = request.GET.get('supplier')
    if sup:
        qs = qs.filter(supplier_id=sup)
    status = request.GET.get('status')
    if status:
        qs = qs.filter(status=status)
    rows = [{'id': po.pk, 'po_number': po.po_number,
             'supplier': po.supplier.name, 'supplier_id': po.supplier_id,
             'order_date': po.order_date.isoformat() if po.order_date else None,
             'ready': (po.expected_ready_date.isoformat()
                       if po.expected_ready_date else None),
             'status': po.status, 'terms': po.payment_terms,
             'ordered': po.ordered_units, 'wastage': po.wastage_units,
             'allocated': po.allocated_units, 'received': po.received_units,
             'remaining': po.remaining_units,
             'fob_value': float(po.fob_value)} for po in qs]
    return JsonResponse({'rows': rows})


@login_required
@permission_required('can_view_inventory')
def api_po_detail(request, pk):
    """Drill 2/3 — a PO with its Production Plans (one per product category)
    and the SKU lines inside each."""
    from .models import PurchaseOrder
    po = PurchaseOrder.objects.select_related('supplier').get(pk=pk)
    lines = _all_lines().filter(po=po)
    rows = _line_rows(lines)
    groups = []
    for g in po.groups.select_related('plan').prefetch_related('lines'):
        gr = [r for r in rows if r['category'] == g.category]
        plan = getattr(g, 'plan', None)
        groups.append({
            'id': g.pk, 'reference': g.reference, 'category': g.category,
            'pp': plan.pp_number if plan else '',
            'plan_id': plan.pk if plan else None,
            'fob': float(g.fob_rate), 'boxes': g.boxes,
            'ordered': g.ordered_units, 'amount': float(g.total_amount),
            'wastage': sum(r['wastage'] for r in gr),
            'allocated': sum(r['allocated'] for r in gr),
            'received': sum(r['received'] for r in gr),
            'remaining': sum(r['remaining'] for r in gr),
            'lines': gr})
    return JsonResponse({
        'po': {'id': po.pk, 'po_number': po.po_number,
               'supplier': po.supplier.name, 'supplier_id': po.supplier_id,
               'order_date': po.order_date.isoformat() if po.order_date else None,
               'ready': (po.expected_ready_date.isoformat()
                         if po.expected_ready_date else None),
               'terms': po.payment_terms, 'status': po.status,
               'fob_value': float(po.fob_value),
               'ordered': po.ordered_units, 'wastage': po.wastage_units,
               'allocated': po.allocated_units, 'received': po.received_units,
               'remaining': po.remaining_units},
        'groups': groups})


@login_required
@permission_required('can_view_inventory')
def api_production_plan(request, pk):
    """Drill 4 — one Production Plan (a product on a PO): its SKU lines and
    every container allocation drawn from them."""
    from .models import ProductionPlan
    p = (ProductionPlan.objects
         .select_related('group__po__supplier')
         .prefetch_related('group__lines__allocations__shipment')
         .get(pk=pk))
    allocs = []
    for l in p.group.lines.all():
        for a in l.allocations.exclude(shipment__status='cancelled'):
            allocs.append({
                'container': a.shipment.container_no or f'#{a.shipment_id}',
                'shipment_id': a.shipment_id, 'sku': l.sku,
                'region': a.shipment.region, 'fnsku': a.fnsku,
                'status': a.shipment.status, 'units': a.units,
                'received': a.received_units,
                'eta': (a.shipment.eta_destination.isoformat()
                        if a.shipment.eta_destination else None)})
    return JsonResponse({
        'plan': {'id': p.pk, 'pp': p.pp_number, 'category': p.category,
                 'po': p.group.po.po_number, 'po_id': p.group.po_id,
                 'supplier': p.group.po.supplier.name,
                 'fob': float(p.group.fob_rate),
                 'ordered': p.ordered_qty, 'wastage': p.wastage_qty,
                 'allocated': p.allocated_qty, 'loaded': p.loaded_qty,
                 'received': p.received_qty, 'transit': p.in_transit_qty,
                 'remaining': p.remaining_qty, 'status': p.status,
                 'ready': (p.expected_ready_date.isoformat()
                           if p.expected_ready_date else None)},
        'lines': _line_rows(p.group.lines.all()),
        'allocations': allocs})


def _post_date(request, key):
    from datetime import datetime
    v = (request.POST.get(key) or '').strip()
    try:
        return datetime.strptime(v, '%Y-%m-%d').date() if v else None
    except ValueError:
        return None


@login_required
@permission_required('can_manage_cogs')
@require_POST
def import_po(request):
    """Upload an ops PO workbook (Summary + Production Plan sheets)."""
    from .procurement import import_po_workbook
    f = request.FILES.get('file')
    if f is None:
        return JsonResponse({'status': 'failed', 'message': 'No file.'},
                            status=400)
    supplier = (request.POST.get('supplier') or '').strip()
    po_number = (request.POST.get('po_number') or '').strip()
    if not supplier or not po_number:
        return JsonResponse({'status': 'failed',
                             'message': 'Supplier and PO number are required.'},
                            status=400)
    try:
        r = import_po_workbook(f, supplier_name=supplier, po_number=po_number,
                               order_date=_post_date(request, 'order_date'),
                               expected_ready_date=_post_date(request, 'ready_date'),
                               payment_terms=request.POST.get('terms', ''),
                               user=request.user)
    except Exception as exc:
        return JsonResponse({'status': 'failed',
                             'message': f'{type(exc).__name__}: {exc}'},
                            status=400)
    msg = (f'{r["po_number"]}: {r["groups"]} products ({r["plans"]} production '
           f'plans), {r["lines"]} SKU lines, {r["ordered_units"]:,} units, '
           f'FOB ${r["fob_value"]:,.2f}.')
    if r['warnings']:
        msg += ' ⚠ ' + ' '.join(r['warnings'])
    return JsonResponse({'status': 'ok', **r, 'message': msg})


@login_required
@permission_required('can_manage_cogs')
@require_POST
def import_wastage_view(request):
    from .procurement import import_wastage
    f = request.FILES.get('file')
    if f is None:
        return JsonResponse({'status': 'failed', 'message': 'No file.'},
                            status=400)
    supplier_id = request.POST.get('supplier') or None
    if not supplier_id:
        return JsonResponse({'status': 'failed',
                             'message': 'Choose the supplier this wastage '
                                        'belongs to.'}, status=400)
    try:
        r = import_wastage(f, supplier_id=supplier_id,
                           po_id=request.POST.get('po') or None)
    except Exception as exc:
        return JsonResponse({'status': 'failed',
                             'message': f'{type(exc).__name__}: {exc}'},
                            status=400)
    msg = f'{r["units"]:,} units of wastage applied across {r["applied"]} SKUs.'
    if r['unmatched']:
        msg += (f' ⚠ {len(r["unmatched"])} SKU(s) had no open balance to '
                f'absorb it — check the supplier/PO selection.')
    return JsonResponse({'status': 'ok', **r, 'message': msg})


@login_required
@permission_required('can_manage_cogs')
@require_POST
def import_opening_view(request):
    """Opening balance = backlog owed by a supplier before the system went
    live. Stored dated, so it back-dates cleanly against existing POs."""
    from datetime import date

    from .procurement import import_opening_balance
    f = request.FILES.get('file')
    if f is None:
        return JsonResponse({'status': 'failed', 'message': 'No file.'},
                            status=400)
    supplier_id = request.POST.get('supplier') or None
    if not supplier_id:
        return JsonResponse({'status': 'failed',
                             'message': 'Choose a supplier.'}, status=400)
    try:
        r = import_opening_balance(f, supplier_id=supplier_id,
                                   as_of=_post_date(request, 'as_of')
                                         or date.today())
    except Exception as exc:
        return JsonResponse({'status': 'failed',
                             'message': f'{type(exc).__name__}: {exc}'},
                            status=400)
    msg = (f'{r["supplier"]}: opening balance of {r["units"]:,} units across '
           f'{r["rows"]} SKUs, as at {r["as_of"]}.')
    return JsonResponse({'status': 'ok', **r, 'message': msg})


# ── Allocation Workbench: packing list → Container Allocations ──────────────

@login_required
@permission_required('can_view_inventory')
def allocation(request):
    return render(request, 'inventory_planning/allocation.html', {
        'can_edit': request.user.has_perm_flag('can_manage_cogs'),
    })


@login_required
@permission_required('can_manage_cogs')
@require_POST
def api_packing_preview(request):
    """Step 1 — parse a packing list and resolve every line to a Production
    Plan. Writes nothing; surfaces all problems for the operator to confirm."""
    from .procurement import parse_packing_list, preview_packing_list
    f = request.FILES.get('file')
    if f is None:
        return JsonResponse({'status': 'failed', 'message': 'No file.'},
                            status=400)
    supplier_id = request.POST.get('supplier')
    region = request.POST.get('region') or 'usa'
    if not supplier_id:
        return JsonResponse({'status': 'failed',
                             'message': 'Choose the supplier this container '
                                        'is loading from.'}, status=400)
    try:
        rows = parse_packing_list(f)
        if not rows:
            raise ValueError('No SKU rows found in the packing list.')
        prev = preview_packing_list(rows, supplier_id=supplier_id,
                                    region=region,
                                    container_size=request.POST.get('size', ''))
    except Exception as exc:
        return JsonResponse({'status': 'failed',
                             'message': f'{type(exc).__name__}: {exc}'},
                            status=400)
    return JsonResponse({'status': 'ok', **prev})


@login_required
@permission_required('can_manage_cogs')
@require_POST
def api_packing_commit(request):
    """Step 2 — write the confirmed allocations onto a container."""
    from .procurement import commit_packing_list
    try:
        d = json.loads(request.body)
        container = d['container']
        allocs = d['allocations']
    except (ValueError, KeyError):
        return JsonResponse({'status': 'failed', 'message': 'Bad payload.'},
                            status=400)
    if not container.get('container_no'):
        return JsonResponse({'status': 'failed',
                             'message': 'Container number is required.'},
                            status=400)
    if not allocs:
        return JsonResponse({'status': 'failed',
                             'message': 'Nothing to allocate.'}, status=400)
    try:
        res = commit_packing_list(container, allocs,
                                  supplier_id=d['supplier_id'],
                                  region=d.get('region') or 'usa')
    except Exception as exc:
        return JsonResponse({'status': 'failed',
                             'message': f'{type(exc).__name__}: {exc}'},
                            status=400)
    msg = (f'Container {res["container_no"]}: {res["units"]:,} units across '
           f'{res["lines"]} SKU lines allocated — balances updated.')
    return JsonResponse({'status': 'ok', **res, 'message': msg})


@login_required
@permission_required('can_view_inventory')
def api_open_plans(request):
    """Open Production Plans for a supplier — the pool a container draws from."""
    from .models import ProductionPlan
    qs = (ProductionPlan.objects.select_related('group__po__supplier')
          .prefetch_related('group__lines__allocations__shipment'))
    sup = request.GET.get('supplier')
    if sup:
        qs = qs.filter(group__po__supplier_id=sup)
    rows = []
    for p in qs:
        if p.remaining_qty <= 0:
            continue
        rows.append({'id': p.pk, 'pp': p.pp_number, 'category': p.category,
                     'po': p.group.po.po_number, 'po_id': p.group.po_id,
                     'supplier': p.group.po.supplier.name,
                     'fob': float(p.group.fob_rate),
                     'ordered': p.ordered_qty, 'remaining': p.remaining_qty,
                     'lines': [{'po_line_id': l.pk, 'sku': l.sku,
                                'name': l.name, 'remaining': l.remaining_units}
                               for l in p.group.lines.all()
                               if l.remaining_units > 0]})
    rows.sort(key=lambda r: (r['po'], r['pp']))
    return JsonResponse({'rows': rows})


# ── Loading Plan: suggested next-container quantities ───────────────────────

@login_required
@permission_required('can_view_inventory')
def loading_plan(request):
    return render(request, 'inventory_planning/loading_plan.html', {
        'can_edit': request.user.has_perm_flag('can_manage_cogs'),
    })


@login_required
@permission_required('can_view_inventory')
def api_loading_plan(request):
    from .planning import build_loading_plan
    region = request.GET.get('region', 'usa')
    cover = request.GET.get('cover_days')
    plan = build_loading_plan(
        region,
        cover_days=int(cover) if (cover or '').isdigit() else None,
        category=request.GET.get('category', ''),
        tier=request.GET.get('tier', ''))
    return JsonResponse(plan)


@login_required
@permission_required('can_view_inventory')
def loading_plan_export(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from .planning import build_loading_plan
    region = request.GET.get('region', 'usa')
    cover = request.GET.get('cover_days')
    plan = build_loading_plan(
        region, cover_days=int(cover) if (cover or '').isdigit() else None)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Loading Plan'
    hdr = ['Tier', 'Category', 'SKU', 'Name', 'Demand/day', 'Basis',
           'On hand', 'In transit', 'Cover now (days)', 'Target days',
           'LOAD (units)', 'Boxes', 'On open PO', 'From PO', 'New PO required',
           'Factory available', 'Stock-out', 'Ship by']
    ws.append(hdr)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in plan['rows']:
        ws.append([r['sku_type'], r['category'], r['sku'], r['name'],
                   r['demand'], r['basis'], r['onhand'], r['transit'],
                   r['cover_now_days'], r['target_days'], r['need'], r['boxes'],
                   r['on_order'], r['from_po'], r['new_po'], r['factory'],
                   r['stockout_date'] or '', r['ship_by'] or ''])
    for i, w in enumerate([7, 18, 20, 26, 10, 7, 9, 9, 13, 11, 12, 8, 11,
                           9, 15, 15, 12, 12], start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    resp = HttpResponse(content_type='application/vnd.openxmlformats-'
                                     'officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = (f'attachment; '
                                   f'filename="loading_plan_{region}.xlsx"')
    wb.save(resp)
    return resp


# ── FBA Transfers: 3PL/AWD → Amazon FC replenishment movements ──────────────

@login_required
@permission_required('can_view_inventory')
def fba_transfers(request):
    return render(request, 'inventory_planning/fba_transfers.html', {
        'can_edit': request.user.has_perm_flag('can_manage_cogs'),
    })


@login_required
@permission_required('can_view_inventory')
def api_fba_transfers(request):
    from django.db.models import Q
    from .models import FBATransfer
    region = request.GET.get('region', 'usa')
    qs = (FBATransfer.objects.filter(region=region)
          .select_related('source', 'created_by')
          .prefetch_related('lines').order_by('-created_at'))
    show = request.GET.get('show', 'all')
    if show == 'open':
        qs = qs.exclude(status__in=['received', 'cancelled'])
    q = (request.GET.get('q') or '').strip()
    if q:
        qs = qs.filter(Q(fba_shipment_id__icontains=q)
                       | Q(reference__icontains=q)
                       | Q(lines__sku__icontains=q)).distinct()
    rows = []
    for t in qs[:400]:
        rows.append({
            'id': t.pk, 'shipment_id': t.fba_shipment_id,
            'source': t.source.name, 'source_id': t.source_id,
            'carrier': t.carrier, 'reference': t.reference,
            'status': t.status,
            'shipped': t.shipped_date.isoformat() if t.shipped_date else None,
            'received': t.received_date.isoformat() if t.received_date else None,
            'created': t.created_at.strftime('%Y-%m-%d'),
            'by': (getattr(t.created_by, 'full_name', '')
                   or getattr(t.created_by, 'email', '')) if t.created_by else '',
            'units': t.total_units, 'units_received': t.total_received,
            'notes': t.notes,
            'lines': [{'sku': l.sku, 'units': l.units,
                       'received': l.received_units} for l in t.lines.all()],
        })
    # KPIs
    open_q = FBATransfer.objects.filter(region=region, status='shipped')
    in_transit = sum(t.total_units - t.total_received for t in
                     open_q.prefetch_related('lines'))
    kpi = {
        'draft': FBATransfer.objects.filter(region=region, status='draft').count(),
        'in_transit_shipments': open_q.count(),
        'in_transit_units': in_transit,
    }
    return JsonResponse({'rows': rows, 'kpi': kpi, 'region': region})


@login_required
@permission_required('can_view_inventory')
def api_source_stock(request):
    """SKUs + available units at a 3PL/AWD warehouse — for the transfer form."""
    from .models import WarehouseStock
    wid = request.GET.get('warehouse')
    rows = [{'sku': s.sku, 'units': s.units}
            for s in WarehouseStock.objects.filter(warehouse_id=wid)
            .order_by('sku')] if wid else []
    return JsonResponse({'rows': rows})


@login_required
@permission_required('can_manage_cogs')
@require_POST
def api_fba_transfer_save(request):
    """Create or update a DRAFT transfer (header + lines)."""
    from django.db import transaction
    from .models import FBATransfer, FBATransferLine, Warehouse
    try:
        d = json.loads(request.body)
    except ValueError:
        return JsonResponse({'status': 'failed', 'message': 'Bad payload.'},
                            status=400)
    lines = [(str(l.get('sku', '')).strip().upper(), int(l.get('units') or 0))
             for l in d.get('lines', [])]
    lines = [(s, u) for s, u in lines if s and u > 0]
    if not d.get('source') or not lines:
        return JsonResponse({'status': 'failed',
                             'message': 'Source warehouse and at least one '
                                        'SKU + quantity are required.'},
                            status=400)
    with transaction.atomic():
        if d.get('id'):
            t = FBATransfer.objects.select_for_update().get(pk=d['id'])
            if t.status != 'draft':
                return JsonResponse({'status': 'failed',
                                     'message': 'Only draft transfers can be '
                                                'edited.'}, status=400)
        else:
            t = FBATransfer(created_by=request.user)
        t.region = d.get('region') or 'usa'
        t.source = Warehouse.objects.get(pk=d['source'])
        t.fba_shipment_id = (d.get('shipment_id') or '').strip()[:64]
        t.carrier = (d.get('carrier') or '').strip()[:64]
        t.reference = (d.get('reference') or '').strip()[:64]
        t.notes = (d.get('notes') or '').strip()[:256]
        t.save()
        t.lines.all().delete()
        for sku, units in lines:
            FBATransferLine.objects.create(transfer=t, sku=sku, units=units)
    return JsonResponse({'status': 'ok', 'id': t.pk,
                         'message': 'Draft saved.'})


@login_required
@permission_required('can_manage_cogs')
@require_POST
def api_fba_transfer_action(request):
    """ship / receive / cancel a transfer, adjusting source stock."""
    from datetime import date, datetime
    from django.db import transaction
    from django.utils import timezone
    from .models import FBATransfer, WarehouseStock
    try:
        d = json.loads(request.body)
        action = d['action']
        t = FBATransfer.objects.select_for_update
    except (ValueError, KeyError):
        return JsonResponse({'status': 'failed', 'message': 'Bad payload.'},
                            status=400)

    def _d(v):
        try:
            return datetime.strptime(v, '%Y-%m-%d').date() if v else date.today()
        except (TypeError, ValueError):
            return date.today()

    with transaction.atomic():
        try:
            t = FBATransfer.objects.select_for_update().prefetch_related(
                'lines').get(pk=d['id'])
        except FBATransfer.DoesNotExist:
            return JsonResponse({'status': 'failed',
                                 'message': 'Transfer not found.'}, status=404)
        now = timezone.now()

        if action == 'ship':
            if t.status != 'draft':
                return JsonResponse({'status': 'failed',
                                     'message': f'Already {t.status}.'})
            short = []
            for l in t.lines.all():
                ws = WarehouseStock.objects.filter(warehouse=t.source,
                                                   sku=l.sku).first()
                have = ws.units if ws else 0
                if l.units > have:
                    short.append(f'{l.sku} (need {l.units:,}, have {have:,})')
            if short and not d.get('force'):
                return JsonResponse({'status': 'confirm',
                                     'message': 'Shipping more than on hand at '
                                     f'{t.source.name}: ' + '; '.join(short)
                                     + '. Ship anyway?'})
            for l in t.lines.all():          # draw down source stock
                ws = WarehouseStock.objects.filter(warehouse=t.source,
                                                   sku=l.sku).first()
                if ws:
                    ws.units = max(0, ws.units - l.units)
                    ws.as_of = now
                    ws.save(update_fields=['units', 'as_of'])
            t.status = 'shipped'
            t.stock_applied = True
            t.shipped_date = _d(d.get('date'))
            t.save(update_fields=['status', 'stock_applied', 'shipped_date'])
            msg = f'Shipped {t.total_units:,} units to Amazon; {t.source.name} drawn down.'

        elif action == 'receive':
            if t.status != 'shipped':
                return JsonResponse({'status': 'failed',
                                     'message': 'Only shipped transfers can be '
                                                'received.'})
            for l in t.lines.all():          # default full receipt
                l.received_units = l.units
                l.save(update_fields=['received_units'])
            t.status = 'received'
            t.received_date = _d(d.get('date'))
            t.save(update_fields=['status', 'received_date'])
            msg = ('Marked received at FC. Amazon’s FBA sync will report '
                   'these as fulfillable.')

        elif action == 'cancel':
            if t.stock_applied and t.status == 'shipped':
                for l in t.lines.all():      # return the units to source
                    ws = WarehouseStock.objects.filter(warehouse=t.source,
                                                       sku=l.sku).first()
                    if ws:
                        ws.units += l.units
                        ws.as_of = now
                        ws.save(update_fields=['units', 'as_of'])
                t.stock_applied = False
            t.status = 'cancelled'
            t.save(update_fields=['status', 'stock_applied'])
            msg = 'Transfer cancelled' + (' and stock returned to source.'
                                          if t.stock_applied is False else '.')
        else:
            return JsonResponse({'status': 'failed',
                                 'message': 'Unknown action.'}, status=400)

    return JsonResponse({'status': 'ok', 'message': msg})


@login_required
@permission_required('can_manage_cogs')
@require_POST
def api_fba_transfer_delete(request):
    from .models import FBATransfer
    try:
        t = FBATransfer.objects.get(pk=json.loads(request.body)['id'])
    except (ValueError, KeyError, FBATransfer.DoesNotExist):
        return JsonResponse({'status': 'failed', 'message': 'Not found.'},
                            status=404)
    if t.status != 'draft':
        return JsonResponse({'status': 'failed',
                             'message': 'Only drafts can be deleted; cancel '
                                        'shipped transfers instead.'})
    t.delete()
    return JsonResponse({'status': 'ok', 'message': 'Draft deleted.'})


@login_required
@permission_required('can_view_inventory')
def fba_transfer_template(request):
    """Downloadable .xlsx template for bulk FBA-transfer upload."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = 'FBA Transfers'
    headers = ['Amazon Shipment ID', 'Source Warehouse', 'SKU', 'Units',
               'Carrier', 'Reference', 'Ship Date', 'Status']
    ws.append(headers)
    fill = PatternFill('solid', fgColor='232F3E')
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = fill
    # example rows — two SKUs on one shipment, one on another
    ws.append(['FBA15EXAMPLE1', 'Jarrett WH', 'BTH-SHT-NBL-600', 480,
               'UPS', 'BOL-1001', '2026-07-24', 'shipped'])
    ws.append(['FBA15EXAMPLE1', 'Jarrett WH', 'TWL-HND-WHT-6', 240,
               'UPS', 'BOL-1001', '2026-07-24', 'shipped'])
    ws.append(['FBA15EXAMPLE2', 'Maverick WH', 'BTH-MT-WHT-900', 300,
               'LTL', '', '', 'draft'])
    ws.append(['# Rows with the SAME Amazon Shipment ID become one transfer. '
               'Source Warehouse = the 3PL/AWD name or code. Status: '
               '"shipped" draws stock now, "draft" (or blank) just records it.',
               '', '', '', '', '', '', ''])
    for i, w in enumerate([20, 20, 22, 9, 12, 14, 12, 10], start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    resp = HttpResponse(content_type='application/vnd.openxmlformats-'
                                     'officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = ('attachment; '
                                   'filename="fba_transfer_template.xlsx"')
    wb.save(resp)
    return resp


@login_required
@permission_required('can_manage_cogs')
@require_POST
def import_fba_transfers_view(request):
    """Bulk-create transfers from an uploaded workbook. Rows sharing an Amazon
    Shipment ID group into one transfer; status 'shipped' draws stock now."""
    import openpyxl
    from datetime import datetime, date as _date
    from django.db import transaction
    from django.utils import timezone
    from .models import (FBATransfer, FBATransferLine, Warehouse,
                         WarehouseStock)
    from .procurement import _col_map, _find_header, _int, _pick, _txt

    f = request.FILES.get('file')
    region = request.POST.get('region', 'usa')
    if f is None:
        return JsonResponse({'status': 'failed', 'message': 'No file.'},
                            status=400)
    try:
        ws = openpyxl.load_workbook(f, data_only=True).active
        hrow, cells = _find_header(ws, ['sku'])
        if hrow is None:
            raise ValueError('No header row containing "SKU".')
        cm = _col_map(cells)
        c_ship = _pick(cm, 'shipment', 'amazon shipment')
        c_src = _pick(cm, 'source', 'warehouse', 'from')
        c_sku = _pick(cm, 'sku')
        c_qty = _pick(cm, 'units', 'qty', 'quantity')
        c_car = _pick(cm, 'carrier')
        c_ref = _pick(cm, 'reference', 'bol', 'ref')
        c_dat = _pick(cm, 'ship date', 'date')
        c_sta = _pick(cm, 'status')
        if not (c_src and c_qty):
            raise ValueError('Need at least "Source Warehouse" and "Units" '
                             'columns.')

        wh_map = {}
        for w in (Warehouse.objects.filter(region=region)
                  .exclude(kind='fba')):
            wh_map[w.name.strip().lower()] = w
            wh_map[w.code.strip().lower()] = w

        def _pdate(v):
            if hasattr(v, 'date'):
                return v.date()
            if isinstance(v, _date):
                return v
            try:
                return datetime.strptime(str(v).strip(), '%Y-%m-%d').date()
            except (TypeError, ValueError):
                return None

        groups: dict = {}
        res = {'rows': 0, 'transfers': 0, 'shipped': 0, 'units': 0,
               'errors': []}
        for r in range(hrow + 1, ws.max_row + 1):
            sku = _txt(ws.cell(r, c_sku).value) if c_sku else ''
            if not sku or sku.startswith('#'):
                continue
            qty = _int(ws.cell(r, c_qty).value)
            if qty <= 0:
                continue
            src_txt = _txt(ws.cell(r, c_src).value)
            w = wh_map.get(src_txt.lower())
            if w is None:
                res['errors'].append(f'Row {r}: unknown source warehouse '
                                     f'"{src_txt}".')
                continue
            res['rows'] += 1
            ship_id = _txt(ws.cell(r, c_ship).value) if c_ship else ''
            key = (ship_id or f'__row{r}', w.pk)
            g = groups.setdefault(key, {
                'wh': w, 'ship_id': ship_id,
                'carrier': _txt(ws.cell(r, c_car).value) if c_car else '',
                'ref': _txt(ws.cell(r, c_ref).value) if c_ref else '',
                'date': _pdate(ws.cell(r, c_dat).value) if c_dat else None,
                'status': (_txt(ws.cell(r, c_sta).value).lower()
                           if c_sta else ''),
                'lines': {}})
            g['lines'][sku.upper()] = g['lines'].get(sku.upper(), 0) + qty

        with transaction.atomic():
            now = timezone.now()
            for g in groups.values():
                t = FBATransfer.objects.create(
                    region=region, source=g['wh'],
                    fba_shipment_id=g['ship_id'][:64],
                    carrier=g['carrier'][:64], reference=g['ref'][:64],
                    created_by=request.user)
                for sku, units in g['lines'].items():
                    FBATransferLine.objects.create(transfer=t, sku=sku,
                                                   units=units)
                    res['units'] += units
                res['transfers'] += 1
                if g['status'] in ('shipped', 'ship', 'yes', 'sent'):
                    for l in t.lines.all():
                        stk = WarehouseStock.objects.filter(
                            warehouse=t.source, sku=l.sku).first()
                        if stk:
                            stk.units = max(0, stk.units - l.units)
                            stk.as_of = now
                            stk.save(update_fields=['units', 'as_of'])
                    t.status = 'shipped'
                    t.stock_applied = True
                    t.shipped_date = g['date'] or _date.today()
                    t.save(update_fields=['status', 'stock_applied',
                                          'shipped_date'])
                    res['shipped'] += 1
    except Exception as exc:
        return JsonResponse({'status': 'failed',
                             'message': f'{type(exc).__name__}: {exc}'},
                            status=400)
    msg = (f'{res["transfers"]} transfer(s) created ({res["units"]:,} units); '
           f'{res["shipped"]} shipped (stock drawn down).')
    if res['errors']:
        msg += f' {len(res["errors"])} row(s) skipped.'
    return JsonResponse({'status': 'ok', **res, 'message': msg})


@login_required
@permission_required('can_manage_cogs')
@require_POST
def api_container_pickup(request):
    """One-click 'freight forwarder picked it up': set a container to In Transit
    and stamp departure + ETA, without opening the full edit form.
    JSON: {id, departure?, eta_port?, eta_destination?}"""
    from datetime import datetime, timedelta
    from .models import InTransitShipment
    try:
        d = json.loads(request.body)
        sh = InTransitShipment.objects.get(pk=d['id'])
    except (ValueError, KeyError, InTransitShipment.DoesNotExist):
        return JsonResponse({'status': 'failed', 'message': 'Container not '
                             'found.'}, status=404)
    if sh.status in ('received', 'cancelled'):
        return JsonResponse({'status': 'failed',
                             'message': f'Container is already {sh.status}.'})

    def _d(v):
        try:
            return datetime.strptime(v, '%Y-%m-%d').date() if v else None
        except (TypeError, ValueError):
            return None

    dep = _d(d.get('departure'))
    eta_port = _d(d.get('eta_port'))
    eta_dest = _d(d.get('eta_destination')) or (
        eta_port + timedelta(days=10) if eta_port else None)
    if dep:
        sh.departure_date = dep
    if eta_port:
        sh.eta_port = eta_port
    if eta_dest:
        sh.eta_destination = eta_dest
    sh.status = 'in_transit'
    sh.save(update_fields=['departure_date', 'eta_port', 'eta_destination',
                           'status'])
    return JsonResponse({'status': 'ok',
                         'message': f'{sh.container_no or "Container"} marked '
                         f'picked up — now In Transit'
                         + (f', ETA {sh.eta_destination}.'
                            if sh.eta_destination else '.')})


# ── Sourcing view: product type → SKU → which suppliers hold PO balance ──────

@login_required
@permission_required('can_view_inventory')
def sourcing(request):
    return render(request, 'inventory_planning/sourcing.html', {
        'can_edit': request.user.has_perm_flag('can_manage_cogs'),
    })


@login_required
@permission_required('can_view_inventory')
def api_sourcing(request):
    """Product type → SKUs, each with total open PO balance and how many
    suppliers currently hold balance for it."""
    from .models import POLine
    lines = (POLine.objects.select_related('po__supplier', 'group')
             .exclude(po__status='cancelled'))
    cats: dict = {}
    for l in lines:
        cat = l.group.category or 'Uncategorised'
        skus = cats.setdefault(cat, {})
        s = skus.setdefault(l.sku, {
            'sku': l.sku, 'name': l.name, 'ordered': 0, 'allocated': 0,
            'remaining': 0, 'suppliers': set(), 'open_suppliers': set()})
        s['ordered'] += l.ordered_units
        s['allocated'] += l.allocated_units
        s['remaining'] += l.remaining_units
        s['suppliers'].add(l.po.supplier_id)
        if l.remaining_units > 0:
            s['open_suppliers'].add(l.po.supplier_id)

    out = []
    for cat, skus in cats.items():
        srows = [{'sku': s['sku'], 'name': s['name'], 'ordered': s['ordered'],
                  'allocated': s['allocated'], 'remaining': s['remaining'],
                  'suppliers': len(s['suppliers']),
                  'open_suppliers': len(s['open_suppliers'])}
                 for s in skus.values()]
        srows.sort(key=lambda r: (-r['remaining'], r['sku']))
        out.append({'category': cat, 'skus': srows, 'sku_count': len(srows),
                    'ordered': sum(r['ordered'] for r in srows),
                    'remaining': sum(r['remaining'] for r in srows),
                    'open_suppliers': len({sid for s in skus.values()
                                           for sid in s['open_suppliers']})})
    out.sort(key=lambda c: (-c['remaining'], c['category']))
    return JsonResponse({'categories': out})


@login_required
@permission_required('can_view_inventory')
def api_sku_sourcing(request):
    """For one SKU: every supplier that has ever supplied it, with their open
    PO balance. Suppliers WITH balance = allocate from these; suppliers with
    history but 0 open = candidates for a new Production Plan."""
    from .models import POLine
    sku = (request.GET.get('sku') or '').strip()
    lines = (POLine.objects.filter(sku__iexact=sku)
             .select_related('po__supplier', 'group')
             .order_by('po__order_date'))
    sup: dict = {}
    for l in lines:
        s = l.po.supplier
        d = sup.setdefault(s.pk, {
            'id': s.pk, 'supplier': s.name, 'ordered': 0, 'allocated': 0,
            'remaining': 0, 'pos': [], 'fobs': [],
            'lead': (s.production_lead_days + s.sea_lead_days
                     + s.port_to_wh_days),
            'capacity': s.monthly_capacity_units})
        d['ordered'] += l.ordered_units
        d['allocated'] += l.allocated_units
        d['remaining'] += l.remaining_units
        if l.group.fob_rate:
            d['fobs'].append(float(l.group.fob_rate))
        d['pos'].append({
            'po': l.po.po_number, 'status': l.po.status,
            'po_line_id': l.pk,
            'ordered': l.ordered_units, 'allocated': l.allocated_units,
            'remaining': l.remaining_units,
            'reserved': l.reserved_units, 'available': l.available_to_promise,
            'reservations': [{'id': rv.pk, 'region': rv.region,
                              'units': rv.units}
                             for rv in l.reservations.all()],
            'fob': float(l.group.fob_rate),
            'order_date': (l.po.order_date.isoformat()
                           if l.po.order_date else None),
            'ready': (l.expected_ready_date.isoformat()
                      if l.expected_ready_date else None)})
    rows = []
    for d in sup.values():
        fobs = d.pop('fobs')
        d['avg_fob'] = round(sum(fobs) / len(fobs), 2) if fobs else 0
        d['has_open'] = d['remaining'] > 0
        rows.append(d)
    # open-balance suppliers first (most balance), then cheapest candidates
    rows.sort(key=lambda r: (0 if r['has_open'] else 1,
                             -r['remaining'], r['avg_fob'] or 9e9))
    total_open = sum(r['remaining'] for r in rows)
    return JsonResponse({'sku': sku, 'suppliers': rows,
                         'total_open': total_open,
                         'open_supplier_count': sum(1 for r in rows
                                                    if r['has_open'])})


# ── Cash Flow Planner (per region) ──────────────────────────────────────────

@login_required
@permission_required('can_view_inventory')
def cashflow(request):
    return render(request, 'inventory_planning/cashflow.html', {
        'can_edit': request.user.has_perm_flag('can_manage_cogs'),
    })


@login_required
@permission_required('can_view_inventory')
def api_cashflow(request):
    from .cashflow import build_ledger
    from .models import CashFlowPlan
    region = request.GET.get('region', 'usa')
    data = build_ledger(region)
    plan = CashFlowPlan.objects.filter(region=region).first()
    data['opening_set'] = bool(plan)
    data['pay_lead_days'] = plan.pay_lead_days if plan else 0
    return JsonResponse(data)


@login_required
@permission_required('can_manage_cogs')
@require_POST
def api_cashflow_refresh(request):
    from .cashflow import refresh_region
    region = json.loads(request.body or '{}').get('region', 'usa')
    r = refresh_region(region)
    return JsonResponse({'status': 'ok', **r,
                         'message': f'{r["containers"]} container payment(s) and '
                         f'{r["amazon"]} Amazon inflow(s) refreshed'
                         + (f'; {r["skipped_locked"]} locked row(s) kept.'
                            if r['skipped_locked'] else '.')})


@login_required
@permission_required('can_manage_cogs')
@require_POST
def api_cashflow_opening(request):
    from datetime import datetime
    from .models import CashFlowPlan
    d = json.loads(request.body)
    region = d.get('region', 'usa')

    def _dt(v):
        try:
            return datetime.strptime(v, '%Y-%m-%d').date() if v else None
        except (TypeError, ValueError):
            return None

    plan, _ = CashFlowPlan.objects.get_or_create(region=region)
    plan.opening_balance = d.get('opening_balance') or 0
    plan.opening_as_of = _dt(d.get('as_of'))
    if 'pay_lead_days' in d:
        try:
            plan.pay_lead_days = max(0, int(d.get('pay_lead_days') or 0))
        except (TypeError, ValueError):
            pass
    plan.save()
    return JsonResponse({'status': 'ok', 'message': 'Settings saved.'})


@login_required
@permission_required('can_manage_cogs')
@require_POST
def api_cashflow_entry(request):
    """Create / update / delete a cash-flow line. Editing an auto row locks it."""
    from datetime import datetime
    from .models import CashFlowEntry
    d = json.loads(request.body)
    action = d.get('action', 'save')

    if action == 'delete':
        CashFlowEntry.objects.filter(pk=d.get('id')).delete()
        return JsonResponse({'status': 'ok', 'message': 'Line removed.'})

    def _dt(v):
        try:
            return datetime.strptime(v, '%Y-%m-%d').date()
        except (TypeError, ValueError):
            return None

    dt = _dt(d.get('date'))
    if not dt:
        return JsonResponse({'status': 'failed', 'message': 'A valid date is '
                             'required.'}, status=400)
    if d.get('id'):
        e = CashFlowEntry.objects.get(pk=d['id'])
        if e.auto_source:                 # user touched an auto row → lock it
            e.locked = True
    else:
        e = CashFlowEntry(region=d.get('region', 'usa'),
                          created_by=request.user)
    e.date = dt
    e.direction = 'in' if d.get('direction') == 'in' else 'out'
    e.category = d.get('category') or 'other'
    e.description = (d.get('description') or '')[:128]
    e.vendor = (d.get('vendor') or '')[:64]
    e.amount = d.get('amount') or 0
    e.note = (d.get('note') or '')[:128]
    e.save()
    return JsonResponse({'status': 'ok', 'message': 'Line saved.'})


@login_required
@permission_required('can_manage_cogs')
@require_POST
def api_container_freight(request):
    """Set a container's freight/duty cost (feeds its cash-flow payment)."""
    from .models import InTransitShipment
    d = json.loads(request.body)
    try:
        sh = InTransitShipment.objects.get(pk=d['id'])
    except (KeyError, InTransitShipment.DoesNotExist):
        return JsonResponse({'status': 'failed', 'message': 'Not found.'},
                            status=404)
    sh.freight_cost = d.get('freight_cost') or 0
    sh.save(update_fields=['freight_cost'])
    return JsonResponse({'status': 'ok', 'message': 'Freight/duty updated.'})


# ── Goods-Receipt Variance ──────────────────────────────────────────────────

@login_required
@permission_required('can_view_inventory')
def goods_receipt(request):
    return render(request, 'inventory_planning/goods_receipt.html', {
        'can_edit': request.user.has_perm_flag('can_manage_cogs'),
    })


@login_required
@permission_required('can_view_inventory')
def api_goods_receipt(request):
    from .procurement import build_variance
    data = build_variance(
        supplier_id=request.GET.get('supplier') or None,
        po_id=request.GET.get('po') or None,
        region=request.GET.get('region', ''),
        variance_only=request.GET.get('variance_only') == '1')
    return JsonResponse(data)


@login_required
@permission_required('can_manage_cogs')
@require_POST
def api_short_close(request):
    """Realise production shortage by short-closing PO line(s).
    JSON: {line_id}  OR  {bulk:true, supplier?, po?}  — bulk closes every open
    line that has nothing left in transit (allocated == received)."""
    from .models import POLine, PurchaseOrder
    d = json.loads(request.body)

    if d.get('line_id'):
        qs = POLine.objects.filter(pk=d['line_id'])
    elif d.get('bulk'):
        qs = POLine.objects.exclude(status__in=['closed', 'short_closed'])
        if d.get('po'):
            qs = qs.filter(po_id=d['po'])
        if d.get('supplier'):
            qs = qs.filter(po__supplier_id=d['supplier'])
        # only lines where everything shipped has landed (nothing still inbound)
        qs = [l for l in qs.prefetch_related('allocations__shipment')
              if l.in_transit_units == 0]
    else:
        return JsonResponse({'status': 'failed', 'message': 'Nothing to close.'},
                            status=400)

    n = short = 0
    touched_pos = set()
    for l in qs:
        short += l.production_shortage or max(
            l.ordered_units - l.wastage_units - l.allocated_units, 0)
        l.status = 'short_closed'
        l.save(update_fields=['status'])
        touched_pos.add(l.po_id)
        n += 1
    # close any PO whose lines are all closed
    for pid in touched_pos:
        po = PurchaseOrder.objects.prefetch_related('lines').get(pk=pid)
        if po.lines.exists() and all(x.is_closed for x in po.lines.all()):
            po.status = 'closed'
            po.save(update_fields=['status'])
    return JsonResponse({'status': 'ok', 'closed': n, 'shortage': short,
                         'message': f'{n} line(s) short-closed — '
                         f'{short:,} units booked as production shortage.'})


@login_required
@permission_required('can_manage_cogs')
@require_POST
def api_receipt_reason(request):
    """Set the variance reason on a received container line."""
    from .models import InTransitLine
    d = json.loads(request.body)
    try:
        ln = InTransitLine.objects.get(pk=d['line_id'])
    except (KeyError, InTransitLine.DoesNotExist):
        return JsonResponse({'status': 'failed', 'message': 'Line not found.'},
                            status=404)
    ln.variance_reason = d.get('reason', '')[:16]
    ln.save(update_fields=['variance_reason'])
    return JsonResponse({'status': 'ok', 'message': 'Reason saved.'})


@login_required
@permission_required('can_view_inventory')
def goods_receipt_export(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from .procurement import build_variance
    d = build_variance(supplier_id=request.GET.get('supplier') or None,
                       po_id=request.GET.get('po') or None)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Variance'
    hdr = ['PO', 'Supplier', 'Category', 'SKU', 'Ordered', 'Wastage',
           'Allocated', 'Loaded', 'Received', 'Remaining',
           'Production short', 'Transit short', 'Over-receipt', 'Status']
    ws.append(hdr)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in d['lines']:
        ws.append([r['po'], r['supplier'], r['category'], r['sku'],
                   r['ordered'], r['wastage'], r['allocated'], r['loaded'],
                   r['received'], r['remaining'], r['prod_short'],
                   r['transit_short'], r['over'], r['status']])
    resp = HttpResponse(content_type='application/vnd.openxmlformats-'
                                     'officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = ('attachment; '
                                   'filename="goods_receipt_variance.xlsx"')
    wb.save(resp)
    return resp


# ── Planner → Reorder Suggestions (Draft PO) ────────────────────────────────

@login_required
@permission_required('can_view_inventory')
def reorder(request):
    return render(request, 'inventory_planning/reorder.html', {
        'can_edit': request.user.has_perm_flag('can_manage_cogs'),
    })


@login_required
@permission_required('can_view_inventory')
def api_reorder(request):
    from .models import ReorderSuggestion
    qs = (ReorderSuggestion.objects.filter(status='suggested')
          .select_related('supplier'))
    rows = [{'id': s.pk, 'sku': s.sku, 'name': s.name, 'category': s.category,
             'supplier': s.supplier.name if s.supplier else '(no supplier)',
             'supplier_id': s.supplier_id,
             'demand': s.demand_per_day, 'qty': s.recommended_qty,
             'open_po': s.open_po_units, 'on_hand': s.on_hand_units,
             'transit': s.transit_units, 'fob': float(s.fob_rate),
             'value': round(float(s.fob_rate) * s.recommended_qty, 2),
             'ready': (s.target_ready_date.isoformat()
                       if s.target_ready_date else None),
             'regions': s.regions_detail} for s in qs]
    tot = {'count': len(rows), 'units': sum(r['qty'] for r in rows),
           'value': round(sum(r['value'] for r in rows), 2),
           'no_supplier': sum(1 for r in rows if not r['supplier_id'])}
    return JsonResponse({'rows': rows, 'totals': tot})


@login_required
@permission_required('can_manage_cogs')
@require_POST
def api_reorder_regenerate(request):
    from .reorder import build_suggestions
    cover = json.loads(request.body or '{}').get('cover_days')
    r = build_suggestions(cover_days=int(cover) if str(cover).isdigit() else None)
    return JsonResponse({'status': 'ok', **r,
                         'message': f'{r["created"]} suggestion(s), '
                         f'{r["units"]:,} units.'})


@login_required
@permission_required('can_manage_cogs')
@require_POST
def api_reorder_action(request):
    from .models import ReorderSuggestion
    from .reorder import approve_suggestions
    d = json.loads(request.body)
    ids = d.get('ids') or ([d['id']] if d.get('id') else [])
    if d.get('action') == 'dismiss':
        ReorderSuggestion.objects.filter(id__in=ids).update(status='dismissed')
        return JsonResponse({'status': 'ok',
                             'message': f'{len(ids)} dismissed.'})
    if d.get('action') == 'approve':
        r = approve_suggestions(ids, user=request.user)
        msg = (f'{r["pos"]} draft PO(s) created, {r["lines"]} lines.')
        if r['skipped_no_supplier']:
            msg += (f' {r["skipped_no_supplier"]} skipped (no supplier — set '
                    f'one first).')
        return JsonResponse({'status': 'ok', **r, 'message': msg})
    return JsonResponse({'status': 'failed', 'message': 'Unknown action.'},
                        status=400)


# ── Multi-supplier reservation ──────────────────────────────────────────────

@login_required
@permission_required('can_manage_cogs')
@require_POST
def api_reserve(request):
    """Reserve open PO-line balance to a region (or delete a reservation).
    JSON: {po_line_id, region, units} | {delete_id}"""
    from .models import POLine, POLineReservation
    d = json.loads(request.body)
    if d.get('delete_id'):
        POLineReservation.objects.filter(pk=d['delete_id']).delete()
        return JsonResponse({'status': 'ok', 'message': 'Reservation removed.'})
    try:
        line = POLine.objects.get(pk=d['po_line_id'])
    except (KeyError, POLine.DoesNotExist):
        return JsonResponse({'status': 'failed', 'message': 'PO line not '
                             'found.'}, status=404)
    units = int(d.get('units') or 0)
    if units <= 0:
        return JsonResponse({'status': 'failed', 'message': 'Units must be '
                             'positive.'}, status=400)
    if units > line.available_to_promise:
        return JsonResponse({'status': 'failed',
                             'message': f'Only {line.available_to_promise:,} '
                             f'units unreserved on this line.'}, status=400)
    POLineReservation.objects.create(
        po_line=line, region=d.get('region', 'usa'), units=units,
        note=d.get('note', '')[:128], created_by=request.user)
    return JsonResponse({'status': 'ok', 'message': f'{units:,} units reserved '
                         f'to {d.get("region", "usa").upper()}.'})
