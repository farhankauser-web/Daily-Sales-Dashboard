"""
Procurement imports & roll-ups (Phase 2).

PO workbook (ops format, e.g. "PP- AKT PO#12.xlsx"):
  · 'Summary'         → one row per CATEGORY: the commercial line where the
                        FOB rate is agreed  → POLineGroup
  · 'Production Plan' → one row per SKU within a category            → POLine
                        each POLine gets one ProductionPlan (the outstanding
                        manufacturing commitment for that SKU)

Re-importing the same PO number REPLACES its lines (the workbook is ops'
source of truth) but never touches allocations already made against it.
"""
from __future__ import annotations

from datetime import date

from django.db import transaction


def _num(v, default=0):
    try:
        if v in (None, ''):
            return default
        return float(str(v).replace(',', '').strip())
    except (TypeError, ValueError):
        return default


def _int(v, default=0):
    return int(round(_num(v, default)))


def _txt(v):
    return str(v).strip() if v not in (None, '') else ''


def _match_key(v) -> str:
    """Fold a supplier label to something typo-tolerant.

    'J.Sons', 'JSons', 'j sons' and 'JSONS' are all the same factory; the
    punctuation and case in a packing list mean nothing.
    """
    return ''.join(ch for ch in str(v or '').lower() if ch.isalnum())


def supplier_index() -> dict:
    """{folded name or code: Supplier} for resolving a Supplier column.

    Never creates a supplier. A packing list that names one we do not know is
    a typo or a genuinely new factory, and both need a person to look — silently
    minting a supplier would strand the units against a duplicate record.
    """
    from .models import Supplier
    idx = {}
    for s in Supplier.objects.all():
        for label in (s.name, s.code):
            k = _match_key(label)
            if k:
                idx.setdefault(k, s)
    return idx


def _find_header(ws, must_have, limit=12):
    """Row index whose cells contain all `must_have` tokens (lowercased)."""
    for r in range(1, min(ws.max_row, limit) + 1):
        cells = [_txt(ws.cell(r, c).value).lower()
                 for c in range(1, ws.max_column + 1)]
        joined = ' | '.join(cells)
        if all(tok in joined for tok in must_have):
            return r, cells
    return None, []


def _col_map(cells):
    """header-text → 1-based column index"""
    return {c.strip(): i + 1 for i, c in enumerate(cells) if c.strip()}


def _pick(cmap, *names):
    """Column for the first matching header. Exact matches win over prefix
    ones, so 'CBM' picks the total column rather than 'CBM/Box'."""
    for want in names:
        for head, idx in cmap.items():
            if head == want:
                return idx
    for want in names:
        for head, idx in cmap.items():
            if head.startswith(want):
                return idx
    return None


@transaction.atomic
def import_po_workbook(file_obj, supplier_name: str, po_number: str,
                       order_date=None, expected_ready_date=None,
                       payment_terms: str = '', user=None) -> dict:
    """Parse an ops PO workbook into Supplier → PO → Groups → Lines → Plans."""
    import openpyxl

    from .models import (POLine, POLineGroup, ProductionPlan, PurchaseOrder,
                         Supplier)

    wb = openpyxl.load_workbook(file_obj, data_only=True)
    res = {'po_number': po_number, 'supplier': supplier_name,
           'groups': 0, 'lines': 0, 'plans': 0, 'ordered_units': 0,
           'fob_value': 0.0, 'unmatched_categories': [], 'unknown_skus': [],
           'warnings': []}

    # ── supplier ────────────────────────────────────────────────────────────
    code = ''.join(ch for ch in supplier_name.upper() if ch.isalnum())[:32]
    supplier, _ = Supplier.objects.get_or_create(
        code=code, defaults={'name': supplier_name})

    # ── purchase order (re-import replaces its lines) ───────────────────────
    po, created = PurchaseOrder.objects.get_or_create(
        po_number=po_number,
        defaults={'supplier': supplier, 'order_date': order_date or date.today(),
                  'expected_ready_date': expected_ready_date,
                  'payment_terms': payment_terms, 'created_by': user})
    if not created:
        allocated = sum(l.allocated_units for l in po.lines.all())
        if allocated:
            res['warnings'].append(
                f'PO already has {allocated:,} units allocated to containers — '
                f're-import kept existing lines; delete allocations first to '
                f'reshape this PO.')
            return _summarize(po, res)
        po.groups.all().delete()          # cascades lines + plans
        po.supplier = supplier
        if order_date:
            po.order_date = order_date
        po.save()

    # ── Summary sheet → POLineGroup (category + FOB) ────────────────────────
    ws = wb['Summary'] if 'Summary' in wb.sheetnames else wb[wb.sheetnames[0]]
    hrow, cells = _find_header(ws, ['category', 'units'])
    if hrow is None:
        raise ValueError('Summary sheet: could not find a header row with '
                         '"Category" and "Units".')
    cm = _col_map(cells)
    c_ref = _pick(cm, 'po number', 'po no', 'reference')
    c_cat = _pick(cm, 'category')
    c_box = _pick(cm, 'boxes', 'box')
    c_per = _pick(cm, 'per box')
    c_uni = _pick(cm, 'units')
    c_fob = _pick(cm, 'fob')
    c_amt = _pick(cm, 'total amount', 'amount')
    # 'Number of Pcs' was always a duplicate of Units and nothing ever read it,
    # so it is no longer part of the upload. Still tolerated if an older
    # workbook carries it, but Units is the figure that counts.
    c_pcs = _pick(cm, 'number of pcs', 'pcs')

    groups: dict[str, object] = {}
    for r in range(hrow + 1, ws.max_row + 1):
        cat = _txt(ws.cell(r, c_cat).value) if c_cat else ''
        units = _int(ws.cell(r, c_uni).value) if c_uni else 0
        if not cat or cat.lower() in ('total', 'grand total') or units <= 0:
            continue
        g = POLineGroup.objects.create(
            po=po,
            reference=_txt(ws.cell(r, c_ref).value) if c_ref else '',
            category=cat,
            fob_rate=round(_num(ws.cell(r, c_fob).value), 4) if c_fob else 0,
            units_per_box=_int(ws.cell(r, c_per).value) if c_per else 0,
            boxes=_int(ws.cell(r, c_box).value) if c_box else 0,
            ordered_units=units,
            total_amount=round(_num(ws.cell(r, c_amt).value), 2) if c_amt else 0,
            # Falls back to Units — they were always the same number.
            pcs=(_int(ws.cell(r, c_pcs).value) if c_pcs else units),
            sort_order=res['groups'])
        groups[cat.lower()] = g
        res['groups'] += 1
        res['fob_value'] += float(g.total_amount)

    # ── Production Plan sheet → POLine + ProductionPlan (per SKU) ───────────
    pp_name = next((s for s in wb.sheetnames
                    if 'production' in s.lower() or s.lower() == 'pp'), None)
    if pp_name is None:
        raise ValueError('Workbook has no "Production Plan" sheet.')
    ws = wb[pp_name]
    hrow, cells = _find_header(ws, ['sku'])
    if hrow is None:
        raise ValueError('Production Plan sheet: no header row containing "SKU".')
    cm = _col_map(cells)
    # Category is no longer part of the Production Plan sheet. It is looked up
    # from the SKU catalogue, because the SKU already knows its category and
    # typing it a second time only created a way for the two sheets to
    # disagree. Still honoured when an older workbook carries the column, so
    # existing files keep importing.
    c_cat = _pick(cm, 'category')
    c_nam = _pick(cm, 'name', 'description')
    c_sku = _pick(cm, 'sku')
    c_per = _pick(cm, 'per box')
    c_box = _pick(cm, 'boxes', 'box')
    c_uni = _pick(cm, 'units')
    c_cbm = _pick(cm, 'cbm per box')
    c_tcb = _pick(cm, 'total cbm')
    c_dat = _pick(cm, 'pp date')
    c_wst = _pick(cm, 'wasteage', 'wastage')

    # SKU → category, from the catalogue. Case-insensitive because a packing
    # list and the catalogue rarely agree on capitalisation.
    from .models import PlanningSku
    cat_of_sku = {s.upper(): c for s, c in
                  PlanningSku.objects.values_list('sku', 'category') if s}

    seq = 0
    for r in range(hrow + 1, ws.max_row + 1):
        sku = _txt(ws.cell(r, c_sku).value) if c_sku else ''
        units = _int(ws.cell(r, c_uni).value) if c_uni else 0
        if not sku or units <= 0:
            continue
        # Typed value wins when present (old workbooks); otherwise the
        # catalogue answers. A SKU the catalogue has never seen leaves this
        # blank and falls into the Uncategorised group below, which is
        # reported back rather than swallowed.
        cat = (_txt(ws.cell(r, c_cat).value) if c_cat else '') \
            or cat_of_sku.get(sku.upper(), '')
        if not cat and sku not in res['unknown_skus']:
            res['unknown_skus'].append(sku)
        g = groups.get(cat.lower())
        if g is None:                      # SKU with no matching Summary row
            if cat and cat not in res['unmatched_categories']:
                res['unmatched_categories'].append(cat)
            g = POLineGroup.objects.create(
                po=po, category=cat or 'Uncategorised', ordered_units=0,
                sort_order=900 + len(res['unmatched_categories']))
            groups[(cat or 'uncategorised').lower()] = g
            res['groups'] += 1

        pp_date = ws.cell(r, c_dat).value if c_dat else None
        if hasattr(pp_date, 'date'):
            pp_date = pp_date.date()
        elif not isinstance(pp_date, date):
            pp_date = None

        seq += 1
        POLine.objects.create(
            po=po, group=g, sku=sku,
            name=_txt(ws.cell(r, c_nam).value) if c_nam else '',
            units_per_box=_int(ws.cell(r, c_per).value) if c_per else 0,
            boxes=_int(ws.cell(r, c_box).value) if c_box else 0,
            ordered_units=units,
            wastage_units=_int(ws.cell(r, c_wst).value) if c_wst else 0,
            cbm_per_box=_num(ws.cell(r, c_cbm).value) if c_cbm else 0,
            total_cbm=_num(ws.cell(r, c_tcb).value) if c_tcb else 0,
            expected_ready_date=pp_date or expected_ready_date,
            sort_order=seq)
        res['lines'] += 1
        res['ordered_units'] += units

    # ── one Production Plan per PRODUCT (category), numbered PP-1, PP-2 … ────
    for i, g in enumerate(po.groups.order_by('sort_order'), start=1):
        ready = [l.expected_ready_date for l in g.lines.all()
                 if l.expected_ready_date]
        ProductionPlan.objects.create(
            group=g, pp_number=f'PP-{i}',
            expected_ready_date=max(ready) if ready else expected_ready_date)
        res['plans'] += 1

    # ── reconcile SKU rows against the category totals ──────────────────────
    for cat, g in groups.items():
        line_sum = sum(l.ordered_units for l in g.lines.all())
        if g.ordered_units and line_sum != g.ordered_units:
            res['warnings'].append(
                f'{g.category}: Summary says {g.ordered_units:,} units but the '
                f'SKU rows total {line_sum:,} ({line_sum - g.ordered_units:+,}).')

    return _summarize(po, res)


def _summarize(po, res):
    res['po_id'] = po.pk
    res['ordered_units'] = po.ordered_units or res['ordered_units']
    res['fob_value'] = float(po.fob_value) or res['fob_value']
    return res


# ── Wastage upload (SKU + qty) ──────────────────────────────────────────────

@transaction.atomic
def import_wastage(file_obj, supplier_id=None, po_id=None) -> dict:
    """Factory fault report: SKU + qty. Reduces the outstanding balance
    permanently (we don't pay for it, the supplier doesn't remake it).

    Scoped to the chosen supplier/PO; within that, matched FIFO across open
    SKU lines, oldest PO first.
    """
    import openpyxl

    from .models import POLine

    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active
    hrow, cells = _find_header(ws, ['sku'])
    if hrow is None:
        raise ValueError('Wastage file: no header row containing "SKU".')
    cm = _col_map(cells)
    c_sku = _pick(cm, 'sku')
    c_qty = _pick(cm, 'units', 'qty', 'quantity', 'wastage', 'wasteage')
    if not c_qty:
        raise ValueError('Wastage file: no quantity column found.')

    res = {'rows': 0, 'applied': 0, 'units': 0, 'unmatched': []}
    for r in range(hrow + 1, ws.max_row + 1):
        sku = _txt(ws.cell(r, c_sku).value)
        qty = _int(ws.cell(r, c_qty).value)
        if not sku or qty <= 0:
            continue
        res['rows'] += 1
        lines = POLine.objects.filter(sku=sku).select_related('po')
        if supplier_id:
            lines = lines.filter(po__supplier_id=supplier_id)
        if po_id:
            lines = lines.filter(po_id=po_id)
        lines = [l for l in lines.order_by('po__order_date', 'pk')
                 if l.remaining_units > 0]
        left = qty
        for l in lines:
            if left <= 0:
                break
            take = min(left, l.remaining_units)
            l.wastage_units += take
            l.save(update_fields=['wastage_units'])
            left -= take
            res['units'] += take
        if left > 0:
            res['unmatched'].append({'sku': sku, 'units': left})
        else:
            res['applied'] += 1
    return res


# ── Opening balance upload (backlog carried in from before the system) ──────

@transaction.atomic
def import_opening_balance(file_obj, supplier_id, as_of) -> dict:
    """SKU (or Category) + Units, as at a date. Replaces any prior upload for
    that supplier/date, so re-uploading a corrected file is safe."""
    import openpyxl

    from .models import PlanningSku, Supplier, SupplierOpeningBalance

    supplier = Supplier.objects.get(pk=supplier_id)
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active
    hrow, cells = _find_header(ws, ['sku'])
    if hrow is None:
        raise ValueError('Opening balance file: no header row containing "SKU".')
    cm = _col_map(cells)
    c_sku = _pick(cm, 'sku')
    c_cat = _pick(cm, 'category')
    c_qty = _pick(cm, 'units', 'qty', 'quantity', 'opening', 'balance')
    if not c_qty:
        raise ValueError('Opening balance file: no quantity column found.')

    SupplierOpeningBalance.objects.filter(supplier=supplier,
                                          as_of=as_of).delete()
    cat_of = dict(PlanningSku.objects.values_list('sku', 'category'))
    res = {'rows': 0, 'units': 0, 'as_of': str(as_of), 'supplier': supplier.name}
    for r in range(hrow + 1, ws.max_row + 1):
        sku = _txt(ws.cell(r, c_sku).value)
        if not sku or sku.lower() in ('total', 'grand total'):
            continue
        qty = _int(ws.cell(r, c_qty).value)
        cat = _txt(ws.cell(r, c_cat).value) if c_cat else ''
        SupplierOpeningBalance.objects.create(
            supplier=supplier, sku=sku, category=cat or cat_of.get(sku, ''),
            units=qty, as_of=as_of)
        res['rows'] += 1
        res['units'] += qty
    return res


def opening_by_supplier() -> dict:
    """{supplier_id: units} — uploaded opening backlog per supplier."""
    from .models import SupplierOpeningBalance
    out: dict[int, int] = {}
    for b in SupplierOpeningBalance.objects.all():
        out[b.supplier_id] = out.get(b.supplier_id, 0) + b.units
    return out


def opening_by_category(supplier_id) -> dict:
    from .models import SupplierOpeningBalance
    out: dict[str, int] = {}
    for b in SupplierOpeningBalance.objects.filter(supplier_id=supplier_id):
        key = b.category or 'Uncategorised'
        out[key] = out.get(key, 0) + b.units
    return out


# ── Product identifiers: SKU × region → FNSKU/UPC/ASIN ──────────────────────
# The SKU string is near-universal across regions; the FNSKU is what differs,
# and it is the label the factory applies once a container's region is known.

REGION_SHEETS = {'amazon usa': 'usa', 'amazon uk': 'uk', 'amazon uae': 'ae',
                 'ksa': 'sa', 'amazon de': 'de'}


@transaction.atomic
def import_product_identifiers(file_obj) -> dict:
    """Load the 'Infinitee Product UPCs' workbook. Only fills identity fields
    on SKUs we already know; unknown SKUs are created so region validation on
    a packing list can work (name/category come from PRODUCT + COLOR)."""
    import openpyxl

    from .models import PlanningSku

    wb = openpyxl.load_workbook(file_obj, data_only=True)
    res = {'regions': {}, 'updated': 0, 'created': 0}
    for sheet in wb.sheetnames:
        region = REGION_SHEETS.get(sheet.strip().lower())
        if not region:
            continue                      # Walmart / Free UPCs — not a region
        ws = wb[sheet]
        hrow, cells = _find_header(ws, ['sku'])
        if hrow is None:
            continue
        cm = _col_map(cells)
        c_prod = _pick(cm, 'product')
        c_col = _pick(cm, 'color', 'colour')
        c_upc = _pick(cm, 'upc')
        c_sku = _pick(cm, 'sku')
        c_fn = _pick(cm, 'fnsku')
        c_asin = _pick(cm, 'asin')
        n = 0
        for r in range(hrow + 1, ws.max_row + 1):
            sku = _txt(ws.cell(r, c_sku).value)
            if not sku:
                continue
            upc = _txt(ws.cell(r, c_upc).value) if c_upc else ''
            obj, created = PlanningSku.objects.get_or_create(
                sku=sku, region=region,
                defaults={'name': _txt(ws.cell(r, c_col).value) if c_col else '',
                          'category': _txt(ws.cell(r, c_prod).value) if c_prod else ''})
            obj.fnsku = (_txt(ws.cell(r, c_fn).value) if c_fn else '')[:16]
            obj.upc = '' if upc.upper() in ('N/A', 'NA') else upc[:20]
            obj.asin = (_txt(ws.cell(r, c_asin).value) if c_asin else '')[:16]
            obj.save(update_fields=['fnsku', 'upc', 'asin'])
            res['created' if created else 'updated'] += 1
            n += 1
        res['regions'][region] = n
    return res


# ── Container packing list → Container Allocations ──────────────────────────
# The packing list is the operational document that (a) commits units to a
# region (the factory applies that region's FNSKU) and (b) draws them off a
# Production Plan. Two-step: preview → operator confirms → commit.

CONTAINER_CBM = {'20ft': 33.0, '40ft': 67.0, '40hc': 76.0}


def parse_packing_list(file_obj) -> list[dict]:
    """Rows of {name, sku, boxes, per_box, units, cbm, po_number, supplier}.

    `supplier` is optional and per row. One container often loads from two
    factories — 17 of the containers on record do — and a packing list is one
    document, so the supplier belongs on the line rather than on the upload.
    Blank falls back to the supplier chosen on the form, which is what every
    existing single-supplier file does.
    """
    import openpyxl
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active
    hrow, cells = _find_header(ws, ['sku'])
    if hrow is None:
        raise ValueError('Packing list: no header row containing "SKU".')
    cm = _col_map(cells)
    c_nam = _pick(cm, 'name', 'description')
    c_sku = _pick(cm, 'sku')
    c_box = _pick(cm, 'box')
    c_per = _pick(cm, 'p/box', 'per box', 'pcs/box')
    c_uni = _pick(cm, 'units', 'qty', 'quantity')
    c_cbm = _pick(cm, 'total cbm', 'cbm')   # the TOTAL, not CBM/Box
    c_po = _pick(cm, 'po number', 'po no', 'po')
    c_sup = _pick(cm, 'supplier', 'vendor', 'factory')

    out = []
    for r in range(hrow + 1, ws.max_row + 1):
        sku = _txt(ws.cell(r, c_sku).value) if c_sku else ''
        if not sku or sku.lower() in ('total', 'grand total'):
            continue
        boxes = _int(ws.cell(r, c_box).value) if c_box else 0
        per = _int(ws.cell(r, c_per).value) if c_per else 0
        units = _int(ws.cell(r, c_uni).value) if c_uni else 0
        if units <= 0:
            units = boxes * per
        if units <= 0:
            continue
        out.append({
            'name': _txt(ws.cell(r, c_nam).value) if c_nam else '',
            'sku': sku, 'boxes': boxes, 'per_box': per, 'units': units,
            'cbm': round(_num(ws.cell(r, c_cbm).value), 3) if c_cbm else 0,
            'po_number': _txt(ws.cell(r, c_po).value) if c_po else '',
            'supplier': _txt(ws.cell(r, c_sup).value) if c_sup else '',
        })
    return out


def _remaining_for(line, release_container_id=None) -> int:
    """Units still open on a PO line.

    release_container_id gives back whatever that container currently holds,
    for the case where those lines are about to be deleted and re-written
    (a REPLACE re-upload). Without it, re-uploading a corrected packing list
    for a container reads its own existing allocation as competing demand and
    reports a false over-allocation.

    Must NOT be passed when APPENDING to a container: those units are staying,
    so they are genuine competing demand.
    """
    rem = line.remaining_units
    if release_container_id:
        rem += sum(a.units for a in line._live_allocations()
                   if a.shipment_id == int(release_container_id))
    return rem


def _open_lines_for(sku, supplier_id, po_number='', release_container_id=None):
    """That supplier's open PO lines for a SKU, oldest PO first (FIFO)."""
    from .models import POLine
    qs = (POLine.objects.filter(sku__iexact=sku, po__supplier_id=supplier_id)
          .select_related('po', 'group__plan')
          .prefetch_related('allocations__shipment'))
    if po_number:
        qs = qs.filter(po__po_number__iexact=po_number)
    return [l for l in qs.order_by('po__order_date', 'pk')
            if _remaining_for(l, release_container_id) > 0]


def preview_packing_list(rows, supplier_id, region, container_size='',
                         exclude_container_id=None) -> dict:
    """Resolve every packing-list row to Production Plan allocations, and
    surface every problem BEFORE anything is written.

    exclude_container_id: only for a REPLACE re-upload, where that container's
    current lines are about to be deleted — its units are released back into
    the open balance first. Leave it None when appending a second supplier to
    a container, because those units are not going anywhere.
    """
    from .models import PlanningSku, Supplier

    supplier = Supplier.objects.get(pk=supplier_id)
    fnskus = dict(PlanningSku.objects.filter(region=region)
                  .values_list('sku', 'fnsku'))
    fn_ci = {k.upper(): v for k, v in fnskus.items()}
    sup_idx = supplier_index()

    out = {'supplier': supplier.name, 'supplier_id': supplier.pk,
           'region': region, 'rows': [], 'errors': 0, 'warnings': 0,
           'total_units': 0, 'total_cbm': 0.0, 'blocking': False,
           'suppliers': []}

    for r in rows:
        row = dict(r, allocs=[], errors=[], warnings=[], fnsku='')

        # 0. whose goods is this line? The row wins over the form, so one file
        #    can carry a whole container. An unknown name is refused rather
        #    than guessed — a typo must not create a duplicate factory.
        row_sup, named, sup_ok = supplier, _txt(r.get('supplier')), True
        if named:
            hit = sup_idx.get(_match_key(named))
            if hit is None:
                sup_ok = False
                near = sorted(
                    {s.name for s in sup_idx.values()
                     if _match_key(named)[:3] and
                     _match_key(named)[:3] in _match_key(s.name)})
                row['errors'].append(
                    f'Supplier "{named}" is not on record'
                    + (f' — did you mean {", ".join(near[:3])}?' if near
                       else ' — add it under Suppliers first.'))
            else:
                row_sup = hit
        row['supplier'] = row_sup.name if sup_ok else named
        row['supplier_id'] = row_sup.pk if sup_ok else None
        if sup_ok and row_sup.name not in out['suppliers']:
            out['suppliers'].append(row_sup.name)

        sku = r['sku']

        # 1. region identity — the factory must have an FNSKU to label with
        fn = fn_ci.get(sku.upper())
        if fn is None:
            row['errors'].append(
                f'{sku} is not listed for {region.upper()} — no FNSKU to label '
                f'with, so it cannot ship to this region.')
        elif not fn:
            row['errors'].append(f'{sku} has no FNSKU recorded for '
                                 f'{region.upper()}.')
        else:
            row['fnsku'] = fn

        # 2. attribute to a Production Plan
        # An unresolved supplier means we do not know whose goods these are.
        # Falling back to the form's supplier here would attribute them to the
        # wrong factory and draw down the wrong PO — silently, since the row
        # would otherwise look allocated. Leave it unallocated instead.
        lines = (_open_lines_for(sku, row_sup.pk, r.get('po_number', ''),
                                 release_container_id=exclude_container_id)
                 if sup_ok else [])
        if not sup_ok:
            pass                      # the supplier error already says why
        elif not lines:
            row['errors'].append(
                f'No open balance for {sku} on {row_sup.name}'
                + (f' PO {r["po_number"]}' if r.get('po_number') else '')
                + ' — nothing to draw these units from.')
        else:
            left = r['units']
            for l in lines:
                if left <= 0:
                    break
                avail = _remaining_for(l, exclude_container_id)
                take = min(left, avail)
                plan = getattr(l.group, 'plan', None)
                row['allocs'].append({
                    'po_line_id': l.pk, 'po': l.po.po_number,
                    'po_id': l.po_id,
                    'pp': plan.pp_number if plan else '',
                    'category': l.group.category, 'units': take,
                    'line_remaining': avail})
                left -= take
            if left > 0:
                row['errors'].append(
                    f'Over-allocation: {left:,} of {r["units"]:,} units have no '
                    f'open balance left on {row_sup.name}.')
            elif len(row['allocs']) > 1:
                row['warnings'].append(
                    f'Split across {len(row["allocs"])} POs (oldest first).')
        if not r.get('po_number'):
            row['warnings'].append('No PO number on the packing list — '
                                   'matched FIFO; check the PO shown.')

        out['errors'] += len(row['errors'])
        out['warnings'] += len(row['warnings'])
        out['total_units'] += r['units']
        out['total_cbm'] += r.get('cbm') or 0
        out['rows'].append(row)

    out['total_cbm'] = round(out['total_cbm'], 2)
    cap = CONTAINER_CBM.get((container_size or '').lower().replace(' ', ''))
    if cap:
        out['capacity_cbm'] = cap
        out['fill_pct'] = round(100 * out['total_cbm'] / cap)
        if out['total_cbm'] > cap:
            out['capacity_warning'] = (
                f'{out["total_cbm"]} CBM exceeds a {container_size} '
                f'container (~{cap} CBM usable).')
            out['warnings'] += 1
    out['blocking'] = out['errors'] > 0
    return out


@transaction.atomic
def commit_packing_list(container, allocs, supplier_id, region,
                        mode='replace') -> dict:
    """Write the confirmed allocations.

    `container` carries container_no, size, destination warehouse code, dates
    and status.

    mode='replace' (default)
        Re-committing the same container number REPLACES its lines, releasing
        the old balance. Right for re-uploading a corrected packing list.

    mode='append'
        Keeps the existing lines and adds these on top. This is how a
        container loaded from two suppliers is recorded: upload supplier A,
        then append supplier B. Vendor and PO number become the combined
        list, so the container shows both. Container dates, destination and
        status are left alone — they were settled by the first upload, and
        the append form does not ask for them again.

    A container is never owned by a supplier; ownership lives on the line, via
    po_line. That is what makes appending safe: two suppliers' lines coexist
    without either losing its attribution.
    """
    from datetime import datetime

    from .models import (InTransitLine, InTransitShipment, POLine, Supplier,
                         Warehouse)

    supplier = Supplier.objects.get(pk=supplier_id)

    def _date(v):
        try:
            return datetime.strptime(v, '%Y-%m-%d').date() if v else None
        except (TypeError, ValueError):
            return None

    wh = None
    if container.get('warehouse'):
        wh = Warehouse.objects.filter(code=container['warehouse']).first()

    def _merge(existing: str, *new: str) -> str:
        """Comma list, de-duplicated, order preserved, trimmed to the column."""
        seen, out = set(), []
        for part in ([p.strip() for p in (existing or '').split(',')]
                     + [str(n or '').strip() for n in new]):
            if part and part.lower() not in seen:
                seen.add(part.lower())
                out.append(part)
        return ', '.join(out)

    sh = InTransitShipment.objects.filter(
        region=region, container_no=container['container_no']).first()
    new_pos = sorted({a['po'] for a in allocs if a.get('po')})

    # Vendor is DERIVED from the PO lines the units are actually drawn from,
    # never from the form. With a Supplier column a single file can carry two
    # factories, and the derived name is the only one that cannot disagree
    # with the lines. Sorting also ends 'J.Sons + Ghazali' vs
    # 'Ghazali + J.Sons' being two different strings for one arrangement.
    _po_lines = POLine.objects.filter(
        pk__in=[int(a['po_line_id']) for a in allocs if a.get('po_line_id')]
    ).select_related('po__supplier')
    new_vendors = sorted({l.po.supplier.name for l in _po_lines})
    if not new_vendors:
        new_vendors = [supplier.name]

    if mode == 'append':
        if sh is None:
            raise ValueError(
                f'No container {container["container_no"]} to append to — '
                f'create it with the first supplier\'s packing list.')
        if sh.status in ('received', 'cancelled'):
            raise ValueError(
                f'{sh.container_no} is {sh.get_status_display().lower()}; '
                f'nothing more can be loaded onto it.')
        # Both suppliers are on the container now, and both PO sets.
        sh.vendor = _merge(sh.vendor, *new_vendors)[:64]
        sh.po_number = _merge(sh.po_number, *new_pos)[:64]
        if wh is not None:
            sh.destination = wh
        sh.save()
    else:
        if sh:
            sh.lines.all().delete()      # releases the previous allocation
        else:
            sh = InTransitShipment(region=region,
                                   container_no=container['container_no'])
        sh.vendor = ', '.join(new_vendors)[:64]
        sh.destination = wh
        sh.po_number = ', '.join(new_pos)[:64]
        sh.departure_date = _date(container.get('departure_date'))
        sh.eta_port = _date(container.get('eta_port'))
        sh.eta_destination = _date(container.get('eta_destination'))
        sh.status = container.get('status') or 'waiting_pickup'
        sh.notes = (container.get('notes') or '')[:256]
        sh.save()

    # merge duplicate (sku, po_line) pairs so one SKU = one container line
    merged: dict[tuple, dict] = {}
    for a in allocs:
        units = int(a['units'])
        if units <= 0:
            continue
        key = (a['sku'].upper(), int(a['po_line_id']))
        if key in merged:
            merged[key]['units'] += units
        else:
            merged[key] = {'sku': a['sku'], 'po_line_id': int(a['po_line_id']),
                           'units': units, 'fnsku': a.get('fnsku', '')}

    res = {'container_no': sh.container_no, 'shipment_id': sh.pk,
           'mode': mode, 'lines': 0, 'units': 0, 'over_allocated': [],
           'vendor': sh.vendor}
    for m in merged.values():
        line = POLine.objects.filter(pk=m['po_line_id']).first()
        if line is None:
            continue
        # On append the container's existing lines are staying, so they are
        # already counted in remaining_units — no release, unlike replace
        # (whose lines were deleted a few lines above).
        avail = line.remaining_units
        if m['units'] > avail:
            res['over_allocated'].append(
                {'sku': m['sku'], 'requested': m['units'], 'available': avail})
            raise ValueError(
                f'{m["sku"]}: {m["units"]:,} units requested but only '
                f'{avail:,} remain open on {line.po.po_number}.')
        # Same SKU drawn from the SAME PO line is genuinely additive; a
        # different po_line stays its own row so each supplier's attribution
        # survives.
        existing = (sh.lines.filter(sku=m['sku'], po_line=line).first()
                    if mode == 'append' else None)
        if existing:
            existing.units += m['units']
            existing.save(update_fields=['units'])
        else:
            InTransitLine.objects.create(
                shipment=sh, sku=m['sku'], units=m['units'],
                po_line=line, fnsku=(m.get('fnsku') or '')[:16])
        res['lines'] += 1
        res['units'] += m['units']
    return res


# ── Goods-Receipt Variance ──────────────────────────────────────────────────
# Reconciles Ordered → Allocated → Received per PO line, isolating:
#   production shortage  = ordered − wastage − allocated  (only once line closed)
#   transit shortage     = allocated − received  (on received containers)
# Built entirely from existing POLine / allocation data — no legacy pollution,
# since it reads only PO-linked allocations.

def build_variance(supplier_id=None, po_id=None, region: str = '',
                   variance_only: bool = False) -> dict:
    from .models import InTransitShipment, POLine

    lines = (POLine.objects.select_related('po__supplier', 'group')
             .prefetch_related('allocations__shipment')
             .exclude(po__status='cancelled'))
    if supplier_id:
        lines = lines.filter(po__supplier_id=supplier_id)
    if po_id:
        lines = lines.filter(po_id=po_id)

    line_rows, tot = [], {'ordered': 0, 'wastage': 0, 'allocated': 0,
                          'received': 0, 'prod_short': 0, 'transit_short': 0,
                          'over': 0, 'value_short': 0.0}
    for l in lines:
        prod = l.production_shortage
        trans = l.transit_shortage
        over = l.over_receipt
        if variance_only and not (prod or trans or over):
            continue
        fob = float(l.group.fob_rate)
        line_rows.append({
            'id': l.pk, 'sku': l.sku, 'name': l.name,
            'category': l.group.category, 'po': l.po.po_number, 'po_id': l.po_id,
            'supplier': l.po.supplier.name, 'fob': fob,
            'ordered': l.ordered_units, 'wastage': l.wastage_units,
            'allocated': l.allocated_units, 'loaded': l.loaded_units,
            'received': l.received_units, 'remaining': l.remaining_units,
            'prod_short': prod, 'transit_short': trans, 'over': over,
            'status': l.status, 'closed': l.is_closed})
        tot['ordered'] += l.ordered_units
        tot['wastage'] += l.wastage_units
        tot['allocated'] += l.allocated_units
        tot['received'] += l.received_units
        tot['prod_short'] += prod
        tot['transit_short'] += trans
        tot['over'] += over
        tot['value_short'] += (prod + trans) * fob
    line_rows.sort(key=lambda r: (-(r['prod_short'] + r['transit_short']),
                                  r['po'], r['sku']))
    tot['value_short'] = round(tot['value_short'], 2)
    delivered = tot['received']
    tot['fill_rate'] = (round(100 * delivered
                              / (tot['ordered'] - tot['wastage']), 1)
                        if (tot['ordered'] - tot['wastage']) else None)

    # Tab B — received containers, shipped vs received per SKU (PO-linked only)
    containers = []
    cqs = (InTransitShipment.objects.filter(status='received')
           .prefetch_related('lines__po_line__po__supplier'))
    if region:
        cqs = cqs.filter(region=region)
    for sh in cqs:
        rows = []
        for l in sh.lines.all():
            if not l.po_line_id:               # skip legacy / non-PO lines
                continue
            disc = l.received_units - l.units
            if variance_only and disc == 0:
                continue
            rows.append({'sku': l.sku, 'shipped': l.units,
                         'received': l.received_units, 'disc': disc,
                         'reason': l.variance_reason, 'line_id': l.pk})
        if not rows:
            continue
        containers.append({
            'id': sh.pk, 'container_no': sh.container_no or sh.shipment_id,
            'region': sh.region, 'vendor': sh.vendor,
            'received_date': (sh.received_date.isoformat()
                              if sh.received_date else None),
            'shipped': sum(r['shipped'] for r in rows),
            'received': sum(r['received'] for r in rows),
            'disc': sum(r['disc'] for r in rows), 'lines': rows})
    containers.sort(key=lambda c: c['disc'])   # biggest shortfall first

    return {'lines': line_rows, 'containers': containers, 'totals': tot}
