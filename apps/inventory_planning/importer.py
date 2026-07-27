"""
Import ops' "Amazon Required Inventory Status Report" workbook.

Main sheet (per-SKU rows, headers on row 4/5) → PlanningSku master +
3PL/AWD WarehouseStock + factory pipeline.
'Transit' sheet (container-per-COLUMN matrix) → InTransitShipment/Lines.

The workbook is ops' source of truth: re-uploading REPLACES 3PL stock and
the transit list for the region (SKU master rows are upserted).
"""
from __future__ import annotations

import io
from datetime import date, timedelta

from django.db import transaction
from django.utils import timezone

# 3PL / destination names as they appear in the workbook → warehouse code
WAREHOUSE_ALIASES = {
    'aftab':    ('AFTAB',    'Aftab WH',    '3pl'),
    'jarrett':  ('JARRETT',  'Jarrett WH',  '3pl'),
    'jarret':   ('JARRETT',  'Jarrett WH',  '3pl'),
    'american': ('AMERICAN', 'American WH', '3pl'),
    'maverick': ('MAVERICK', 'Maverick WH', '3pl'),
    'awd':      ('AWD-USA',  'Amazon AWD USA', 'awd'),
    'amazon':   ('FBA-USA',  'Amazon FBA USA', 'fba'),
}



def _wh(region: str, alias: str):
    from .models import Warehouse
    key = (alias or '').strip().lower()
    for prefix, (code, name, kind) in WAREHOUSE_ALIASES.items():
        if key.startswith(prefix):
            # region-scope the generic codes
            if code in ('AWD-USA', 'FBA-USA') and region != 'usa':
                code = code.replace('USA', region.upper())
                name = name.replace('USA', region.upper())
            wh, _ = Warehouse.objects.get_or_create(
                code=code, defaults={'name': name, 'region': region,
                                     'kind': kind})
            return wh
    if not key:
        return None
    wh, _ = Warehouse.objects.get_or_create(
        code=key.upper()[:32],
        defaults={'name': (alias or '').strip()[:64], 'region': region,
                  'kind': '3pl'})
    return wh


def _num(v) -> int:
    try:
        return int(float(v or 0))
    except (TypeError, ValueError):
        return 0


def _to_date(v):
    if hasattr(v, 'date') and callable(v.date):   # datetime → date
        return v.date()
    if isinstance(v, date):
        return v
    return None


def import_status_workbook(file_bytes: bytes, region: str = 'usa',
                           user=None) -> dict:
    import openpyxl
    from .models import (InTransitLine, InTransitShipment, PlanningSku,
                         Warehouse, WarehouseStock)

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # ── locate the main per-SKU sheet: 'Product Status' in A4 ──────────────
    main = None
    for ws in wb.worksheets:
        if (ws.cell(4, 1).value or '').strip() == 'Product Status':
            main = ws
            break
    if main is None:
        raise ValueError("Couldn't find the main sheet (expects 'Product "
                         "Status' in cell A4) — is this the ops status report?")

    now = timezone.now()
    res = {'skus': 0, 'stock_rows': 0, 'shipments': 0, 'shipment_units': 0,
           'pds_rows': 0, 'warnings': []}

    # ── 'Inventory' sheet: sales tier (Alpha/Beta/Ceta) + PDS per SKU ───────
    # header row has 'SKU Type' in col C; SKU in F, tier in C, PDS in J.
    tier_pds = {}          # sku → {'tier': str, 'pds': float|None}
    inv = None
    for ws in wb.worksheets:
        if ws.title.strip().lower() == 'inventory':
            inv = ws
            break
    if inv is not None:
        hr = None
        for ri in range(1, 12):
            if str(inv.cell(ri, 3).value or '').strip().lower() == 'sku type':
                hr = ri
                break
        if hr:
            for ri in range(hr + 1, inv.max_row + 1):
                s = str(inv.cell(ri, 6).value or '').strip().upper()
                if not s:
                    continue
                pds_raw = inv.cell(ri, 10).value
                try:
                    pds = float(pds_raw) if pds_raw not in (None, '') else None
                except (TypeError, ValueError):
                    pds = None
                tier_pds[s] = {
                    'tier': str(inv.cell(ri, 3).value or '').strip(),
                    'pds': pds}

    # column layout of the main sheet (1-indexed, fixed per ops format)
    COL = dict(status=1, manager=2, sku_type=3, category=4, name=5, sku=6,
               pk_stock=12, pk_production=13, aftab=14, jarrett=15, awd=16,
               american=17, per_box=26, msq=27)
    WH_COLS = [('aftab', 'aftab'), ('jarrett', 'jarrett'), ('awd', 'awd'),
               ('american', 'american')]

    with transaction.atomic():
        # replace 3PL/AWD imported stock for the region
        wh_objs = {}
        for _, alias in WH_COLS:
            wh_objs[alias] = _wh(region, alias)
        WarehouseStock.objects.filter(
            warehouse__in=[w for w in wh_objs.values() if w],
            source='import').delete()

        seen_skus = set()
        for ri in range(6, main.max_row + 1):
            sku = (main.cell(ri, COL['sku']).value or '')
            sku = str(sku).strip().upper()
            if not sku or sku in seen_skus:
                continue
            seen_skus.add(sku)
            tp = tier_pds.get(sku, {})
            status = str(main.cell(ri, COL['status']).value or '').strip()
            # Drop discontinued SKUs from planning (kept in DB but inactive so
            # they never appear in the planner).
            active = 'discontinu' not in status.lower()
            if not active:
                res['discontinued'] = res.get('discontinued', 0) + 1
            PlanningSku.objects.update_or_create(
                sku=sku, region=region,
                defaults={
                    'name': str(main.cell(ri, COL['name']).value or '')[:128],
                    'category': str(main.cell(ri, COL['category']).value or '')[:64],
                    'sku_type': (tp.get('tier') or '')[:16],   # Alpha/Beta/Ceta
                    'product_status': status[:32],
                    'product_manager': str(main.cell(ri, COL['manager']).value or '')[:64],
                    'units_per_box': _num(main.cell(ri, COL['per_box']).value),
                    'msq': _num(main.cell(ri, COL['msq']).value),
                    'factory_stock': _num(main.cell(ri, COL['pk_stock']).value),
                    'factory_production': _num(main.cell(ri, COL['pk_production']).value),
                    'is_active': active,
                })
            res['skus'] += 1
            # seed PDS from the sheet if the sales team filled it and we don't
            # already have a current entry
            pds = tp.get('pds')
            if pds is not None and pds > 0:
                from .models import DemandInput
                if not DemandInput.objects.filter(sku=sku, region=region,
                                                  note='imported from workbook',
                                                  pds=pds).exists():
                    DemandInput.objects.create(
                        sku=sku, region=region, pds=pds,
                        effective_from=date.today(),
                        note='imported from workbook', entered_by=user)
                    res['pds_rows'] += 1
            for col_key, alias in WH_COLS:
                units = _num(main.cell(ri, COL[col_key]).value)
                if units and wh_objs[alias]:
                    WarehouseStock.objects.update_or_create(
                        warehouse=wh_objs[alias], sku=sku,
                        defaults={'units': units, 'as_of': now,
                                  'source': 'import'})
                    res['stock_rows'] += 1

        # ── Transit sheet: containers are COLUMNS ──────────────────────────
        tr = None
        for ws in wb.worksheets:
            if ws.title.strip().lower() == 'transit':
                tr = ws
                break
        if tr is not None:
            # row labels: 1 Vendor, 2 Destination, 3 Departure, 4 ETA (port),
            # 5 Container No., 6 Shipment ID; SKU rows start where col C says
            # 'SKU' on the row above the first SKU.
            sku_start = None
            for ri in range(5, 12):
                if str(tr.cell(ri, 3).value or '').strip().upper() == 'SKU':
                    sku_start = ri + 1
                    break
            if sku_start is None:
                res['warnings'].append('Transit sheet: SKU header row not found')
            else:
                sku_rows = []
                for ri in range(sku_start, tr.max_row + 1):
                    s = str(tr.cell(ri, 3).value or '').strip().upper()
                    if s:
                        sku_rows.append((ri, s))
                # full replace for the region
                InTransitShipment.objects.filter(region=region).delete()
                today = date.today()
                for ci in range(5, tr.max_column + 1):
                    units_by_sku = {s: _num(tr.cell(ri, ci).value)
                                    for ri, s in sku_rows
                                    if _num(tr.cell(ri, ci).value) > 0}
                    if not units_by_sku:
                        continue
                    vendor = str(tr.cell(1, ci).value or '').strip()
                    dest = str(tr.cell(2, ci).value or '').strip()
                    dep = _to_date(tr.cell(3, ci).value)
                    eta = _to_date(tr.cell(4, ci).value)
                    container = str(tr.cell(5, ci).value or '').strip()
                    shipid = str(tr.cell(6, ci).value or '').strip()
                    # skip summary columns ("Total Transit" etc.) — they'd
                    # double-count every real container
                    joined = f'{vendor} {dest} {container} {shipid}'.lower()
                    if 'total' in joined or 'grand' in joined:
                        continue
                    eta_dest = eta + timedelta(days=10) if eta else None
                    if eta and eta < today - timedelta(days=21):
                        status = 'received'
                    elif eta and eta <= today:
                        status = 'at_port'
                    elif dep and dep <= today:
                        status = 'departed'
                    else:
                        status = 'pending'
                    sh = InTransitShipment.objects.create(
                        region=region, container_no=container[:32],
                        shipment_id=shipid[:64], vendor=vendor[:64],
                        destination=_wh(region, dest),
                        departure_date=dep, eta_port=eta,
                        eta_destination=eta_dest,
                        received_date=eta_dest if status == 'received' else None,
                        status=status)
                    InTransitLine.objects.bulk_create([
                        InTransitLine(shipment=sh, sku=s, units=u)
                        for s, u in units_by_sku.items()])
                    res['shipments'] += 1
                    res['shipment_units'] += sum(units_by_sku.values())
        else:
            res['warnings'].append('No Transit sheet found — transit skipped')

    return res


# ── File 1: SKU + PDS + 3PL inventory (recurring master file) ────────────────
# Non-warehouse header names; every OTHER column is treated as a 3PL warehouse.
_MASTER_KNOWN = {
    'sku', 'name', 'product name', 'category', 'sku type', 'tier', 'status',
    'product status', 'pds', 'potential daily sale', 'in hand pakistan',
    'in hand stock pakistan', 'pakistan', 'in production', 'production',
    'product manager', 'manager', 'units per box', 'msq',
}


def _hmap(ws, header_row=1):
    return {str(ws.cell(header_row, c).value or '').strip().lower(): c
            for c in range(1, ws.max_column + 1)
            if ws.cell(header_row, c).value not in (None, '')}


def import_master_file(file_bytes: bytes, region: str = 'usa', user=None) -> dict:
    """
    Single recurring file: SKU + PDS + 3PL warehouse inventory (+ optional
    Name/Category/SKU Type/Status/Pakistan/Production). Upserts the SKU master,
    sets PDS, and REPLACES imported 3PL stock for the region.
    """
    import openpyxl
    from .models import DemandInput, PlanningSku, WarehouseStock

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    H = _hmap(ws)
    if 'sku' not in H:
        raise ValueError("File needs a 'SKU' column header on row 1.")

    def g(row, *names):
        for nm in names:
            c = H.get(nm)
            if c is not None and len(row) >= c:
                return row[c - 1]
        return None

    wh_cols = {name: col for name, col in H.items() if name not in _MASTER_KNOWN}
    wh_objs = {name: _wh(region, name) for name in wh_cols}

    now = timezone.now()
    res = {'skus': 0, 'pds_rows': 0, 'stock_rows': 0, 'discontinued': 0,
           'warehouses': sorted({w.name for w in wh_objs.values() if w}),
           'warnings': []}

    with transaction.atomic():
        WarehouseStock.objects.filter(
            warehouse__in=[w for w in wh_objs.values() if w],
            source='import').delete()
        for row in ws.iter_rows(min_row=2, values_only=True):
            sku = str(g(row, 'sku') or '').strip().upper()
            if not sku:
                continue
            status = str(g(row, 'status', 'product status') or '').strip()
            active = 'discontinu' not in status.lower()
            if not active:
                res['discontinued'] += 1
            defaults = {'is_active': active}
            nm = g(row, 'name', 'product name')
            if nm is not None:
                defaults['name'] = str(nm)[:128]
            cat = g(row, 'category')
            if cat is not None:
                defaults['category'] = str(cat)[:64]
            tier = g(row, 'sku type', 'tier')
            if tier is not None and str(tier).strip():
                defaults['sku_type'] = str(tier).strip()[:16]
            if status:
                defaults['product_status'] = status[:32]
            pak = g(row, 'in hand pakistan', 'in hand stock pakistan', 'pakistan')
            if pak is not None:
                defaults['factory_stock'] = _num(pak)
            prod = g(row, 'in production', 'production')
            if prod is not None:
                defaults['factory_production'] = _num(prod)
            PlanningSku.objects.update_or_create(
                sku=sku, region=region, defaults=defaults)
            res['skus'] += 1

            pds_raw = g(row, 'pds', 'potential daily sale')
            try:
                pds = float(pds_raw) if pds_raw not in (None, '') else None
            except (TypeError, ValueError):
                pds = None
            if pds is not None and pds >= 0:
                # newest entry wins; skip if identical current value exists
                latest = (DemandInput.objects
                          .filter(sku=sku, region=region)
                          .order_by('-effective_from', '-created_at').first())
                if not latest or latest.pds != pds:
                    DemandInput.objects.create(
                        sku=sku, region=region, pds=pds,
                        effective_from=date.today(),
                        note='master file upload', entered_by=user)
                    res['pds_rows'] += 1

            for name, col in wh_cols.items():
                wh = wh_objs.get(name)
                units = _num(row[col - 1]) if len(row) >= col else 0
                if wh and units:
                    WarehouseStock.objects.update_or_create(
                        warehouse=wh, sku=sku,
                        defaults={'units': units, 'as_of': now,
                                  'source': 'import'})
                    res['stock_rows'] += 1
    return res


# ── File 2: container details → transit inventory ───────────────────────────
def import_containers(file_bytes: bytes, region: str = 'usa', user=None) -> dict:
    """
    Container manifest (long format, one row per SKU-in-container):
      Container No | Vendor | Destination | Departure Date | ETA Port |
      Status | SKU | Units
    Upserts each container by its number (replaces that container's lines);
    other containers are left untouched, so it composes with the Add-Container
    form. Rows for a container marked 'received'/'cancelled' still update it.
    """
    import openpyxl
    from collections import defaultdict
    from .models import InTransitLine, InTransitShipment
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    H = _hmap(ws)

    def col(*names):
        for nm in names:
            if nm in H:
                return H[nm]
        return None

    c_container = col('container no', 'container', 'container number')
    c_sku = col('sku')
    c_units = col('units', 'quantity', 'qty')
    if not (c_container and c_sku and c_units):
        raise ValueError("File needs 'Container No', 'SKU' and 'Units' columns.")
    c_vendor = col('vendor', 'supplier')
    c_dest = col('destination', 'reaching destination', 'warehouse')
    c_dep = col('departure date', 'departure', 'etd')
    c_eta = col('eta port', 'eta', 'reaching date')
    c_status = col('status')

    def cell(row, c):
        return row[c - 1] if c and len(row) >= c else None

    groups = defaultdict(lambda: {'meta': {}, 'lines': {}})
    for row in ws.iter_rows(min_row=2, values_only=True):
        cont = str(cell(row, c_container) or '').strip()
        sku = str(cell(row, c_sku) or '').strip().upper()
        units = _num(cell(row, c_units))
        if not cont or not sku or units <= 0:
            continue
        g = groups[cont]
        g['lines'][sku] = g['lines'].get(sku, 0) + units
        if not g['meta']:
            g['meta'] = {
                'vendor': str(cell(row, c_vendor) or '').strip()[:64],
                'dest': str(cell(row, c_dest) or '').strip(),
                'dep': _to_date(cell(row, c_dep)),
                'eta': _to_date(cell(row, c_eta)),
                'status': str(cell(row, c_status) or '').strip().lower(),
            }

    today = date.today()
    res = {'containers': 0, 'lines': 0, 'units': 0, 'warnings': []}
    with transaction.atomic():
        for cont, g in groups.items():
            m = g['meta']
            eta = m.get('eta')
            eta_dest = eta + timedelta(days=10) if eta else None
            status = m.get('status')
            if status not in ('pending', 'departed', 'at_port', 'received',
                              'cancelled'):
                if eta and eta < today - timedelta(days=21):
                    status = 'received'
                elif eta and eta <= today:
                    status = 'at_port'
                elif m.get('dep') and m['dep'] <= today:
                    status = 'departed'
                else:
                    status = 'pending'
            sh, _ = InTransitShipment.objects.update_or_create(
                region=region, container_no=cont[:32],
                defaults={'vendor': m.get('vendor', ''),
                          'destination': _wh(region, m.get('dest', '')),
                          'departure_date': m.get('dep'),
                          'eta_port': eta, 'eta_destination': eta_dest,
                          'received_date': eta_dest if status == 'received' else None,
                          'status': status})
            sh.lines.all().delete()
            InTransitLine.objects.bulk_create([
                InTransitLine(shipment=sh, sku=s, units=u)
                for s, u in g['lines'].items()])
            res['containers'] += 1
            res['lines'] += len(g['lines'])
            res['units'] += sum(g['lines'].values())
    return res
