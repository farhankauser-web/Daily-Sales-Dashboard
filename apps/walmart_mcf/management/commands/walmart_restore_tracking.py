"""
Restore tracking numbers from a Walmart-orders export (WLM.xlsx) whose
per-order tracking/MCF links were lost from the DB. For every row that carries
a tracking number (Walmart already has it — the "✓"), re-create the MCF link +
a confirmed ShipmentPackage so the order shows its tracking and archives.

    python manage.py walmart_restore_tracking [--file ~/Downloads/WLM.xlsx] [--dry-run]

Columns expected (as exported by the orders page):
    Walmart PO | Status | Order date | Customer | Ship to | SKUs | Units |
    MCF order | Amazon status | Tracking
"""
from __future__ import annotations

import hashlib
import io
import os

from django.core.management.base import BaseCommand
from django.utils import timezone


def _carrier(tracking: str) -> str:
    t = tracking.strip().upper()
    if t.startswith(('92', '93', '94', '95')) or 'USPS' in t:
        return 'USPS'
    if t.startswith('1Z'):
        return 'UPS'
    if len(t) == 12 and t.isdigit():
        return 'FedEx'
    return 'USPS'


class Command(BaseCommand):
    help = 'Restore per-order tracking numbers from a WLM.xlsx export.'

    def add_arguments(self, parser):
        parser.add_argument('--file', default=os.path.expanduser('~/Downloads/WLM.xlsx'))
        parser.add_argument('--dry-run', action='store_true')

    def handle(self, *args, **opts):
        import openpyxl
        from apps.walmart_mcf.models import (AmazonMCFOrder, ShipmentPackage,
                                             WalmartOrder,
                                             WalmartOrderState as S)

        wb = openpyxl.load_workbook(opts['file'], data_only=True)
        ws = wb.active
        # locate columns by header
        hdr = {str(ws.cell(1, c).value or '').strip().lower(): c
               for c in range(1, ws.max_column + 1)}
        c_po = hdr.get('walmart po', 1)
        c_mcf = hdr.get('mcf order') or hdr.get('mcf order id') or 8
        c_trk = hdr.get('tracking', 10)

        now = timezone.now()
        res = {'rows': 0, 'restored': 0, 'already_ok': 0, 'no_tracking': 0,
               'order_not_found': 0, 'mcf_conflict': 0}
        dry = opts['dry_run']

        for ri in range(2, ws.max_row + 1):
            po = ws.cell(ri, c_po).value
            if po in (None, ''):
                continue
            res['rows'] += 1
            po = str(po).strip()
            trk = str(ws.cell(ri, c_trk).value or '').replace('✓', '').strip()
            trk = trk.split()[0] if trk else ''          # drop trailing marks
            if not trk or trk in ('—', '-'):
                res['no_tracking'] += 1
                continue
            order = (WalmartOrder.objects.filter(purchase_order_id=po)
                     .select_related('mcf').first())
            if order is None:
                res['order_not_found'] += 1
                continue
            # already has this tracking uploaded?
            mcf = getattr(order, 'mcf', None)
            if mcf and mcf.packages.filter(tracking_number=trk,
                                           uploaded_to_walmart_at__isnull=False
                                           ).exists():
                res['already_ok'] += 1
                continue
            if dry:
                res['restored'] += 1
                continue

            # ensure MCF link
            if mcf is None:
                mcf_id = str(ws.cell(ri, c_mcf).value or '').strip() \
                    or f'RESTORED-{po}'
                existing = AmazonMCFOrder.objects.filter(
                    fulfillment_order_id=mcf_id).first()
                if existing and existing.order_id != order.id:
                    # id belongs to another order — use a unique restore id
                    mcf_id = f'RESTORED-{po}'
                mcf, _ = AmazonMCFOrder.objects.get_or_create(
                    fulfillment_order_id=mcf_id[:48],
                    defaults={'order': order, 'amazon_status': 'Shipped'})
                if mcf.order_id != order.id:
                    res['mcf_conflict'] += 1
                    continue
            # confirmed package
            h = hashlib.sha1(f'{po}|restore|{trk}'.encode()).hexdigest()
            car = _carrier(trk)
            ShipmentPackage.objects.get_or_create(
                upload_hash=h,
                defaults={'mcf_order': mcf, 'carrier_code': car,
                          'carrier_walmart': car, 'tracking_number': trk,
                          'uploaded_to_walmart_at': now,
                          'upload_error': 'restored from Walmart export'})
            # make sure the order is in a terminal (archivable) state
            if order.status not in (S.TRACKING_UPLOADED, S.COMPLETED):
                WalmartOrder.objects.filter(pk=order.pk).update(
                    status=S.TRACKING_UPLOADED)
            res['restored'] += 1

        self.stdout.write(str(res))
