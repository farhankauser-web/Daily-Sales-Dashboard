"""
Walmart → MCF pages: credentials form (linked from API Configuration)
and the Walmart Orders operations page in the dashboard.
"""
from __future__ import annotations

import json

from django import forms
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
from django.http import JsonResponse
from django.shortcuts import redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST

from apps.core.decorators import permission_required

from .models import (SkuMapping, WalmartAPIConfig, WalmartOrder,
                     WalmartOrderState as S)
from .state import IllegalTransition, transition


# ── Credentials (API Configuration page) ────────────────────────────────────

class WalmartConfigForm(forms.ModelForm):
    class Meta:
        model = WalmartAPIConfig
        fields = ['label', 'client_id', 'client_secret', 'is_active']
        widgets = {
            'label': forms.TextInput(attrs={'class': 'form-control'}),
            'client_id': forms.TextInput(attrs={'class': 'form-control',
                                                'autocomplete': 'off'}),
            'client_secret': forms.PasswordInput(render_value=True,
                                                 attrs={'class': 'form-control',
                                                        'autocomplete': 'off'}),
        }


@login_required
@permission_required('can_configure_api')
def walmart_config(request):
    instance = WalmartAPIConfig.objects.first()
    form = WalmartConfigForm(request.POST or None, instance=instance)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Walmart API credentials saved (encrypted).')
        return redirect('amazon_api:list')
    return render(request, 'walmart_mcf/config_form.html', {'form': form,
                                                            'instance': instance})


@login_required
@permission_required('can_configure_api')
@require_POST
def walmart_config_test(request):
    """Round-trip test: fetch an OAuth token from Walmart."""
    try:
        from .walmart_client import WalmartClient
        WalmartClient()._access_token()
        return JsonResponse({'status': 'ok',
                             'message': 'Token issued — credentials work.'})
    except Exception as exc:
        return JsonResponse({'status': 'failed',
                             'error': f'{type(exc).__name__}: {exc}'},
                            status=200)   # 200 so the page JS can render it


# ── Walmart Orders page ──────────────────────────────────────────────────────

@login_required
@permission_required('can_view_dashboard')
def walmart_orders(request):
    statuses = [c[0] for c in S.choices]
    return render(request, 'walmart_mcf/orders.html', {'statuses': statuses})


# Terminal-success states.
ARCHIVE_STATUSES = [S.TRACKING_UPLOADED, S.COMPLETED]
# An order is ARCHIVED only at a genuinely final state:
#   • tracking-uploaded / completed AND at least one confirmed Walmart package
#     (the pipeline only reaches these states once every SKU has shipped, so a
#      partially-shipped multi-SKU order is never here — it stays SHIPPED); or
#   • CANCELLED — terminal on both Walmart and Amazon, no packages expected.
# Anything else stays in the Active list, whatever its status.
ARCHIVED_Q = (
    (Q(status__in=ARCHIVE_STATUSES)
     & Q(mcf__packages__uploaded_to_walmart_at__isnull=False))
    | Q(status=S.CANCELLED)
)


@login_required
@permission_required('can_view_dashboard')
def api_walmart_orders(request):
    qs = (WalmartOrder.objects
          .select_related('mcf')
          .prefetch_related('items', 'mcf__packages', 'audit_events'))
    view = request.GET.get('view') or 'active'
    if view == 'archive':
        qs = qs.filter(ARCHIVED_Q).distinct()
    elif view == 'active':
        qs = qs.exclude(ARCHIVED_Q)
    status = request.GET.get('status') or ''
    if status:
        qs = qs.filter(status=status)
    q = (request.GET.get('q') or '').strip()
    if q:
        qs = qs.filter(
            Q(purchase_order_id__icontains=q) |
            Q(customer_name__icontains=q) |
            Q(items__walmart_sku__icontains=q) |
            Q(mcf__fulfillment_order_id__icontains=q) |
            Q(mcf__packages__tracking_number__icontains=q)).distinct()

    today = timezone.now().date()
    # archive count = orders with confirmed Walmart tracking upload
    archived_count = (WalmartOrder.objects.filter(ARCHIVED_Q)
                      .distinct().count())
    kpi = WalmartOrder.objects.aggregate(
        imported_today=Count('pk', filter=Q(imported_at__date=today)),
        submitted=Count('pk', filter=Q(status__in=[
            S.MCF_CREATED, S.SHIPPED, S.TRACKING_UPLOADED, S.COMPLETED])),
        pending=Count('pk', filter=Q(status__in=[
            S.NEW, S.VALIDATED, S.PROCESSING, S.HOLD])),
        shipped=Count('pk', filter=Q(status__in=[
            S.SHIPPED, S.TRACKING_UPLOADED, S.COMPLETED])),
        errors=Count('pk', filter=Q(status=S.ERROR)),
    )
    kpi['tracking'] = archived_count

    rows = []
    for o in qs.order_by('-order_date')[:300]:
        mcf = getattr(o, 'mcf', None)
        rows.append({
            'id': o.pk,
            'po': o.purchase_order_id,
            'status': o.status,
            'order_date': o.order_date.strftime('%Y-%m-%d %H:%M'),
            'customer': o.customer_name,
            'city': (o.shipping_address or {}).get('city', ''),
            'state': (o.shipping_address or {}).get('state', ''),
            'method': o.shipping_method,
            'skus': [{'sku': i.walmart_sku, 'qty': i.quantity,
                      'name': i.product_name} for i in o.items.all()],
            'units': sum(i.quantity for i in o.items.all()),
            'mcf_id': mcf.fulfillment_order_id if mcf else '',
            'mcf_status': mcf.amazon_status if mcf else '',
            'packages': [{'tracking': p.tracking_number,
                          'carrier': p.carrier_code,
                          'uploaded': bool(p.uploaded_to_walmart_at)}
                         for p in (mcf.packages.all() if mcf else [])],
            'error': o.error_reason,
            'audit': [{'at': a.created_at.strftime('%m-%d %H:%M'),
                       'from': a.from_state, 'to': a.to_state,
                       'actor': a.actor}
                      for a in o.audit_events.all()],
        })
    mapping_count = SkuMapping.objects.filter(enabled=True).count()
    creds = bool(WalmartAPIConfig.objects.filter(is_active=True)
                 .exclude(client_id='').exists())
    return JsonResponse({'rows': rows, 'kpi': kpi,
                         'mappings_enabled': mapping_count,
                         'credentials_configured': creds})


@login_required
@permission_required('can_manage_cogs')
@require_POST
def api_walmart_run(request):
    """Run one pipeline step now (buttons on the orders page)."""
    from . import pipeline
    from .core import JobAlreadyRunning, job_lock
    try:
        payload = json.loads(request.body)
        action = payload.get('action')
    except ValueError:
        return JsonResponse({'error': 'bad payload'}, status=400)
    order_ids = payload.get('order_ids') or None
    runners = {'import': pipeline.import_orders,
               'submit': (lambda: pipeline.submit_orders(order_ids=order_ids)),
               'status': (lambda: pipeline.check_status(order_ids=order_ids)),
               'tracking': (lambda: pipeline.upload_tracking(order_ids=order_ids)),
               'backfill': pipeline.backfill_manual_tracking,
               'inventory': pipeline.sync_inventory}
    fn = runners.get(action)
    if not fn:
        return JsonResponse({'error': 'unknown action'}, status=400)
    try:
        with job_lock(f'web_{action}'):
            res = fn()
    except JobAlreadyRunning:
        return JsonResponse({'status': 'failed',
                             'error': 'This job is already running.'})
    except Exception as exc:
        return JsonResponse({'status': 'failed',
                             'error': f'{type(exc).__name__}: {exc}'},
                            status=500)
    return JsonResponse({'status': 'ok', 'action': action, 'result': res})


@login_required
@permission_required('can_manage_cogs')
@require_POST
def api_walmart_reprocess(request):
    """Reprocess an ERROR/HOLD/CANCELLED order (→ NEW)."""
    try:
        pk = json.loads(request.body).get('id')
        order = WalmartOrder.objects.get(pk=pk)
    except (ValueError, WalmartOrder.DoesNotExist):
        return JsonResponse({'error': 'order not found'}, status=404)
    try:
        # This project's User is email-based and has no `username` — reading it
        # raised AttributeError, Django returned its HTML 500 page, and the
        # page reported "Session expired" because the response was not JSON.
        # The button had therefore never worked.
        if order.status in (S.ERROR, S.HOLD, S.CANCELLED) and \
                transition(order, S.NEW, f'web:{request.user.email}',
                           {'action': 'reprocess'}, error_reason=''):
            return JsonResponse({'status': 'ok'})
    except IllegalTransition:
        pass
    return JsonResponse({'status': 'failed',
                         'error': f'Order is in {order.status} — only ERROR/'
                                  f'HOLD/CANCELLED can be reprocessed.'})


@login_required
@permission_required('can_view_dashboard')
def walmart_export_xlsx(request):
    """Excel export of Walmart orders, honouring the page's status/q filters."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from django.http import HttpResponse

    qs = (WalmartOrder.objects.select_related('mcf')
          .prefetch_related('items', 'mcf__packages').order_by('-order_date'))
    view = request.GET.get('view') or 'active'
    if view == 'archive':
        qs = qs.filter(ARCHIVED_Q).distinct()
    elif view == 'active':
        qs = qs.exclude(ARCHIVED_Q)
    status = request.GET.get('status') or ''
    if status:
        qs = qs.filter(status=status)
    q = (request.GET.get('q') or '').strip()
    if q:
        qs = qs.filter(
            Q(purchase_order_id__icontains=q) |
            Q(customer_name__icontains=q) |
            Q(items__walmart_sku__icontains=q) |
            Q(mcf__fulfillment_order_id__icontains=q) |
            Q(mcf__packages__tracking_number__icontains=q)).distinct()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Walmart Orders'
    headers = ['Walmart PO', 'Customer Order ID', 'Status', 'Order Date',
               'Customer', 'City', 'State', 'Ship Method', 'SKUs', 'Units',
               'MCF Order ID', 'Amazon Status', 'Carrier(s)', 'Tracking',
               'Tracking Uploaded', 'Error']
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for o in qs[:2000]:
        mcf = getattr(o, 'mcf', None)
        pkgs = list(mcf.packages.all()) if mcf else []
        ws.append([
            o.purchase_order_id, o.customer_order_id, o.status,
            o.order_date.strftime('%Y-%m-%d %H:%M'),
            o.customer_name,
            (o.shipping_address or {}).get('city', ''),
            (o.shipping_address or {}).get('state', ''),
            o.shipping_method,
            '; '.join(f'{i.walmart_sku} x{i.quantity}' for i in o.items.all()),
            sum(i.quantity for i in o.items.all()),
            mcf.fulfillment_order_id if mcf else '',
            mcf.amazon_status if mcf else '',
            '; '.join(p.carrier_code for p in pkgs),
            '; '.join(p.tracking_number for p in pkgs),
            'yes' if pkgs and all(p.uploaded_to_walmart_at for p in pkgs)
            else ('partial' if any(p.uploaded_to_walmart_at for p in pkgs)
                  else ''),
            o.error_reason[:200],
        ])
    for i, w in enumerate([18, 18, 14, 16, 20, 14, 8, 10, 40, 7,
                           20, 12, 14, 26, 10, 30], start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument'
                     '.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="walmart_orders.xlsx"'
    wb.save(resp)
    return resp
